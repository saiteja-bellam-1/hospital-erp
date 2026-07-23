"""Tests for billing Excel/CSV export (filter-aware, hospital-branded)."""
from datetime import datetime, date, time
from io import BytesIO


def _paid_consultation(db_session, seed_data, fee=500.0):
    from app.models.outpatient import Appointment
    ts = datetime.now().timestamp()
    a = Appointment(
        appointment_number=f"XLS-{ts}",
        patient_id=seed_data["patient_id"],
        doctor_id=seed_data["doctor_user_id"],
        appointment_date=datetime.now(),
        appointment_time=time(10, 0),
        consultation_fee=fee,
        registration_fee=0,
        final_amount=fee,
        payment_status="paid",
        payment_method="cash",
    )
    db_session.add(a)
    db_session.commit()
    return a


def _find_header_row(ws, expected="Date"):
    """Locate the data header row after the branding block."""
    for r in range(1, min(ws.max_row or 1, 40) + 1):
        if ws.cell(row=r, column=1).value == expected:
            return r
    return None


class TestBillingExcelExport:

    def test_export_returns_xlsx(self, client, auth_headers, db_session, seed_data):
        _paid_consultation(db_session, seed_data)
        today = date.today().isoformat()
        r = client.get(
            f"/api/hospital/billing/export.xlsx?date_from={today}&date_to={today}",
            headers=auth_headers,
        )
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers.get("content-type", "")
        assert "billing_" in r.headers.get("content-disposition", "")
        assert r.content[:2] == b"PK"  # zip/xlsx magic

        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(r.content))
        assert "Bills" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        bills = wb["Bills"]
        header_row = _find_header_row(bills)
        assert header_row is not None
        assert bills.cell(row=header_row, column=1).value == "Date"
        assert bills.max_row >= header_row + 1

        # Hospital brand should appear above the table
        assert any(
            bills.cell(row=i, column=1).value or bills.cell(row=i, column=2).value
            for i in range(1, header_row)
        )
        # Filter / totals metadata
        labels = {
            bills.cell(row=i, column=1).value
            for i in range(1, header_row)
        }
        assert "Date range" in labels
        assert "Module" in labels
        assert "Total billed" in labels

    def test_export_includes_hospital_name(self, client, auth_headers, db_session, seed_data):
        from app.models.hospital import Hospital
        hospital = db_session.query(Hospital).filter(
            Hospital.id == seed_data["hospital_id"]
        ).first()
        hospital.name = "Branded Test Hospital"
        hospital.address = "42 Care Lane"
        hospital.city = "Hyderabad"
        hospital.state = "TS"
        hospital.postal_code = "500001"
        db_session.commit()

        today = date.today().isoformat()
        r = client.get(
            f"/api/hospital/billing/export.xlsx?date_from={today}&date_to={today}",
            headers=auth_headers,
        )
        assert r.status_code == 200
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(r.content))
        bills = wb["Bills"]
        values = []
        for row in bills.iter_rows(min_row=1, max_row=8, max_col=2, values_only=True):
            values.extend([v for v in row if v])
        assert "Branded Test Hospital" in values
        assert any("42 Care Lane" in str(v) for v in values)

    def test_export_respects_bill_type_filter(self, client, auth_headers, db_session, seed_data):
        _paid_consultation(db_session, seed_data, fee=400)
        today = date.today().isoformat()
        # Lab-only filter should not include the consultation row
        r = client.get(
            f"/api/hospital/billing/export.xlsx?date_from={today}&date_to={today}&bill_type=lab",
            headers=auth_headers,
        )
        assert r.status_code == 200, r.text
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(r.content))
        bills = wb["Bills"]
        header_row = _find_header_row(bills)
        assert header_row is not None
        # Module label in meta
        module_val = None
        for i in range(1, header_row):
            if bills.cell(row=i, column=1).value == "Module":
                module_val = bills.cell(row=i, column=2).value
                break
        assert module_val == "Lab"
        types = [
            bills.cell(row=i, column=2).value
            for i in range(header_row + 1, bills.max_row + 1)
            if bills.cell(row=i, column=2).value
        ]
        assert "consultation" not in types

    def test_export_empty_range_still_xlsx(self, client, auth_headers):
        r = client.get(
            "/api/hospital/billing/export.xlsx?date_from=2000-01-01&date_to=2000-01-02",
            headers=auth_headers,
        )
        assert r.status_code == 200, r.text
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(r.content))
        bills = wb["Bills"]
        header_row = _find_header_row(bills)
        assert header_row is not None
        assert bills.max_row == header_row  # header only, no data rows
        # Date range present in meta
        found_range = False
        for i in range(1, header_row):
            if bills.cell(row=i, column=1).value == "Date range":
                assert "2000-01-01" in str(bills.cell(row=i, column=2).value)
                assert "2000-01-02" in str(bills.cell(row=i, column=2).value)
                found_range = True
        assert found_range

    def test_export_allows_receptionist(self, client, db_session, seed_data):
        from app.utils.auth import create_access_token, get_password_hash
        from app.models.user import User, UserRole
        role = db_session.query(UserRole).filter_by(name="receptionist").first()
        if role is None:
            role = UserRole(name="receptionist", is_system_role=True)
            db_session.add(role)
            db_session.flush()
        user = db_session.query(User).filter_by(username="testreceptionist").first()
        if user is None:
            user = User(
                username="testreceptionist",
                password_hash=get_password_hash("recv123"),
                email="recv@test.com",
                first_name="Front",
                last_name="Desk",
                role_id=role.id,
                hospital_id=seed_data["hospital_id"],
                is_active=True,
            )
            db_session.add(user)
            db_session.commit()
        token = create_access_token(data={"sub": "testreceptionist"})
        hdr = {"Authorization": f"Bearer {token}"}
        r = client.get("/api/hospital/billing/export.xlsx", headers=hdr)
        assert r.status_code == 200

    def test_export_rejects_unprivileged(self, client, db_session, seed_data):
        from app.utils.auth import create_access_token
        token = create_access_token(data={"sub": "testdoctor"})
        hdr = {"Authorization": f"Bearer {token}"}
        r = client.get("/api/hospital/billing/export.xlsx", headers=hdr)
        assert r.status_code == 403


class TestBillingCsvExport:

    def test_export_returns_csv_with_branding(self, client, auth_headers, db_session, seed_data):
        from app.models.hospital import Hospital
        hospital = db_session.query(Hospital).filter(
            Hospital.id == seed_data["hospital_id"]
        ).first()
        hospital.name = "CSV Brand Hospital"
        hospital.address = "9 Export Road"
        db_session.commit()

        _paid_consultation(db_session, seed_data)
        today = date.today().isoformat()
        r = client.get(
            f"/api/hospital/billing/export.csv?date_from={today}&date_to={today}&bill_type=consultation",
            headers=auth_headers,
        )
        assert r.status_code == 200, r.text
        assert "text/csv" in r.headers.get("content-type", "")
        text = r.content.decode("utf-8-sig")
        assert "CSV Brand Hospital" in text
        assert "9 Export Road" in text
        assert "Date range" in text
        assert "Module" in text
        assert "Outpatient" in text
        assert "Total billed" in text
        assert "Date,Type,Reference,Patient" in text.replace('"', '')
