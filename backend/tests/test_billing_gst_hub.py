"""Billing hub sales summary, GST registers, and CA Excel export."""
from datetime import datetime, date, timedelta
from io import BytesIO

import pytest
from openpyxl import load_workbook


def _bill(db_session, seed_data, *, bill_type="admission", total=500.0, tax=0.0,
          subtotal=None, status="paid", item_type=None, item_name="Room charge",
          sgst_pct=0.0, cgst_pct=0.0, igst_pct=0.0, tax_percentage=None,
          hsn_sac=None, tax_category=None):
    from app.models.billing import Bill, BillItem, Payment
    ts = datetime.now().timestamp()
    subtotal = total if subtotal is None else subtotal
    line_tax = tax
    if not line_tax and (sgst_pct or cgst_pct or igst_pct):
        taxable = total / (1 + (sgst_pct + cgst_pct + igst_pct) / 100.0) if (sgst_pct or cgst_pct or igst_pct) else total
        line_tax = round(total - taxable, 2)
        unit = round(taxable, 2)
    else:
        unit = total
        taxable = total - line_tax if line_tax else total
    sgst_amt = round(taxable * (sgst_pct or 0) / 100.0, 2)
    cgst_amt = round(taxable * (cgst_pct or 0) / 100.0, 2)
    igst_amt = round(taxable * (igst_pct or 0) / 100.0, 2)
    bill = Bill(
        bill_number=f"HUB-{bill_type}-{ts}",
        patient_id=seed_data["patient_id"],
        bill_type=bill_type,
        subtotal=subtotal,
        tax_amount=line_tax or tax,
        discount_amount=0,
        total_amount=total,
        status=status,
        bill_date=datetime.now(),
        created_by_id=seed_data["admin_user_id"],
        hospital_id=seed_data["hospital_id"],
    )
    db_session.add(bill)
    db_session.flush()
    db_session.add(BillItem(
        bill_id=bill.id,
        item_type=item_type or bill_type,
        item_name=item_name,
        quantity=1,
        unit_price=unit if (sgst_pct or cgst_pct or igst_pct or tax) else total,
        total_price=total,
        tax_percentage=tax_percentage if tax_percentage is not None else (sgst_pct + cgst_pct + igst_pct),
        hsn_sac=hsn_sac,
        tax_category=tax_category,
        sgst_pct=sgst_pct,
        cgst_pct=cgst_pct,
        igst_pct=igst_pct,
        sgst_amount=sgst_amt,
        cgst_amount=cgst_amt,
        igst_amount=igst_amt,
    ))
    if status == "paid":
        db_session.add(Payment(
            payment_number=f"PAY-HUB-{ts}",
            bill_id=bill.id,
            amount_paid=total,
            payment_method_name="cash",
            payment_date=datetime.now(),
            received_by_id=seed_data["admin_user_id"],
        ))
    db_session.commit()
    return bill


def _pharmacy_sale(db_session, seed_data, *, grand=118.0, tax=18.0, billing_mode="cash_at_pharmacy",
                   gstin=None, sgst=9, cgst=9):
    from app.models.pharmacy import (
        PharmacySale, PharmacySaleItem, Medicine, PharmacyHSN, MedicineCategory,
        PharmacyInventory,
    )
    ts = int(datetime.now().timestamp() * 1000)
    hsn = PharmacyHSN(
        code=f"3004{ts % 10000}",
        sgst_pct=sgst, cgst_pct=cgst, igst_pct=sgst + cgst,
        hospital_id=seed_data["hospital_id"],
        is_active=True,
    )
    db_session.add(hsn)
    db_session.flush()
    cat = db_session.query(MedicineCategory).filter_by(hospital_id=seed_data["hospital_id"]).first()
    if cat is None:
        cat = MedicineCategory(name=f"Cat-{ts}", hospital_id=seed_data["hospital_id"], is_active=True)
        db_session.add(cat)
        db_session.flush()
    med = Medicine(
        medicine_code=f"M{ts}",
        name=f"Med-{ts}",
        unit_price=100.0,
        hospital_id=seed_data["hospital_id"],
        hsn_id=hsn.id,
        category_id=cat.id,
        is_active=True,
    )
    db_session.add(med)
    db_session.flush()
    batch = PharmacyInventory(
        medicine_id=med.id,
        batch_number=f"B{ts}",
        expiry_date=date.today() + timedelta(days=365),
        quantity_in_stock=100,
        cost_price=10,
        selling_price=100,
        hospital_id=seed_data["hospital_id"],
        is_active=True,
    )
    db_session.add(batch)
    db_session.flush()
    sale = PharmacySale(
        sale_number=f"SL-{ts}",
        sale_date=datetime.now(),
        payment_type="cash",
        subtotal=100,
        tax_total=tax,
        grand_total=grand,
        status="completed",
        billing_mode=billing_mode,
        customer_gstin=gstin,
        hospital_id=seed_data["hospital_id"],
        created_by=seed_data["admin_user_id"],
    )
    db_session.add(sale)
    db_session.flush()
    db_session.add(PharmacySaleItem(
        sale_id=sale.id,
        medicine_id=med.id,
        batch_id=batch.id,
        quantity=1,
        rate=100,
        tax_pct=sgst + cgst,
        sgst_pct=sgst,
        cgst_pct=cgst,
        igst_pct=0,
        line_total=grand,
    ))
    db_session.commit()
    return sale, hsn


def _pharmacy_purchase(db_session, seed_data, *, qty=1, rate=100.0, sgst=9, cgst=9, gstin="36CCCCC0000C1Z5"):
    from app.models.pharmacy import (
        PharmacyPurchase, PharmacyPurchaseItem, PharmacySupplier, Medicine, PharmacyHSN, MedicineCategory,
    )
    ts = int(datetime.now().timestamp() * 1000)
    hsn = PharmacyHSN(
        code=f"3004{ts % 10000}",
        sgst_pct=sgst, cgst_pct=cgst, igst_pct=sgst + cgst,
        hospital_id=seed_data["hospital_id"],
        is_active=True,
    )
    db_session.add(hsn)
    db_session.flush()
    cat = db_session.query(MedicineCategory).filter_by(hospital_id=seed_data["hospital_id"]).first()
    if cat is None:
        cat = MedicineCategory(name=f"PCat-{ts}", hospital_id=seed_data["hospital_id"], is_active=True)
        db_session.add(cat)
        db_session.flush()
    med = Medicine(
        medicine_code=f"PM{ts}",
        name=f"PMed-{ts}",
        unit_price=rate,
        hospital_id=seed_data["hospital_id"],
        hsn_id=hsn.id,
        category_id=cat.id,
        is_active=True,
    )
    db_session.add(med)
    db_session.flush()
    supplier = PharmacySupplier(
        name=f"Supplier-{ts}",
        gstin_no=gstin,
        hospital_id=seed_data["hospital_id"],
        is_active=True,
        ledger_type="registered" if gstin else "unregistered",
    )
    db_session.add(supplier)
    db_session.flush()
    taxable = qty * rate
    tax = taxable * (sgst + cgst) / 100.0
    purch = PharmacyPurchase(
        purchase_number=f"GRN-{ts}",
        entry_date=date.today(),
        bill_date=date.today(),
        invoice_number=f"SINV-{ts}",
        supplier_id=supplier.id,
        status="confirmed",
        subtotal=taxable,
        total_tax=tax,
        grand_total=taxable + tax,
        hospital_id=seed_data["hospital_id"],
    )
    db_session.add(purch)
    db_session.flush()
    db_session.add(PharmacyPurchaseItem(
        purchase_id=purch.id,
        medicine_id=med.id,
        batch_number=f"PB{ts}",
        expiry_date=date.today() + timedelta(days=365),
        quantity=qty,
        purchase_rate=rate,
        sgst_pct=sgst,
        cgst_pct=cgst,
        igst_pct=0,
        tax_amount=tax,
        line_total=taxable + tax,
        hsn_id=hsn.id,
    ))
    db_session.commit()
    return purch


def _period():
    today = date.today()
    return {"year": today.year, "month": today.month}


def _set_module_setting(db_session, seed_data, category, key, value):
    from app.models.permissions import HospitalSettings
    row = db_session.query(HospitalSettings).filter_by(
        setting_category=category, setting_key=key,
    ).first()
    if row:
        row.setting_value = value
    else:
        db_session.add(HospitalSettings(
            setting_category=category,
            setting_key=key,
            setting_value=value,
            setting_type="string",
            created_by=seed_data["admin_user_id"],
        ))


class TestBillingHubReports:

    def test_sales_summary_all_equals_module_slices(self, client, auth_headers, db_session, seed_data):
        _bill(db_session, seed_data, bill_type="day_care", total=200, item_type="procedure")
        _bill(db_session, seed_data, bill_type="catch_up", total=150, item_type="misc")
        r = client.get("/api/hospital/billing/reports/sales-summary", headers=auth_headers)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["totals"]["billed"] >= 350
        modules = {m["module"]: m["billed"] for m in body["by_module"]}
        slice_sum = sum(modules.values())
        assert abs(slice_sum - body["totals"]["billed"]) < 0.05

        day = client.get("/api/hospital/billing/reports/sales-summary?module=day_care", headers=auth_headers)
        assert day.status_code == 200
        assert day.json()["totals"]["billed"] >= 200
        for inv in day.json().get("invoices") or []:
            assert inv["module"] == "day_care"

    def test_ip_pharmacy_not_double_counted(self, client, auth_headers, db_session, seed_data):
        from app.models.billing import Bill, BillItem

        before = client.get("/api/hospital/billing/reports/sales-summary", headers=auth_headers)
        assert before.status_code == 200, before.text
        before_mod = {m["module"]: m["billed"] for m in before.json()["by_module"]}

        sale, _ = _pharmacy_sale(db_session, seed_data, grand=118, billing_mode="inpatient_bill")
        ts = datetime.now().timestamp()
        bill = Bill(
            bill_number=f"ADM-HUB-{ts}",
            patient_id=seed_data["patient_id"],
            bill_type="admission",
            subtotal=1118,
            tax_amount=0,
            discount_amount=0,
            total_amount=1118,
            status="paid",
            bill_date=datetime.now(),
            created_by_id=seed_data["admin_user_id"],
            hospital_id=seed_data["hospital_id"],
        )
        db_session.add(bill)
        db_session.flush()
        db_session.add(BillItem(
            bill_id=bill.id, item_type="room_charge", item_name="Room",
            quantity=1, unit_price=1000, total_price=1000,
        ))
        db_session.add(BillItem(
            bill_id=bill.id, item_type="pharmacy", item_name="Rx med",
            quantity=1, unit_price=118, total_price=118,
            source_ref_type="pharmacy_sale_item", source_ref_id=1,
        ))
        db_session.commit()

        r = client.get("/api/hospital/billing/reports/sales-summary", headers=auth_headers)
        assert r.status_code == 200, r.text
        by_mod = {m["module"]: m["billed"] for m in r.json()["by_module"]}
        ip_delta = by_mod.get("inpatient", 0) - before_mod.get("inpatient", 0)
        ph_ip_delta = by_mod.get("pharmacy_ip", 0) - before_mod.get("pharmacy_ip", 0)
        # Inpatient billed excludes pharmacy lines; pharmacy_ip has the POS sale.
        assert ip_delta == pytest.approx(1000.0)
        assert ph_ip_delta >= 118
        # Pharmacy 118 must not sit on both inpatient and pharmacy_ip.
        assert ip_delta + ph_ip_delta == pytest.approx(1000.0 + ph_ip_delta)
        assert ip_delta + ph_ip_delta < 1118 + 118

    def test_gst_outward_hsn_matches_sale_snapshot(self, client, auth_headers, db_session, seed_data):
        _pharmacy_sale(db_session, seed_data, grand=118, tax=18, sgst=9, cgst=9)
        r = client.get("/api/hospital/billing/reports/gst/outward-hsn", headers=auth_headers)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["totals"]["taxable_value"] >= 100
        assert body["totals"]["sgst_amount"] >= 9
        assert body["totals"]["cgst_amount"] >= 9
        assert abs(body["totals"]["igst_amount"]) < 0.05

    def test_b2b_requires_gstin(self, client, auth_headers, db_session, seed_data):
        _pharmacy_sale(db_session, seed_data, gstin="36AAAAA0000A1Z5")
        _pharmacy_sale(db_session, seed_data, gstin=None)
        r = client.get("/api/hospital/billing/reports/gst/b2b-b2c", headers=auth_headers)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["totals"]["b2b_count"] >= 1
        assert all(row["gstin"] for row in body["b2b"])

    def test_purchase_summary_empty_ok(self, client, auth_headers):
        r = client.get("/api/hospital/billing/reports/purchase-summary", headers=auth_headers)
        assert r.status_code == 200, r.text
        assert "totals" in r.json()

    def test_excel_pack_has_module_and_gstr_sheets(self, client, auth_headers, db_session, seed_data):
        _bill(db_session, seed_data, bill_type="day_care", total=250)
        _pharmacy_sale(db_session, seed_data, gstin="29BBBBB0000B1Z5")
        r = client.get("/api/hospital/billing/reports/gst/audit.xlsx", headers=auth_headers)
        assert r.status_code == 200, r.text
        wb = load_workbook(BytesIO(r.content))
        names = set(wb.sheetnames)
        for required in (
            "Cover", "GSTR-1 B2B", "GSTR-1 B2CS", "GSTR-1 HSN", "GSTR-1 CDNR",
            "Exempt", "Purchase register", "GSTR-3B summary", "Sales by module",
            "Legacy tax",
        ):
            assert required in names, names
        assert any(n.startswith("Day Care") or n == "Day Care" for n in names)

    def test_reports_still_forbid_doctor(self, client):
        from app.utils.auth import create_access_token
        token = create_access_token(data={"sub": "testdoctor"})
        hdr = {"Authorization": f"Bearer {token}"}
        r = client.get("/api/hospital/billing/reports/sales-summary", headers=hdr)
        assert r.status_code == 403
        r = client.get("/api/hospital/billing/reports/gst/gstr1", headers=hdr)
        assert r.status_code == 403

    def test_gstr1_excel_has_gstn_sheet_names(self, client, auth_headers, db_session, seed_data):
        _pharmacy_sale(db_session, seed_data, gstin=None)
        r = client.get("/api/hospital/billing/reports/gst/gstr1.xlsx", headers=auth_headers, params=_period())
        assert r.status_code == 200, r.text
        wb = load_workbook(BytesIO(r.content))
        names = set(wb.sheetnames)
        for required in (
            "Cover", "b2b,sez,de", "b2cl", "b2cs", "cdnr", "cdnur",
            "exp", "at", "atadj", "exemp", "hsn(b2b)", "hsn(b2c)", "docs",
        ):
            assert required in names, names

    def test_gstr1_splits_b2b_and_b2cs(self, client, auth_headers, db_session, seed_data):
        _pharmacy_sale(db_session, seed_data, gstin="36AAAAA0000A1Z5")
        _pharmacy_sale(db_session, seed_data, gstin=None)
        r = client.get("/api/hospital/billing/reports/gst/gstr1", headers=auth_headers, params=_period())
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["kind"] == "gstr1"
        assert len(body["b2b"]["rows"]) >= 1
        assert all(row["gstin"] for row in body["b2b"]["rows"])
        assert any(row["rate"] == 18 for row in body["b2cs"]["rows"]) or body["b2cs"]["summary"]["taxable_value"] >= 100

    def test_exempt_inpatient_in_gstr1_table_8_and_3b(self, client, auth_headers, db_session, seed_data):
        _bill(db_session, seed_data, bill_type="admission", total=500, item_type="room_charge")
        r1 = client.get("/api/hospital/billing/reports/gst/gstr1", headers=auth_headers, params=_period())
        assert r1.status_code == 200, r1.text
        intra_unreg = next(
            row for row in r1.json()["exemp"]["rows"]
            if "unregistered" in row["description"].lower() and "Intra" in row["description"]
        )
        assert intra_unreg["nil"] >= 500
        r3 = client.get("/api/hospital/billing/reports/gst/gstr3b", headers=auth_headers, params=_period())
        assert r3.status_code == 200, r3.text
        assert r3.json()["table_3_1"]["c"]["taxable"] >= 500
        assert r3.json()["exempt_value"] >= 500

    def test_gstr3b_ties_to_gstr1_outward_tax(self, client, auth_headers, db_session, seed_data):
        _pharmacy_sale(db_session, seed_data, gstin=None, grand=118, tax=18, sgst=9, cgst=9)
        params = _period()
        r1 = client.get("/api/hospital/billing/reports/gst/gstr1", headers=auth_headers, params=params)
        r3 = client.get("/api/hospital/billing/reports/gst/gstr3b", headers=auth_headers, params=params)
        assert r1.status_code == 200 and r3.status_code == 200
        t1 = r1.json()["totals"]
        t31a = r3.json()["table_3_1"]["a"]
        assert abs(t31a["taxable"] - t1["outward_taxable"]) < 0.05
        assert abs(t31a["cgst"] + t31a["sgst"] + t31a["igst"] - t1["outward_tax"]) < 0.05
        assert r3.json()["check"]["gstr1_3b_taxable_match"] is True
        assert r3.json()["check"]["gstr1_3b_tax_match"] is True

    def test_gstr3b_itc_matches_purchases(self, client, auth_headers, db_session, seed_data):
        _pharmacy_purchase(db_session, seed_data, qty=1, rate=100, sgst=9, cgst=9)
        params = _period()
        r2 = client.get("/api/hospital/billing/reports/gst/gstr2", headers=auth_headers, params=params)
        r3 = client.get("/api/hospital/billing/reports/gst/gstr3b", headers=auth_headers, params=params)
        assert r2.status_code == 200 and r3.status_code == 200, r2.text + r3.text
        itc = r2.json()["itc"]
        a5 = r3.json()["table_4"]["a5_all_other"]
        assert itc["cgst"] >= 9
        assert itc["sgst"] >= 9
        assert abs(a5["cgst"] - itc["cgst"]) < 0.05
        assert abs(a5["sgst"] - itc["sgst"]) < 0.05
        assert r3.json()["check"]["itc_matches_gstr2"] is True

    def test_gstr3b_pdf_and_gstr9_ok(self, client, auth_headers, db_session, seed_data):
        _pharmacy_sale(db_session, seed_data, gstin=None)
        params = _period()
        pdf = client.get("/api/hospital/billing/reports/gst/gstr3b.pdf", headers=auth_headers, params=params)
        assert pdf.status_code == 200, pdf.text
        assert pdf.content[:4] == b"%PDF"
        fy = date.today().year if date.today().month >= 4 else date.today().year - 1
        r9 = client.get("/api/hospital/billing/reports/gst/gstr9", headers=auth_headers, params={"fy_start": fy})
        assert r9.status_code == 200, r9.text
        assert r9.json()["kind"] == "gstr9"
        x9 = client.get("/api/hospital/billing/reports/gst/gstr9.xlsx", headers=auth_headers, params={"fy_start": fy})
        assert x9.status_code == 200, x9.text
        wb = load_workbook(BytesIO(x9.content))
        assert "Table 4 Outward" in wb.sheetnames

    def test_use_hospital_gstin_when_module_has_no_gst(self, client, auth_headers, db_session, seed_data):
        from app.models.hospital import Hospital

        hospital = db_session.query(Hospital).filter_by(id=seed_data["hospital_id"]).first()
        hospital.gstin = "36HOSP0000H1Z5"
        _set_module_setting(db_session, seed_data, "pharmacy_config", "provider_name", "Outsourced Pharmacy")
        _set_module_setting(db_session, seed_data, "pharmacy_config", "gst_number", "")
        _set_module_setting(db_session, seed_data, "pharmacy_config", "use_hospital_gstin", "true")
        db_session.commit()

        ph = client.get(
            "/api/hospital/billing/reports/gst/outward-hsn",
            headers=auth_headers,
            params={"module": "pharmacy"},
        )
        assert ph.status_code == 200, ph.text
        assert ph.json()["gstin"] == "36HOSP0000H1Z5"
        assert ph.json()["gstin_source"] == "hospital"

    def test_in_house_without_flag_uses_hospital_gstin(self, client, auth_headers, db_session, seed_data):
        from app.models.hospital import Hospital

        hospital = db_session.query(Hospital).filter_by(id=seed_data["hospital_id"]).first()
        hospital.gstin = "36HOSP0000H1Z5"
        _set_module_setting(db_session, seed_data, "pharmacy_config", "provider_name", "")
        _set_module_setting(db_session, seed_data, "pharmacy_config", "gst_number", "")
        _set_module_setting(db_session, seed_data, "pharmacy_config", "use_hospital_gstin", "")
        db_session.commit()

        ph = client.get(
            "/api/hospital/billing/reports/gst/outward-hsn",
            headers=auth_headers,
            params={"module": "pharmacy"},
        )
        assert ph.status_code == 200, ph.text
        assert ph.json()["gstin"] == "36HOSP0000H1Z5"
        assert ph.json()["gstin_source"] == "hospital"

    def test_gstr1_module_pharmacy_excludes_inpatient_exempt(self, client, auth_headers, db_session, seed_data):
        _bill(db_session, seed_data, bill_type="admission", total=500, item_type="room_charge")
        _pharmacy_sale(db_session, seed_data, gstin=None, grand=118, tax=18, sgst=9, cgst=9)
        params = {**_period(), "module": "pharmacy"}
        r = client.get("/api/hospital/billing/reports/gst/gstr1", headers=auth_headers, params=params)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["hospital"]["module"] == "pharmacy"
        assert body["totals"]["exempt_value"] < 500
        assert body["totals"]["outward_taxable"] >= 100




    def test_gstr2_pharmacy_ip_does_not_reuse_grn_itc(self, client, auth_headers, db_session, seed_data):
        _pharmacy_purchase(db_session, seed_data, qty=1, rate=100, sgst=9, cgst=9)
        ip = client.get(
            "/api/hospital/billing/reports/gst/gstr2",
            headers=auth_headers,
            params={**_period(), "module": "pharmacy_ip"},
        )
        ph = client.get(
            "/api/hospital/billing/reports/gst/gstr2",
            headers=auth_headers,
            params={**_period(), "module": "pharmacy"},
        )
        assert ip.status_code == 200 and ph.status_code == 200
        assert abs(ip.json()["itc"]["total"]) < 0.05
        assert ph.json()["itc"]["cgst"] >= 9

    def test_invalid_gst_module_treated_as_all(self, client, auth_headers, db_session, seed_data):
        _pharmacy_sale(db_session, seed_data, gstin=None, grand=118, tax=18, sgst=9, cgst=9)
        params = _period()
        good = client.get("/api/hospital/billing/reports/gst/gstr1", headers=auth_headers, params=params)
        bad = client.get(
            "/api/hospital/billing/reports/gst/gstr1",
            headers=auth_headers,
            params={**params, "module": "not_a_real_module"},
        )
        assert good.status_code == 200 and bad.status_code == 200
        assert bad.json()["hospital"]["module"] == "all"
        assert abs(good.json()["totals"]["outward_taxable"] - bad.json()["totals"]["outward_taxable"]) < 0.05

    def test_pharmacy_gst_group_includes_ip_sales(self, client, auth_headers, db_session, seed_data):
        _pharmacy_sale(db_session, seed_data, gstin=None, grand=118, tax=18, sgst=9, cgst=9,
                       billing_mode="inpatient_bill")
        r = client.get(
            "/api/hospital/billing/reports/gst/gstr1",
            headers=auth_headers,
            params={**_period(), "module": "pharmacy"},
        )
        assert r.status_code == 200, r.text
        assert r.json()["totals"]["outward_taxable"] >= 100

    def test_hospital_gstr2_has_no_pharmacy_itc(self, client, auth_headers, db_session, seed_data):
        _pharmacy_purchase(db_session, seed_data, qty=1, rate=100, sgst=9, cgst=9)
        r = client.get(
            "/api/hospital/billing/reports/gst/gstr2",
            headers=auth_headers,
            params={**_period(), "module": "hospital"},
        )
        assert r.status_code == 200, r.text
        assert abs(r.json()["itc"]["total"]) < 0.05
        assert r.json()["inward_note"]

    def test_gst_reports_excel_and_pdf_export(self, client, auth_headers, db_session, seed_data):
        _pharmacy_sale(db_session, seed_data, gstin="36AAAAA0000A1Z5")
        _pharmacy_sale(db_session, seed_data, gstin=None)
        _bill(db_session, seed_data, bill_type="admission", total=500, item_type="room_charge")

        cases = [
            ("/api/hospital/billing/reports/gst/outward-hsn", ["Outward HSN"]),
            ("/api/hospital/billing/reports/gst/inward-hsn", ["Inward HSN"]),
            ("/api/hospital/billing/reports/gst/b2b-b2c", ["B2B invoices", "B2C rate-wise"]),
            ("/api/hospital/billing/reports/gst/exempt", ["Exempt by module (SAC 9993)", "Exempt invoices"]),
        ]
        for path, sheets in cases:
            xlsx = client.get(f"{path}.xlsx", headers=auth_headers)
            assert xlsx.status_code == 200, f"{path}.xlsx {xlsx.text}"
            assert xlsx.headers["content-type"].startswith(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            wb = load_workbook(BytesIO(xlsx.content))
            for name in sheets:
                assert name in wb.sheetnames, wb.sheetnames

            pdf = client.get(f"{path}.pdf", headers=auth_headers)
            assert pdf.status_code == 200, f"{path}.pdf {pdf.text}"
            assert pdf.content[:4] == b"%PDF"

        g3b_xlsx = client.get(
            "/api/hospital/billing/reports/gst/gstr3b.xlsx",
            headers=auth_headers,
            params=_period(),
        )
        assert g3b_xlsx.status_code == 200, g3b_xlsx.text
        g3b_pdf = client.get(
            "/api/hospital/billing/reports/gst/gstr3b.pdf",
            headers=auth_headers,
            params=_period(),
        )
        assert g3b_pdf.status_code == 200, g3b_pdf.text
        assert g3b_pdf.content[:4] == b"%PDF"
