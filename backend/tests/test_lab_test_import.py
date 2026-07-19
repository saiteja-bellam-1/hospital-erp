"""Tests for the lab test bulk-import endpoints.

Covers:
- template (.xlsx) and sample (.csv) downloads
- dry-run preview writes nothing
- real import creates tests, auto-creates categories + sample types
- Parameters sheet rows collapse into per-parameter reference_ranges
- invalid rows are reported while valid rows still import
- duplicate handling (skip vs update)

Reuses the session-scoped seed_data fixture (hospital + super_admin). Test
codes are made unique per-run so they don't collide in the shared session DB.
"""
from __future__ import annotations

import io
import uuid

import openpyxl

from app.models.lab import LabTest, LabTestCategory, LabTestParameter, SampleType


def _uc(prefix: str) -> str:
    return f"{prefix}{uuid.uuid4().hex[:6]}".upper()


def _xlsx_bytes(test_rows, param_rows=None, tests_header=None, params_header=None):
    """Build an in-memory .xlsx with a Tests sheet and optional Parameters sheet."""
    tests_header = tests_header or [
        "test_code", "name", "category", "sample_type", "cost",
        "method", "description", "preparation_instructions",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tests"
    ws.append(tests_header)
    for r in test_rows:
        ws.append(r)
    if param_rows is not None:
        params_header = params_header or [
            "test_code", "section", "parameter_name", "unit", "method", "field_type",
            "ref_min", "ref_max", "gender", "age_min", "age_max", "description",
            "possible_values", "normal_value", "abnormal_values",
            "critical_low", "critical_high",
        ]
        ws2 = wb.create_sheet("Parameters")
        ws2.append(params_header)
        for r in param_rows:
            ws2.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _upload(client, auth_headers, content, filename, *, dry_run=False, on_duplicate="skip"):
    media = ("text/csv" if filename.endswith(".csv")
             else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return client.post(
        "/api/lab/tests/import",
        headers=auth_headers,
        files={"file": (filename, io.BytesIO(content), media)},
        data={"dry_run": str(dry_run).lower(), "on_duplicate": on_duplicate},
    )


# --------------------------------------------------------------------------
# Template / sample downloads
# --------------------------------------------------------------------------

def test_import_template_downloads_xlsx(client, auth_headers):
    resp = client.get("/api/lab/tests/import/template", headers=auth_headers)
    assert resp.status_code == 200, resp.text
    assert "spreadsheet" in resp.headers.get("content-type", "")
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Tests" in wb.sheetnames
    assert "Parameters" in wb.sheetnames


def test_import_sample_csv_downloads(client, auth_headers):
    resp = client.get("/api/lab/tests/import/sample-csv", headers=auth_headers)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith("text/csv")
    assert "test_code,name,category" in resp.text


# --------------------------------------------------------------------------
# Import — dry run
# --------------------------------------------------------------------------

def test_dry_run_previews_without_writing(client, auth_headers, db_session, seed_data):
    code = _uc("DRY")
    cat = f"DryCat-{uuid.uuid4().hex[:5]}"
    content = _xlsx_bytes([[code, "Dry Run Test", cat, "Blood", 250, "", "", ""]])

    resp = _upload(client, auth_headers, content, "tests.xlsx", dry_run=True)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["dry_run"] is True
    assert body["created"] == 1
    assert cat in body["categories_created"]
    assert body["preview"][0]["status"] == "new"

    # Nothing persisted
    assert db_session.query(LabTest).filter_by(test_code=code).first() is None
    assert db_session.query(LabTestCategory).filter_by(name=cat).first() is None


# --------------------------------------------------------------------------
# Import — real, with auto-created category/sample type + parameters
# --------------------------------------------------------------------------

def test_import_creates_test_category_sampletype_and_params(client, auth_headers, db_session, seed_data):
    code = _uc("CBC")
    cat = f"Hematology-{uuid.uuid4().hex[:5]}"
    st = f"Blood-EDTA-{uuid.uuid4().hex[:5]}"
    tests = [[code, "Complete Blood Count", cat, st, 300, "Analyzer", "", ""]]
    params = [
        [code, "", "Hemoglobin", "g/dL", "", "numeric", 13, 17, "male", "", "", "", "", "", "", "", ""],
        [code, "", "Hemoglobin", "g/dL", "", "numeric", 12, 15, "female", "", "", "", "", "", "", "", ""],
        [code, "", "WBC", "10^3/uL", "", "numeric", 4, 11, "common", "", "", "", "", "", "", "", ""],
    ]
    content = _xlsx_bytes(tests, params)

    resp = _upload(client, auth_headers, content, "tests.xlsx", dry_run=False)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["created"] == 1
    assert cat in body["categories_created"]
    assert st in body["sample_types_created"]

    test = db_session.query(LabTest).filter_by(test_code=code).one()
    assert test.name == "Complete Blood Count"
    assert test.cost == 300
    assert db_session.query(LabTestCategory).filter_by(id=test.category_id).one().name == cat
    assert db_session.query(SampleType).filter_by(name=st).first() is not None

    params_db = (
        db_session.query(LabTestParameter)
        .filter_by(test_id=test.id)
        .order_by(LabTestParameter.display_order)
        .all()
    )
    # Two params: Hemoglobin (2 ranges collapsed) + WBC (1 range)
    assert [p.parameter_name for p in params_db] == ["Hemoglobin", "WBC"]
    hb = params_db[0]
    assert len(hb.reference_ranges) == 2
    assert {r["gender"] for r in hb.reference_ranges} == {"male", "female"}


def test_import_reuses_existing_category(client, auth_headers, db_session, seed_data):
    """A category that already exists should NOT be reported as created."""
    cat_name = f"ExistingCat-{uuid.uuid4().hex[:5]}"
    db_session.add(LabTestCategory(name=cat_name, hospital_id=seed_data["hospital_id"]))
    db_session.commit()

    code = _uc("REU")
    content = _xlsx_bytes([[code, "Reuse Cat Test", cat_name, "", 100, "", "", ""]])
    resp = _upload(client, auth_headers, content, "tests.xlsx", dry_run=False)
    body = resp.json()
    assert body["created"] == 1
    assert cat_name not in body["categories_created"]


# --------------------------------------------------------------------------
# Import — CSV (tests only)
# --------------------------------------------------------------------------

def test_import_csv_tests_only(client, auth_headers, db_session, seed_data):
    code = _uc("TSH")
    cat = f"Endocrine-{uuid.uuid4().hex[:5]}"
    csv = (
        "test_code,name,category,sample_type,cost\n"
        f"{code},Thyroid Stimulating Hormone,{cat},Blood,400\n"
    )
    resp = _upload(client, auth_headers, csv.encode("utf-8"), "tests.csv", dry_run=False)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["created"] == 1
    assert db_session.query(LabTest).filter_by(test_code=code).first() is not None


# --------------------------------------------------------------------------
# Import — error reporting (valid rows import, bad rows reported)
# --------------------------------------------------------------------------

def test_invalid_rows_reported_valid_rows_import(client, auth_headers, db_session, seed_data):
    good = _uc("GOOD")
    bad_cost = _uc("BADC")
    no_cat = _uc("NOCAT")
    cat = f"MixCat-{uuid.uuid4().hex[:5]}"
    tests = [
        [good, "Good Test", cat, "Blood", 200, "", "", ""],
        [bad_cost, "Bad Cost Test", cat, "Blood", "not-a-number", "", "", ""],
        [no_cat, "No Category Test", "", "Blood", 100, "", "", ""],
    ]
    content = _xlsx_bytes(tests)
    resp = _upload(client, auth_headers, content, "tests.xlsx", dry_run=False)
    assert resp.status_code == 200, resp.text
    body = resp.json()

    assert body["created"] == 1
    assert body["error_count"] == 2
    reported_rows = {(e["sheet"], e["row"]) for e in body["errors"]}
    # bad_cost is spreadsheet row 3, no_cat is row 4 (header is row 1)
    assert ("Tests", 3) in reported_rows
    assert ("Tests", 4) in reported_rows

    assert db_session.query(LabTest).filter_by(test_code=good).first() is not None
    assert db_session.query(LabTest).filter_by(test_code=bad_cost).first() is None
    assert db_session.query(LabTest).filter_by(test_code=no_cat).first() is None


# --------------------------------------------------------------------------
# Import — duplicate handling
# --------------------------------------------------------------------------

def test_duplicate_skip_then_update(client, auth_headers, db_session, seed_data):
    code = _uc("DUP")
    cat = f"DupCat-{uuid.uuid4().hex[:5]}"

    # First import creates it
    first = _upload(
        client, auth_headers,
        _xlsx_bytes([[code, "Original Name", cat, "Blood", 500, "", "", ""]]),
        "tests.xlsx", dry_run=False,
    )
    assert first.json()["created"] == 1

    # Re-import with skip -> skipped, unchanged
    skip = _upload(
        client, auth_headers,
        _xlsx_bytes([[code, "Changed Name", cat, "Blood", 999, "", "", ""]]),
        "tests.xlsx", dry_run=False, on_duplicate="skip",
    )
    sbody = skip.json()
    assert sbody["skipped"] == 1
    assert sbody["created"] == 0
    test = db_session.query(LabTest).filter_by(test_code=code).one()
    db_session.refresh(test)
    assert test.name == "Original Name"
    assert test.cost == 500

    # Re-import with update -> updated
    upd = _upload(
        client, auth_headers,
        _xlsx_bytes([[code, "Updated Name", cat, "Blood", 750, "", "", ""]]),
        "tests.xlsx", dry_run=False, on_duplicate="update",
    )
    ubody = upd.json()
    assert ubody["updated"] == 1
    db_session.expire_all()
    test = db_session.query(LabTest).filter_by(test_code=code).one()
    assert test.name == "Updated Name"
    assert test.cost == 750


# --------------------------------------------------------------------------
# Import — bad file type
# --------------------------------------------------------------------------

def test_import_rejects_unsupported_file_type(client, auth_headers):
    resp = _upload(client, auth_headers, b"whatever", "tests.txt", dry_run=True)
    assert resp.status_code == 400
