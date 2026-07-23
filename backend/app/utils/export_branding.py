"""Hospital branding helpers for spreadsheet exports (Excel / CSV)."""
from __future__ import annotations

import os
from typing import Any, Optional

from sqlalchemy.orm import Session

from app.models.hospital import Hospital
from app.utils.paths import get_uploads_dir

BILL_TYPE_LABELS = {
    None: "All Bills",
    "": "All Bills",
    "all": "All Bills",
    "consultation": "Outpatient",
    "lab": "Lab",
    "pharmacy": "Pharmacy",
    "admission": "Inpatient",
}


def hospital_brand_dict(db: Session, hospital_id: Optional[int] = None) -> dict:
    """Load hospital fields used in export headers."""
    hospital = None
    if hospital_id:
        hospital = db.query(Hospital).filter(Hospital.id == hospital_id).first()
    if not hospital:
        hospital = db.query(Hospital).first()
    if not hospital:
        return {
            "name": "Hospital",
            "address": "",
            "city": "",
            "state": "",
            "postal_code": "",
            "phone": "",
            "email": "",
            "logo_url": "",
        }
    return {
        "name": hospital.name or "Hospital",
        "address": hospital.address or "",
        "city": hospital.city or "",
        "state": hospital.state or "",
        "postal_code": hospital.postal_code or "",
        "phone": hospital.phone or "",
        "email": hospital.email or "",
        "logo_url": getattr(hospital, "logo_url", "") or "",
    }


def format_hospital_address(hospital: dict) -> str:
    """Single-line address from hospital brand fields."""
    parts = []
    if hospital.get("address"):
        parts.append(str(hospital["address"]).strip())
    city_line = ", ".join(
        p for p in [
            (hospital.get("city") or "").strip(),
            (hospital.get("state") or "").strip(),
            (hospital.get("postal_code") or "").strip(),
        ] if p
    )
    if city_line:
        parts.append(city_line)
    return ", ".join(parts)


def format_hospital_contact(hospital: dict) -> str:
    bits = []
    if hospital.get("phone"):
        bits.append(f"Phone: {hospital['phone']}")
    if hospital.get("email"):
        bits.append(f"Email: {hospital['email']}")
    return "  |  ".join(bits)


def resolve_logo_path(logo_url: str) -> Optional[str]:
    """Resolve logo_url (/uploads/...) to an absolute filesystem path, if present."""
    if not logo_url:
        return None
    relative = str(logo_url).lstrip("/")
    if relative.startswith("uploads/"):
        relative = relative[len("uploads/"):]
    full = os.path.join(get_uploads_dir(), relative)
    return full if os.path.isfile(full) else None


def bill_type_label(bill_type: Optional[str]) -> str:
    if bill_type is None:
        return BILL_TYPE_LABELS[None]
    return BILL_TYPE_LABELS.get(bill_type, bill_type.replace("_", " ").title())


def build_export_meta(
    *,
    date_from: str,
    date_to: str,
    bill_type: Optional[str] = None,
    payment_status: Optional[str] = None,
    patient_search: Optional[str] = None,
    doctor_name: Optional[str] = None,
    referred_by: Optional[str] = None,
    summary: Optional[dict] = None,
) -> list[tuple[str, Any]]:
    """Label/value rows describing the export filters and totals."""
    summary = summary or {}
    rows: list[tuple[str, Any]] = [
        ("Report", "Billing Export"),
        ("Date range", f"{date_from} to {date_to}"),
        ("Module", bill_type_label(bill_type)),
    ]
    if payment_status and payment_status != "all":
        rows.append(("Payment status", payment_status.replace("_", " ").title()))
    else:
        rows.append(("Payment status", "All"))
    if patient_search:
        rows.append(("Patient search", patient_search))
    if doctor_name:
        rows.append(("Doctor", doctor_name))
    if referred_by and referred_by != "all":
        rows.append(("Referred by", referred_by))

    rows.append(("", ""))  # spacer
    rows.append(("Total bills", summary.get("total_bills", 0)))
    rows.append(("Total billed", float(summary.get("total_billed") or 0)))
    rows.append(("Collected", float(summary.get("total_paid") or 0)))
    rows.append(("Pending", float(summary.get("total_pending") or 0)))
    cancelled = summary.get("cancelled_count", 0)
    if cancelled:
        rows.append(("Cancelled", cancelled))
    return rows


def apply_workbook_branding(
    ws,
    hospital: dict,
    meta_rows: list[tuple[str, Any]],
    *,
    start_row: int = 1,
) -> int:
    """
    Write hospital logo + name/address (top-left) and export metadata.

    Returns the next free row index (1-based) after the branding block,
    suitable as the header row for tabular data.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    title_font = Font(bold=True, size=14, color="1A365D")
    address_font = Font(size=10, color="4A5568")
    label_font = Font(bold=True, size=10, color="2D3748")
    value_font = Font(size=10)
    meta_fill = PatternFill("solid", fgColor="F7FAFC")

    logo_path = resolve_logo_path(hospital.get("logo_url") or "")
    name = (hospital.get("name") or "Hospital").strip()
    address = format_hospital_address(hospital)
    contact = format_hospital_contact(hospital)

    text_col = 1
    row = start_row

    if logo_path:
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(logo_path)
            # Keep a compact square thumbnail in the top-left.
            max_px = 64
            w, h = float(img.width or max_px), float(img.height or max_px)
            scale = min(max_px / w, max_px / h, 1.0)
            img.width = max(1, int(w * scale))
            img.height = max(1, int(h * scale))
            ws.add_image(img, f"A{row}")
            ws.row_dimensions[row].height = 52
            text_col = 2
            # Give logo column a little room
            ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 12)
        except Exception:
            text_col = 1

    name_cell = ws.cell(row=row, column=text_col, value=name)
    name_cell.font = title_font
    name_cell.alignment = Alignment(vertical="center")
    row += 1

    if address:
        cell = ws.cell(row=row, column=text_col, value=address)
        cell.font = address_font
        row += 1
    if contact:
        cell = ws.cell(row=row, column=text_col, value=contact)
        cell.font = address_font
        row += 1

    # Blank spacer between brand and filter/totals block
    row += 1

    for label, value in meta_rows:
        if label == "" and value == "":
            row += 1
            continue
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = label_font
        label_cell.fill = meta_fill
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.font = value_font
        value_cell.fill = meta_fill
        if isinstance(value, float):
            value_cell.number_format = "#,##0.00"
        row += 1

    # Widen label/value columns used by the meta block
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 18)
    ws.column_dimensions[get_column_letter(2)].width = max(
        ws.column_dimensions[get_column_letter(2)].width or 0, 28
    )

    # Extra blank row before the data table
    row += 1
    return row


def csv_brand_lines(hospital: dict, meta_rows: list[tuple[str, Any]]) -> list[list[Any]]:
    """Header rows (as lists of cells) to prepend before CSV column headers."""
    lines: list[list[Any]] = [[hospital.get("name") or "Hospital"]]
    address = format_hospital_address(hospital)
    if address:
        lines.append([address])
    contact = format_hospital_contact(hospital)
    if contact:
        lines.append([contact])
    lines.append([])
    for label, value in meta_rows:
        if label == "" and value == "":
            lines.append([])
            continue
        lines.append([label, value])
    lines.append([])
    return lines
