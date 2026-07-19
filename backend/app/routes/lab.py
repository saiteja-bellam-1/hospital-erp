from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func as sqlfunc
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime, date, timedelta
import uuid
import io

from config.database import get_db
from app.utils.pdf_settings import bill_pdf_gen_kwargs, pdf_gen_kwargs
from app.models.user import User
from app.models.patient import Patient
from app.models.hospital import Hospital
from app.models.permissions import HospitalSettings
from app.models.lab import (
    SampleType, LabTestCategory, LabTest, LabTestParameter,
    PatientLabOrder, LabReport,
    LabTestPackageCategory, LabTestPackage, LabTestPackageItem
)
from app.utils.dependencies import get_current_user, require_permission
from app.utils.auth import Modules
from app.utils.pdf_service import pdf_service
from app.utils.lab_reference import (
    match_reference_range as _match_reference_range,
    filter_reference_ranges,
    format_reference_ranges_display,
    find_normal_tier,
    is_value_abnormal_for_tiers,
    is_value_abnormal_for_bounds,
    uses_tiered_abnormal_check,
    _coerce_float,
)

router = APIRouter()

# Double-click / parallel submit window for reception book + package book.
_RAPID_BOOK_WINDOW_SECONDS = 3.0


def _new_lab_bill_group(prefix: str = "LB") -> tuple:
    """Allocate a unique lab bill group id and display number.

    The list endpoint groups by ``lab_bill_group_id`` but shows
    ``lab_bill_number`` as the Reference. Timestamp+patient alone collides
    when two bookings land in the same second, so the number always embeds
    a short UUID fragment derived from the group id.
    """
    group_id = str(uuid.uuid4())
    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    short = group_id.replace("-", "")[:8].upper()
    return group_id, f"{prefix}-{stamp}-{short}"


def _reject_rapid_repeat_booking(db: Session, patient_id: int, test_ids: list) -> None:
    """Block identical bookings submitted within a few seconds (double-submit).

    Runs even when ``force=True`` so "Proceed Anyway" double-clicks cannot
    create a second paid bill for the same tests.
    """
    if not test_ids:
        return
    # created_at uses the DB server clock (UTC on SQLite); keep the cutoff in
    # the same frame so IST-vs-UTC skew cannot miss a just-inserted row.
    cutoff = datetime.utcnow() - timedelta(seconds=_RAPID_BOOK_WINDOW_SECONDS)
    recent = (
        db.query(PatientLabOrder)
        .filter(
            PatientLabOrder.patient_id == patient_id,
            PatientLabOrder.test_id.in_(list(test_ids)),
            PatientLabOrder.status != "cancelled",
            PatientLabOrder.created_at >= cutoff,
        )
        .all()
    )
    if not recent:
        return
    recent_ids = {o.test_id for o in recent}
    if set(test_ids).issubset(recent_ids):
        raise HTTPException(
            status_code=409,
            detail={
                "message": "Identical lab booking already submitted moments ago",
                "duplicates": [
                    {
                        "test_id": o.test_id,
                        "order_number": o.order_number,
                        "lab_bill_number": o.lab_bill_number,
                    }
                    for o in recent
                ],
            },
        )


# ============================================================
# Pydantic Models
# ============================================================

class SampleTypeCreate(BaseModel):
    name: str = Field(..., max_length=100)
    description: Optional[str] = None

class SampleTypeResponse(BaseModel):
    id: int
    name: str
    description: Optional[str]
    is_active: bool
    test_count: Optional[int] = 0
    class Config:
        from_attributes = True

class CategoryCreate(BaseModel):
    name: str = Field(..., max_length=100)
    description: Optional[str] = None

class CategoryResponse(BaseModel):
    id: int
    name: str
    description: Optional[str]
    is_active: bool
    test_count: Optional[int] = 0
    class Config:
        from_attributes = True

class ParameterCreate(BaseModel):
    parameter_name: str = Field(..., max_length=200)
    unit: Optional[str] = None
    method: Optional[str] = None
    section: Optional[str] = None
    field_type: str = Field(default="numeric", pattern="^(numeric|tiered_numeric|less_than|greater_than|positive_negative|reactive|presence_absence|cloudy_clear|colour|manual|text|select)$")
    reference_ranges: Optional[list] = None  # [{min, max, gender, age_min, age_max, description, is_normal}]
    possible_values: Optional[list] = None
    abnormal_values: Optional[list] = None
    normal_value: Optional[str] = None
    notes: Optional[str] = None
    display_order: int = 0
    # Legacy fields — kept for backward compat
    reference_min_male: Optional[float] = None
    reference_max_male: Optional[float] = None
    reference_min_female: Optional[float] = None
    reference_max_female: Optional[float] = None
    reference_min_default: Optional[float] = None
    reference_max_default: Optional[float] = None
    reference_min_child: Optional[float] = None
    reference_max_child: Optional[float] = None

class ParameterResponse(BaseModel):
    id: int
    parameter_name: str
    unit: Optional[str]
    method: Optional[str] = None
    section: Optional[str] = None
    field_type: str
    reference_ranges: Optional[list] = None
    possible_values: Optional[list]
    abnormal_values: Optional[list] = None
    normal_value: Optional[str] = None
    notes: Optional[str] = None
    display_order: int
    is_active: bool
    # Legacy
    reference_min_male: Optional[float] = None
    reference_max_male: Optional[float] = None
    reference_min_female: Optional[float] = None
    reference_max_female: Optional[float] = None
    reference_min_default: Optional[float] = None
    reference_max_default: Optional[float] = None
    reference_min_child: Optional[float] = None
    reference_max_child: Optional[float] = None
    class Config:
        from_attributes = True

class TestCreate(BaseModel):
    test_code: str = Field(..., max_length=20)
    name: str = Field(..., max_length=200)
    description: Optional[str] = None
    category_id: int
    cost: float = Field(..., ge=0)
    sample_type: Optional[str] = None  # Legacy free-text
    sample_type_id: Optional[int] = None
    method: Optional[str] = None
    preparation_instructions: Optional[str] = None
    parameters: Optional[List[ParameterCreate]] = None

class TestUpdate(BaseModel):
    test_code: Optional[str] = None
    name: Optional[str] = None
    description: Optional[str] = None
    category_id: Optional[int] = None
    cost: Optional[float] = None
    sample_type: Optional[str] = None  # Legacy free-text
    sample_type_id: Optional[int] = None
    method: Optional[str] = None
    preparation_instructions: Optional[str] = None
    is_active: Optional[bool] = None

class TestResponse(BaseModel):
    id: int
    test_code: str
    name: str
    description: Optional[str]
    category_id: int
    category_name: Optional[str] = None
    cost: float
    sample_type: Optional[str]
    sample_type_id: Optional[int] = None
    sample_type_name: Optional[str] = None
    method: Optional[str]
    preparation_instructions: Optional[str]
    is_active: bool
    parameters: List[ParameterResponse] = []
    class Config:
        from_attributes = True

class OrderCreate(BaseModel):
    patient_id: int
    test_ids: List[int] = Field(..., min_length=1)
    appointment_id: Optional[int] = None
    admission_id: Optional[int] = None
    priority: str = Field(default="normal", pattern="^(normal|urgent|stat)$")
    force: bool = False
    notes: Optional[str] = None

class OrderResponse(BaseModel):
    id: int
    order_number: str
    patient_id: int
    patient_name: Optional[str] = None
    test_id: int
    test_name: Optional[str] = None
    test_code: Optional[str] = None
    doctor_id: Optional[int]
    doctor_name: Optional[str] = None
    status: str
    priority: str
    order_date: datetime
    collection_date: Optional[datetime]
    completion_date: Optional[datetime]
    notes: Optional[str]
    consultation_id: Optional[int] = None
    appointment_id: Optional[int] = None
    admission_id: Optional[int] = None
    order_source: Optional[str] = None  # "appointment", "package", "direct", "inpatient"
    has_report: bool = False
    report_id: Optional[int] = None
    amount: float = 0.0
    payment_status: str = "pending"
    payment_method: Optional[str] = None
    payment_date: Optional[datetime] = None
    package_id: Optional[int] = None
    package_name: Optional[str] = None
    package_booking_id: Optional[str] = None
    sample_id: Optional[str] = None
    sample_type_name: Optional[str] = None
    class Config:
        from_attributes = True

class ResultEntry(BaseModel):
    parameter_id: int
    value: str
    remarks: Optional[str] = None
    manual_abnormal: bool = False

class ResultSubmit(BaseModel):
    results: List[ResultEntry]
    interpretation: Optional[str] = None

class ReportParameterResult(BaseModel):
    parameter_id: int
    parameter_name: str
    value: str
    unit: Optional[str]
    reference_min: Optional[float]
    reference_max: Optional[float]
    is_abnormal: bool = False
    field_type: str = "numeric"

class ReportResponse(BaseModel):
    id: int
    order_id: int
    order_number: str
    patient_id: int
    patient_name: str
    patient_gender: Optional[str] = None
    patient_age: Optional[float] = None
    patient_age_display: Optional[str] = None
    test_id: int
    test_name: str
    test_code: str
    method: Optional[str] = None
    doctor_name: Optional[str] = None
    technician_name: Optional[str] = None
    report_date: datetime
    interpretation: Optional[str]
    results: List[ReportParameterResult] = []

class StatsResponse(BaseModel):
    total_tests: int
    total_categories: int
    total_orders: int
    pending_orders: int
    completed_today: int

# --- Import Schemas ---

class ImportRowError(BaseModel):
    sheet: str
    row: int
    message: str

class ImportPreviewRow(BaseModel):
    row: int
    test_code: str
    name: str
    category: Optional[str] = None
    status: str  # new | update | skip | error
    message: Optional[str] = None
    parameter_count: int = 0

class ImportSummary(BaseModel):
    dry_run: bool
    total_rows: int
    created: int
    updated: int
    skipped: int
    error_count: int
    categories_created: List[str] = []
    sample_types_created: List[str] = []
    errors: List[ImportRowError] = []
    preview: List[ImportPreviewRow] = []

# --- Package Schemas ---

class PackageCategoryCreate(BaseModel):
    name: str = Field(..., max_length=100)
    description: Optional[str] = None

class PackageCategoryResponse(BaseModel):
    id: int
    name: str
    description: Optional[str]
    is_active: bool
    package_count: int = 0
    class Config:
        from_attributes = True

class PackageCreate(BaseModel):
    package_code: str = Field(..., max_length=20)
    name: str = Field(..., max_length=200)
    description: Optional[str] = None
    category_id: int
    package_price: float = Field(..., ge=0)
    test_ids: List[int] = Field(..., min_length=1)

class PackageUpdate(BaseModel):
    package_code: Optional[str] = None
    name: Optional[str] = None
    description: Optional[str] = None
    category_id: Optional[int] = None
    package_price: Optional[float] = None
    test_ids: Optional[List[int]] = None
    is_active: Optional[bool] = None

class PackageTestInfo(BaseModel):
    id: int
    test_code: str
    name: str
    cost: float
    sample_type: Optional[str] = None

class PackageResponse(BaseModel):
    id: int
    package_code: str
    name: str
    description: Optional[str]
    category_id: int
    category_name: Optional[str] = None
    package_price: float
    actual_price: float
    discount_percentage: float = 0.0
    is_active: bool
    tests: List[PackageTestInfo] = []
    class Config:
        from_attributes = True

class PackageBooking(BaseModel):
    patient_id: int
    priority: str = Field(default="normal", pattern="^(normal|urgent|stat)$")
    notes: Optional[str] = None
    referred_by: Optional[str] = Field(None, max_length=100)
    payment_method: str = Field(..., pattern="^(cash|card|upi|cheque|online|insurance)$")
    discount_amount: float = Field(default=0.0, ge=0)
    force: bool = False

# ============================================================
# Helper
# ============================================================

def _require_lab_admin(current_user: User):
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin', 'lab_admin']):
        raise HTTPException(status_code=403, detail="Lab admin access required")

def _require_lab_access(current_user: User):
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin', 'lab_admin', 'lab_technician', 'doctor']):
        raise HTTPException(status_code=403, detail="Lab access required")

from app.utils.patient_age import format_patient_age, patient_age_years_float, patient_age_years_int


def _patient_age(patient):
    """Age in fractional years for lab reference-range matching."""
    return patient_age_years_float(patient)


def _patient_age_display(patient):
    return format_patient_age(patient)

def _sync_parameter_reference_fields(param: LabTestParameter, data: ParameterCreate) -> None:
    """Persist reference_ranges and mirror the first demographic row into legacy columns."""
    param.reference_ranges = data.reference_ranges
    legacy = _legacy_from_reference_ranges(data.reference_ranges)
    for field, val in legacy.items():
        setattr(param, field, val)


def _build_test_response(test: LabTest, db: Session) -> dict:
    category = db.query(LabTestCategory).filter(LabTestCategory.id == test.category_id).first()
    sample_type_obj = db.query(SampleType).filter(SampleType.id == test.sample_type_id).first() if test.sample_type_id else None
    params = db.query(LabTestParameter).filter(
        LabTestParameter.test_id == test.id,
        LabTestParameter.is_active == True
    ).order_by(LabTestParameter.display_order).all()
    return {
        "id": test.id,
        "test_code": test.test_code,
        "name": test.name,
        "description": test.description,
        "category_id": test.category_id,
        "category_name": category.name if category else None,
        "cost": test.cost,
        "sample_type": sample_type_obj.name if sample_type_obj else test.sample_type,
        "sample_type_id": test.sample_type_id,
        "sample_type_name": sample_type_obj.name if sample_type_obj else test.sample_type,
        "method": test.method,
        "preparation_instructions": test.preparation_instructions,
        "is_active": test.is_active,
        "parameters": [
            {
                "id": p.id,
                "parameter_name": p.parameter_name,
                "unit": p.unit,
                "method": p.method,
                "section": p.section,
                "field_type": p.field_type,
                "reference_ranges": p.reference_ranges,
                "reference_min_male": p.reference_min_male,
                "reference_max_male": p.reference_max_male,
                "reference_min_female": p.reference_min_female,
                "reference_max_female": p.reference_max_female,
                "reference_min_default": p.reference_min_default,
                "reference_max_default": p.reference_max_default,
                "reference_min_child": p.reference_min_child,
                "reference_max_child": p.reference_max_child,
                "possible_values": p.possible_values,
                "abnormal_values": p.abnormal_values,
                "normal_value": p.normal_value,
                "notes": p.notes,
                "display_order": p.display_order,
                "is_active": p.is_active,
            } for p in params
        ]
    }

def _build_order_response(order: PatientLabOrder, db: Session) -> dict:
    patient = db.query(Patient).filter(Patient.id == order.patient_id).first()
    test = db.query(LabTest).filter(LabTest.id == order.test_id).first()
    doctor = db.query(User).filter(User.id == order.doctor_id).first() if order.doctor_id else None
    report = db.query(LabReport).filter(LabReport.order_id == order.id).first()
    return {
        "id": order.id,
        "order_number": order.order_number,
        "patient_id": order.patient_id,
        "patient_name": f"{patient.first_name} {patient.last_name}" if patient else None,
        "test_id": order.test_id,
        "test_name": test.name if test else None,
        "test_code": test.test_code if test else None,
        "doctor_id": order.doctor_id,
        "doctor_name": f"Dr. {doctor.first_name} {doctor.last_name}" if doctor else None,
        "status": order.status,
        "priority": order.priority,
        "order_date": order.order_date,
        "collection_date": order.collection_date,
        "completion_date": order.completion_date,
        "notes": order.notes,
        "consultation_id": order.consultation_id,
        "appointment_id": order.appointment_id,
        "admission_id": getattr(order, 'admission_id', None),
        "order_source": "package" if order.package_id else ("inpatient" if getattr(order, 'admission_id', None) else ("appointment" if order.appointment_id else "direct")),
        "has_report": report is not None,
        "report_id": report.id if report else None,
        "amount": order.amount or (test.cost if test else 0.0),
        "payment_status": order.payment_status or "pending",
        "payment_method": order.payment_method,
        "payment_date": order.payment_date,
        "package_id": order.package_id,
        "package_name": order.package.name if order.package_id and order.package else None,
        "package_booking_id": order.package_booking_id,
        "sample_id": order.sample_id,
        "lab_bill_group_id": getattr(order, "lab_bill_group_id", None),
        "lab_bill_number": getattr(order, "lab_bill_number", None),
        "sample_type_name": (
            test.sample_type_ref.name if test and test.sample_type_id and test.sample_type_ref
            else (test.sample_type if test else None)
        ),
    }

def _build_report_response(report: LabReport, db: Session) -> dict:
    order = db.query(PatientLabOrder).filter(PatientLabOrder.id == report.order_id).first()
    patient = db.query(Patient).filter(Patient.id == order.patient_id).first()
    test = db.query(LabTest).filter(LabTest.id == order.test_id).first()
    doctor = db.query(User).filter(User.id == order.doctor_id).first() if order.doctor_id else None
    tech = db.query(User).filter(User.id == report.technician_id).first() if report.technician_id else None

    gender = patient.gender.lower() if patient and patient.gender else None
    age = _patient_age(patient)

    # Build results with abnormal flags
    result_values = report.result_values or []
    results = []
    for rv in result_values:
        param = db.query(LabTestParameter).filter(LabTestParameter.id == rv.get("parameter_id")).first()
        if not param:
            continue

        # Determine reference range — use new reference_ranges if available, else legacy columns
        ref_min = None
        ref_max = None
        matched_desc = ""
        matched_ranges = []
        reference_range_display = ""
        if param.reference_ranges:
            matched_ranges = filter_reference_ranges(param.reference_ranges, gender, age)
            if matched_ranges:
                reference_range_display = format_reference_ranges_display(
                    matched_ranges, param.unit or "", html=True
                )
            if uses_tiered_abnormal_check(param.field_type, matched_ranges):
                normal_tier = find_normal_tier(matched_ranges)
                if normal_tier:
                    ref_min = _coerce_float(normal_tier.get("min"))
                    ref_max = _coerce_float(normal_tier.get("max"))
            else:
                ref_min, ref_max, matched_desc = _match_reference_range(param.reference_ranges, gender, age)
        else:
            # Legacy fallback
            if gender == "male" and param.reference_min_male is not None:
                ref_min = param.reference_min_male
                ref_max = param.reference_max_male
            elif gender == "female" and param.reference_min_female is not None:
                ref_min = param.reference_min_female
                ref_max = param.reference_max_female
            else:
                ref_min = param.reference_min_default
                ref_max = param.reference_max_default

        # Check abnormal
        is_abnormal = False
        raw_value = rv.get("value", "")
        if param.field_type in ("numeric", "less_than", "greater_than", "tiered_numeric") and raw_value:
            try:
                clean_val = raw_value.strip().lstrip('<>').strip()
                val = float(clean_val)
                if uses_tiered_abnormal_check(param.field_type, matched_ranges):
                    is_abnormal = is_value_abnormal_for_tiers(val, matched_ranges)
                else:
                    is_abnormal = is_value_abnormal_for_bounds(
                        val, raw_value, param.field_type, ref_min, ref_max
                    )
            except (ValueError, TypeError):
                pass
        elif param.field_type in ("select", "text", "colour", "manual",
                                   "positive_negative", "reactive", "presence_absence", "cloudy_clear") and raw_value:
            # Check against abnormal_values list
            abnormal_list = param.abnormal_values or []
            if abnormal_list and raw_value.strip() in abnormal_list:
                is_abnormal = True

        # Additive: technician can force-mark any parameter as abnormal via checkbox
        if rv.get("manual_abnormal", False):
            is_abnormal = True

        results.append({
            "parameter_id": param.id,
            "parameter_name": param.parameter_name,
            "value": raw_value,
            "unit": param.unit,
            "method": param.method or "",
            "section": param.section or "",
            "reference_min": ref_min,
            "reference_max": ref_max,
            "reference_range_display": reference_range_display,
            "normal_value": param.normal_value,
            "is_abnormal": is_abnormal,
            "field_type": param.field_type,
            "remarks": rv.get("remarks", ""),
            "notes": param.notes or "",
        })

    # Determine referral label: doctor = "Prescribed By", referral/self = "Referred By"
    referral_label = "Referred By"
    referral_name = "Self"
    if doctor:
        referral_label = "Prescribed By"
        referral_name = f"Dr. {doctor.first_name} {doctor.last_name}"
    elif order.referred_by:
        referral_name = order.referred_by
    elif patient and patient.referred_by:
        referral_name = patient.referred_by

    return {
        "id": report.id,
        "order_id": order.id,
        "order_number": order.order_number,
        "patient_id": patient.id if patient else 0,
        "patient_uuid": patient.patient_id if patient else "",
        "patient_name": f"{patient.first_name} {patient.last_name}" if patient else "Unknown",
        "patient_phone": patient.primary_phone if patient else "",
        "mrn": (patient.mrn or "") if patient else "",
        "patient_gender": patient.gender if patient else None,
        "patient_age": patient_age_years_int(patient),
        "patient_age_display": _patient_age_display(patient),
        "village": (patient.village or "") if patient else "",
        "district": (patient.district or "") if patient else "",
        "test_id": test.id if test else 0,
        "test_name": test.name if test else "Unknown",
        "test_code": test.test_code if test else "",
        "test_description": test.description if test else None,
        "method": test.method if test else None,
        "doctor_name": f"Dr. {doctor.first_name} {doctor.last_name}" if doctor else None,
        "referral_label": referral_label,
        "referral_name": referral_name,
        "technician_name": f"{tech.first_name} {tech.last_name}" if tech else None,
        "order_date": order.order_date,
        "collection_date": order.collection_date,
        "report_date": report.report_date,
        "sample_id": order.sample_id or "",
        "interpretation": report.interpretation,
        "results": results
    }

# ============================================================
# Sample Type Endpoints
# ============================================================

@router.get("/sample-types", response_model=List[SampleTypeResponse])
async def list_sample_types(
    current_user: User = Depends(require_permission(Modules.LAB, "read")),
    db: Session = Depends(get_db)
):
    types = db.query(SampleType).filter(
        SampleType.hospital_id == current_user.hospital_id,
        SampleType.is_active == True
    ).order_by(SampleType.name).all()
    result = []
    for st in types:
        count = db.query(LabTest).filter(LabTest.sample_type_id == st.id, LabTest.is_active == True).count()
        result.append({
            "id": st.id, "name": st.name, "description": st.description,
            "is_active": st.is_active, "test_count": count
        })
    return result

@router.post("/sample-types", response_model=SampleTypeResponse)
async def create_sample_type(
    data: SampleTypeCreate,
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    # Check duplicate name
    existing = db.query(SampleType).filter(
        SampleType.name == data.name,
        SampleType.hospital_id == current_user.hospital_id,
        SampleType.is_active == True
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Sample type '{data.name}' already exists")
    st = SampleType(
        name=data.name, description=data.description,
        hospital_id=current_user.hospital_id
    )
    db.add(st)
    db.commit()
    db.refresh(st)
    return {"id": st.id, "name": st.name, "description": st.description, "is_active": True, "test_count": 0}

@router.put("/sample-types/{sample_type_id}", response_model=SampleTypeResponse)
async def update_sample_type(
    sample_type_id: int, data: SampleTypeCreate,
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    st = db.query(SampleType).filter(
        SampleType.id == sample_type_id,
        SampleType.hospital_id == current_user.hospital_id
    ).first()
    if not st:
        raise HTTPException(status_code=404, detail="Sample type not found")
    st.name = data.name
    if data.description is not None:
        st.description = data.description
    db.commit()
    count = db.query(LabTest).filter(LabTest.sample_type_id == st.id, LabTest.is_active == True).count()
    return {"id": st.id, "name": st.name, "description": st.description, "is_active": st.is_active, "test_count": count}

@router.delete("/sample-types/{sample_type_id}")
async def delete_sample_type(
    sample_type_id: int,
    current_user: User = Depends(require_permission(Modules.LAB, "delete")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    st = db.query(SampleType).filter(
        SampleType.id == sample_type_id,
        SampleType.hospital_id == current_user.hospital_id
    ).first()
    if not st:
        raise HTTPException(status_code=404, detail="Sample type not found")
    st.is_active = False
    db.commit()
    return {"message": "Sample type deleted"}

# ============================================================
# Category Endpoints
# ============================================================

@router.get("/categories", response_model=List[CategoryResponse])
async def list_categories(
    current_user: User = Depends(require_permission(Modules.LAB, "read")),
    db: Session = Depends(get_db)
):
    cats = db.query(LabTestCategory).filter(
        LabTestCategory.hospital_id == current_user.hospital_id,
        LabTestCategory.is_active == True
    ).order_by(LabTestCategory.name).all()

    result = []
    for c in cats:
        count = db.query(LabTest).filter(LabTest.category_id == c.id, LabTest.is_active == True).count()
        result.append({
            "id": c.id, "name": c.name, "description": c.description,
            "is_active": c.is_active, "test_count": count
        })
    return result

@router.post("/categories", response_model=CategoryResponse)
async def create_category(
    data: CategoryCreate,
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    cat = LabTestCategory(
        name=data.name, description=data.description,
        hospital_id=current_user.hospital_id
    )
    db.add(cat)
    db.commit()
    db.refresh(cat)
    return {"id": cat.id, "name": cat.name, "description": cat.description, "is_active": True, "test_count": 0}

@router.put("/categories/{category_id}", response_model=CategoryResponse)
async def update_category(
    category_id: int, data: CategoryCreate,
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    cat = db.query(LabTestCategory).filter(
        LabTestCategory.id == category_id,
        LabTestCategory.hospital_id == current_user.hospital_id
    ).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    cat.name = data.name
    if data.description is not None:
        cat.description = data.description
    db.commit()
    count = db.query(LabTest).filter(LabTest.category_id == cat.id, LabTest.is_active == True).count()
    return {"id": cat.id, "name": cat.name, "description": cat.description, "is_active": cat.is_active, "test_count": count}

@router.delete("/categories/{category_id}")
async def delete_category(
    category_id: int,
    current_user: User = Depends(require_permission(Modules.LAB, "delete")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    cat = db.query(LabTestCategory).filter(
        LabTestCategory.id == category_id,
        LabTestCategory.hospital_id == current_user.hospital_id
    ).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    cat.is_active = False
    db.commit()
    return {"message": "Category deleted"}

# ============================================================
# Lab Test Endpoints
# ============================================================

@router.get("/tests", response_model=List[TestResponse])
async def list_tests(
    category_id: Optional[int] = None,
    search: Optional[str] = None,
    include_inactive: bool = False,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    _assert_can_view_lab_test_catalog(current_user, db)
    from app.utils.auth import UserRoles
    user_roles = set(current_user.role_names)
    can_manage_catalog = bool(user_roles & {
        UserRoles.SUPER_ADMIN, UserRoles.HOSPITAL_ADMIN,
        UserRoles.LAB_ADMIN,
    })
    if include_inactive and not can_manage_catalog:
        include_inactive = False
    query = db.query(LabTest).filter(LabTest.hospital_id == current_user.hospital_id)
    if not include_inactive:
        query = query.filter(LabTest.is_active == True)
    if category_id:
        query = query.filter(LabTest.category_id == category_id)
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            (LabTest.name.ilike(search_term)) | (LabTest.test_code.ilike(search_term))
        )
    tests = query.order_by(LabTest.name).all()
    return [_build_test_response(t, db) for t in tests]

@router.get("/tests/{test_id}", response_model=TestResponse)
async def get_test(
    test_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    _assert_can_view_lab_test_catalog(current_user, db)
    test = db.query(LabTest).filter(
        LabTest.id == test_id,
        LabTest.hospital_id == current_user.hospital_id
    ).first()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    return _build_test_response(test, db)

@router.post("/tests", response_model=TestResponse)
async def create_test(
    data: TestCreate,
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)

    # Check category
    cat = db.query(LabTestCategory).filter(
        LabTestCategory.id == data.category_id,
        LabTestCategory.hospital_id == current_user.hospital_id
    ).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")

    # Check duplicate code
    existing = db.query(LabTest).filter(
        LabTest.test_code == data.test_code,
        LabTest.hospital_id == current_user.hospital_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Test code '{data.test_code}' already exists")

    test = LabTest(
        test_code=data.test_code, name=data.name, description=data.description,
        category_id=data.category_id, cost=data.cost,
        sample_type=data.sample_type, sample_type_id=data.sample_type_id,
        method=data.method, preparation_instructions=data.preparation_instructions,
        hospital_id=current_user.hospital_id
    )
    db.add(test)
    db.flush()

    # Add parameters if provided
    if data.parameters:
        for i, p in enumerate(data.parameters):
            param = LabTestParameter(
                test_id=test.id, parameter_name=p.parameter_name, unit=p.unit,
                field_type=p.field_type,
                reference_min_male=p.reference_min_male, reference_max_male=p.reference_max_male,
                reference_min_female=p.reference_min_female, reference_max_female=p.reference_max_female,
                reference_min_default=p.reference_min_default, reference_max_default=p.reference_max_default,
                possible_values=p.possible_values, display_order=p.display_order or i
            )
            db.add(param)

    db.commit()
    db.refresh(test)
    return _build_test_response(test, db)

@router.put("/tests/{test_id}", response_model=TestResponse)
async def update_test(
    test_id: int, data: TestUpdate,
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    test = db.query(LabTest).filter(
        LabTest.id == test_id, LabTest.hospital_id == current_user.hospital_id
    ).first()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")

    for field in ['test_code', 'name', 'description', 'category_id', 'cost', 'sample_type', 'sample_type_id', 'method', 'preparation_instructions', 'is_active']:
        val = getattr(data, field, None)
        if val is not None:
            setattr(test, field, val)

    db.commit()
    db.refresh(test)
    return _build_test_response(test, db)

@router.delete("/tests/{test_id}")
async def delete_test(
    test_id: int,
    current_user: User = Depends(require_permission(Modules.LAB, "delete")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    test = db.query(LabTest).filter(
        LabTest.id == test_id, LabTest.hospital_id == current_user.hospital_id
    ).first()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    test.is_active = False
    db.commit()
    return {"message": "Test deleted"}

# ============================================================
# Parameter Endpoints
# ============================================================

@router.post("/tests/{test_id}/parameters", response_model=ParameterResponse)
async def add_parameter(
    test_id: int, data: ParameterCreate,
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    test = db.query(LabTest).filter(
        LabTest.id == test_id, LabTest.hospital_id == current_user.hospital_id
    ).first()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")

    param = LabTestParameter(test_id=test_id, **{k: v for k, v in data.dict().items() if k != "reference_ranges"})
    _sync_parameter_reference_fields(param, data)
    db.add(param)
    db.commit()
    db.refresh(param)
    return param

# NOTE: reorder and bulk routes MUST come before {param_id} to avoid path conflict
@router.put("/tests/{test_id}/parameters/reorder")
async def reorder_parameters(
    test_id: int, order: List[int],
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    """Update display_order for parameters. Body is a list of parameter IDs in desired order."""
    _require_lab_admin(current_user)
    for idx, param_id in enumerate(order):
        param = db.query(LabTestParameter).filter(
            LabTestParameter.id == param_id, LabTestParameter.test_id == test_id
        ).first()
        if param:
            param.display_order = idx
    db.commit()
    return {"message": f"Reordered {len(order)} parameters"}

@router.put("/tests/{test_id}/parameters/bulk")
async def bulk_upsert_parameters(
    test_id: int, parameters: List[ParameterCreate],
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    """Replace all parameters for a test"""
    _require_lab_admin(current_user)
    test = db.query(LabTest).filter(
        LabTest.id == test_id, LabTest.hospital_id == current_user.hospital_id
    ).first()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")

    # Delete existing
    db.query(LabTestParameter).filter(LabTestParameter.test_id == test_id).delete()

    # Add new
    for i, p in enumerate(parameters):
        param = LabTestParameter(
            test_id=test_id, parameter_name=p.parameter_name, unit=p.unit,
            method=p.method, section=p.section,
            field_type=p.field_type,
            reference_min_male=p.reference_min_male, reference_max_male=p.reference_max_male,
            reference_min_female=p.reference_min_female, reference_max_female=p.reference_max_female,
            reference_min_default=p.reference_min_default, reference_max_default=p.reference_max_default,
            possible_values=p.possible_values, display_order=p.display_order or i
        )
        db.add(param)

    db.commit()
    return {"message": f"Updated {len(parameters)} parameters"}

@router.put("/tests/{test_id}/parameters/{param_id}", response_model=ParameterResponse)
async def update_parameter(
    test_id: int, param_id: int, data: ParameterCreate,
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    param = db.query(LabTestParameter).filter(
        LabTestParameter.id == param_id, LabTestParameter.test_id == test_id
    ).first()
    if not param:
        raise HTTPException(status_code=404, detail="Parameter not found")

    # Update all fields from data (including nullable fields like section, method)
    update_data = data.dict()
    for field, val in update_data.items():
        if field == "reference_ranges":
            continue
        setattr(param, field, val)
    _sync_parameter_reference_fields(param, data)

    db.commit()
    db.refresh(param)
    return param

@router.delete("/tests/{test_id}/parameters/{param_id}")
async def delete_parameter(
    test_id: int, param_id: int,
    current_user: User = Depends(require_permission(Modules.LAB, "delete")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    param = db.query(LabTestParameter).filter(
        LabTestParameter.id == param_id, LabTestParameter.test_id == test_id
    ).first()
    if not param:
        raise HTTPException(status_code=404, detail="Parameter not found")
    db.delete(param)
    db.commit()
    return {"message": "Parameter deleted"}


def _check_duplicate_orders(db, patient_id: int, test_ids: list) -> list:
    """Check if any tests were already ordered today for this patient (paid or pending)."""
    from sqlalchemy import func as sql_func
    today = date.today()
    duplicates = []
    for test_id in test_ids:
        existing = db.query(PatientLabOrder).filter(
            PatientLabOrder.patient_id == patient_id,
            PatientLabOrder.test_id == test_id,
            PatientLabOrder.status != "cancelled",
            sql_func.date(PatientLabOrder.order_date) == today,
        ).first()
        if existing:
            test = db.query(LabTest).filter(LabTest.id == test_id).first()
            duplicates.append({
                "test_id": test_id,
                "test_name": test.name if test else "Unknown",
                "order_number": existing.order_number,
                "order_time": existing.order_date.strftime('%I:%M %p') if existing.order_date else "",
                "status": existing.status,
                "payment_status": existing.payment_status,
            })
    return duplicates


def _get_lab_hospital_info(db, hospital):
    """Build hospital_info dict using lab config as primary, hospital as fallback.
    Includes hospital_subname for showing hospital name below lab name in PDFs."""
    lab_settings = db.query(HospitalSettings).filter(
        HospitalSettings.setting_category == "lab_config"
    ).all()
    lab_config = {s.setting_key: s.setting_value for s in lab_settings}

    lab_name = lab_config.get('provider_name', '')
    hosp_name = hospital.name if hospital else 'Hospital'
    # If lab has its own name different from hospital, include hospital as subname
    hospital_subname = hosp_name if lab_name and lab_name.strip().upper() != hosp_name.strip().upper() else ''

    return {
        "name": lab_name or hosp_name,
        "hospital_subname": hospital_subname,
        "address": lab_config.get('provider_address') or (hospital.address if hospital else ''),
        "phone": lab_config.get('provider_phone') or (hospital.phone if hospital else ''),
        "email": lab_config.get('provider_email') or (hospital.email if hospital else ''),
        "logo_url": lab_config.get('provider_logo') or (hospital.logo_url if hospital else ''),
    }


# ============================================================
# Lab Order Endpoints
# ============================================================

@router.post("/orders/check-duplicates")
async def check_duplicate_orders(
    data: dict,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Check if any of the given tests were already booked today for this patient.
    Accessible by any authenticated user (no module gate). Uses the same helper as backend enforcement."""
    patient_id = data.get("patient_id")
    test_ids = data.get("test_ids", [])

    if not patient_id or not test_ids:
        return {"duplicates": []}

    duplicates = _check_duplicate_orders(db, patient_id, test_ids)
    return {"duplicates": duplicates}


_RECEPTION_LAB_BOOK_ROLES = frozenset({"receptionist", "frontdesk", "hospital_admin", "super_admin"})

_LAB_READ_PERMISSIONS = frozenset({
    "view_appointments", "view_patients", "view_schedules", "read",
    "view_reports", "view_records", "view_history", "view_prescriptions",
    "view_occupancy", "view_financial_reports", "view_system_reports",
    "view_vitals", "view_mar",
})


def _assert_reception_lab_book_role(current_user: User) -> None:
    if not _RECEPTION_LAB_BOOK_ROLES & set(current_user.role_names):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only reception or admin can book lab tests",
        )


def _can_view_lab_test_catalog(current_user: User, db: Session) -> bool:
    """May list active lab tests for booking/ordering (not full lab admin access)."""
    from app.utils.auth import UserRoles
    from app.models.permissions import RoleModulePermission

    user_roles = set(current_user.role_names)
    if user_roles & {
        UserRoles.SUPER_ADMIN, UserRoles.HOSPITAL_ADMIN,
        UserRoles.LAB_ADMIN,
    }:
        return True
    if user_roles & _RECEPTION_LAB_BOOK_ROLES:
        return True

    role_ids = [r.id for r in (current_user.roles or [])]
    if current_user.role_id and current_user.role_id not in role_ids:
        role_ids.append(current_user.role_id)

    for rid in role_ids:
        rp = db.query(RoleModulePermission).filter(
            RoleModulePermission.role_id == rid,
            RoleModulePermission.module_name == Modules.LAB,
        ).first()
        if rp and rp.permissions and any(p in _LAB_READ_PERMISSIONS for p in rp.permissions):
            return True

    return False


def _assert_can_view_lab_test_catalog(current_user: User, db: Session) -> None:
    if not _can_view_lab_test_catalog(current_user, db):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Insufficient permissions to view lab tests",
        )


def _can_create_lab_order(current_user: User, db: Session, admission_id: Optional[int]) -> None:
    """Permission gate for creating a lab order.

    Allowed when EITHER:
      • the user has the legacy `LAB:write` action permission (doctors, lab admins,
        lab techs, super_admin/hospital_admin), OR
      • the order is for an inpatient admission AND the user has the granular
        `inpatient:order_labs` permission (lets nurses order labs from the
        bedside without giving them full lab-module write access).
    Raises 403 if neither path passes.
    """
    from fastapi import HTTPException, status as _status
    from app.utils.auth import UserRoles
    from app.models.permissions import RoleModulePermission

    user_roles = set(current_user.role_names)
    if {UserRoles.SUPER_ADMIN, UserRoles.HOSPITAL_ADMIN} & user_roles:
        return
    # Lab roles always allowed
    if {UserRoles.LAB_ADMIN, UserRoles.LAB_TECHNICIAN} & user_roles:
        return

    role_ids = [r.id for r in (current_user.roles or [])]
    if current_user.role_id and current_user.role_id not in role_ids:
        role_ids.append(current_user.role_id)

    # Path 1 — legacy LAB:write
    write_perms = {
        "schedule_appointments", "register_patients", "manage_queues", "write",
        "update_appointments", "create_reports", "edit_records", "create_prescriptions",
        "process_payments", "generate_invoices", "dispense_medications",
        "admit_patients", "discharge_patients", "manage_tests", "set_rates",
        "manage_inventory", "manage_templates", "manage_equipment", "generate_reports",
    }
    for rid in role_ids:
        rp = db.query(RoleModulePermission).filter(
            RoleModulePermission.role_id == rid,
            RoleModulePermission.module_name == Modules.LAB,
        ).first()
        if rp and rp.permissions and any(p in write_perms for p in rp.permissions):
            return

    # Path 2 — inpatient.order_labs (only for admission-scoped orders)
    if admission_id is not None:
        for rid in role_ids:
            rp = db.query(RoleModulePermission).filter(
                RoleModulePermission.role_id == rid,
                RoleModulePermission.module_name == Modules.INPATIENT,
            ).first()
            if rp and rp.permissions and "order_labs" in rp.permissions:
                return

    raise HTTPException(
        status_code=_status.HTTP_403_FORBIDDEN,
        detail="Permission required: lab.write or inpatient.order_labs (for admission orders)",
    )


@router.post("/orders", response_model=List[OrderResponse])
async def create_orders(
    data: OrderCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create lab orders for a patient (outpatient or inpatient)."""
    _can_create_lab_order(current_user, db, data.admission_id)
    patient = db.query(Patient).filter(
        Patient.id == data.patient_id,
        Patient.hospital_id == current_user.hospital_id
    ).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    # Duplicate check
    if not data.force:
        duplicates = _check_duplicate_orders(db, data.patient_id, data.test_ids)
        if duplicates:
            raise HTTPException(status_code=409, detail={"message": "Duplicate orders found", "duplicates": duplicates})

    orders = []
    for test_id in data.test_ids:
        test = db.query(LabTest).filter(
            LabTest.id == test_id,
            LabTest.hospital_id == current_user.hospital_id,
            LabTest.is_active == True
        ).first()
        if not test:
            raise HTTPException(status_code=404, detail=f"Test ID {test_id} not found")

        order = PatientLabOrder(
            order_number=f"LAB-{str(uuid.uuid4())[:8].upper()}",
            patient_id=data.patient_id,
            test_id=test_id,
            doctor_id=current_user.id,
            appointment_id=data.appointment_id,
            admission_id=data.admission_id,
            priority=data.priority,
            notes=data.notes,
            status="ordered",
            amount=test.cost or 0.0,
            payment_status="pending"
        )
        db.add(order)
        orders.append(order)

    db.commit()

    # Audit log
    try:
        from app.services.audit_service import log_action
        test_names = ", ".join(o.test.name for o in orders if o.test)
        patient_name = f"{patient.first_name} {patient.last_name}"
        log_action(db, current_user, "order_lab_tests", "lab", "LabOrder", orders[0].id if orders else None,
            f"Ordered {len(orders)} lab test(s) for {patient_name}: {test_names}",
            details={"patient": patient_name, "tests": test_names, "count": len(orders)})
    except Exception:
        pass

    return [_build_order_response(o, db) for o in orders]

class ReceptionLabBooking(BaseModel):
    patient_id: int
    test_ids: List[int] = Field(..., min_length=1)
    payment_method: str = Field(..., pattern="^(cash|card|upi|cheque|online|insurance)$")
    doctor_id: Optional[int] = None
    referred_by: Optional[str] = None
    discount_amount: float = Field(default=0.0, ge=0)
    force: bool = False
    notes: Optional[str] = None


@router.post("/orders/reception-book")
async def reception_book_lab_tests(
    data: ReceptionLabBooking,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Reception books individual lab tests directly for a patient with payment."""
    _assert_reception_lab_book_role(current_user)

    patient = db.query(Patient).filter(
        Patient.id == data.patient_id,
        Patient.hospital_id == current_user.hospital_id
    ).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    booking_doctor = None
    if data.doctor_id is not None:
        booking_doctor = db.query(User).join(User.role).filter(
            User.id == data.doctor_id,
            User.role.has(name="doctor"),
            User.hospital_id == current_user.hospital_id,
            User.is_active == True,
        ).first()
        if not booking_doctor:
            raise HTTPException(status_code=400, detail="Invalid or inactive doctor")

    # Duplicate check
    if not data.force:
        duplicates = _check_duplicate_orders(db, data.patient_id, data.test_ids)
        if duplicates:
            raise HTTPException(status_code=409, detail={"message": "Duplicate orders found", "duplicates": duplicates})

    # Always block true double-submits (same tests within a few seconds),
    # even when the operator chose "Proceed Anyway" for same-day duplicates.
    _reject_rapid_repeat_booking(db, data.patient_id, data.test_ids)

    now = datetime.now()
    orders = []
    total = 0.0
    bill_group_id, bill_number = _new_lab_bill_group("LB")

    for test_id in data.test_ids:
        test = db.query(LabTest).filter(
            LabTest.id == test_id,
            LabTest.hospital_id == current_user.hospital_id,
            LabTest.is_active == True
        ).first()
        if not test:
            raise HTTPException(status_code=404, detail=f"Test ID {test_id} not found")

        order = PatientLabOrder(
            order_number=f"LAB-{str(uuid.uuid4())[:8].upper()}",
            patient_id=data.patient_id,
            test_id=test_id,
            doctor_id=booking_doctor.id if booking_doctor else None,
            referred_by=data.referred_by,
            priority="normal",
            notes=data.notes,
            status="ordered",
            order_date=now,
            amount=test.cost or 0.0,
            payment_status="paid",
            payment_method=data.payment_method,
            payment_date=now,
            lab_bill_group_id=bill_group_id,
            lab_bill_number=bill_number,
        )
        db.add(order)
        orders.append(order)
        total += test.cost or 0.0

    db.commit()

    # Audit log
    try:
        from app.services.audit_service import log_action
        test_names = ", ".join(o.test.name for o in orders if o.test)
        patient_name = f"{patient.first_name} {patient.last_name}"
        log_action(db, current_user, "reception_book_lab", "lab", "LabOrder", orders[0].id if orders else None,
            f"Reception booked {len(orders)} lab test(s) for {patient_name}: {test_names}, Total: ₹{total}",
            details={"patient": patient_name, "tests": test_names, "total": total, "method": data.payment_method})
    except Exception:
        pass

    # Generate bill PDF
    from app.utils.pdf_service import pdf_service
    from app.models.hospital import Hospital

    hospital = db.query(Hospital).filter(Hospital.id == current_user.hospital_id).first()
    hospital_info = _get_lab_hospital_info(db, hospital)

    age = _patient_age(patient)

    bill_data = {
        "bill_number": bill_number,
        "bill_date": now.isoformat(),
        "patient_name": f"{patient.first_name} {patient.last_name}",
        "patient_age": patient_age_years_int(patient),
        "patient_age_display": _patient_age_display(patient),
        "patient_gender": patient.gender,
        "patient_phone": patient.primary_phone or "",
        "mrn": patient.mrn or "",
        "patient_id": patient.patient_id,
        "village": patient.village or "",
        "district": patient.district or "",
        "reg_no": patient.patient_id,
        "doctor_name": (
            f"Dr. {booking_doctor.first_name} {booking_doctor.last_name}"
            if booking_doctor else ""
        ),
        "referred_by": data.referred_by or "",
        "payment_method": data.payment_method.capitalize(),
        "items": [{"item_name": o.test.name, "item_code": o.test.test_code, "total_price": o.amount} for o in orders if o.test],
        "subtotal": total,
        "discount_amount": data.discount_amount,
        "amount_paid": round(total - data.discount_amount, 2),
        "balance_due": 0,
        "prepared_by": f"{current_user.first_name} {current_user.last_name}",
    }

    from fastapi.responses import StreamingResponse
    order_ids = ",".join(str(o.id) for o in orders)
    pdf_buffer = pdf_service.generate_bill_pdf(bill_data, hospital_info, **bill_pdf_gen_kwargs(db, current_user.hospital_id, 'lab_bill'))
    filename = f"lab_bill_{bill_data['bill_number']}.pdf"
    return StreamingResponse(pdf_buffer, media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "X-Order-Ids": order_ids,
            "Access-Control-Expose-Headers": "X-Order-Ids",
        })


@router.get("/orders", response_model=List[OrderResponse])
async def list_orders(
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    patient_id: Optional[int] = None,
    appointment_id: Optional[int] = None,
    reception_view: bool = False,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    query = db.query(PatientLabOrder).join(Patient).filter(
        Patient.hospital_id == current_user.hospital_id
    )

    if status:
        query = query.filter(PatientLabOrder.status == status)
    if patient_id:
        query = query.filter(PatientLabOrder.patient_id == patient_id)
    if appointment_id:
        query = query.filter(PatientLabOrder.appointment_id == appointment_id)
    if date_from:
        query = query.filter(PatientLabOrder.order_date >= date_from)
    if date_to:
        query = query.filter(PatientLabOrder.order_date <= date_to + " 23:59:59")

    # For doctors, only show their orders
    if current_user.has_role('doctor'):
        query = query.filter(PatientLabOrder.doctor_id == current_user.id)

    # Lab technicians only see paid orders on the lab workflow (payment gate).
    # IPD orders bypass this gate — they're billed on the admission account
    # and stay 'pending' until the IPD bill is settled.
    # Reception pages pass reception_view=true so dual-role users still see
    # unpaid OPD orders when collecting payment at the front desk.
    skip_payment_gate = (
        reception_view
        and bool(_RECEPTION_LAB_BOOK_ROLES & set(current_user.role_names))
    )
    if current_user.has_role('lab_technician') and not skip_payment_gate:
        query = query.filter(
            (PatientLabOrder.payment_status == 'paid')
            | (PatientLabOrder.admission_id.isnot(None))
        )

    orders = query.order_by(PatientLabOrder.order_date.desc()).limit(200).all()
    return [_build_order_response(o, db) for o in orders]

@router.get("/orders/{order_id}", response_model=OrderResponse)
async def get_order(
    order_id: int,
    current_user: User = Depends(require_permission(Modules.LAB, "read")),
    db: Session = Depends(get_db)
):
    order = db.query(PatientLabOrder).join(Patient).filter(
        PatientLabOrder.id == order_id,
        Patient.hospital_id == current_user.hospital_id
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return _build_order_response(order, db)

def _generate_sample_id(db: Session) -> str:
    """Generate next sample ID in format S-YYMMDD-NNNN."""
    today_prefix = f"S-{datetime.now().strftime('%y%m%d')}-"
    last = db.query(PatientLabOrder).filter(
        PatientLabOrder.sample_id.like(f"{today_prefix}%")
    ).order_by(PatientLabOrder.sample_id.desc()).first()
    if last and last.sample_id:
        try:
            seq = int(last.sample_id.split('-')[-1]) + 1
        except ValueError:
            seq = 1
    else:
        seq = 1
    return f"{today_prefix}{seq:04d}"

@router.put("/orders/{order_id}/status")
async def update_order_status(
    order_id: int,
    status: str,
    force_new_sample: bool = False,
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    if status not in ['ordered', 'collected', 'processing', 'completed', 'cancelled']:
        raise HTTPException(status_code=400, detail="Invalid status")

    order = db.query(PatientLabOrder).join(Patient).filter(
        PatientLabOrder.id == order_id,
        Patient.hospital_id == current_user.hospital_id
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    order.status = status
    sample_id = None
    grouped_orders = []
    if status == "collected":
        order.collection_date = datetime.now()
        if not order.sample_id:
            test = db.query(LabTest).filter(LabTest.id == order.test_id).first()
            sample_type_id = test.sample_type_id if test else None
            sample_type_text = test.sample_type if test else None

            # Determine if we can group: need either sample_type_id or legacy sample_type text
            can_group = (sample_type_id or sample_type_text) and not force_new_sample

            if can_group:
                # Find other "ordered" orders for same patient + same sample type
                sibling_query = db.query(PatientLabOrder).join(LabTest).filter(
                    PatientLabOrder.patient_id == order.patient_id,
                    PatientLabOrder.id != order.id,
                    PatientLabOrder.status == "ordered",
                    PatientLabOrder.sample_id.is_(None),
                )
                # Match by sample_type_id (preferred) or legacy text fallback
                if sample_type_id:
                    sibling_query = sibling_query.filter(LabTest.sample_type_id == sample_type_id)
                else:
                    sibling_query = sibling_query.filter(LabTest.sample_type == sample_type_text)
                sibling_orders = sibling_query.all()

                new_sample_id = _generate_sample_id(db)
                order.sample_id = new_sample_id

                for sibling in sibling_orders:
                    sibling.status = "collected"
                    sibling.collection_date = datetime.now()
                    sibling.sample_id = new_sample_id
                    grouped_orders.append({
                        "id": sibling.id,
                        "order_number": sibling.order_number,
                        "test_name": sibling.test.name if sibling.test else "",
                    })
            else:
                order.sample_id = _generate_sample_id(db)
        sample_id = order.sample_id
    elif status == "completed":
        order.completion_date = datetime.now()

    db.commit()
    return {
        "message": f"Order status updated to {status}",
        "sample_id": sample_id,
        "order_number": order.order_number,
        "patient_name": f"{order.patient.first_name} {order.patient.last_name}" if order.patient else "",
        "test_name": order.test.name if order.test else "",
        "grouped_orders": grouped_orders,
    }


@router.get("/orders/{order_id}/bill")
async def download_order_bill(
    order_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Download/regenerate bill PDF for a single lab order."""
    order = db.query(PatientLabOrder).join(Patient).filter(
        PatientLabOrder.id == order_id,
        Patient.hospital_id == current_user.hospital_id
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    patient = db.query(Patient).filter(Patient.id == order.patient_id).first()
    test = db.query(LabTest).filter(LabTest.id == order.test_id).first()
    hospital = db.query(Hospital).filter(Hospital.id == current_user.hospital_id).first()
    hospital_info = _get_lab_hospital_info(db, hospital)

    age = _patient_age(patient)

    doctor = db.query(User).filter(User.id == order.doctor_id).first() if order.doctor_id else None
    doctor_name = f"Dr. {doctor.first_name} {doctor.last_name}" if doctor else ""
    referred_by = order.referred_by or ""

    bill_data = {
        "bill_number": f"LB-{order.order_number}",
        "bill_date": order.order_date.isoformat() if order.order_date else datetime.now().isoformat(),
        "patient_name": f"{patient.first_name} {patient.last_name}" if patient else "Unknown",
        "patient_age": patient_age_years_int(patient),
        "patient_age_display": _patient_age_display(patient),
        "patient_gender": patient.gender if patient else "",
        "patient_phone": patient.primary_phone if patient else "",
        "mrn": (patient.mrn or "") if patient else "",
        "patient_id": patient.patient_id if patient else "",
        "village": (patient.village or "") if patient else "",
        "district": (patient.district or "") if patient else "",
        "doctor_name": doctor_name,
        "referred_by": referred_by,
        "payment_method": (order.payment_method or "cash").capitalize(),
        "items": [{"item_name": test.name if test else "Lab Test", "item_code": test.test_code if test else "", "total_price": order.amount or 0}],
        "subtotal": order.amount or 0,
        "discount_amount": 0,
        "amount_paid": order.amount or 0,
        "balance_due": 0,
        "prepared_by": "",
    }

    from app.utils.pdf_service import pdf_service
    from fastapi.responses import StreamingResponse
    pdf_buffer = pdf_service.generate_bill_pdf(bill_data, hospital_info, **bill_pdf_gen_kwargs(db, current_user.hospital_id, 'lab_bill'))
    filename = f"lab_bill_{order.order_number}.pdf"
    return StreamingResponse(pdf_buffer, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


class RegenerateBillRequest(BaseModel):
    order_ids: List[int] = Field(..., min_length=1)


@router.post("/orders/regenerate-bill")
async def regenerate_lab_bill(
    data: RegenerateBillRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Regenerate bill PDF for given order IDs (no payment side effects). Used for preview with header toggle."""
    orders = db.query(PatientLabOrder).join(Patient).filter(
        PatientLabOrder.id.in_(data.order_ids),
        Patient.hospital_id == current_user.hospital_id
    ).all()
    if not orders:
        raise HTTPException(status_code=404, detail="Orders not found")

    patient = db.query(Patient).filter(Patient.id == orders[0].patient_id).first()
    hospital = db.query(Hospital).filter(Hospital.id == current_user.hospital_id).first()
    hospital_info = _get_lab_hospital_info(db, hospital)

    age = _patient_age(patient)

    doctor = db.query(User).filter(User.id == orders[0].doctor_id).first() if orders[0].doctor_id else None
    doctor_name = f"Dr. {doctor.first_name} {doctor.last_name}" if doctor else ""
    referred_by = orders[0].referred_by or ""

    items = []
    total = 0.0
    for order in orders:
        test = db.query(LabTest).filter(LabTest.id == order.test_id).first()
        items.append({
            "item_name": test.name if test else "Lab Test",
            "item_code": test.test_code if test else "",
            "total_price": order.amount or 0,
        })
        total += order.amount or 0

    # Package bills: one line at actual_price; paid = sum of stored order amounts
    # (package price minus any extra operator discount applied at booking).
    discount = 0.0
    pkg = None
    if orders[0].package_id:
        pkg = db.query(LabTestPackage).filter(LabTestPackage.id == orders[0].package_id).first()
        if pkg:
            paid = round(sum(o.amount or 0 for o in orders), 2)
            items = [{"item_name": pkg.name, "item_code": pkg.package_code, "total_price": pkg.actual_price}]
            total = pkg.actual_price
            discount = round(pkg.actual_price - paid, 2)

    now = orders[0].payment_date or orders[0].order_date or datetime.now()
    stored_number = next(
        (o.lab_bill_number for o in orders if o.lab_bill_number),
        None,
    )
    bill_data = {
        "bill_number": stored_number or (orders[0].order_number if orders else "LB-UNKNOWN"),
        "bill_date": now.isoformat(),
        "patient_name": f"{patient.first_name} {patient.last_name}" if patient else "Unknown",
        "patient_age": patient_age_years_int(patient),
        "patient_age_display": _patient_age_display(patient),
        "patient_gender": patient.gender if patient else "",
        "patient_phone": patient.primary_phone if patient else "",
        "mrn": (patient.mrn or "") if patient else "",
        "patient_id": patient.patient_id if patient else "",
        "village": (patient.village or "") if patient else "",
        "district": (patient.district or "") if patient else "",
        "reg_no": patient.patient_id if patient else "",
        "doctor_name": doctor_name,
        "referred_by": referred_by,
        "payment_method": (orders[0].payment_method or "cash").capitalize(),
        "items": items,
        "subtotal": total,
        "discount_amount": discount,
        "amount_paid": round(total - discount, 2),
        "balance_due": 0,
        "prepared_by": "",
    }

    from app.utils.pdf_service import pdf_service
    from fastapi.responses import StreamingResponse
    pdf_buffer = pdf_service.generate_bill_pdf(bill_data, hospital_info, **bill_pdf_gen_kwargs(db, current_user.hospital_id, 'lab_bill'))
    return StreamingResponse(pdf_buffer, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=lab_bill.pdf"})


@router.get("/bills/{group_id}/pdf")
async def download_grouped_lab_bill(
    group_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Re-render the combined lab bill PDF for every order that shares a
    lab_bill_group_id. This is what the centralised Billing dashboard's
    Download button calls — it reproduces the originally-issued bill
    (multi-test or package) without any payment side effects.
    """
    orders = db.query(PatientLabOrder).join(Patient).filter(
        PatientLabOrder.lab_bill_group_id == group_id,
        Patient.hospital_id == current_user.hospital_id,
    ).order_by(PatientLabOrder.id.asc()).all()
    if not orders:
        raise HTTPException(status_code=404, detail="Bill not found")

    patient = db.query(Patient).filter(Patient.id == orders[0].patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    hospital = db.query(Hospital).filter(Hospital.id == current_user.hospital_id).first()
    hospital_info = _get_lab_hospital_info(db, hospital)
    age = _patient_age(patient)

    # Doctor / referral come from the first order that has them — same
    # heuristic the original bill generators use.
    doctor_name = ""
    referred_by = ""
    for o in orders:
        if not doctor_name and o.doctor_id:
            doc = db.query(User).filter(User.id == o.doctor_id).first()
            if doc:
                doctor_name = f"Dr. {doc.first_name} {doc.last_name}"
        if not referred_by and o.referred_by:
            referred_by = o.referred_by

    # If every order in the group came from the same package, render the
    # bill the way book_package() did — single line with package pricing
    # and the actual_price/package_price gap as discount. Otherwise it's a
    # multi-test bill with one item per order.
    pkg = None
    pkg_ids = {o.package_id for o in orders if o.package_id}
    if len(pkg_ids) == 1 and None not in pkg_ids:
        pkg = db.query(LabTestPackage).filter(LabTestPackage.id == orders[0].package_id).first()

    if pkg:
        items = [{
            "item_name": pkg.name,
            "item_code": pkg.package_code,
            "total_price": pkg.actual_price,
        }]
        subtotal = pkg.actual_price
        # Paid is whatever was stored on orders (package price − any extra discount).
        amount_paid = round(sum(o.amount or 0.0 for o in orders), 2)
        discount = round(pkg.actual_price - amount_paid, 2)
    else:
        items = []
        subtotal = 0.0
        for o in orders:
            test = db.query(LabTest).filter(LabTest.id == o.test_id).first()
            items.append({
                "item_name": test.name if test else "Lab Test",
                "item_code": test.test_code if test else "",
                "total_price": o.amount or 0.0,
            })
            subtotal += o.amount or 0.0
        discount = 0.0
        amount_paid = round(subtotal, 2)

    bill_date = orders[0].payment_date or orders[0].order_date or datetime.now()
    bill_number = orders[0].lab_bill_number or f"LB-GROUP-{group_id[:8]}"

    bill_data = {
        "bill_number": bill_number,
        "bill_date": bill_date.isoformat() if hasattr(bill_date, "isoformat") else str(bill_date),
        "patient_name": f"{patient.first_name} {patient.last_name}",
        "patient_age": patient_age_years_int(patient),
        "patient_age_display": _patient_age_display(patient),
        "patient_gender": patient.gender,
        "patient_phone": patient.primary_phone or "",
        "mrn": patient.mrn or "",
        "patient_id": patient.patient_id,
        "village": patient.village or "",
        "district": patient.district or "",
        "reg_no": patient.patient_id,
        "doctor_name": doctor_name,
        "referred_by": referred_by,
        "payment_method": (orders[0].payment_method or "cash").capitalize(),
        "items": items,
        "subtotal": subtotal,
        "discount_amount": discount,
        "amount_paid": amount_paid,
        "balance_due": 0,
        "prepared_by": "",
    }
    if pkg:
        bill_data["package_name"] = pkg.name

    from app.utils.pdf_service import pdf_service
    from fastapi.responses import StreamingResponse
    pdf_buffer = pdf_service.generate_bill_pdf(bill_data, hospital_info, **bill_pdf_gen_kwargs(db, current_user.hospital_id, 'lab_bill'))
    filename = f"lab_bill_{bill_number}.pdf"
    return StreamingResponse(pdf_buffer, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


class LabPaymentUpdate(BaseModel):
    payment_method: str = Field(..., pattern="^(cash|card|upi|cheque|online|insurance)$")
    discount_amount: float = Field(default=0.0, ge=0)
@router.put("/orders/{order_id}/payment")
async def update_order_payment(
    order_id: int,
    data: LabPaymentUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Mark a lab order as paid. Accessible by receptionist, hospital_admin, super_admin."""
    _assert_reception_lab_book_role(current_user)

    order = db.query(PatientLabOrder).join(Patient).filter(
        PatientLabOrder.id == order_id,
        Patient.hospital_id == current_user.hospital_id
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    if order.payment_status == 'paid':
        raise HTTPException(status_code=400, detail="Payment already collected")

    now = datetime.now()
    order.payment_status = "paid"
    order.payment_method = data.payment_method
    order.payment_date = now
    # Single-order payment still gets a bill grouping so the Billing
    # dashboard can render it as one row and the regenerate endpoint can
    # reproduce a bill PDF for it.
    if not order.lab_bill_group_id:
        order.lab_bill_group_id, order.lab_bill_number = _new_lab_bill_group("LB")

    db.commit()

    # Audit log
    try:
        from app.services.audit_service import log_action
        patient = db.query(Patient).filter(Patient.id == order.patient_id).first()
        test = db.query(LabTest).filter(LabTest.id == order.test_id).first()
        patient_name = f"{patient.first_name} {patient.last_name}" if patient else "Unknown"
        log_action(db, current_user, "collect_lab_payment", "billing", "LabOrder", order.id,
            f"Collected lab payment ₹{order.amount} for {patient_name} — {test.name if test else 'test'}, Method: {data.payment_method}",
            details={"patient": patient_name, "test": test.name if test else "", "amount": order.amount, "method": data.payment_method})
    except Exception:
        pass

    return _build_order_response(order, db)


@router.post("/orders/patient/{patient_id}/bill")
async def generate_lab_bill(
    patient_id: int,
    data: LabPaymentUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Pay all pending lab orders for a patient and generate a combined PDF bill.
    Uses lab config provider details if available, falls back to hospital info."""
    _assert_reception_lab_book_role(current_user)

    # Get pending orders
    orders = db.query(PatientLabOrder).join(Patient).filter(
        PatientLabOrder.patient_id == patient_id,
        Patient.hospital_id == current_user.hospital_id,
        PatientLabOrder.payment_status == "pending",
        PatientLabOrder.status != "cancelled"
    ).order_by(PatientLabOrder.order_date.desc()).all()

    if not orders:
        raise HTTPException(status_code=404, detail="No pending lab orders found")

    patient = db.query(Patient).filter(Patient.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    # Mark all orders as paid and stamp them with a shared bill grouping
    # so the Billing dashboard collapses them into one row and the bill PDF
    # can be regenerated later from the group id.
    now = datetime.now()
    bill_group_id, bill_number = _new_lab_bill_group("LB")
    for order in orders:
        order.payment_status = "paid"
        order.payment_method = data.payment_method
        order.payment_date = now
        order.lab_bill_group_id = bill_group_id
        order.lab_bill_number = bill_number

    db.commit()

    # Build bill data
    items = []
    total = 0.0
    for order in orders:
        test = db.query(LabTest).filter(LabTest.id == order.test_id).first()
        doctor = db.query(User).filter(User.id == order.doctor_id).first() if order.doctor_id else None
        amount = order.amount or (test.cost if test else 0.0)
        total += amount
        items.append({
            "item_name": test.name if test else "Lab Test",
            "item_code": test.test_code if test else "",
            "total_price": amount,
        })

    # Get hospital info
    hospital = db.query(Hospital).filter(Hospital.id == current_user.hospital_id).first()

    hospital_info = _get_lab_hospital_info(db, hospital)

    # Calculate patient age
    age = _patient_age(patient)

    # bill_number was assigned above when the orders were stamped with their
    # shared lab_bill_group_id. Reuse it here so the PDF carries the same
    # number that's persisted on every order in the group.

    # Determine referral info from orders
    referral_name = ""
    doctor_name = ""
    for order in orders:
        if order.doctor_id:
            doc = db.query(User).filter(User.id == order.doctor_id).first()
            if doc:
                doctor_name = f"Dr. {doc.first_name} {doc.last_name}"
                break
        if order.referred_by and not referral_name:
            referral_name = order.referred_by

    bill_data = {
        "bill_number": bill_number,
        "bill_date": now.isoformat(),
        "patient_name": f"{patient.first_name} {patient.last_name}",
        "patient_age": patient_age_years_int(patient),
        "patient_age_display": _patient_age_display(patient),
        "patient_gender": patient.gender,
        "patient_phone": patient.primary_phone or "",
        "mrn": patient.mrn or "",
        "patient_id": patient.patient_id,
        "village": patient.village or "",
        "district": patient.district or "",
        "reg_no": patient.patient_id,
        "doctor_name": doctor_name,
        "referred_by": referral_name,
        "payment_method": data.payment_method.capitalize(),
        "items": items,
        "subtotal": total,
        "discount_amount": data.discount_amount,
        "amount_paid": round(total - data.discount_amount, 2),
        "balance_due": 0,
        "prepared_by": f"{current_user.first_name} {current_user.last_name}",
    }

    try:
        order_ids_str = ",".join(str(o.id) for o in orders)
        pdf_buffer = pdf_service.generate_bill_pdf(bill_data, hospital_info, **bill_pdf_gen_kwargs(db, current_user.hospital_id, 'lab_bill'))
        filename = f"lab_bill_{bill_number}.pdf"
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Order-Ids": order_ids_str,
                "Access-Control-Expose-Headers": "X-Order-Ids",
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF generation error: {str(e)}")


@router.get("/orders/patient/{patient_id}/pending-payment", response_model=List[OrderResponse])
async def get_pending_payment_orders(
    patient_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get lab orders pending payment for a patient. For reception to collect fees."""
    orders = db.query(PatientLabOrder).join(Patient).filter(
        PatientLabOrder.patient_id == patient_id,
        Patient.hospital_id == current_user.hospital_id,
        PatientLabOrder.payment_status == "pending",
        PatientLabOrder.status != "cancelled"
    ).order_by(PatientLabOrder.order_date.desc()).all()
    return [_build_order_response(o, db) for o in orders]

@router.get("/orders/patient/{patient_id}", response_model=List[OrderResponse])
async def get_patient_orders(
    patient_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    orders = db.query(PatientLabOrder).join(Patient).filter(
        PatientLabOrder.patient_id == patient_id,
        Patient.hospital_id == current_user.hospital_id
    ).order_by(PatientLabOrder.order_date.desc()).all()
    return [_build_order_response(o, db) for o in orders]

# ============================================================
# Lab Result Entry (for technicians)
# ============================================================

@router.get("/orders/{order_id}/entry-form")
async def get_entry_form(
    order_id: int,
    current_user: User = Depends(require_permission(Modules.LAB, "read")),
    db: Session = Depends(get_db)
):
    """Get test parameters for result entry form"""
    order = db.query(PatientLabOrder).join(Patient).filter(
        PatientLabOrder.id == order_id,
        Patient.hospital_id == current_user.hospital_id
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    test = db.query(LabTest).filter(LabTest.id == order.test_id).first()
    patient = db.query(Patient).filter(Patient.id == order.patient_id).first()
    params = db.query(LabTestParameter).filter(
        LabTestParameter.test_id == test.id,
        LabTestParameter.is_active == True
    ).order_by(LabTestParameter.display_order).all()

    gender = patient.gender.lower() if patient and patient.gender else None
    age = _patient_age(patient)

    def _resolve_range(p):
        if p.reference_ranges:
            rmin, rmax, _ = _match_reference_range(p.reference_ranges, gender, age)
            return rmin, rmax
        # Legacy fallback
        if gender == "male" and p.reference_min_male is not None:
            return p.reference_min_male, p.reference_max_male
        if gender == "female" and p.reference_min_female is not None:
            return p.reference_min_female, p.reference_max_female
        return p.reference_min_default, p.reference_max_default

    return {
        "order_id": order.id,
        "order_number": order.order_number,
        "test_name": test.name,
        "test_code": test.test_code,
        "patient_name": f"{patient.first_name} {patient.last_name}" if patient else "Unknown",
        "patient_gender": patient.gender if patient else None,
        "parameters": [
            {
                "id": p.id,
                "parameter_name": p.parameter_name,
                "unit": p.unit,
                "field_type": p.field_type,
                "reference_min": _resolve_range(p)[0],
                "reference_max": _resolve_range(p)[1],
                "possible_values": p.possible_values,
                "abnormal_values": p.abnormal_values,
                "normal_value": p.normal_value,
                "notes": p.notes,
                "display_order": p.display_order
            } for p in params
        ]
    }

@router.post("/orders/{order_id}/results")
async def submit_results(
    order_id: int, data: ResultSubmit,
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    """Lab technician submits test results"""
    order = db.query(PatientLabOrder).join(Patient).filter(
        PatientLabOrder.id == order_id,
        Patient.hospital_id == current_user.hospital_id
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Check if report already exists
    existing = db.query(LabReport).filter(LabReport.order_id == order_id).first()
    if existing:
        raise HTTPException(status_code=400, detail="Report already exists for this order")

    # Legacy catch-up marked completed without a LabReport — allow late entry.
    # If completed with a report, the existing check above already blocked.
    if order.status == "completed":
        # no report → fall through (admin/lab can backfill results)
        pass
    elif order.status == "cancelled":
        raise HTTPException(status_code=400, detail="Cannot submit results for a cancelled order")

    result_values = [{"parameter_id": r.parameter_id, "value": r.value, "remarks": r.remarks or "", "manual_abnormal": r.manual_abnormal} for r in data.results]

    report = LabReport(
        order_id=order_id,
        result_values=result_values,
        interpretation=data.interpretation,
        technician_id=current_user.id,
        report_date=datetime.now()
    )
    db.add(report)

    order.status = "completed"
    order.completion_date = datetime.now()

    db.commit()
    db.refresh(report)

    return {"message": "Results submitted successfully", "report_id": report.id}

@router.put("/reports/{report_id}")
async def update_report(
    report_id: int, data: ResultSubmit,
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    """Update existing lab report"""
    report = db.query(LabReport).filter(LabReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    report.result_values = [{"parameter_id": r.parameter_id, "value": r.value} for r in data.results]
    if data.interpretation is not None:
        report.interpretation = data.interpretation

    db.commit()
    return {"message": "Report updated successfully"}

# ============================================================
# Lab Report Viewing
# ============================================================

@router.get("/reports/patient/{patient_id}", response_model=List[ReportResponse])
async def get_patient_reports(
    patient_id: int,
    current_user: User = Depends(require_permission(Modules.LAB, "read")),
    db: Session = Depends(get_db)
):
    orders = db.query(PatientLabOrder).join(Patient).filter(
        PatientLabOrder.patient_id == patient_id,
        Patient.hospital_id == current_user.hospital_id,
        PatientLabOrder.status == "completed"
    ).all()

    reports = []
    for order in orders:
        report = db.query(LabReport).filter(LabReport.order_id == order.id).first()
        if report:
            reports.append(_build_report_response(report, db))

    return reports

@router.get("/reports/{report_id}", response_model=ReportResponse)
async def get_report(
    report_id: int,
    current_user: User = Depends(require_permission(Modules.LAB, "read")),
    db: Session = Depends(get_db)
):
    report = db.query(LabReport).filter(LabReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    return _build_report_response(report, db)

@router.get("/reports/{report_id}/download")
async def download_report_pdf(
    report_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Download lab report as PDF. Accessible by any authenticated user."""
    report = db.query(LabReport).filter(LabReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    report_data = _build_report_response(report, db)

    # Get hospital info (fallback)
    hospital = db.query(Hospital).filter(Hospital.id == current_user.hospital_id).first()
    hospital_info = {
        "name": hospital.name if hospital else "Hospital",
        "address": hospital.address if hospital else "",
        "phone": hospital.phone if hospital else "",
        "email": hospital.email if hospital else "",
        "logo_url": hospital.logo_url if hospital else ""
    }

    # Get lab-specific config from HospitalSettings
    lab_settings = db.query(HospitalSettings).filter(
        HospitalSettings.setting_category == "lab_config"
    ).all()
    lab_config = {s.setting_key: s.setting_value for s in lab_settings}

    try:
        pdf_buffer = pdf_service.generate_lab_report_pdf(report_data, hospital_info, lab_config, **pdf_gen_kwargs(db, current_user.hospital_id, 'lab_report'))
        filename = f"lab_report_{report_data['order_number']}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF generation error: {str(e)}")

@router.get("/reports/package/{package_booking_id}/download")
async def download_package_report_pdf(
    package_booking_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Download combined PDF for all completed tests in a package booking."""
    orders = db.query(PatientLabOrder).filter(
        PatientLabOrder.package_booking_id == package_booking_id
    ).all()
    if not orders:
        raise HTTPException(status_code=404, detail="No orders found for this package")

    # Collect all completed reports
    reports_data = []
    for order in orders:
        report = db.query(LabReport).filter(LabReport.order_id == order.id).first()
        if report:
            reports_data.append(_build_report_response(report, db))

    if not reports_data:
        raise HTTPException(status_code=404, detail="No completed reports found for this package")

    hospital = db.query(Hospital).filter(Hospital.id == current_user.hospital_id).first()
    hospital_info = {
        "name": hospital.name if hospital else "Hospital",
        "address": hospital.address if hospital else "",
        "phone": hospital.phone if hospital else "",
        "email": hospital.email if hospital else "",
        "logo_url": hospital.logo_url if hospital else ""
    }

    lab_settings = db.query(HospitalSettings).filter(
        HospitalSettings.setting_category == "lab_config"
    ).all()
    lab_config = {s.setting_key: s.setting_value for s in lab_settings}

    try:
        pdf_buffer = pdf_service.generate_combined_lab_report_pdf(
            reports_data, hospital_info, lab_config, **pdf_gen_kwargs(db, current_user.hospital_id, 'lab_report')
        )
        pkg_name = orders[0].package.name if orders[0].package else "package"
        patient = db.query(Patient).filter(Patient.id == orders[0].patient_id).first()
        patient_name = f"{patient.first_name}_{patient.last_name}" if patient else "patient"
        filename = f"{patient_name}_{pkg_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"PDF generation error: {str(e)}")

# ============================================================
# Stats & Seed
# ============================================================

@router.get("/stats", response_model=StatsResponse)
async def get_lab_stats(
    current_user: User = Depends(require_permission(Modules.LAB, "read")),
    db: Session = Depends(get_db)
):
    total_tests = db.query(LabTest).filter(
        LabTest.hospital_id == current_user.hospital_id, LabTest.is_active == True
    ).count()
    total_categories = db.query(LabTestCategory).filter(
        LabTestCategory.hospital_id == current_user.hospital_id, LabTestCategory.is_active == True
    ).count()
    total_orders = db.query(PatientLabOrder).join(Patient).filter(
        Patient.hospital_id == current_user.hospital_id
    ).count()
    pending_orders = db.query(PatientLabOrder).join(Patient).filter(
        Patient.hospital_id == current_user.hospital_id,
        PatientLabOrder.status.in_(["ordered", "collected", "processing"])
    ).count()
    today = date.today()
    completed_today = db.query(PatientLabOrder).join(Patient).filter(
        Patient.hospital_id == current_user.hospital_id,
        PatientLabOrder.status == "completed",
        sqlfunc.date(PatientLabOrder.completion_date) == today
    ).count()

    return {
        "total_tests": total_tests,
        "total_categories": total_categories,
        "total_orders": total_orders,
        "pending_orders": pending_orders,
        "completed_today": completed_today
    }

@router.get("/tests/{test_id}/sample-report")
async def generate_sample_report(
    test_id: int,
    current_user: User = Depends(require_permission(Modules.LAB, "read")),
    db: Session = Depends(get_db)
):
    """Generate a sample/blank lab report PDF for a test to preview the report layout."""
    test = db.query(LabTest).filter(
        LabTest.id == test_id, LabTest.hospital_id == current_user.hospital_id
    ).first()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")

    params = db.query(LabTestParameter).filter(
        LabTestParameter.test_id == test_id, LabTestParameter.is_active == True
    ).order_by(LabTestParameter.display_order).all()

    # Build sample report data
    sample_results = []
    for p in params:
        ref_min = p.reference_min_default
        ref_max = p.reference_max_default
        # Use a sample value in the middle of the range if numeric
        sample_value = ""
        if p.field_type == "numeric" and ref_min is not None and ref_max is not None:
            sample_value = str(round((ref_min + ref_max) / 2, 2))
        elif p.field_type == "select" and p.possible_values:
            sample_value = p.possible_values[0] if isinstance(p.possible_values, list) else ""
        else:
            sample_value = "—"

        sample_results.append({
            "parameter_name": p.parameter_name,
            "value": sample_value,
            "unit": p.unit or "",
            "method": p.method or "",
            "section": p.section or "",
            "reference_min": ref_min,
            "reference_max": ref_max,
            "is_abnormal": False,
            "field_type": p.field_type,
            "notes": p.notes or "",
        })

    report_data = {
        "test_name": test.name,
        "test_code": test.test_code,
        "method": test.method or "",
        "patient_name": "SAMPLE PATIENT",
        "patient_gender": "Male",
        "patient_age": 30,
        "doctor_name": "Dr. Sample Doctor",
        "order_number": "SAMPLE-001",
        "report_date": datetime.now().isoformat(),
        "results": sample_results,
        "interpretation": "This is a sample report for preview purposes only.",
    }

    # Get hospital info
    hospital = db.query(Hospital).filter(Hospital.id == current_user.hospital_id).first()
    hospital_info = {
        "name": hospital.name if hospital else "Hospital",
        "address": hospital.address if hospital else "",
        "phone": hospital.phone if hospital else "",
        "email": hospital.email if hospital else "",
        "logo_url": hospital.logo_url if hospital else "",
    }

    # Get lab config
    lab_config = {}
    settings = db.query(HospitalSettings).filter(
        HospitalSettings.setting_category == "lab_config"
    ).all()
    for s in settings:
        lab_config[s.setting_key] = s.setting_value

    try:
        pdf_buffer = pdf_service.generate_lab_report_pdf(report_data, hospital_info, lab_config, **pdf_gen_kwargs(db, current_user.hospital_id, 'lab_report'))
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")

    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=sample_report_{test.test_code}.pdf"}
    )


# ============================================================
# Package Categories
# ============================================================

def _build_package_response(pkg: LabTestPackage, db) -> dict:
    category = db.query(LabTestPackageCategory).filter(LabTestPackageCategory.id == pkg.category_id).first()
    items = db.query(LabTestPackageItem).filter(LabTestPackageItem.package_id == pkg.id).all()
    tests = []
    for item in items:
        test = db.query(LabTest).filter(LabTest.id == item.test_id).first()
        if test:
            tests.append({
                "id": test.id,
                "test_code": test.test_code,
                "name": test.name,
                "cost": test.cost,
                "sample_type": test.sample_type,
            })
    discount_pct = round((1 - pkg.package_price / pkg.actual_price) * 100, 1) if pkg.actual_price > 0 else 0
    return {
        "id": pkg.id,
        "package_code": pkg.package_code,
        "name": pkg.name,
        "description": pkg.description,
        "category_id": pkg.category_id,
        "category_name": category.name if category else None,
        "package_price": pkg.package_price,
        "actual_price": pkg.actual_price,
        "discount_percentage": discount_pct,
        "is_active": pkg.is_active,
        "tests": tests,
    }


@router.get("/packages/categories", response_model=List[PackageCategoryResponse])
async def list_package_categories(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    cats = db.query(LabTestPackageCategory).filter(
        LabTestPackageCategory.hospital_id == current_user.hospital_id
    ).order_by(LabTestPackageCategory.name).all()
    result = []
    for c in cats:
        count = db.query(sqlfunc.count(LabTestPackage.id)).filter(
            LabTestPackage.category_id == c.id, LabTestPackage.is_active == True
        ).scalar() or 0
        result.append({
            "id": c.id, "name": c.name, "description": c.description,
            "is_active": c.is_active, "package_count": count,
        })
    return result


@router.post("/packages/categories", response_model=PackageCategoryResponse)
async def create_package_category(
    data: PackageCategoryCreate,
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    cat = LabTestPackageCategory(
        name=data.name, description=data.description,
        hospital_id=current_user.hospital_id
    )
    db.add(cat)
    db.commit()
    db.refresh(cat)
    return {"id": cat.id, "name": cat.name, "description": cat.description,
            "is_active": cat.is_active, "package_count": 0}


@router.put("/packages/categories/{cat_id}", response_model=PackageCategoryResponse)
async def update_package_category(
    cat_id: int,
    data: PackageCategoryCreate,
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    cat = db.query(LabTestPackageCategory).filter(
        LabTestPackageCategory.id == cat_id,
        LabTestPackageCategory.hospital_id == current_user.hospital_id
    ).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Package category not found")
    cat.name = data.name
    cat.description = data.description
    db.commit()
    count = db.query(sqlfunc.count(LabTestPackage.id)).filter(
        LabTestPackage.category_id == cat.id, LabTestPackage.is_active == True
    ).scalar() or 0
    return {"id": cat.id, "name": cat.name, "description": cat.description,
            "is_active": cat.is_active, "package_count": count}


@router.delete("/packages/categories/{cat_id}")
async def delete_package_category(
    cat_id: int,
    current_user: User = Depends(require_permission(Modules.LAB, "delete")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    cat = db.query(LabTestPackageCategory).filter(
        LabTestPackageCategory.id == cat_id,
        LabTestPackageCategory.hospital_id == current_user.hospital_id
    ).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Package category not found")
    cat.is_active = False
    db.commit()
    return {"message": "Package category deactivated"}


# ============================================================
# Packages CRUD
# ============================================================

@router.get("/packages", response_model=List[PackageResponse])
async def list_packages(
    category_id: Optional[int] = None,
    search: Optional[str] = None,
    active_only: bool = True,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    q = db.query(LabTestPackage).filter(LabTestPackage.hospital_id == current_user.hospital_id)
    if active_only:
        q = q.filter(LabTestPackage.is_active == True)
    if category_id:
        q = q.filter(LabTestPackage.category_id == category_id)
    if search:
        q = q.filter(LabTestPackage.name.ilike(f"%{search}%"))
    packages = q.order_by(LabTestPackage.name).all()
    return [_build_package_response(p, db) for p in packages]


@router.get("/packages/{pkg_id}", response_model=PackageResponse)
async def get_package(
    pkg_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    pkg = db.query(LabTestPackage).filter(
        LabTestPackage.id == pkg_id,
        LabTestPackage.hospital_id == current_user.hospital_id
    ).first()
    if not pkg:
        raise HTTPException(status_code=404, detail="Package not found")
    return _build_package_response(pkg, db)


@router.post("/packages", response_model=PackageResponse)
async def create_package(
    data: PackageCreate,
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    # Validate category
    cat = db.query(LabTestPackageCategory).filter(
        LabTestPackageCategory.id == data.category_id,
        LabTestPackageCategory.hospital_id == current_user.hospital_id
    ).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Package category not found")

    # Validate tests and compute actual price
    actual_price = 0.0
    tests = []
    for tid in data.test_ids:
        test = db.query(LabTest).filter(
            LabTest.id == tid, LabTest.hospital_id == current_user.hospital_id, LabTest.is_active == True
        ).first()
        if not test:
            raise HTTPException(status_code=404, detail=f"Test ID {tid} not found or inactive")
        actual_price += test.cost
        tests.append(test)

    pkg = LabTestPackage(
        package_code=data.package_code, name=data.name, description=data.description,
        category_id=data.category_id, package_price=data.package_price,
        actual_price=actual_price, hospital_id=current_user.hospital_id
    )
    db.add(pkg)
    db.flush()

    for test in tests:
        db.add(LabTestPackageItem(package_id=pkg.id, test_id=test.id))
    db.commit()
    db.refresh(pkg)
    return _build_package_response(pkg, db)


@router.put("/packages/{pkg_id}", response_model=PackageResponse)
async def update_package(
    pkg_id: int,
    data: PackageUpdate,
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    pkg = db.query(LabTestPackage).filter(
        LabTestPackage.id == pkg_id,
        LabTestPackage.hospital_id == current_user.hospital_id
    ).first()
    if not pkg:
        raise HTTPException(status_code=404, detail="Package not found")

    if data.package_code is not None:
        pkg.package_code = data.package_code
    if data.name is not None:
        pkg.name = data.name
    if data.description is not None:
        pkg.description = data.description
    if data.category_id is not None:
        pkg.category_id = data.category_id
    if data.is_active is not None:
        pkg.is_active = data.is_active
    if data.package_price is not None:
        pkg.package_price = data.package_price

    # Update test list if provided
    if data.test_ids is not None:
        actual_price = 0.0
        for tid in data.test_ids:
            test = db.query(LabTest).filter(
                LabTest.id == tid, LabTest.hospital_id == current_user.hospital_id, LabTest.is_active == True
            ).first()
            if not test:
                raise HTTPException(status_code=404, detail=f"Test ID {tid} not found or inactive")
            actual_price += test.cost
        pkg.actual_price = actual_price
        # Replace items
        db.query(LabTestPackageItem).filter(LabTestPackageItem.package_id == pkg.id).delete()
        for tid in data.test_ids:
            db.add(LabTestPackageItem(package_id=pkg.id, test_id=tid))

    db.commit()
    db.refresh(pkg)
    return _build_package_response(pkg, db)


@router.delete("/packages/{pkg_id}")
async def delete_package(
    pkg_id: int,
    current_user: User = Depends(require_permission(Modules.LAB, "delete")),
    db: Session = Depends(get_db)
):
    _require_lab_admin(current_user)
    pkg = db.query(LabTestPackage).filter(
        LabTestPackage.id == pkg_id,
        LabTestPackage.hospital_id == current_user.hospital_id
    ).first()
    if not pkg:
        raise HTTPException(status_code=404, detail="Package not found")
    pkg.is_active = False
    db.commit()
    return {"message": "Package deactivated"}


# ============================================================
# Package Booking
# ============================================================

@router.post("/packages/{pkg_id}/book")
async def book_package(
    pkg_id: int,
    data: PackageBooking,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Book a test package for a patient. Creates individual lab orders, marks as paid, returns bill PDF."""
    _assert_reception_lab_book_role(current_user)

    # Validate package
    pkg = db.query(LabTestPackage).filter(
        LabTestPackage.id == pkg_id,
        LabTestPackage.hospital_id == current_user.hospital_id,
        LabTestPackage.is_active == True
    ).first()
    if not pkg:
        raise HTTPException(status_code=404, detail="Package not found or inactive")

    # Validate patient
    patient = db.query(Patient).filter(
        Patient.id == data.patient_id,
        Patient.hospital_id == current_user.hospital_id
    ).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    # Get package tests
    items = db.query(LabTestPackageItem).filter(LabTestPackageItem.package_id == pkg.id).all()
    if not items:
        raise HTTPException(status_code=400, detail="Package has no tests")

    tests = []
    for item in items:
        test = db.query(LabTest).filter(LabTest.id == item.test_id).first()
        if test:
            tests.append(test)

    if not tests:
        raise HTTPException(status_code=400, detail="No valid tests in package")

    # Duplicate check
    if not data.force:
        duplicates = _check_duplicate_orders(db, data.patient_id, [t.id for t in tests])
        if duplicates:
            raise HTTPException(status_code=409, detail={"message": "Duplicate orders found", "duplicates": duplicates})

    _reject_rapid_repeat_booking(db, data.patient_id, [t.id for t in tests])

    # Net payable = package price minus any extra operator discount.
    # Extra discount is folded into stored order amounts so reprints / billing
    # list stay consistent without a separate discount column.
    extra_discount = round(min(float(data.discount_amount or 0), float(pkg.package_price or 0)), 2)
    net_payable = round(float(pkg.package_price or 0) - extra_discount, 2)

    # Proportional distribution of net payable across package tests
    booking_id = f"PKG-{str(uuid.uuid4())[:8].upper()}"
    now = datetime.now()
    bill_group_id, bill_number = _new_lab_bill_group("PKG")
    orders = []
    distributed_total = 0.0
    actual_price = float(pkg.actual_price or 0)

    for i, test in enumerate(tests):
        if i == len(tests) - 1:
            # Last test gets the remainder to avoid floating-point rounding
            amount = round(net_payable - distributed_total, 2)
        elif actual_price > 0:
            amount = round((float(test.cost or 0) / actual_price) * net_payable, 2)
            distributed_total += amount
        else:
            amount = 0.0

        order = PatientLabOrder(
            order_number=f"LAB-{str(uuid.uuid4())[:8].upper()}",
            patient_id=data.patient_id,
            test_id=test.id,
            doctor_id=None,
            package_id=pkg.id,
            package_booking_id=booking_id,
            referred_by=data.referred_by,
            priority=data.priority,
            notes=data.notes,
            status="ordered",
            order_date=now,
            amount=amount,
            payment_status="paid",
            payment_method=data.payment_method,
            payment_date=now,
            lab_bill_group_id=bill_group_id,
            lab_bill_number=bill_number,
        )
        db.add(order)
        orders.append(order)

    db.commit()

    # Build bill PDF — single line for the package
    bill_items = [{
        "item_name": pkg.name,
        "item_code": pkg.package_code,
        "total_price": pkg.actual_price,
    }]

    hospital = db.query(Hospital).filter(Hospital.id == current_user.hospital_id).first()
    hospital_info = _get_lab_hospital_info(db, hospital)

    age = _patient_age(patient)

    # Catalog savings + any extra operator discount, as a single discount line
    discount = round(float(pkg.actual_price or 0) - net_payable, 2)
    # bill_number was set above when stamping the orders so the persisted
    # number matches the PDF.

    bill_data = {
        "bill_number": bill_number,
        "bill_date": now.isoformat(),
        "patient_name": f"{patient.first_name} {patient.last_name}",
        "patient_age": patient_age_years_int(patient),
        "patient_age_display": _patient_age_display(patient),
        "patient_gender": patient.gender,
        "patient_phone": patient.primary_phone or "",
        "mrn": patient.mrn or "",
        "village": patient.village or "",
        "district": patient.district or "",
        "reg_no": patient.patient_id,
        "doctor_name": "",
        "payment_method": data.payment_method.capitalize(),
        "items": bill_items,
        "subtotal": pkg.actual_price,
        "discount_amount": discount,
        "amount_paid": net_payable,
        "balance_due": 0,
        "prepared_by": f"{current_user.first_name} {current_user.last_name}",
        "package_name": pkg.name,
    }

    try:
        order_ids = ",".join(str(o.id) for o in orders)
        pdf_buffer = pdf_service.generate_bill_pdf(bill_data, hospital_info, **bill_pdf_gen_kwargs(db, current_user.hospital_id, 'lab_bill'))
        filename = f"lab_package_bill_{bill_number}.pdf"
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Order-Ids": order_ids,
                "Access-Control-Expose-Headers": "X-Order-Ids",
            }
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"PDF generation error: {str(e)}")


def _legacy_from_seed_param(param: dict) -> dict:
    """Map legacy seed keys to LabTestParameter columns."""
    return {
        "reference_min_male": param.get("min_m"),
        "reference_max_male": param.get("max_m"),
        "reference_min_female": param.get("min_f"),
        "reference_max_female": param.get("max_f"),
        "reference_min_default": param.get("min_d"),
        "reference_max_default": param.get("max_d"),
        "reference_min_child": param.get("min_c"),
        "reference_max_child": param.get("max_c"),
    }


def _legacy_from_reference_ranges(reference_ranges: Optional[list]) -> dict:
    legacy = {
        "reference_min_male": None,
        "reference_max_male": None,
        "reference_min_female": None,
        "reference_max_female": None,
        "reference_min_default": None,
        "reference_max_default": None,
        "reference_min_child": None,
        "reference_max_child": None,
    }
    if not reference_ranges:
        return legacy

    for row in reference_ranges:
        gender = row.get("gender") or "common"
        if gender == "male":
            legacy["reference_min_male"] = row.get("min")
            legacy["reference_max_male"] = row.get("max")
        elif gender == "female":
            legacy["reference_min_female"] = row.get("min")
            legacy["reference_max_female"] = row.get("max")
        elif gender == "child":
            legacy["reference_min_child"] = row.get("min")
            legacy["reference_max_child"] = row.get("max")
        else:
            legacy["reference_min_default"] = row.get("min")
            legacy["reference_max_default"] = row.get("max")
    return legacy


def _create_seed_parameter(test_id: int, param: dict, display_order: int) -> LabTestParameter:
    reference_ranges = param.get("reference_ranges")
    legacy = (
        _legacy_from_reference_ranges(reference_ranges)
        if reference_ranges is not None
        else _legacy_from_seed_param(param)
    )
    return LabTestParameter(
        test_id=test_id,
        parameter_name=param["name"],
        unit=param.get("unit"),
        method=param.get("method"),
        section=param.get("section"),
        field_type=param.get("field_type", "numeric"),
        reference_ranges=reference_ranges,
        possible_values=param.get("possible_values"),
        abnormal_values=param.get("abnormal_values"),
        normal_value=param.get("normal_value"),
        notes=param.get("notes"),
        critical_low=param.get("critical_low"),
        critical_high=param.get("critical_high"),
        display_order=display_order,
        **legacy,
    )


# ============================================================
# Bulk Import (Excel / CSV)
# ============================================================

_IMPORT_TEST_HEADERS = [
    "test_code", "name", "category", "sample_type", "cost",
    "method", "description", "preparation_instructions",
]
_IMPORT_PARAM_HEADERS = [
    "test_code", "section", "parameter_name", "unit", "method", "field_type",
    "ref_min", "ref_max", "gender", "age_min", "age_max", "description",
    "possible_values", "normal_value", "abnormal_values",
    "critical_low", "critical_high",
]
_VALID_FIELD_TYPES = {
    "numeric", "tiered_numeric", "less_than", "greater_than",
    "positive_negative", "reactive", "presence_absence", "cloudy_clear",
    "colour", "manual", "text", "select",
}


def _norm_header(h) -> str:
    if h is None:
        return ""
    return str(h).strip().lower().replace(" ", "_")


def _cell_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def _cell_float(v):
    s = _cell_str(v)
    if s is None:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        raise ValueError(f"'{s}' is not a valid number")


def _cell_list(v):
    s = _cell_str(v)
    if s is None:
        return None
    parts = [p.strip() for p in s.split(",") if p.strip()]
    return parts or None


def _row_is_empty(r) -> bool:
    return not any(c is not None and str(c).strip() != "" for c in r)


def _read_xlsx_sheet(wb, preferred_names, fallback_first=False):
    """Return list of row dicts keyed by normalized header, with 1-based '_row'."""
    lower_map = {name.lower(): name for name in wb.sheetnames}
    ws = None
    for pn in preferred_names:
        if pn.lower() in lower_map:
            ws = wb[lower_map[pn.lower()]]
            break
    if ws is None:
        if fallback_first and wb.sheetnames:
            ws = wb[wb.sheetnames[0]]
        else:
            return []
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next((i for i, r in enumerate(rows) if not _row_is_empty(r)), None)
    if header_idx is None:
        return []
    headers = [_norm_header(c) for c in rows[header_idx]]
    out = []
    for j in range(header_idx + 1, len(rows)):
        r = rows[j]
        if _row_is_empty(r):
            continue
        rowdict = {}
        for k, h in enumerate(headers):
            if not h:
                continue
            rowdict[h] = r[k] if k < len(r) else None
        rowdict["_row"] = j + 1  # 1-based, matches spreadsheet numbering
        out.append(rowdict)
    return out


def _parse_import_xlsx(content: bytes):
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    test_rows = _read_xlsx_sheet(wb, ["Tests"], fallback_first=True)
    param_rows = _read_xlsx_sheet(wb, ["Parameters"])
    return test_rows, param_rows


def _parse_import_csv(content: bytes):
    """CSV is tests-only (single sheet). Parameters require the .xlsx template."""
    import csv
    import io
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    test_rows = []
    for i, raw in enumerate(reader):
        rowdict = {_norm_header(k): v for k, v in raw.items() if k is not None}
        rowdict["_row"] = i + 2  # header is line 1
        test_rows.append(rowdict)
    return test_rows, []


def _build_params_for_test(param_rows_for_test, errors):
    """Collapse parameter rows into parameter dicts; rows sharing a
    parameter_name accumulate multiple reference_ranges (in order)."""
    ordered = []
    by_name = {}
    for pr in param_rows_for_test:
        pname = _cell_str(pr.get("parameter_name"))
        rownum = pr.get("_row", 0)
        if not pname:
            errors.append(ImportRowError(sheet="Parameters", row=rownum, message="Missing parameter_name"))
            continue
        key = pname.lower()
        if key not in by_name:
            field_type = (_cell_str(pr.get("field_type")) or "numeric").lower()
            if field_type not in _VALID_FIELD_TYPES:
                errors.append(ImportRowError(
                    sheet="Parameters", row=rownum,
                    message=f"Invalid field_type '{field_type}' for '{pname}' — using 'numeric'",
                ))
                field_type = "numeric"
            by_name[key] = {
                "name": pname,
                "unit": _cell_str(pr.get("unit")),
                "method": _cell_str(pr.get("method")),
                "section": _cell_str(pr.get("section")),
                "field_type": field_type,
                "possible_values": _cell_list(pr.get("possible_values")),
                "normal_value": _cell_str(pr.get("normal_value")),
                "abnormal_values": _cell_list(pr.get("abnormal_values")),
                "reference_ranges": [],
            }
            try:
                by_name[key]["critical_low"] = _cell_float(pr.get("critical_low"))
                by_name[key]["critical_high"] = _cell_float(pr.get("critical_high"))
            except ValueError as e:
                errors.append(ImportRowError(sheet="Parameters", row=rownum, message=f"{pname}: {e}"))
                by_name[key]["critical_low"] = None
                by_name[key]["critical_high"] = None
            ordered.append(key)
        entry = by_name[key]
        try:
            rmin = _cell_float(pr.get("ref_min"))
            rmax = _cell_float(pr.get("ref_max"))
            age_min = _cell_float(pr.get("age_min"))
            age_max = _cell_float(pr.get("age_max"))
        except ValueError as e:
            errors.append(ImportRowError(sheet="Parameters", row=rownum, message=f"{pname}: {e}"))
            continue
        gender = (_cell_str(pr.get("gender")) or "common").lower()
        desc = _cell_str(pr.get("description"))
        if rmin is not None or rmax is not None or desc:
            entry["reference_ranges"].append({
                "min": rmin, "max": rmax, "gender": gender,
                "age_min": age_min, "age_max": age_max,
                "description": desc or "",
            })
    result = []
    for key in ordered:
        e = by_name[key]
        if not e["reference_ranges"]:
            e["reference_ranges"] = None
        result.append(e)
    return result


@router.get("/tests/import/template")
async def download_import_template(
    current_user: User = Depends(require_permission(Modules.LAB, "read")),
    db: Session = Depends(get_db),
):
    """Download a ready-to-fill .xlsx template (Tests + Parameters sheets)."""
    _require_lab_admin(current_user)
    import io
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tests"
    ws.append(_IMPORT_TEST_HEADERS)
    ws.append(["CBC", "Complete Blood Count", "Hematology", "Blood (EDTA)", 300,
               "Automated Analyzer", "Full blood count", "None"])
    ws.append(["LFT", "Liver Function Test", "Biochemistry", "Blood (Serum)", 600,
               "Colorimetric", "Liver panel", "Fasting 8-12 hours"])

    ws2 = wb.create_sheet("Parameters")
    ws2.append(_IMPORT_PARAM_HEADERS)
    ws2.append(["CBC", "", "Hemoglobin", "g/dL", "", "numeric", 13, 17, "male", "", "",
                "", "", "", "", "", ""])
    ws2.append(["CBC", "", "Hemoglobin", "g/dL", "", "numeric", 12, 15, "female", "", "",
                "", "", "", "", "", ""])
    ws2.append(["LFT", "", "SGPT (ALT)", "U/L", "", "numeric", 7, 56, "common", "", "",
                "", "", "", "", "", ""])

    ws3 = wb.create_sheet("Instructions")
    notes = [
        ["KT HEALTH ERP — Lab Test Import Template"],
        [""],
        ["Fill the 'Tests' sheet — one row per test."],
        ["  Required columns: test_code, name, category, cost."],
        ["  Categories and sample types are created automatically if they don't exist."],
        [""],
        ["Fill the 'Parameters' sheet (optional) — one row per reference range."],
        ["  Link to a test using the same test_code."],
        ["  Repeat parameter_name on multiple rows to add several reference ranges"],
        ["  (e.g. one row per gender or age band)."],
        [""],
        ["gender values: male, female, common"],
        ["field_type values: " + ", ".join(sorted(_VALID_FIELD_TYPES))],
        ["possible_values / abnormal_values: comma-separated (e.g. Positive, Negative)"],
    ]
    for row in notes:
        ws3.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=lab_tests_import_template.xlsx"},
    )


@router.get("/tests/import/sample-csv")
async def download_import_sample_csv(
    current_user: User = Depends(require_permission(Modules.LAB, "read")),
    db: Session = Depends(get_db),
):
    """Download a sample CSV (tests only) that can be imported as-is."""
    _require_lab_admin(current_user)
    import csv
    import io
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_IMPORT_TEST_HEADERS)
    writer.writerow(["CBC", "Complete Blood Count", "Hematology", "Blood (EDTA)", "300",
                     "Automated Analyzer", "Full blood count", "None"])
    writer.writerow(["LFT", "Liver Function Test", "Biochemistry", "Blood (Serum)", "600",
                     "Colorimetric", "Liver panel", "Fasting 8-12 hours"])
    writer.writerow(["RFT", "Renal Function Test", "Biochemistry", "Blood (Serum)", "550",
                     "Colorimetric", "Kidney panel", "None"])
    writer.writerow(["TSH", "Thyroid Stimulating Hormone", "Endocrinology", "Blood (Serum)",
                     "400", "CLIA", "Thyroid screen", "None"])
    data = buf.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=lab_tests_sample.csv"},
    )


@router.post("/tests/import", response_model=ImportSummary)
async def import_tests(
    file: UploadFile = File(...),
    dry_run: bool = Form(False),
    on_duplicate: str = Form("skip"),
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db),
):
    """Bulk-import lab tests (and optional parameters) from .xlsx or .csv.

    - Auto-creates missing categories and sample types.
    - Validates each row; valid rows import while invalid rows are reported.
    - ``dry_run=True`` validates and previews without writing anything.
    - ``on_duplicate`` = ``skip`` (default) or ``update`` for existing test_code.
    """
    _require_lab_admin(current_user)
    if on_duplicate not in ("skip", "update"):
        on_duplicate = "skip"

    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    try:
        if filename.endswith(".csv"):
            test_rows, param_rows = _parse_import_csv(content)
        elif filename.endswith(".xlsx"):
            test_rows, param_rows = _parse_import_xlsx(content)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type. Upload a .xlsx or .csv file.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read file: {e}")

    hospital_id = current_user.hospital_id
    categories = {
        c.name.strip().lower(): c
        for c in db.query(LabTestCategory).filter(LabTestCategory.hospital_id == hospital_id).all()
    }
    sample_types = {
        s.name.strip().lower(): s
        for s in db.query(SampleType).filter(SampleType.hospital_id == hospital_id).all()
    }

    # Group parameter rows by test_code (case-insensitive)
    params_by_test = {}
    for pr in param_rows:
        tc = _cell_str(pr.get("test_code"))
        if not tc or tc.startswith("#"):
            continue
        params_by_test.setdefault(tc.lower(), []).append(pr)

    errors: List[ImportRowError] = []
    preview: List[ImportPreviewRow] = []
    categories_created: List[str] = []
    sample_types_created: List[str] = []
    created = updated = skipped = 0
    total_rows = 0

    for tr in test_rows:
        rownum = tr.get("_row", 0)
        test_code = _cell_str(tr.get("test_code"))
        if not test_code or test_code.startswith("#"):
            continue
        total_rows += 1

        name = _cell_str(tr.get("name"))
        cat_name = _cell_str(tr.get("category"))
        row_errs = []
        if not name:
            row_errs.append("Missing test name")
        if not cat_name:
            row_errs.append("Missing category")
        cost = 0.0
        try:
            parsed_cost = _cell_float(tr.get("cost"))
            if parsed_cost is not None:
                cost = parsed_cost
        except ValueError:
            row_errs.append("Cost must be a number")
        if cost < 0:
            row_errs.append("Cost cannot be negative")

        if row_errs:
            msg = "; ".join(row_errs)
            errors.append(ImportRowError(sheet="Tests", row=rownum, message=msg))
            preview.append(ImportPreviewRow(
                row=rownum, test_code=test_code, name=name or "",
                category=cat_name, status="error", message=msg,
            ))
            continue

        existing = db.query(LabTest).filter(
            LabTest.test_code == test_code,
            LabTest.hospital_id == hospital_id,
        ).first()
        if existing and on_duplicate == "skip":
            skipped += 1
            preview.append(ImportPreviewRow(
                row=rownum, test_code=test_code, name=name, category=cat_name,
                status="skip", message="Test code already exists",
            ))
            continue

        # Resolve / auto-create category
        cat = categories.get(cat_name.lower())
        if not cat:
            cat = LabTestCategory(name=cat_name, hospital_id=hospital_id)
            db.add(cat)
            db.flush()
            categories[cat_name.lower()] = cat
            categories_created.append(cat_name)

        # Resolve / auto-create sample type
        st_name = _cell_str(tr.get("sample_type"))
        st = None
        if st_name:
            st = sample_types.get(st_name.lower())
            if not st:
                st = SampleType(name=st_name, hospital_id=hospital_id)
                db.add(st)
                db.flush()
                sample_types[st_name.lower()] = st
                sample_types_created.append(st_name)

        p_list = _build_params_for_test(params_by_test.get(test_code.lower(), []), errors)

        if existing:  # on_duplicate == "update"
            existing.name = name
            existing.description = _cell_str(tr.get("description"))
            existing.category_id = cat.id
            existing.cost = cost
            existing.method = _cell_str(tr.get("method"))
            existing.preparation_instructions = _cell_str(tr.get("preparation_instructions"))
            if st:
                existing.sample_type_id = st.id
                existing.sample_type = st.name
            if p_list:
                db.query(LabTestParameter).filter(LabTestParameter.test_id == existing.id).delete()
                for i, p in enumerate(p_list):
                    db.add(_create_seed_parameter(existing.id, p, i))
            updated += 1
            preview.append(ImportPreviewRow(
                row=rownum, test_code=test_code, name=name, category=cat_name,
                status="update", parameter_count=len(p_list),
            ))
        else:
            test = LabTest(
                test_code=test_code, name=name, description=_cell_str(tr.get("description")),
                category_id=cat.id, cost=cost,
                sample_type=(st.name if st else None),
                sample_type_id=(st.id if st else None),
                method=_cell_str(tr.get("method")),
                preparation_instructions=_cell_str(tr.get("preparation_instructions")),
                hospital_id=hospital_id,
            )
            db.add(test)
            db.flush()
            for i, p in enumerate(p_list):
                db.add(_create_seed_parameter(test.id, p, i))
            created += 1
            preview.append(ImportPreviewRow(
                row=rownum, test_code=test_code, name=name, category=cat_name,
                status="new", parameter_count=len(p_list),
            ))

    if dry_run:
        db.rollback()
    else:
        db.commit()
        try:
            from app.services.audit_service import log_action
            log_action(
                db, current_user, "import_lab_tests", "lab", "LabTest", None,
                description=f"Imported lab tests: {created} created, {updated} updated, {skipped} skipped",
                details={"created": created, "updated": updated, "skipped": skipped,
                         "categories_created": categories_created,
                         "sample_types_created": sample_types_created},
            )
        except Exception:
            pass

    return ImportSummary(
        dry_run=dry_run,
        total_rows=total_rows,
        created=created,
        updated=updated,
        skipped=skipped,
        error_count=len(errors),
        categories_created=categories_created,
        sample_types_created=sample_types_created,
        errors=errors,
        preview=preview,
    )


@router.post("/seed-defaults")
async def seed_default_tests(
    current_user: User = Depends(require_permission(Modules.LAB, "write")),
    db: Session = Depends(get_db)
):
    """Seed default lab tests with parameters for the hospital"""
    _require_lab_admin(current_user)
    hospital_id = current_user.hospital_id

    # Check if already seeded
    existing = db.query(LabTest).filter(LabTest.hospital_id == hospital_id).count()
    if existing > 0:
        raise HTTPException(status_code=400, detail="Lab tests already exist for this hospital. Delete existing tests first or add manually.")

    seed_data = _get_seed_data()
    created_tests = 0

    for cat_name, tests in seed_data.items():
        cat = LabTestCategory(name=cat_name, hospital_id=hospital_id)
        db.add(cat)
        db.flush()

        for test_info in tests:
            test = LabTest(
                test_code=test_info["code"], name=test_info["name"],
                description=test_info.get("description", ""),
                category_id=cat.id, cost=test_info.get("cost", 0),
                sample_type=test_info.get("sample_type", "Blood"),
                method=test_info.get("method"),
                preparation_instructions=test_info.get("instructions", ""),
                hospital_id=hospital_id
            )
            db.add(test)
            db.flush()

            for i, param in enumerate(test_info.get("parameters", [])):
                db.add(_create_seed_parameter(test.id, param, i))
            created_tests += 1

    db.commit()
    return {"message": f"Seeded {created_tests} lab tests with parameters"}


def _get_seed_data():
    return {
        "Biochemistry": [
            {
                "code": "LFT", "name": "Liver Function Test (LFT)",
                "description": "Assesses liver health",
                "cost": 600,
                "sample_type": "Blood (Serum)",
                "instructions": "Fasting 8-12 hours preferred",
                "parameters": [
                    {"name": "Total Bilirubin", "unit": "mg/dL", "reference_ranges": [{"min": 0.1, "max": 1.2}]},
                    {"name": "Direct Bilirubin", "unit": "mg/dL", "reference_ranges": [{"min": 0.0, "max": 0.3}]},
                    {"name": "Indirect Bilirubin", "unit": "mg/dL", "reference_ranges": [{"min": 0.1, "max": 0.9}]},
                    {"name": "SGOT (AST)", "unit": "U/L", "reference_ranges": [{"min": 5.0, "max": 40.0}]},
                    {"name": "SGPT (ALT)", "unit": "U/L", "reference_ranges": [{"min": 7.0, "max": 56.0}]},
                    {"name": "Alkaline Phosphatase (ALP)", "unit": "U/L", "reference_ranges": [{"min": 44.0, "max": 147.0}]},
                    {"name": "Total Protein", "unit": "g/dL", "reference_ranges": [{"min": 6.0, "max": 8.3}]},
                    {"name": "Albumin", "unit": "g/dL", "reference_ranges": [{"min": 3.5, "max": 5.5}]},
                    {"name": "Globulin", "unit": "g/dL", "reference_ranges": [{"min": 2.0, "max": 3.5}]},
                ]
            },
            {
                "code": "LIPID", "name": "Lipid Profile",
                "description": "Measures cholesterol and triglyceride levels",
                "cost": 500,
                "sample_type": "Blood (Serum)",
                "instructions": "Fasting 12 hours required",
                "parameters": [
                    {"name": "Total Cholesterol", "unit": "mg/dL", "reference_ranges": [{"min": 0.0, "max": 200.0}]},
                    {"name": "Triglycerides", "unit": "mg/dL", "reference_ranges": [{"min": 0.0, "max": 150.0}]},
                    {"name": "HDL Cholesterol", "unit": "mg/dL", "reference_ranges": [{"min": 40.0, "max": 999.0, "gender": "male"}, {"min": 50.0, "max": 999.0, "gender": "female"}]},
                    {"name": "LDL Cholesterol", "unit": "mg/dL", "reference_ranges": [{"min": 0.0, "max": 100.0}]},
                    {"name": "VLDL Cholesterol", "unit": "mg/dL", "reference_ranges": [{"min": 5.0, "max": 40.0}]},
                ]
            },
            {
                "code": "RFT", "name": "Renal Function Test (RFT)",
                "description": "Evaluates kidney function",
                "cost": 500,
                "sample_type": "Blood (Serum)",
                "instructions": "Fasting 8-12 hours preferred",
                "parameters": [
                    {"name": "Blood Urea", "unit": "mg/dL", "reference_ranges": [{"min": 15.0, "max": 40.0}]},
                    {"name": "Serum Creatinine", "unit": "mg/dL", "reference_ranges": [{"min": 0.7, "max": 1.3, "gender": "male"}, {"min": 0.6, "max": 1.1, "gender": "female"}]},
                    {"name": "Uric Acid", "unit": "mg/dL", "reference_ranges": [{"min": 3.5, "max": 7.2, "gender": "male"}, {"min": 2.6, "max": 6.0, "gender": "female"}]},
                    {"name": "BUN", "unit": "mg/dL", "reference_ranges": [{"min": 7.0, "max": 20.0}]},
                    {"name": "Sodium", "unit": "mEq/L", "reference_ranges": [{"min": 136.0, "max": 145.0}]},
                    {"name": "Potassium", "unit": "mEq/L", "reference_ranges": [{"min": 3.5, "max": 5.1}]},
                    {"name": "Chloride", "unit": "mEq/L", "reference_ranges": [{"min": 98.0, "max": 106.0}]},
                ]
            },
        ],
        "Blood Sugar": [
            {
                "code": "FBS", "name": "Fasting Blood Sugar",
                "description": "Measures blood glucose after fasting",
                "cost": 100,
                "sample_type": "Blood (Fluoride)",
                "instructions": "Fasting 8-12 hours required",
                "parameters": [
                    {"name": "Fasting Blood Glucose", "unit": "mg/dL", "reference_ranges": [{"min": 70.0, "max": 100.0}]},
                ]
            },
            {
                "code": "HBA1C", "name": "Glycated Hemoglobin (HbA1c)",
                "description": "Average blood sugar over 2-3 months",
                "cost": 450,
                "sample_type": "Blood (EDTA)",
                "instructions": "No special preparation",
                "parameters": [
                    {"name": "HbA1c", "unit": "%", "reference_ranges": [{"min": 4.0, "max": 5.6}]},
                ]
            },
            {
                "code": "PPBS", "name": "Post Prandial Blood Sugar",
                "description": "Measures blood glucose 2 hours after eating",
                "cost": 100,
                "sample_type": "Blood (Fluoride)",
                "instructions": "2 hours after meal",
                "parameters": [
                    {"name": "PP Blood Glucose", "unit": "mg/dL", "reference_ranges": [{"min": 70.0, "max": 140.0}]},
                ]
            },
            {
                "code": "RBS", "name": "Random Blood Sugar",
                "description": "Random blood glucose measurement",
                "cost": 80,
                "sample_type": "Blood (Fluoride)",
                "instructions": "No special preparation",
                "parameters": [
                    {"name": "Random Blood Glucose", "unit": "mg/dL", "reference_ranges": [{"min": 70.0, "max": 140.0}]},
                ]
            },
        ],
        "Hematology": [
            {
                "code": "CBC", "name": "Complete Blood Count (CBC)",
                "description": "Measures different components of blood",
                "cost": 350,
                "sample_type": "Blood (EDTA)",
                "instructions": "No special preparation",
                "parameters": [
                    {"name": "Hemoglobin", "unit": "g/dL", "reference_ranges": [{"min": 13.0, "max": 17.0, "gender": "male"}, {"min": 12.0, "max": 16.0, "gender": "female"}]},
                    {"name": "RBC Count", "unit": "million/\u00b5L", "reference_ranges": [{"min": 4.5, "max": 5.5, "gender": "male"}, {"min": 4.0, "max": 5.0, "gender": "female"}]},
                    {"name": "WBC Count", "unit": "cells/\u00b5L", "reference_ranges": [{"min": 4000.0, "max": 11000.0}]},
                    {"name": "Platelet Count", "unit": "lakh/\u00b5L", "reference_ranges": [{"min": 1.5, "max": 4.0}]},
                    {"name": "PCV / Hematocrit", "unit": "%", "reference_ranges": [{"min": 40.0, "max": 50.0, "gender": "male"}, {"min": 36.0, "max": 44.0, "gender": "female"}]},
                    {"name": "MCV", "unit": "fL", "reference_ranges": [{"min": 80.0, "max": 100.0}]},
                    {"name": "MCH", "unit": "pg", "reference_ranges": [{"min": 27.0, "max": 33.0}]},
                    {"name": "MCHC", "unit": "g/dL", "reference_ranges": [{"min": 32.0, "max": 36.0}]},
                    {"name": "ESR", "unit": "mm/hr", "reference_ranges": [{"min": 0.0, "max": 15.0, "gender": "male"}, {"min": 0.0, "max": 20.0, "gender": "female"}]},
                ]
            },
        ],
        "Thyroid": [
            {
                "code": "THYROID", "name": "Thyroid Profile",
                "description": "Evaluates thyroid gland function",
                "cost": 700,
                "sample_type": "Blood (Serum)",
                "instructions": "No special preparation",
                "parameters": [
                    {"name": "T3 (Triiodothyronine)", "unit": "ng/dL", "reference_ranges": [{"min": 80.0, "max": 200.0}]},
                    {"name": "T4 (Thyroxine)", "unit": "\u00b5g/dL", "reference_ranges": [{"min": 4.5, "max": 12.5}]},
                    {"name": "TSH", "unit": "\u00b5IU/mL", "reference_ranges": [{"min": 0.4, "max": 4.0}]},
                ]
            },
        ],
        "Urine Analysis": [
            {
                "code": "URINE-R", "name": "Urine Routine & Microscopy",
                "description": "Physical, chemical and microscopic examination of urine",
                "cost": 150,
                "sample_type": "Urine (Mid-stream)",
                "instructions": "Mid-stream clean catch sample",
                "parameters": [
                    {"name": "Color", "field_type": "text"},
                    {"name": "Appearance", "field_type": "select", "possible_values": ["Clear", "Slightly Turbid", "Turbid"]},
                    {"name": "pH", "reference_ranges": [{"min": 4.5, "max": 8.0}]},
                    {"name": "Specific Gravity", "reference_ranges": [{"min": 1.005, "max": 1.03}]},
                    {"name": "Protein", "field_type": "select", "possible_values": ["Nil", "Trace", "+", "++", "+++"]},
                    {"name": "Glucose", "field_type": "select", "possible_values": ["Nil", "Trace", "+", "++", "+++"]},
                    {"name": "Ketones", "field_type": "select", "possible_values": ["Nil", "Trace", "+", "++", "+++"]},
                    {"name": "RBC", "unit": "/HPF", "reference_ranges": [{"min": 0.0, "max": 2.0}]},
                    {"name": "Pus Cells (WBC)", "unit": "/HPF", "reference_ranges": [{"min": 0.0, "max": 5.0}]},
                    {"name": "Epithelial Cells", "unit": "/HPF", "field_type": "select", "possible_values": ["Few", "Moderate", "Many"]},
                ]
            },
        ],
    }
