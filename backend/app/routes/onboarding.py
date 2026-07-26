"""Optional, guided post-install hospital onboarding."""
from __future__ import annotations

import io
import zipfile
from typing import Literal

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.models.permissions import HospitalSettings
from app.models.user import User
from app.services import onboarding_import as importers
from app.services.onboarding_state import (
    SETUP_CATEGORY,
    get_enabled_module_names,
    get_onboarding_status,
    mark_step,
    set_json_setting,
)
from app.utils.dependencies import get_current_user
from config.database import get_db


router = APIRouter()
UPLOAD_MAX_BYTES = 5 * 1024 * 1024


def _require_admin(user: User) -> None:
    if not any(role in user.role_names for role in ("super_admin", "hospital_admin")):
        raise HTTPException(status_code=403, detail="Admin access required")


def _require_module(db: Session, module_name: str) -> None:
    """Reject setup actions for modules that aren't currently available.

    Templates and steps are already hidden in the UI when a module is off, but
    the import endpoints must also refuse direct calls so the setup flow stays
    consistent with the hospital's licensed/enabled modules.
    """
    if module_name not in get_enabled_module_names(db):
        raise HTTPException(
            status_code=403,
            detail=f"The '{module_name}' module is not enabled for this hospital",
        )


class StepUpdate(BaseModel):
    status: Literal["completed", "skipped", "pending"]


class DepartmentUpdate(BaseModel):
    names: list[str] = Field(default_factory=list, max_length=100)


class NursingRateItem(BaseModel):
    room_type: str
    nursing_charge_per_visit: float = Field(ge=0)


class NursingRatesBulk(BaseModel):
    rates: list[NursingRateItem]


TEMPLATES = {
    "doctors": {
        "label": "Doctors",
        "filename": "doctors_setup_template.xlsx",
        "module": None,
        "headers": [
            "username", "email", "first_name", "last_name", "password", "phone",
            "specialization", "license_number", "qualification",
            "consultation_fee_inr", "inpatient_fee_inr", "emergency_fee_inr",
            "experience_years", "default_consultation_duration",
        ],
        "examples": [
            ["drravi", "drravi@hospital.in", "Ravi", "Kumar", "Welcome@123",
             "9876543210", "Cardiology", "MCI-12345", "MBBS MD", 800, 2000,
             1500, 12, 15],
        ],
        "instructions": [
            "Required: username, email, first_name, last_name, password, specialization, license_number.",
            "Set inpatient_fee_inr for inpatient doctor visits (used as the default IP visit rate).",
            "Upload from Users & Roles > Bulk Import.",
        ],
    },
    "nurses": {
        "label": "Nurses",
        "filename": "nurses_setup_template.xlsx",
        "module": None,
        "headers": ["username", "email", "first_name", "last_name", "password", "phone"],
        "examples": [
            ["asha", "asha.n@hospital.in", "Asha", "Menon", "Welcome@123", "9876543220"],
        ],
        "instructions": [
            "All columns except phone are required.",
            "Upload from Users & Roles > Bulk Import.",
        ],
    },
    "staff": {
        "label": "Other staff",
        "filename": "staff_setup_template.xlsx",
        "module": None,
        "headers": [
            "username", "email", "first_name", "last_name", "role", "password",
            "phone", "additional_roles",
        ],
        "examples": [
            ["reception1", "reception@hospital.in", "Front", "Desk", "receptionist",
             "Welcome@123", "9876543211", "frontdesk"],
        ],
        "instructions": [
            "Required: username, email, first_name, last_name, role and password.",
            "Separate multiple additional_roles with semicolons.",
        ],
    },
    "lab_tests": {
        "label": "Laboratory tests",
        "filename": "lab_tests_setup_template.xlsx",
        "module": "lab",
        "headers": [
            "test_code", "name", "category", "sample_type", "cost", "method",
            "description", "preparation_instructions",
        ],
        "examples": [
            ["CBC", "Complete Blood Count", "Hematology", "Blood (EDTA)", 300,
             "Automated Analyzer", "Full blood count", "None"],
        ],
        "instructions": [
            "Required: test_code, name, category and cost.",
            "Import from the Laboratory module or the detailed lab template.",
        ],
    },
    "rooms": {
        "label": "Rooms and beds",
        "filename": "rooms_setup_template.xlsx",
        "module": "inpatient",
        "multi_sheet": True,
        "sheets": {
            "Rooms": {
                "headers": [
                    "room_number", "room_type", "floor", "department", "ward",
                    "bed_count", "room_charge_per_day", "nursing_charge_per_visit",
                    "amenities", "is_isolation", "gender_policy",
                ],
                "examples": [
                    ["G-101", "general", "1", "Medicine", "Male Ward", 4, 1500, 100,
                     "fan;call_bell", "false", "male"],
                    ["ICU-1", "icu", "2", "Critical Care", "ICU", 1, 8000, 350,
                     "ac;cardiac_monitor;oxygen_point", "false", "mixed"],
                ],
            },
            "Beds": {
                "headers": ["room_number", "bed_label"],
                "examples": [
                    ["G-101", "A"],
                    ["G-101", "B"],
                    ["G-101", "C"],
                    ["G-101", "D"],
                    ["ICU-1", "Bed-1"],
                ],
            },
        },
        "instructions": [
            "Fill the Rooms sheet. room_type must be one of: general, semi_private, private, suite, icu, hdu, nicu, picu, isolation, labour, recovery, daycare, emergency, operation.",
            "If the Beds sheet is empty for a room, beds are created automatically as Bed-1..Bed-N from bed_count.",
            "amenities: semicolon-separated keys such as ac, tv, wifi, oxygen_point.",
            "gender_policy: mixed, male or female.",
            "Upload from Guided Setup > Rooms and beds.",
        ],
    },
    "nursing_rates": {
        "label": "Nursing rates by room type",
        "filename": "nursing_rates_setup_template.xlsx",
        "module": "inpatient",
        "headers": ["room_type", "nursing_charge_per_visit"],
        "examples": [
            ["general", 100],
            ["semi_private", 150],
            ["private", 200],
            ["icu", 350],
            ["hdu", 250],
        ],
        "instructions": [
            "room_type must match the hospital room-type keys.",
            "nursing_charge_per_visit is the default nurse visit charge for that room type.",
            "Upload from Guided Setup, or edit the on-screen grid.",
        ],
    },
    "ancillary_services": {
        "label": "Ancillary services",
        "filename": "ancillary_services_setup_template.xlsx",
        "module": "inpatient",
        "headers": [
            "service_name", "service_code", "category", "default_charge",
            "charge_unit", "description",
        ],
        "examples": [
            ["Chest X-Ray", "XR-CHEST", "imaging", 500, "per_session", "PA view"],
            ["Oxygen", "O2-DAY", "oxygen", 800, "per_day", "Cylinder / piped oxygen"],
            ["Physiotherapy session", "PT-01", "physiotherapy", 400, "per_session", ""],
        ],
        "instructions": [
            "category: imaging, physiotherapy, dialysis, oxygen, equipment, consumable, procedure, other.",
            "charge_unit: per_session, per_hour, per_day, per_unit.",
            "Upload from Guided Setup > Ancillary services.",
        ],
    },
    "doctor_room_rates": {
        "label": "Doctor room-type visit rates",
        "filename": "doctor_room_rates_setup_template.xlsx",
        "module": "inpatient",
        "headers": ["doctor_username", "room_type", "visit_rate"],
        "examples": [
            ["drravi", "general", 800],
            ["drravi", "icu", 1500],
            ["drmeera", "private", 1000],
        ],
        "instructions": [
            "doctor_username must match an existing doctor account.",
            "Optional overrides when a doctor's inpatient_fee_inr is not enough.",
            "Upload from Guided Setup > Doctor inpatient visit rates.",
        ],
    },
    "opd_procedures": {
        "label": "OPD / day-care procedures",
        "filename": "opd_procedures_setup_template.xlsx",
        "module": "outpatient",
        "headers": ["name", "code", "category", "default_price", "description", "is_active"],
        "examples": [
            ["Dressing", "OPD-DR", "Nursing", 200, "Simple wound dressing", "true"],
            ["IV fluids", "OPD-IV", "Infusion", 350, "Including set", "true"],
        ],
        "instructions": [
            "Required: name and default_price.",
            "Upload from Guided Setup > OPD / day-care procedures.",
        ],
    },
    "pharmacy_medicines": {
        "label": "Pharmacy medicines",
        "filename": "pharmacy_medicines_setup_template.xlsx",
        "module": "pharmacy",
        "builder": "medicines",
    },
    "pharmacy_suppliers": {
        "label": "Pharmacy suppliers",
        "filename": "pharmacy_suppliers_setup_template.xlsx",
        "module": "pharmacy",
        "builder": "suppliers",
    },
    "pharmacy_masters": {
        "label": "Pharmacy masters",
        "filename": "pharmacy_masters_setup_template.xlsx",
        "module": "pharmacy",
        "builder": "masters",
    },
    "pharmacy_opening_stock": {
        "label": "Pharmacy opening stock",
        "filename": "pharmacy_opening_stock_setup_template.xlsx",
        "module": "pharmacy",
        "builder": "opening_stock",
    },
}


def _visible_templates(db: Session) -> dict:
    enabled = get_enabled_module_names(db)
    return {
        key: value
        for key, value in TEMPLATES.items()
        if not value.get("module") or value["module"] in enabled
    }


def _workbook_bytes(template_key: str) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    template = TEMPLATES[template_key]
    builder_key = template.get("builder")
    if builder_key:
        from app.services.pharmacy_import import (
            build_masters_template,
            build_medicines_template,
            build_opening_stock_template,
            build_suppliers_template,
        )
        builders = {
            "medicines": build_medicines_template,
            "suppliers": build_suppliers_template,
            "masters": build_masters_template,
            "opening_stock": build_opening_stock_template,
        }
        return builders[builder_key]()

    workbook = openpyxl.Workbook()

    if template.get("multi_sheet"):
        first = True
        for sheet_name, sheet_def in template["sheets"].items():
            sheet = workbook.active if first else workbook.create_sheet(sheet_name)
            if first:
                sheet.title = sheet_name
                first = False
            sheet.append(sheet_def["headers"])
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2563EB")
            for row in sheet_def["examples"]:
                sheet.append(row)
            sheet.freeze_panes = "A2"
            for column in sheet.columns:
                letter = column[0].column_letter
                sheet.column_dimensions[letter].width = min(
                    36, max(14, max(len(str(cell.value or "")) for cell in column) + 2)
                )
    else:
        sheet = workbook.active
        sheet.title = "Data"
        sheet.append(template["headers"])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
        for row in template["examples"]:
            sheet.append(row)
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(
                36, max(14, max(len(str(cell.value or "")) for cell in column) + 2)
            )

    notes = workbook.create_sheet("Instructions")
    notes.append([f"KT HEALTH ERP — {template['label']} setup template"])
    notes["A1"].font = Font(bold=True, size=14)
    notes.append([])
    for instruction in template["instructions"]:
        notes.append([instruction])
    notes.column_dimensions["A"].width = 110

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


async def _read_upload(file: UploadFile) -> bytes:
    raw = await file.read()
    if len(raw) > UPLOAD_MAX_BYTES:
        raise HTTPException(status_code=413, detail="Import file too large (max 5 MB)")
    name = (file.filename or "").lower()
    if not (name.endswith(".xlsx") or name.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Upload an .xlsx or .csv file")
    return raw


@router.get("/status")
async def onboarding_status(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)
    return get_onboarding_status(db)


@router.put("/steps/{step_key}")
async def update_step(
    step_key: str,
    data: StepUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)
    try:
        mark_step(
            db,
            step_key,
            completed=data.status == "completed",
            skipped=data.status == "skipped",
            user_id=current_user.id,
        )
    except KeyError:
        raise HTTPException(status_code=404, detail="Unknown onboarding step")
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="This step is completed automatically when its required data is configured",
        )
    db.commit()
    return get_onboarding_status(db)


@router.put("/departments")
async def update_departments(
    data: DepartmentUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)
    names = []
    seen = set()
    for raw in data.names:
        name = raw.strip()
        if name and name.casefold() not in seen:
            names.append(name)
            seen.add(name.casefold())
    set_json_setting(
        db,
        "departments",
        names,
        user_id=current_user.id,
        description="Managed department and ward name suggestions",
    )
    db.commit()
    return get_onboarding_status(db)


@router.put("/nursing-rates")
async def save_nursing_rates(
    data: NursingRatesBulk,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Bulk upsert room-type nursing rates from the setup wizard grid."""
    _require_admin(current_user)
    _require_module(db, "inpatient")
    from app.models.hospital import Hospital
    from app.models.inpatient import RoomTypeRateConfig

    hospital = db.query(Hospital).filter(Hospital.is_active.is_(True)).first()
    if not hospital:
        raise HTTPException(status_code=404, detail="Hospital not found")

    for item in data.rates:
        room_type = item.room_type.strip().lower()
        if room_type not in importers.ROOM_TYPES:
            raise HTTPException(status_code=400, detail=f"Invalid room_type '{room_type}'")
        row = (
            db.query(RoomTypeRateConfig)
            .filter(
                RoomTypeRateConfig.hospital_id == hospital.id,
                RoomTypeRateConfig.room_type == room_type,
            )
            .first()
        )
        if row:
            row.nursing_charge_per_visit = item.nursing_charge_per_visit
        else:
            db.add(
                RoomTypeRateConfig(
                    hospital_id=hospital.id,
                    room_type=room_type,
                    nursing_charge_per_visit=item.nursing_charge_per_visit,
                )
            )
    db.commit()
    return get_onboarding_status(db)


@router.post("/dismiss")
async def dismiss_onboarding(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)
    set_json_setting(db, "dismissed", True, user_id=current_user.id)
    db.commit()
    return {"dismissed": True}


@router.post("/reset")
async def reset_onboarding(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)
    db.query(HospitalSettings).filter(
        HospitalSettings.setting_category == SETUP_CATEGORY,
        HospitalSettings.setting_key != "departments",
    ).delete(synchronize_session=False)
    db.commit()
    return get_onboarding_status(db)


@router.get("/templates")
async def list_templates(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)
    return [
        {"key": key, "label": value["label"], "filename": value["filename"], "module": value.get("module")}
        for key, value in _visible_templates(db).items()
    ]


@router.get("/templates/all.zip")
async def download_all_templates(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for key, template in _visible_templates(db).items():
            archive.writestr(template["filename"], _workbook_bytes(key))
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="kthealth_setup_templates.zip"'},
    )


@router.get("/templates/{template_key}")
async def download_template(
    template_key: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)
    visible = _visible_templates(db)
    template = visible.get(template_key)
    if not template:
        raise HTTPException(status_code=404, detail="Unknown setup template")
    return StreamingResponse(
        io.BytesIO(_workbook_bytes(template_key)),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{template["filename"]}"'},
    )


@router.post("/import/rooms")
async def import_rooms(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)
    _require_module(db, "inpatient")
    raw = await _read_upload(file)
    try:
        return importers.import_rooms(db, raw, file.filename or "")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/import/nursing-rates")
async def import_nursing_rates(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)
    _require_module(db, "inpatient")
    raw = await _read_upload(file)
    try:
        return importers.import_nursing_rates(db, raw, file.filename or "")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/import/ancillary-services")
async def import_ancillary_services(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)
    _require_module(db, "inpatient")
    raw = await _read_upload(file)
    try:
        return importers.import_ancillary(db, raw, file.filename or "")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/import/doctor-room-rates")
async def import_doctor_room_rates(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)
    _require_module(db, "inpatient")
    raw = await _read_upload(file)
    try:
        return importers.import_doctor_room_rates(db, raw, file.filename or "")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/import/opd-procedures")
async def import_opd_procedures(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _require_admin(current_user)
    _require_module(db, "outpatient")
    raw = await _read_upload(file)
    try:
        return importers.import_opd_procedures(
            db, raw, file.filename or "", user_id=current_user.id
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
