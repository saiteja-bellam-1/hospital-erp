from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func as sql_func, cast, Date
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime, date, timedelta
from io import BytesIO
import os
import uuid
import json

from config.database import get_db
from app.models.user import User
from app.models.hospital import Hospital
from app.models.permissions import HospitalSettings
from app.models.patient import Patient
from app.models.outpatient import Appointment
from app.models.ehr import Consultation
from app.models.lab import PatientLabOrder, LabTest, LabTestCategory, LabTestPackage
from app.models.billing import Bill, BillItem, Payment
from app.models.inpatient import Admission
from app.models.pharmacy import PharmacySale
from app.utils.dependencies import get_current_user
from app.utils.pdf_settings import bill_pdf_gen_kwargs, pdf_gen_kwargs

from app.utils.paths import get_uploads_dir

UPLOAD_DIR = os.path.join(get_uploads_dir(), "module-config")

router = APIRouter()

# Pydantic models for API requests/responses
class HospitalInfoResponse(BaseModel):
    id: int
    hospital_id: str
    name: str
    address: Optional[str]
    city: Optional[str]
    state: Optional[str]
    postal_code: Optional[str]
    country: Optional[str]
    phone: Optional[str]
    fax: Optional[str]
    email: Optional[str]
    website: Optional[str]
    license_number: Optional[str]
    registration_number: Optional[str]
    tax_id: Optional[str]
    logo_url: Optional[str]
    description: Optional[str]
    established_date: Optional[datetime]
    mrn_prefix: Optional[str] = None
    is_active: bool

    class Config:
        from_attributes = True

class HospitalInfoUpdateRequest(BaseModel):
    name: Optional[str]
    address: Optional[str]
    city: Optional[str]
    state: Optional[str]
    postal_code: Optional[str]
    country: Optional[str]
    phone: Optional[str]
    fax: Optional[str]
    email: Optional[str]
    website: Optional[str]
    license_number: Optional[str]
    registration_number: Optional[str]
    tax_id: Optional[str]
    logo_url: Optional[str]
    description: Optional[str]
    established_date: Optional[datetime]

class DoctorProfileUpdateRequest(BaseModel):
    license_number: Optional[str]
    consultation_fee_inr: Optional[str]
    inpatient_fee_inr: Optional[str]
    emergency_fee_inr: Optional[str]
    specialization: Optional[str]
    qualification: Optional[str]
    experience_years: Optional[int]

class DoctorProfileResponse(BaseModel):
    id: int
    user_id: str
    username: str
    first_name: str
    last_name: str
    email: str
    phone: Optional[str]
    license_number: Optional[str]
    consultation_fee_inr: Optional[str]
    inpatient_fee_inr: Optional[str]
    emergency_fee_inr: Optional[str]
    specialization: Optional[str]
    qualification: Optional[str]
    experience_years: Optional[int]
    is_active: bool

def require_hospital_admin(current_user: User = Depends(get_current_user)):
    """Dependency to ensure only hospital admin or super admin can access these endpoints"""
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin']):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Hospital admin access required"
        )
    return current_user


def require_print_settings_editor(current_user: User = Depends(get_current_user)):
    """Hospital admin, super admin, or receptionist may edit print settings."""
    if not any(
        r in current_user.role_names
        for r in ['super_admin', 'hospital_admin', 'receptionist']
    ):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Print settings access required",
        )
    return current_user

# HOSPITAL INFORMATION ENDPOINTS
@router.get("/info", response_model=HospitalInfoResponse)
async def get_hospital_info(
    current_user: User = Depends(require_hospital_admin),
    db: Session = Depends(get_db)
):
    """Get hospital information"""
    # For single hospital system, get the first hospital
    hospital = db.query(Hospital).first()
    if not hospital:
        # Create default hospital if none exists
        hospital = Hospital(
            hospital_id="HOSP-001",
            name="General Hospital",
            address="123 Main Street",
            city="New York",
            state="NY",
            country="USA",
            phone="+1-555-0123",
            email="admin@generalhospital.com",
            license_number="LIC-2024-001",
            is_active=True
        )
        db.add(hospital)
        db.commit()
        db.refresh(hospital)
    
    return hospital

class HospitalInfoPartialUpdateRequest(BaseModel):
    name: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    postal_code: Optional[str] = None
    country: Optional[str] = None
    phone: Optional[str] = None
    fax: Optional[str] = None
    email: Optional[str] = None
    website: Optional[str] = None
    license_number: Optional[str] = None
    registration_number: Optional[str] = None
    tax_id: Optional[str] = None
    logo_url: Optional[str] = None
    description: Optional[str] = None
    established_date: Optional[datetime] = None
    mrn_prefix: Optional[str] = None

@router.put("/info", response_model=HospitalInfoResponse)
async def update_hospital_info(
    hospital_data: HospitalInfoPartialUpdateRequest,
    current_user: User = Depends(require_hospital_admin),
    db: Session = Depends(get_db)
):
    """Update hospital information"""
    hospital = db.query(Hospital).first()
    if not hospital:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Hospital not found"
        )
    
    # Update fields that are provided
    if hospital_data.name is not None:
        hospital.name = hospital_data.name
    if hospital_data.address is not None:
        hospital.address = hospital_data.address
    if hospital_data.city is not None:
        hospital.city = hospital_data.city
    if hospital_data.state is not None:
        hospital.state = hospital_data.state
    if hospital_data.postal_code is not None:
        hospital.postal_code = hospital_data.postal_code
    if hospital_data.country is not None:
        hospital.country = hospital_data.country
    if hospital_data.phone is not None:
        hospital.phone = hospital_data.phone
    if hospital_data.fax is not None:
        hospital.fax = hospital_data.fax
    if hospital_data.email is not None:
        hospital.email = hospital_data.email
    if hospital_data.website is not None:
        hospital.website = hospital_data.website
    if hospital_data.license_number is not None:
        hospital.license_number = hospital_data.license_number
    if hospital_data.registration_number is not None:
        hospital.registration_number = hospital_data.registration_number
    if hospital_data.tax_id is not None:
        hospital.tax_id = hospital_data.tax_id
    if hospital_data.logo_url is not None:
        hospital.logo_url = hospital_data.logo_url
    if hospital_data.description is not None:
        hospital.description = hospital_data.description
    if hospital_data.established_date is not None:
        hospital.established_date = hospital_data.established_date
    if hospital_data.mrn_prefix is not None:
        import re
        new_prefix = (hospital_data.mrn_prefix or "").strip().upper()
        if new_prefix and not re.fullmatch(r"[A-Z]{2,8}", new_prefix):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="MRN prefix must be 2-8 uppercase letters (A-Z only)",
            )
        # Note: prefix change is non-retroactive — existing MRNs are immutable.
        # Only patients registered after this point use the new prefix.
        hospital.mrn_prefix = new_prefix or None

    db.commit()
    db.refresh(hospital)
    return hospital

# DOCTOR MANAGEMENT ENDPOINTS
@router.get("/doctors", response_model=List[DoctorProfileResponse])
async def get_all_doctors(
    current_user: User = Depends(require_hospital_admin),
    db: Session = Depends(get_db)
):
    """Get all doctors in the hospital"""
    # Super admin can see all doctors, hospital admin sees only their hospital's doctors
    if current_user.has_role('super_admin'):
        doctors = db.query(User).join(User.role).filter(
            User.role.has(name='doctor')
        ).all()
    else:
        doctors = db.query(User).join(User.role).filter(
            User.role.has(name='doctor'),
            User.hospital_id == current_user.hospital_id
        ).all()
    
    return [
        DoctorProfileResponse(
            id=doctor.id,
            user_id=doctor.user_id,
            username=doctor.username,
            first_name=doctor.first_name,
            last_name=doctor.last_name,
            email=doctor.email,
            phone=doctor.phone,
            license_number=doctor.license_number,
            consultation_fee_inr=doctor.consultation_fee_inr,
            inpatient_fee_inr=doctor.inpatient_fee_inr,
            emergency_fee_inr=doctor.emergency_fee_inr,
            specialization=doctor.specialization,
            qualification=doctor.qualification,
            experience_years=doctor.experience_years,
            is_active=doctor.is_active
        )
        for doctor in doctors
    ]

@router.put("/doctors/{doctor_id}/profile", response_model=DoctorProfileResponse)
async def update_doctor_profile(
    doctor_id: int,
    profile_data: DoctorProfileUpdateRequest,
    current_user: User = Depends(require_hospital_admin),
    db: Session = Depends(get_db)
):
    """Update doctor profile including consultation fees"""
    doctor = db.query(User).filter(User.id == doctor_id).first()
    if not doctor:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Doctor not found"
        )
    
    # Verify it's a doctor
    if doctor.role.name != 'doctor':
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="User is not a doctor"
        )
    
    # Update profile fields
    if profile_data.license_number is not None:
        doctor.license_number = profile_data.license_number
    if profile_data.consultation_fee_inr is not None:
        doctor.consultation_fee_inr = profile_data.consultation_fee_inr
    if profile_data.inpatient_fee_inr is not None:
        doctor.inpatient_fee_inr = profile_data.inpatient_fee_inr
    if profile_data.emergency_fee_inr is not None:
        doctor.emergency_fee_inr = profile_data.emergency_fee_inr
    if profile_data.specialization is not None:
        doctor.specialization = profile_data.specialization
    if profile_data.qualification is not None:
        doctor.qualification = profile_data.qualification
    if profile_data.experience_years is not None:
        doctor.experience_years = profile_data.experience_years
    
    db.commit()
    db.refresh(doctor)
    
    return DoctorProfileResponse(
        id=doctor.id,
        user_id=doctor.user_id,
        username=doctor.username,
        first_name=doctor.first_name,
        last_name=doctor.last_name,
        email=doctor.email,
        phone=doctor.phone,
        consultation_fee=doctor.consultation_fee,
        specialization=doctor.specialization,
        qualification=doctor.qualification,
        experience_years=doctor.experience_years,
        is_active=doctor.is_active
    )

# MODULE CONFIGURATION ENDPOINTS
@router.get("/module-settings/{module_name}")
async def get_module_settings(
    module_name: str,
    current_user: User = Depends(require_hospital_admin),
    db: Session = Depends(get_db)
):
    """Get configuration settings for a specific module"""
    settings = db.query(HospitalSettings).filter(
        HospitalSettings.setting_category == module_name
    ).all()
    
    return [
        {
            "id": setting.id,
            "setting_key": setting.setting_key,
            "setting_value": setting.setting_value,
            "setting_type": setting.setting_type,
            "description": setting.description
        }
        for setting in settings
    ]

class ModuleSettingRequest(BaseModel):
    setting_key: str
    setting_value: str
    setting_type: str = "string"
    description: Optional[str] = None

@router.post("/module-settings/{module_name}")
async def create_or_update_module_setting(
    module_name: str,
    setting_data: ModuleSettingRequest,
    current_user: User = Depends(require_hospital_admin),
    db: Session = Depends(get_db)
):
    """Create or update a module setting"""
    # Check if setting already exists
    existing_setting = db.query(HospitalSettings).filter(
        HospitalSettings.setting_category == module_name,
        HospitalSettings.setting_key == setting_data.setting_key
    ).first()
    
    if existing_setting:
        # Update existing
        existing_setting.setting_value = setting_data.setting_value
        existing_setting.setting_type = setting_data.setting_type
        existing_setting.description = setting_data.description
        db.commit()
        db.refresh(existing_setting)
        return {
            "id": existing_setting.id,
            "setting_key": existing_setting.setting_key,
            "setting_value": existing_setting.setting_value,
            "setting_type": existing_setting.setting_type,
            "description": existing_setting.description,
            "message": "Setting updated successfully"
        }
    else:
        # Create new
        setting = HospitalSettings(
            setting_category=module_name,
            setting_key=setting_data.setting_key,
            setting_value=setting_data.setting_value,
            setting_type=setting_data.setting_type,
            description=setting_data.description,
            created_by=current_user.id
        )
        
        db.add(setting)
        db.commit()
        db.refresh(setting)
        return {
            "id": setting.id,
            "setting_key": setting.setting_key,
            "setting_value": setting.setting_value,
            "setting_type": setting.setting_type,
            "description": setting.description,
            "message": "Setting created successfully"
        }


# FILE UPLOAD ENDPOINT
@router.post("/upload-file")
async def upload_module_file(
    file: UploadFile = File(...),
    current_user: User = Depends(require_hospital_admin),
):
    """Upload a file (logo, signature image) for module config."""
    allowed_types = ["image/png", "image/jpeg", "image/jpg", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="Only PNG, JPEG, and WebP images are allowed")

    # Max 2MB
    content = await file.read()
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size must be under 2MB")

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    ext = file.filename.rsplit(".", 1)[-1] if "." in file.filename else "png"
    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)

    with open(filepath, "wb") as f:
        f.write(content)

    file_url = f"/uploads/module-config/{filename}"
    return {"url": file_url, "filename": filename}


# STRUCTURED MODULE CONFIG ENDPOINTS

# Lab config fields
LAB_CONFIG_FIELDS = [
    "provider_name", "provider_address", "provider_city", "provider_state",
    "provider_pincode", "provider_phone", "provider_email", "provider_logo",
    "registration_number", "nabl_number", "license_number",
    "pathologist_name", "pathologist_qualification", "signature_image",
]

# Pharmacy config fields (same as lab + pharmacy-specific)
PHARMACY_CONFIG_FIELDS = LAB_CONFIG_FIELDS + [
    "drug_license_number", "pharmacist_name", "gst_number",
]


class ModuleConfigUpdate(BaseModel):
    config: dict  # Key-value pairs of config fields


@router.get("/module-config/{module_name}")
async def get_module_config(
    module_name: str,
    current_user: User = Depends(require_hospital_admin),
    db: Session = Depends(get_db),
):
    """Get structured config for a module (lab, pharmacy)."""
    if module_name not in ("lab", "pharmacy"):
        raise HTTPException(status_code=400, detail="Config only available for lab and pharmacy")

    settings = db.query(HospitalSettings).filter(
        HospitalSettings.setting_category == f"{module_name}_config"
    ).all()

    config = {}
    for s in settings:
        config[s.setting_key] = s.setting_value

    return {"module": module_name, "config": config}


@router.put("/module-config/{module_name}")
async def update_module_config(
    module_name: str,
    data: ModuleConfigUpdate,
    current_user: User = Depends(require_hospital_admin),
    db: Session = Depends(get_db),
):
    """Save structured config for a module (lab, pharmacy)."""
    if module_name not in ("lab", "pharmacy"):
        raise HTTPException(status_code=400, detail="Config only available for lab and pharmacy")

    allowed_fields = LAB_CONFIG_FIELDS if module_name == "lab" else PHARMACY_CONFIG_FIELDS
    category = f"{module_name}_config"

    for key, value in data.config.items():
        if key not in allowed_fields:
            continue

        existing = db.query(HospitalSettings).filter(
            HospitalSettings.setting_category == category,
            HospitalSettings.setting_key == key,
        ).first()

        if existing:
            existing.setting_value = str(value)
        else:
            db.add(HospitalSettings(
                setting_category=category,
                setting_key=key,
                setting_value=str(value),
                setting_type="string",
                created_by=current_user.id,
            ))

    db.commit()
    return {"message": f"{module_name.title()} configuration saved successfully"}


# REGISTRATION FEE ENDPOINTS
@router.get("/registration-fee")
async def get_registration_fee(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get the current patient registration fee"""
    setting = db.query(HospitalSettings).filter(
        HospitalSettings.setting_category == "billing",
        HospitalSettings.setting_key == "registration_fee"
    ).first()
    return {"registration_fee": float(setting.setting_value) if setting else 0.0}


class RegistrationFeeRequest(BaseModel):
    registration_fee: float


@router.put("/registration-fee")
async def set_registration_fee(
    data: RegistrationFeeRequest,
    current_user: User = Depends(require_hospital_admin),
    db: Session = Depends(get_db)
):
    """Set the patient registration fee (hospital admin only)"""
    existing = db.query(HospitalSettings).filter(
        HospitalSettings.setting_category == "billing",
        HospitalSettings.setting_key == "registration_fee"
    ).first()

    if existing:
        existing.setting_value = str(data.registration_fee)
    else:
        db.add(HospitalSettings(
            setting_category="billing",
            setting_key="registration_fee",
            setting_value=str(data.registration_fee),
            setting_type="number",
            description="One-time registration fee for new patients",
            created_by=current_user.id,
        ))

    db.commit()
    return {"message": "Registration fee updated", "registration_fee": data.registration_fee}


# PRINT / PDF SETTINGS
@router.get("/print-settings")
async def get_print_settings(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Hospital print settings: letterhead default, gap, per-report overrides."""
    from app.utils.pdf_settings import get_print_settings_payload

    return get_print_settings_payload(db, current_user.hospital_id)


class PrintSettingsUpdate(BaseModel):
    include_header_on_pdfs: Optional[bool] = None
    include_footer_on_pdfs: Optional[bool] = None
    detailed_billing_on_pdfs: Optional[bool] = None
    letterhead_gap_mm: Optional[float] = None
    report_header_overrides: Optional[dict[str, str]] = None
    report_footer_overrides: Optional[dict[str, str]] = None


class PrintSettingsPreviewRequest(BaseModel):
    report_type: str = "opd_bill"
    include_header_on_pdfs: bool = True
    include_footer_on_pdfs: bool = True
    detailed_billing_on_pdfs: bool = True
    letterhead_gap_mm: float = 35.0
    report_header_overrides: Optional[dict[str, str]] = None
    report_footer_overrides: Optional[dict[str, str]] = None


@router.post("/print-settings/preview")
async def preview_print_settings(
    data: PrintSettingsPreviewRequest,
    current_user: User = Depends(require_print_settings_editor),
    db: Session = Depends(get_db),
):
    """Return a sample PDF using draft settings (no save required)."""
    from fastapi.responses import Response
    from app.utils.pdf_settings import MAX_LETTERHEAD_GAP_MM, MIN_LETTERHEAD_GAP_MM
    from app.utils.print_preview import generate_print_preview_pdf

    if not (MIN_LETTERHEAD_GAP_MM <= data.letterhead_gap_mm <= MAX_LETTERHEAD_GAP_MM):
        raise HTTPException(
            status_code=400,
            detail=f"letterhead_gap_mm must be between {MIN_LETTERHEAD_GAP_MM} and {MAX_LETTERHEAD_GAP_MM}",
        )
    buf = generate_print_preview_pdf(
        db,
        current_user.hospital_id,
        report_type=data.report_type,
        include_header_on_pdfs=data.include_header_on_pdfs,
        include_footer_on_pdfs=data.include_footer_on_pdfs,
        detailed_billing_on_pdfs=data.detailed_billing_on_pdfs,
        letterhead_gap_mm=data.letterhead_gap_mm,
        report_header_overrides=data.report_header_overrides,
        report_footer_overrides=data.report_footer_overrides,
    )
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": 'inline; filename="print-preview.pdf"'},
    )


@router.put("/print-settings")
async def update_print_settings(
    data: PrintSettingsUpdate,
    current_user: User = Depends(require_print_settings_editor),
    db: Session = Depends(get_db),
):
    from app.utils.pdf_settings import (
        MAX_LETTERHEAD_GAP_MM,
        MIN_LETTERHEAD_GAP_MM,
        update_print_settings as save_print_settings,
    )

    if data.letterhead_gap_mm is not None:
        if not (MIN_LETTERHEAD_GAP_MM <= data.letterhead_gap_mm <= MAX_LETTERHEAD_GAP_MM):
            raise HTTPException(
                status_code=400,
                detail=f"letterhead_gap_mm must be between {MIN_LETTERHEAD_GAP_MM} and {MAX_LETTERHEAD_GAP_MM}",
            )
    payload = save_print_settings(
        db,
        current_user.hospital_id,
        include_header_on_pdfs=data.include_header_on_pdfs,
        include_footer_on_pdfs=data.include_footer_on_pdfs,
        detailed_billing_on_pdfs=data.detailed_billing_on_pdfs,
        letterhead_gap_mm=data.letterhead_gap_mm,
        report_header_overrides=data.report_header_overrides,
        report_footer_overrides=data.report_footer_overrides,
        created_by=current_user.id,
    )
    db.commit()
    from app.services.audit_service import log_action
    log_action(db, current_user, "update_print_settings", "hospital", details=payload)
    db.commit()
    return {"message": "Print settings updated", **payload}


@router.get("/dashboard-overview")
async def get_dashboard_overview(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Aggregated dashboard data for hospital admin."""
    import traceback
    if not any(r in current_user.role_names for r in ("super_admin", "hospital_admin")):
        raise HTTPException(status_code=403, detail="Not authorized")

    try:
        return _get_dashboard_data(current_user, db)
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


def _get_dashboard_data(current_user, db):
    hospital_id = current_user.hospital_id
    today = date.today()
    week_ago = today - timedelta(days=7)
    month_start = today.replace(day=1)

    # --- Patient Stats ---
    total_patients = db.query(sql_func.count(Patient.id)).filter(
        Patient.hospital_id == hospital_id
    ).scalar() or 0

    new_patients_today = db.query(sql_func.count(Patient.id)).filter(
        Patient.hospital_id == hospital_id,
        cast(Patient.created_at, Date) == today
    ).scalar() or 0

    new_patients_this_month = db.query(sql_func.count(Patient.id)).filter(
        Patient.hospital_id == hospital_id,
        cast(Patient.created_at, Date) >= month_start
    ).scalar() or 0

    # --- Appointment Stats ---
    today_appointments = db.query(Appointment).filter(
        Appointment.doctor.has(hospital_id=hospital_id),
        cast(Appointment.appointment_date, Date) == today
    ).all()

    total_today = len(today_appointments)
    appt_by_status = {}
    for a in today_appointments:
        appt_by_status[a.status] = appt_by_status.get(a.status, 0) + 1

    # Revenue today from appointments
    revenue_today = sum(a.final_amount or 0 for a in today_appointments if a.payment_status == 'paid')
    revenue_pending = sum(a.final_amount or 0 for a in today_appointments if a.payment_status in ('pending', 'partial'))

    # Monthly revenue
    month_appointments = db.query(Appointment).filter(
        Appointment.doctor.has(hospital_id=hospital_id),
        cast(Appointment.appointment_date, Date) >= month_start,
        Appointment.payment_status == 'paid'
    ).all()
    revenue_this_month = sum(a.final_amount or 0 for a in month_appointments)

    # Doctor performance today
    doctor_stats = {}
    for a in today_appointments:
        doc_name = f"Dr. {a.doctor.first_name} {a.doctor.last_name}" if a.doctor else "Unknown"
        if doc_name not in doctor_stats:
            doctor_stats[doc_name] = {"appointments": 0, "completed": 0, "revenue": 0, "specialization": getattr(a.doctor, 'specialization', '') or ''}
        doctor_stats[doc_name]["appointments"] += 1
        if a.status == 'completed':
            doctor_stats[doc_name]["completed"] += 1
        if a.payment_status == 'paid':
            doctor_stats[doc_name]["revenue"] += a.final_amount or 0

    # --- Lab Stats ---
    lab_orders_today = db.query(sql_func.count(PatientLabOrder.id)).filter(
        PatientLabOrder.patient.has(hospital_id=hospital_id),
        cast(PatientLabOrder.order_date, Date) == today
    ).scalar() or 0

    lab_pending = db.query(sql_func.count(PatientLabOrder.id)).filter(
        PatientLabOrder.patient.has(hospital_id=hospital_id),
        PatientLabOrder.status.in_(["ordered", "collected", "processing"])
    ).scalar() or 0

    lab_completed_today = db.query(sql_func.count(PatientLabOrder.id)).filter(
        PatientLabOrder.patient.has(hospital_id=hospital_id),
        PatientLabOrder.status == "completed",
        cast(PatientLabOrder.completion_date, Date) == today
    ).scalar() or 0

    lab_revenue_today = db.query(sql_func.coalesce(sql_func.sum(PatientLabOrder.amount), 0)).filter(
        PatientLabOrder.patient.has(hospital_id=hospital_id),
        PatientLabOrder.payment_status == "paid",
        cast(PatientLabOrder.payment_date, Date) == today
    ).scalar() or 0

    lab_revenue_month = db.query(sql_func.coalesce(sql_func.sum(PatientLabOrder.amount), 0)).filter(
        PatientLabOrder.patient.has(hospital_id=hospital_id),
        PatientLabOrder.payment_status == "paid",
        cast(PatientLabOrder.payment_date, Date) >= month_start
    ).scalar() or 0

    # --- Consultation Stats ---
    consultations_today = db.query(sql_func.count(Consultation.id)).join(
        User, Consultation.doctor_id == User.id
    ).filter(
        User.hospital_id == hospital_id,
        cast(Consultation.consultation_date, Date) == today
    ).scalar() or 0

    # --- Staff Stats ---
    total_doctors = db.query(sql_func.count(User.id)).filter(
        User.hospital_id == hospital_id, User.role.has(name="doctor"), User.is_active == True
    ).scalar() or 0

    total_staff = db.query(sql_func.count(User.id)).filter(
        User.hospital_id == hospital_id, User.is_active == True
    ).scalar() or 0

    # --- Recent appointments (last 5 completed) ---
    recent_appointments = db.query(Appointment).filter(
        Appointment.doctor.has(hospital_id=hospital_id),
        Appointment.status == "completed",
    ).order_by(Appointment.updated_at.desc()).limit(5).all()

    recent_activity = []
    for a in recent_appointments:
        recent_activity.append({
            "type": "appointment",
            "patient": f"{a.patient.first_name} {a.patient.last_name}" if a.patient else "Unknown",
            "doctor": f"Dr. {a.doctor.first_name} {a.doctor.last_name}" if a.doctor else "Unknown",
            "status": a.status,
            "time": a.updated_at.isoformat() if a.updated_at else a.created_at.isoformat(),
            "amount": a.final_amount or 0,
        })

    # --- Pending lab orders (last 5) ---
    pending_labs = db.query(PatientLabOrder).filter(
        PatientLabOrder.patient.has(hospital_id=hospital_id),
        PatientLabOrder.status.in_(["ordered", "collected", "processing"])
    ).order_by(PatientLabOrder.order_date.desc()).limit(5).all()

    pending_lab_list = []
    for lo in pending_labs:
        pending_lab_list.append({
            "order_number": lo.order_number,
            "patient": f"{lo.patient.first_name} {lo.patient.last_name}" if lo.patient else "Unknown",
            "test": lo.test.name if lo.test else "Unknown",
            "status": lo.status,
            "payment_status": lo.payment_status,
            "ordered_at": lo.order_date.isoformat() if lo.order_date else "",
        })

    # --- Weekly trend (appointments per day for last 7 days) ---
    weekly_trend = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        count = db.query(sql_func.count(Appointment.id)).filter(
            Appointment.doctor.has(hospital_id=hospital_id),
            cast(Appointment.appointment_date, Date) == d
        ).scalar() or 0
        weekly_trend.append({"date": d.isoformat(), "day": d.strftime("%a"), "count": count})

    return {
        "patients": {
            "total": total_patients,
            "new_today": new_patients_today,
            "new_this_month": new_patients_this_month,
        },
        "appointments": {
            "total_today": total_today,
            "by_status": appt_by_status,
            "consultations_today": consultations_today,
        },
        "revenue": {
            "today": revenue_today,
            "today_pending": revenue_pending,
            "this_month": revenue_this_month,
            "lab_today": float(lab_revenue_today),
            "lab_this_month": float(lab_revenue_month),
        },
        "lab": {
            "orders_today": lab_orders_today,
            "pending": lab_pending,
            "completed_today": lab_completed_today,
        },
        "staff": {
            "total_doctors": total_doctors,
            "total_staff": total_staff,
        },
        "doctor_performance": [
            {"name": name, **stats} for name, stats in doctor_stats.items()
        ],
        "recent_activity": recent_activity,
        "pending_labs": pending_lab_list,
        "weekly_trend": weekly_trend,
    }


class CancelBillRequest(BaseModel):
    reason: str = Field(..., min_length=1, max_length=500)


@router.post("/billing/cancel/{bill_type}/{bill_id}")
async def cancel_bill(
    bill_type: str,
    bill_id: int,
    data: CancelBillRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Cancel a bill (appointment or lab order). Sets payment_status to 'cancelled'."""
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin']):
        raise HTTPException(status_code=403, detail="Only admins can cancel bills")

    from datetime import datetime

    if bill_type == "consultation":
        record = db.query(Appointment).join(Patient).filter(
            Appointment.id == bill_id,
            Patient.hospital_id == current_user.hospital_id
        ).first()
        if not record:
            raise HTTPException(status_code=404, detail="Appointment not found")
        if record.payment_status == "cancelled":
            raise HTTPException(status_code=400, detail="Bill is already cancelled")
        record.payment_status = "cancelled"
        record.bill_cancelled_reason = data.reason
        record.bill_cancelled_by = current_user.id
        record.bill_cancelled_at = datetime.now()
        label = f"Appointment {record.appointment_number}"

    elif bill_type == "lab":
        record = db.query(PatientLabOrder).join(Patient).filter(
            PatientLabOrder.id == bill_id,
            Patient.hospital_id == current_user.hospital_id
        ).first()
        if not record:
            raise HTTPException(status_code=404, detail="Lab order not found")
        if record.payment_status == "cancelled":
            raise HTTPException(status_code=400, detail="Bill is already cancelled")
        record.payment_status = "cancelled"
        record.bill_cancelled_reason = data.reason
        record.bill_cancelled_by = current_user.id
        record.bill_cancelled_at = datetime.now()
        label = f"Lab order {record.order_number}"
    else:
        raise HTTPException(status_code=400, detail="Invalid bill type. Use 'consultation' or 'lab'.")

    db.commit()

    # Audit log
    try:
        from app.services.audit_service import log_action
        log_action(db, current_user, "cancel_bill", "billing", bill_type, bill_id,
            f"Cancelled bill for {label}: {data.reason}",
            details={"bill_type": bill_type, "bill_id": bill_id, "reason": data.reason})
    except Exception:
        pass

    return {"message": f"Bill cancelled: {label}"}


@router.get("/billing")
async def get_all_bills(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    patient_search: Optional[str] = None,
    bill_type: Optional[str] = None,
    payment_status: Optional[str] = None,
    doctor_id: Optional[int] = None,
    referred_by: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Centralised billing view — consultation, lab, pharmacy, and admission bills."""
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin', 'receptionist']):
        raise HTTPException(status_code=403, detail="Not authorized")

    hospital_id = current_user.hospital_id
    today = date.today()
    d_from = date_from or today.isoformat()
    d_to = date_to or today.isoformat()

    bills = []

    # --- Appointment bills ---
    # Sources rolled into a consolidated bill stay on the appointment row with
    # payment_status='consolidated'; hide them so the CB-* row is the only
    # ledger entry for those charges.
    apt_query = db.query(Appointment).join(Patient).filter(
        Patient.hospital_id == hospital_id,
        sql_func.date(Appointment.created_at) >= d_from,
        sql_func.date(Appointment.created_at) <= d_to,
        Appointment.payment_status != "consolidated",
    )
    if payment_status:
        apt_query = apt_query.filter(Appointment.payment_status == payment_status)
    if patient_search:
        q = f"%{patient_search}%"
        apt_query = apt_query.filter(
            (Patient.first_name.ilike(q)) | (Patient.last_name.ilike(q)) | (Patient.primary_phone.ilike(q))
        )
    if bill_type and bill_type not in ('appointment', 'consultation'):
        apt_query = apt_query.filter(False)  # skip
    if doctor_id:
        apt_query = apt_query.filter(Appointment.doctor_id == doctor_id)
    if referred_by:
        apt_query = apt_query.filter(Appointment.referred_by.ilike(f"%{referred_by}%"))

    for apt in apt_query.order_by(Appointment.created_at.desc()).all():
        # Catch-up consultations also create a Bill (CU-CONS-*); show that ledger
        # row instead so Service Date drives the dashboard date.
        cu_bill = db.query(Bill).filter(
            Bill.bill_type == "consultation",
            Bill.reference_id == apt.id,
            Bill.bill_number.like("CU-%"),
        ).first()
        if cu_bill:
            continue
        p = apt.patient
        doctor = db.query(User).filter(User.id == apt.doctor_id).first() if apt.doctor_id else None
        cancelled_by_user = db.query(User).filter(User.id == apt.bill_cancelled_by).first() if getattr(apt, 'bill_cancelled_by', None) else None
        bills.append({
            "id": f"APT-{apt.id}",
            "bill_id": apt.id,
            "type": "consultation",
            "date": apt.created_at.isoformat() if apt.created_at else "",
            "patient_name": f"{p.first_name} {p.last_name}" if p else "Unknown",
            "patient_phone": p.primary_phone if p else "",
            "patient_id": p.patient_id if p else "",
            "_doctor_id": apt.doctor_id,
            "doctor_name": f"Dr. {doctor.first_name} {doctor.last_name}" if doctor else "",
            "reference": apt.appointment_number,
            "items": f"Consultation{' + Registration' if apt.registration_fee else ''}",
            "subtotal": (apt.consultation_fee or 0) + (apt.registration_fee or 0),
            "discount": apt.discount_amount or 0,
            "amount": apt.final_amount or 0,
            "payment_status": apt.payment_status or "pending",
            "payment_method": apt.payment_method or "",
            "referred_by": apt.referred_by or "",
            "cancel_reason": getattr(apt, 'bill_cancelled_reason', None) or "",
            "cancelled_by": f"{cancelled_by_user.first_name} {cancelled_by_user.last_name}" if cancelled_by_user else "",
            "cancelled_at": apt.bill_cancelled_at.isoformat() if getattr(apt, 'bill_cancelled_at', None) else "",
        })

    # --- Lab order bills (OPD only) ---
    # Inpatient lab orders are charged on the admission bill; consolidated
    # sources are represented by the CB-* row. Exclude both so the Lab tab
    # does not double-list the same charge.
    lab_query = db.query(PatientLabOrder).join(Patient).filter(
        Patient.hospital_id == hospital_id,
        sql_func.date(PatientLabOrder.order_date) >= d_from,
        sql_func.date(PatientLabOrder.order_date) <= d_to,
        PatientLabOrder.admission_id.is_(None),
        PatientLabOrder.payment_status != "consolidated",
    )
    if payment_status:
        lab_query = lab_query.filter(PatientLabOrder.payment_status == payment_status)
    if patient_search:
        q = f"%{patient_search}%"
        lab_query = lab_query.filter(
            (Patient.first_name.ilike(q)) | (Patient.last_name.ilike(q)) | (Patient.primary_phone.ilike(q))
        )
    if bill_type and bill_type != 'lab':
        lab_query = lab_query.filter(False)
    if doctor_id:
        lab_query = lab_query.filter(PatientLabOrder.doctor_id == doctor_id)
    if referred_by:
        lab_query = lab_query.filter(PatientLabOrder.referred_by.ilike(f"%{referred_by}%"))

    # Coalesce lab orders by lab_bill_group_id so each real bill renders as
    # ONE row. Orders with no group (legacy rows from before the grouping
    # columns existed) keep the historical one-row-per-test behavior.
    lab_orders_all = lab_query.order_by(PatientLabOrder.order_date.desc()).all()
    grouped: dict = {}
    ungrouped: list = []
    for lo in lab_orders_all:
        gid = getattr(lo, "lab_bill_group_id", None)
        if gid:
            grouped.setdefault(gid, []).append(lo)
        else:
            ungrouped.append(lo)

    def _row_for_group(group_orders):
        first = group_orders[0]
        p = db.query(Patient).filter(Patient.id == first.patient_id).first()
        lab_doctor = db.query(User).filter(User.id == first.doctor_id).first() if first.doctor_id else None
        lab_cancelled_by = db.query(User).filter(User.id == first.bill_cancelled_by).first() if getattr(first, 'bill_cancelled_by', None) else None
        # Description: package name if all orders share the same package,
        # otherwise a comma-joined list of test names (truncated).
        pkg_ids = {o.package_id for o in group_orders if o.package_id}
        items_text = ""
        if len(pkg_ids) == 1 and None not in pkg_ids:
            pkg_obj = db.query(LabTestPackage).filter(LabTestPackage.id == first.package_id).first()
            if pkg_obj:
                items_text = f"{pkg_obj.name} (Package, {len(group_orders)} test{'s' if len(group_orders) > 1 else ''})"
        if not items_text:
            test_names = []
            for o in group_orders[:3]:
                t = db.query(LabTest).filter(LabTest.id == o.test_id).first()
                if t:
                    test_names.append(t.name)
            items_text = ", ".join(test_names)
            if len(group_orders) > 3:
                items_text += f" +{len(group_orders) - 3} more"
        # Subtotal/amount: package mode shows package_price; otherwise sum.
        subtotal = sum((o.amount or 0) for o in group_orders)
        amount = subtotal
        discount = 0
        # Status: paid only if every order in the group is paid; cancelled
        # if every order is cancelled.
        statuses = {o.payment_status for o in group_orders}
        if statuses == {"paid"}:
            payment_status = "paid"
        elif statuses == {"cancelled"}:
            payment_status = "cancelled"
        elif "pending" in statuses:
            payment_status = "pending"
        else:
            payment_status = "partial"
        # Reference / source label.
        source = "Package" if any(o.package_id for o in group_orders) else (
            "Appointment" if any(o.appointment_id for o in group_orders) else "Direct"
        )
        return {
            "id": f"LBG-{first.lab_bill_group_id}",
            "bill_id": first.id,
            "lab_bill_group_id": first.lab_bill_group_id,
            "type": "lab",
            "date": first.order_date.isoformat() if first.order_date else "",
            "patient_name": f"{p.first_name} {p.last_name}" if p else "Unknown",
            "patient_phone": p.primary_phone if p else "",
            "patient_id": p.patient_id if p else "",
            "_doctor_id": first.doctor_id,
            "doctor_name": f"Dr. {lab_doctor.first_name} {lab_doctor.last_name}" if lab_doctor else "",
            "reference": first.lab_bill_number or first.order_number,
            "items": f"{items_text} ({source})",
            "subtotal": subtotal,
            "discount": discount,
            "amount": amount,
            "payment_status": payment_status,
            "payment_method": first.payment_method or "",
            "referred_by": first.referred_by or "",
            "cancel_reason": getattr(first, 'bill_cancelled_reason', None) or "",
            "cancelled_by": f"{lab_cancelled_by.first_name} {lab_cancelled_by.last_name}" if lab_cancelled_by else "",
            "cancelled_at": first.bill_cancelled_at.isoformat() if getattr(first, 'bill_cancelled_at', None) else "",
        }

    for gid, group_orders in grouped.items():
        bills.append(_row_for_group(group_orders))

    # Backward-compat: legacy ungrouped orders still surface as one row each.
    for lo in ungrouped:
        p = db.query(Patient).filter(Patient.id == lo.patient_id).first()
        test = db.query(LabTest).filter(LabTest.id == lo.test_id).first()
        lab_doctor = db.query(User).filter(User.id == lo.doctor_id).first() if lo.doctor_id else None
        lab_cancelled_by = db.query(User).filter(User.id == lo.bill_cancelled_by).first() if getattr(lo, 'bill_cancelled_by', None) else None
        source = "Package" if lo.package_id else "Appointment" if lo.appointment_id else "Direct"
        bills.append({
            "id": f"LAB-{lo.id}",
            "bill_id": lo.id,
            "lab_bill_group_id": None,
            "type": "lab",
            "date": lo.order_date.isoformat() if lo.order_date else "",
            "patient_name": f"{p.first_name} {p.last_name}" if p else "Unknown",
            "patient_phone": p.primary_phone if p else "",
            "patient_id": p.patient_id if p else "",
            "_doctor_id": lo.doctor_id,
            "doctor_name": f"Dr. {lab_doctor.first_name} {lab_doctor.last_name}" if lab_doctor else "",
            "reference": lo.order_number,
            "items": f"{test.name if test else 'Lab Test'} ({source})",
            "subtotal": lo.amount or 0,
            "discount": 0,
            "amount": lo.amount or 0,
            "payment_status": lo.payment_status or "pending",
            "payment_method": lo.payment_method or "",
            "referred_by": lo.referred_by or "",
            "cancel_reason": getattr(lo, 'bill_cancelled_reason', None) or "",
            "cancelled_by": f"{lab_cancelled_by.first_name} {lab_cancelled_by.last_name}" if lab_cancelled_by else "",
            "cancelled_at": lo.bill_cancelled_at.isoformat() if getattr(lo, 'bill_cancelled_at', None) else "",
        })

    # --- Pharmacy counter sales ---
    # Deferred inpatient sales are consumed by admission bills and must not be
    # listed separately here, otherwise the same medicines are double-counted.
    # Pharmacy stores patient/doctor details as sale-time free text, so the
    # ID-based doctor and referral filters intentionally exclude these rows.
    if bill_type in (None, 'pharmacy') and not doctor_id and not referred_by:
        pharmacy_query = db.query(PharmacySale).filter(
            PharmacySale.hospital_id == hospital_id,
            PharmacySale.billing_mode == "cash_at_pharmacy",
            sql_func.date(PharmacySale.sale_date) >= d_from,
            sql_func.date(PharmacySale.sale_date) <= d_to,
        )
        if patient_search:
            q = f"%{patient_search}%"
            pharmacy_query = pharmacy_query.filter(
                (PharmacySale.patient_name.ilike(q))
                | (PharmacySale.patient_phone.ilike(q))
                | (PharmacySale.patient_ip_id.ilike(q))
            )
        if payment_status:
            if payment_status == "cancelled":
                pharmacy_query = pharmacy_query.filter(PharmacySale.status == "voided")
            elif payment_status == "paid":
                pharmacy_query = pharmacy_query.filter(
                    PharmacySale.status == "completed",
                    PharmacySale.payment_type == "cash",
                )
            elif payment_status == "pending":
                pharmacy_query = pharmacy_query.filter(
                    PharmacySale.status == "completed",
                    PharmacySale.payment_type == "credit",
                )
            else:
                pharmacy_query = pharmacy_query.filter(False)

        for sale in pharmacy_query.order_by(PharmacySale.sale_date.desc()).all():
            medicine_names = [
                item.medicine.name
                for item in sale.items[:3]
                if item.medicine and item.medicine.name
            ]
            items_text = ", ".join(medicine_names)
            if len(sale.items) > 3:
                items_text += f" +{len(sale.items) - 3} more"
            sale_status = (
                "cancelled" if sale.status == "voided"
                else "pending" if sale.payment_type == "credit"
                else "paid"
            )
            bills.append({
                "id": f"PHARM-{sale.id}",
                "bill_id": sale.id,
                "type": "pharmacy",
                "date": sale.sale_date.isoformat() if sale.sale_date else "",
                "patient_name": sale.patient_name or "Walk-in customer",
                "patient_phone": sale.patient_phone or "",
                "patient_id": sale.patient_ip_id or "",
                "_doctor_id": None,
                "doctor_name": sale.doctor_name or "",
                "reference": sale.sale_number,
                "items": items_text or "Pharmacy medicines",
                "subtotal": float(sale.subtotal or 0),
                "discount": float(sale.discount_total or 0),
                "amount": float(sale.grand_total or 0),
                "payment_status": sale_status,
                "payment_method": sale.payment_type or "",
                "referred_by": "",
                "cancel_reason": sale.void_reason or "",
                "cancelled_by": "",
                "cancelled_at": sale.voided_at.isoformat() if sale.voided_at else "",
            })

    # --- Admission (inpatient) bills — one summary row per admission ---
    # Each admission is represented by a single row whose amount = total of all
    # non-cancelled bills for that admission (no double-counting across interim
    # and final bills). Deposits/refunds appear as inline child rows beneath it.
    if bill_type in (None, 'admission', 'inpatient', 'deposit'):
        from app.models.inpatient import AdmissionDeposit

        # Gather all admission bills in the date window (any subtype).
        adm_bill_q = db.query(Bill).join(Patient, Bill.patient_id == Patient.id).filter(
            Patient.hospital_id == hospital_id,
            Bill.bill_type == "admission",
            sql_func.date(Bill.bill_date) >= d_from,
            sql_func.date(Bill.bill_date) <= d_to,
        )
        # Also gather admissions with deposits in the date window so deposit-only
        # admissions (no bill yet) still surface when filter is 'deposit'.
        dep_adm_ids_q = db.query(AdmissionDeposit.admission_id).join(
            Admission, Admission.id == AdmissionDeposit.admission_id
        ).join(Patient, Patient.id == Admission.patient_id).filter(
            Patient.hospital_id == hospital_id,
            sql_func.date(AdmissionDeposit.received_at) >= d_from,
            sql_func.date(AdmissionDeposit.received_at) <= d_to,
        )
        if patient_search:
            q = f"%{patient_search}%"
            adm_bill_q = adm_bill_q.filter(
                (Patient.first_name.ilike(q)) | (Patient.last_name.ilike(q)) | (Patient.primary_phone.ilike(q))
            )
            dep_adm_ids_q = dep_adm_ids_q.filter(
                (Patient.first_name.ilike(q)) | (Patient.last_name.ilike(q)) | (Patient.primary_phone.ilike(q))
            )

        # Group bills by admission_id.
        adm_bills_by_adm: dict = {}
        for b in adm_bill_q.all():
            adm_bills_by_adm.setdefault(b.reference_id, []).append(b)

        # Union of admission IDs from bills and from deposits.
        adm_ids_from_bills = set(adm_bills_by_adm.keys())
        adm_ids_from_deps = {row[0] for row in dep_adm_ids_q.all()}
        all_adm_ids = adm_ids_from_bills | adm_ids_from_deps

        for adm_id in all_adm_ids:
            if adm_id is None:
                continue
            admission = db.query(Admission).filter(Admission.id == adm_id).first()
            if not admission:
                continue
            p = db.query(Patient).filter(Patient.id == admission.patient_id).first()
            admitting_doc = None
            if admission.admitting_doctor_id:
                admitting_doc = db.query(User).filter(User.id == admission.admitting_doctor_id).first()

            adm_bill_list = adm_bills_by_adm.get(adm_id, [])
            active_bills = [b for b in adm_bill_list if b.status != "cancelled"]
            final_bill = next((b for b in active_bills if b.bill_subtype == "final"), None)
            # Representative bill number: final bill if exists, else latest interim.
            rep_bill = final_bill or (sorted(active_bills, key=lambda b: b.id, reverse=True)[0] if active_bills else None)
            bill_subtype = "final" if final_bill else ("interim" if active_bills else "none")

            # Total charges: when a comprehensive final bill exists, its total
            # already includes prior interim charges — use it alone to avoid
            # double-counting (interim ₹X + final ₹X = ₹2X). Fall back to
            # summing all active bills when only interim bills exist.
            if final_bill:
                total_charges = float(final_bill.total_amount or 0)
            else:
                total_charges = sum(float(b.total_amount or 0) for b in active_bills)

            # Build items description from the representative bill's items.
            if rep_bill:
                items_list = db.query(BillItem).filter(BillItem.bill_id == rep_bill.id).all()
                items_text = ", ".join(it.item_name for it in items_list[:3])
                if len(items_list) > 3:
                    items_text += f" +{len(items_list) - 3} more"
                if not items_text:
                    items_text = "Admission charges"
            else:
                items_text = "Admission charges"

            # Deposits and refunds for this admission.
            dep_rows = db.query(AdmissionDeposit).filter(
                AdmissionDeposit.admission_id == adm_id,
            ).order_by(AdmissionDeposit.received_at).all()
            net_deposits = sum(
                float(d.amount or 0) if d.deposit_type != "refund" else -abs(float(d.amount or 0))
                for d in dep_rows
            )
            balance_due = round(total_charges - net_deposits, 2)

            # Admission-level status — based on financial balance, not bill status.
            if total_charges <= 0 or balance_due <= 0:
                adm_status = "paid"
            elif net_deposits > 0:
                adm_status = "partial"
            else:
                adm_status = "pending"

            # Apply payment_status filter against the derived status.
            if payment_status and payment_status != adm_status:
                continue
            # Skip admission-only rows when filter is 'deposit' (deposits handle separately).
            if bill_type == 'deposit' and not dep_rows:
                continue

            # Representative date: final bill date, else latest active bill date, else admission date.
            rep_date = (
                (rep_bill.bill_date.isoformat() if rep_bill and rep_bill.bill_date else None)
                or (admission.admission_date.isoformat() if admission.admission_date else "")
            )

            deposit_children = []
            for d in dep_rows:
                deposit_children.append({
                    "deposit_number": d.deposit_number or "",
                    "date": d.received_at.isoformat() if d.received_at else "",
                    "deposit_type": d.deposit_type or "initial",
                    "amount": float(d.amount or 0),
                    "method": d.payment_method or "cash",
                    "reference": d.reference_number or "",
                })

            bills.append({
                "id": f"ADM-{adm_id}",
                "bill_id": rep_bill.id if rep_bill else None,
                "type": "admission",
                "bill_subtype": bill_subtype,
                "date": rep_date,
                "patient_name": f"{p.first_name} {p.last_name}" if p else "Unknown",
                "patient_phone": p.primary_phone if p else "",
                "patient_id": p.patient_id if p else "",
                "_doctor_id": admission.admitting_doctor_id,
                "doctor_name": f"Dr. {admitting_doc.first_name} {admitting_doc.last_name}" if admitting_doc else "",
                "reference": rep_bill.bill_number if rep_bill else admission.admission_number or "",
                "items": items_text,
                "subtotal": total_charges,
                "discount": 0,
                "amount": total_charges,
                "payment_status": adm_status,
                "payment_method": "",
                "referred_by": "",
                "cancel_reason": "",
                "cancelled_by": "",
                "cancelled_at": "",
                "amount_paid": net_deposits,
                "balance_due": balance_due,
                "admission_id": adm_id,
                "deposits": deposit_children,
                "net_deposits": round(net_deposits, 2),
            })

    # --- Day-care bills from bills table (outpatient day-care services) ---
    # Backwards-compat: still accept the old 'procedure' filter value so any
    # external callers/links keep working after the rename.
    if bill_type in (None, 'day_care', 'procedure'):
        proc_query = db.query(Bill).join(Patient, Bill.patient_id == Patient.id).filter(
            Patient.hospital_id == hospital_id,
            Bill.bill_type == "day_care",
            sql_func.date(Bill.bill_date) >= d_from,
            sql_func.date(Bill.bill_date) <= d_to,
        )
        if payment_status:
            mapped = {"paid": "paid", "pending": "pending", "cancelled": "cancelled"}.get(payment_status, payment_status)
            proc_query = proc_query.filter(Bill.status == mapped)
        if patient_search:
            q = f"%{patient_search}%"
            proc_query = proc_query.filter(
                (Patient.first_name.ilike(q)) | (Patient.last_name.ilike(q)) | (Patient.primary_phone.ilike(q))
            )
        for b in proc_query.order_by(Bill.bill_date.desc()).all():
            p = db.query(Patient).filter(Patient.id == b.patient_id).first()
            items_list = db.query(BillItem).filter(BillItem.bill_id == b.id).all()
            items_text = ", ".join(it.item_name for it in items_list[:3])
            if len(items_list) > 3:
                items_text += f" +{len(items_list) - 3} more"
            paid_total = sum(float(pay.amount_paid) for pay in (b.payments or []))
            bills.append({
                "id": f"DC-{b.id}",
                "bill_id": b.id,
                "type": "day_care",
                "date": b.bill_date.isoformat() if b.bill_date else "",
                "patient_name": f"{p.first_name} {p.last_name}" if p else "Unknown",
                "patient_phone": p.primary_phone if p else "",
                "patient_id": p.patient_id if p else "",
                "_doctor_id": None,
                "doctor_name": "",
                "reference": b.bill_number,
                "items": items_text or "Day care services",
                "subtotal": float(b.subtotal or 0),
                "discount": float(b.discount_amount or 0),
                "amount": float(b.total_amount or 0),
                "payment_status": b.status or "pending",
                "payment_method": "",
                "referred_by": b.referred_by or "",
                "cancel_reason": "",
                "cancelled_by": "",
                "cancelled_at": "",
                "amount_paid": paid_total,
                "balance_due": float(b.total_amount or 0) - paid_total,
                "admission_id": None,
            })

    # --- Consolidated bills from bills table ---
    if bill_type in (None, 'consolidated'):
        cons_query = db.query(Bill).join(Patient, Bill.patient_id == Patient.id).filter(
            Patient.hospital_id == hospital_id,
            Bill.bill_type == "consolidated",
            sql_func.date(Bill.bill_date) >= d_from,
            sql_func.date(Bill.bill_date) <= d_to,
        )
        if payment_status:
            mapped = {"paid": "paid", "pending": "pending", "cancelled": "cancelled"}.get(payment_status, payment_status)
            cons_query = cons_query.filter(Bill.status == mapped)
        if patient_search:
            q = f"%{patient_search}%"
            cons_query = cons_query.filter(
                (Patient.first_name.ilike(q)) | (Patient.last_name.ilike(q)) | (Patient.primary_phone.ilike(q))
            )
        for b in cons_query.order_by(Bill.bill_date.desc()).all():
            p = db.query(Patient).filter(Patient.id == b.patient_id).first()
            items_list = db.query(BillItem).filter(BillItem.bill_id == b.id).all()
            items_text = ", ".join(it.item_name for it in items_list[:3])
            if len(items_list) > 3:
                items_text += f" +{len(items_list) - 3} more"
            paid_total = sum(float(pay.amount_paid) for pay in (b.payments or []))
            bills.append({
                "id": f"CONS-{b.id}",
                "bill_id": b.id,
                "type": "consolidated",
                "date": b.bill_date.isoformat() if b.bill_date else "",
                "patient_name": f"{p.first_name} {p.last_name}" if p else "Unknown",
                "patient_phone": p.primary_phone if p else "",
                "patient_id": p.patient_id if p else "",
                "_doctor_id": None,
                "doctor_name": "",
                "reference": b.bill_number,
                "items": items_text or "Consolidated charges",
                "subtotal": float(b.subtotal or 0),
                "discount": float(b.discount_amount or 0),
                "amount": float(b.total_amount or 0),
                "payment_status": b.status or "pending",
                "payment_method": "",
                "referred_by": "",
                "cancel_reason": "",
                "cancelled_by": "",
                "cancelled_at": "",
                "amount_paid": paid_total,
                "balance_due": float(b.total_amount or 0) - paid_total,
                "admission_id": None,
            })

    # --- Catch-up / POS ledger bills (Bill rows with Service Date on bill_date) ---
    # Consultation catch-up: CU-* bills (source Appointment skipped above).
    # Misc / pharmacy / canteen catch-up: bill_type catch_up|pharmacy|canteen.
    # Lab catch-up already surfaces via PatientLabOrder.order_date — skip Bill.lab.
    if bill_type in (None, 'catch_up', 'pharmacy', 'canteen', 'consultation', 'misc'):
        ledger_types = []
        if bill_type in (None, 'catch_up', 'misc'):
            ledger_types.append("catch_up")
        if bill_type in (None, 'pharmacy'):
            ledger_types.append("pharmacy")
        if bill_type in (None, 'canteen'):
            ledger_types.append("canteen")
        if bill_type in (None, 'consultation'):
            ledger_types.append("consultation")

        if ledger_types:
            ledger_q = db.query(Bill).join(Patient, Bill.patient_id == Patient.id).filter(
                Patient.hospital_id == hospital_id,
                Bill.bill_type.in_(ledger_types),
                sql_func.date(Bill.bill_date) >= d_from,
                sql_func.date(Bill.bill_date) <= d_to,
                Bill.status != "cancelled",
            )
            # Consultation ledger: only CU-* catch-up bills (normal consults stay on Appointment)
            if "consultation" in ledger_types and bill_type != "consultation":
                pass  # filter per-row below
            if patient_search:
                q = f"%{patient_search}%"
                ledger_q = ledger_q.filter(
                    (Patient.first_name.ilike(q)) | (Patient.last_name.ilike(q)) | (Patient.primary_phone.ilike(q))
                )
            if payment_status:
                mapped = {"paid": "paid", "pending": "pending", "cancelled": "cancelled"}.get(
                    payment_status, payment_status
                )
                ledger_q = ledger_q.filter(Bill.status == mapped)

            for b in ledger_q.order_by(Bill.bill_date.desc()).all():
                if b.bill_type == "consultation" and not (b.bill_number or "").startswith("CU-"):
                    continue
                p = db.query(Patient).filter(Patient.id == b.patient_id).first()
                items_list = db.query(BillItem).filter(BillItem.bill_id == b.id).all()
                items_text = ", ".join(it.item_name for it in items_list[:3])
                if len(items_list) > 3:
                    items_text += f" +{len(items_list) - 3} more"
                paid_total = sum(float(pay.amount_paid) for pay in (b.payments or []))
                row_type = b.bill_type if b.bill_type != "catch_up" else "catch_up"
                bills.append({
                    "id": f"CU-{b.id}",
                    "bill_id": b.id,
                    "type": row_type,
                    "date": b.bill_date.isoformat() if b.bill_date else "",
                    "patient_name": f"{p.first_name} {p.last_name}" if p else "Unknown",
                    "patient_phone": p.primary_phone if p else "",
                    "patient_id": p.patient_id if p else "",
                    "_doctor_id": None,
                    "doctor_name": "",
                    "reference": b.bill_number,
                    "items": items_text or "Catch-up bill",
                    "subtotal": float(b.subtotal or 0),
                    "discount": float(b.discount_amount or 0),
                    "amount": float(b.total_amount or 0),
                    "payment_status": b.status or "pending",
                    "payment_method": "",
                    "referred_by": b.referred_by or "",
                    "cancel_reason": "",
                    "cancelled_by": "",
                    "cancelled_at": "",
                    "amount_paid": paid_total,
                    "balance_due": float(b.total_amount or 0) - paid_total,
                    "admission_id": None,
                    "is_catch_up": True,
                })

    # Sort all by date descending
    bills.sort(key=lambda b: b["date"], reverse=True)

    # Summary — exclude cancelled from totals.
    # Admission rows already carry the consolidated total (no double-counting
    # across interim+final). For admissions, "paid" = balance_due <= 0.
    active_bills = [b for b in bills if b["payment_status"] != "cancelled"]
    total_billed = sum(b["amount"] for b in active_bills)
    total_paid = sum(
        (b["net_deposits"] if b["type"] == "admission" else b["amount"])
        for b in active_bills if b["payment_status"] == "paid"
    )
    total_pending = sum(b["amount"] for b in active_bills if b["payment_status"] in ("pending", "partial"))
    cancelled_count = sum(1 for b in bills if b["payment_status"] == "cancelled")
    apt_count = sum(1 for b in bills if b["type"] == "consultation")
    lab_count = sum(1 for b in bills if b["type"] == "lab")
    pharmacy_count = sum(1 for b in bills if b["type"] == "pharmacy")
    adm_count = sum(1 for b in bills if b["type"] == "admission")

    # Extract unique doctors from bills for filter dropdown
    doctor_ids_seen = set()
    doctor_list = []
    for b in bills:
        if b.get("doctor_name") and b.get("_doctor_id") and b["_doctor_id"] not in doctor_ids_seen:
            doctor_ids_seen.add(b["_doctor_id"])
            doctor_list.append({"id": b["_doctor_id"], "name": b["doctor_name"]})
    # Also fetch doctors with doctor role who may not have bills yet
    from app.models.user import UserRole
    doctor_role = db.query(UserRole).filter(UserRole.name == 'doctor').first()
    if doctor_role:
        doc_users = db.query(User).filter(
            User.hospital_id == hospital_id,
            User.is_active == True,
            User.roles.any(id=doctor_role.id)
        ).all()
        for d in doc_users:
            if d.id not in doctor_ids_seen:
                doctor_ids_seen.add(d.id)
                doctor_list.append({"id": d.id, "name": f"Dr. {d.first_name} {d.last_name}"})

    # Fetch referrals for filter dropdown
    from app.models.referral import Referral
    try:
        referral_list = [{"id": r.id, "name": r.name} for r in
            db.query(Referral).filter(Referral.hospital_id == hospital_id, Referral.is_active == True).all()]
    except Exception:
        referral_list = []

    # Remove internal _doctor_id from bill output
    for b in bills:
        b.pop("_doctor_id", None)

    return {
        "bills": bills,
        "summary": {
            "total_bills": len(bills),
            "total_billed": total_billed,
            "total_paid": total_paid,
            "total_pending": total_pending,
            "appointment_count": apt_count,
            "lab_count": lab_count,
            "pharmacy_count": pharmacy_count,
            "admission_count": adm_count,
            "cancelled_count": cancelled_count,
        },
        "doctors": doctor_list,
        "referrals": referral_list,
    }


def _format_export_date(value) -> str:
    """Format bill date strings for spreadsheet cells (YYYY-MM-DD when possible)."""
    if not value:
        return ""
    s = str(value)
    if "T" in s:
        return s.split("T", 1)[0]
    return s[:10] if len(s) >= 10 else s


def _build_billing_export_xlsx(data: dict, date_from: str, date_to: str) -> bytes:
    """Build a two-sheet Excel workbook from the unified billing list payload."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Bills"

    headers = [
        "Date", "Type", "Reference", "Patient", "Phone", "Items",
        "Amount", "Discount", "Final", "Doctor", "Referred By",
        "Status", "Payment Method",
    ]
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8EEF5")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin

    for row_idx, b in enumerate(data.get("bills") or [], 2):
        values = [
            _format_export_date(b.get("date")),
            b.get("type") or "",
            b.get("reference") or "",
            b.get("patient_name") or "",
            b.get("patient_phone") or "",
            b.get("items") or "",
            float(b.get("subtotal") or 0),
            float(b.get("discount") or 0),
            float(b.get("amount") or 0),
            b.get("doctor_name") or "",
            b.get("referred_by") or "",
            b.get("payment_status") or "",
            b.get("payment_method") or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin
            if col in (7, 8, 9):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    from openpyxl.utils import get_column_letter
    col_widths = [12, 14, 16, 22, 14, 36, 12, 10, 12, 20, 16, 12, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{max(1, len(data.get('bills') or []) + 1)}"

    # Summary sheet
    summary = data.get("summary") or {}
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Billing Export Summary"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "Date from"
    ws2["B3"] = date_from
    ws2["A4"] = "Date to"
    ws2["B4"] = date_to
    ws2["A6"] = "Metric"
    ws2["B6"] = "Value"
    ws2["A6"].font = header_font
    ws2["B6"].font = header_font
    ws2["A6"].fill = header_fill
    ws2["B6"].fill = header_fill

    metrics = [
        ("Total bills", summary.get("total_bills", 0)),
        ("Total billed", float(summary.get("total_billed") or 0)),
        ("Collected", float(summary.get("total_paid") or 0)),
        ("Pending", float(summary.get("total_pending") or 0)),
        ("Consultation bills", summary.get("appointment_count", 0)),
        ("Lab bills", summary.get("lab_count", 0)),
        ("Pharmacy bills", summary.get("pharmacy_count", 0)),
        ("Admission bills", summary.get("admission_count", 0)),
        ("Cancelled", summary.get("cancelled_count", 0)),
    ]
    # By-type rollup from rows
    by_type: dict = {}
    for b in data.get("bills") or []:
        if b.get("payment_status") == "cancelled":
            continue
        t = b.get("type") or "other"
        by_type[t] = by_type.get(t, 0.0) + float(b.get("amount") or 0)

    for i, (label, value) in enumerate(metrics, 7):
        ws2.cell(row=i, column=1, value=label)
        cell = ws2.cell(row=i, column=2, value=value)
        if isinstance(value, float):
            cell.number_format = "#,##0.00"

    start = 7 + len(metrics) + 1
    ws2.cell(row=start, column=1, value="Amount by type").font = header_font
    ws2.cell(row=start + 1, column=1, value="Type").font = header_font
    ws2.cell(row=start + 1, column=2, value="Amount").font = header_font
    for i, (t, amt) in enumerate(sorted(by_type.items()), start + 2):
        ws2.cell(row=i, column=1, value=t)
        cell = ws2.cell(row=i, column=2, value=amt)
        cell.number_format = "#,##0.00"

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 16

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/billing/export.xlsx")
async def export_billing_xlsx(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    patient_search: Optional[str] = None,
    bill_type: Optional[str] = None,
    payment_status: Optional[str] = None,
    doctor_id: Optional[int] = None,
    referred_by: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Excel export of the unified billing list using the same filters as GET /billing."""
    data = await get_all_bills(
        date_from=date_from,
        date_to=date_to,
        patient_search=patient_search,
        bill_type=bill_type,
        payment_status=payment_status,
        doctor_id=doctor_id,
        referred_by=referred_by,
        current_user=current_user,
        db=db,
    )
    today = date.today()
    d_from = date_from or today.isoformat()
    d_to = date_to or today.isoformat()
    content = _build_billing_export_xlsx(data, d_from, d_to)
    filename = f"billing_{d_from}_to_{d_to}.xlsx"
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --- Bill Detail (from bills table) ---
@router.get("/billing/bills/{bill_id}")
async def get_bill_detail(
    bill_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get full bill detail with items and payments."""
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin', 'receptionist']):
        raise HTTPException(status_code=403, detail="Not authorized")

    bill = db.query(Bill).filter(Bill.id == bill_id).first()
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")

    patient = db.query(Patient).filter(Patient.id == bill.patient_id).first()
    items = db.query(BillItem).filter(BillItem.bill_id == bill.id).all()
    payments = db.query(Payment).filter(Payment.bill_id == bill.id).order_by(Payment.payment_date.desc()).all()

    total_paid = sum(float(p.amount_paid) for p in payments)

    # For admission bills, fold in deposits already collected against the
    # admission (allocated oldest-bill-first across sibling bills).
    deposit_alloc = 0.0
    if (bill.bill_type or "") == "admission":
        from app.routes.inpatient import allocate_deposits_to_bill
        deposit_alloc = allocate_deposits_to_bill(db, bill)

    effective_paid = total_paid + deposit_alloc
    total_amt = float(bill.total_amount or 0)

    return {
        "id": bill.id,
        "bill_number": bill.bill_number,
        "bill_type": bill.bill_type,
        "bill_date": bill.bill_date.isoformat() if bill.bill_date else None,
        "patient_name": f"{patient.first_name} {patient.last_name}" if patient else "Unknown",
        "patient_phone": patient.primary_phone if patient else "",
        "status": bill.status,
        "subtotal": float(bill.subtotal or 0),
        "tax_amount": float(bill.tax_amount or 0),
        "discount_amount": float(bill.discount_amount or 0),
        "total_amount": total_amt,
        "amount_paid": round(effective_paid, 2),
        "deposit_applied": round(deposit_alloc, 2),
        "payments_recorded": round(total_paid, 2),
        "balance_due": round(max(0.0, total_amt - effective_paid), 2),
        "notes": bill.notes,
        "items": [
            {
                "id": it.id,
                "item_type": it.item_type,
                "item_name": it.item_name,
                "item_code": it.item_code,
                "quantity": it.quantity,
                "unit_price": float(it.unit_price),
                "total_price": float(it.total_price),
            }
            for it in items
        ],
        "payments": [
            {
                "id": p.id,
                "payment_number": p.payment_number,
                "amount_paid": float(p.amount_paid),
                "payment_method_name": p.payment_method_name or "cash",
                "payment_date": p.payment_date.isoformat() if p.payment_date else None,
                "transaction_reference": p.transaction_reference,
                "notes": p.notes,
                "parent_payment_id": p.parent_payment_id,
                "is_refund": float(p.amount_paid or 0) < 0,
                "reversed_at": p.reversed_at.isoformat() if p.reversed_at else None,
                "reversal_reason": p.reversal_reason,
            }
            for p in payments
        ],
    }


class BillPaymentRequest(BaseModel):
    amount_paid: float = Field(..., gt=0)
    payment_method: str = "cash"
    transaction_reference: Optional[str] = None
    notes: Optional[str] = None


def _sync_admission_item_payment_status(bill: Bill, db: Session) -> None:
    """For an IPD admission bill, cascade Bill.status to the payment_status of
    every PatientLabOrder consumed by it AND run the admission-level
    reconciliation so the deposit pool is folded in correctly. No-op for
    non-admission bills."""
    if (bill.bill_type or "") != "admission":
        return
    # Run the admission-wide reconciler so deposit pool + payments on sibling
    # bills are accounted for. This overrides any locally-computed bill.status.
    if bill.reference_id:
        from app.routes.inpatient import reconcile_admission_bill_statuses
        reconcile_admission_bill_statuses(db, bill.reference_id)
        db.flush()
    # Reconciler already cascades PatientLabOrder.payment_status, but ensure
    # this bill's labs are consistent even if reference_id was missing.
    target = "paid" if (bill.status == "paid") else "pending"
    db.query(PatientLabOrder).filter(
        PatientLabOrder.inpatient_bill_id == bill.id
    ).update({PatientLabOrder.payment_status: target}, synchronize_session=False)


@router.post("/billing/bills/{bill_id}/payment")
async def record_bill_payment(
    bill_id: int,
    req: BillPaymentRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Record a payment against a bill in the bills table."""
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin', 'receptionist']):
        raise HTTPException(status_code=403, detail="Not authorized")

    bill = db.query(Bill).filter(Bill.id == bill_id).first()
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    if bill.status == "cancelled":
        raise HTTPException(status_code=400, detail="Cannot pay a cancelled bill")
    if bill.status == "paid":
        raise HTTPException(status_code=400, detail="Bill is already fully paid")

    existing_paid = sum(float(p.amount_paid) for p in (bill.payments or []))
    balance = float(bill.total_amount or 0) - existing_paid

    if req.amount_paid > balance + 0.01:
        raise HTTPException(status_code=400, detail=f"Amount exceeds balance due (₹{balance:.2f})")

    # Generate payment number
    today_str = datetime.now().strftime("%Y%m%d")
    pay_prefix = f"PAY-{today_str}-"
    last_pay = db.query(Payment).filter(
        Payment.payment_number.like(f"{pay_prefix}%")
    ).order_by(Payment.id.desc()).first()
    seq = (int(last_pay.payment_number.split("-")[-1]) + 1) if last_pay else 1
    payment_number = f"{pay_prefix}{seq:04d}"

    payment = Payment(
        payment_number=payment_number,
        bill_id=bill.id,
        amount_paid=req.amount_paid,
        payment_method_name=req.payment_method,
        transaction_reference=req.transaction_reference,
        notes=req.notes,
        received_by_id=current_user.id,
    )
    db.add(payment)

    new_paid = existing_paid + req.amount_paid
    if new_paid >= float(bill.total_amount or 0) - 0.01:
        bill.status = "paid"
    else:
        bill.status = "partial"

    _sync_admission_item_payment_status(bill, db)

    db.commit()

    from app.services.audit_service import log_action
    log_action(db, current_user, "record_payment", "billing", "Payment", payment.id,
               f"Recorded payment {payment_number} of ₹{req.amount_paid:.2f} for bill {bill.bill_number}")

    return {
        "payment_id": payment.id,
        "payment_number": payment_number,
        "amount_paid": req.amount_paid,
        "bill_status": bill.status,
        "total_paid": new_paid,
        "balance_due": float(bill.total_amount or 0) - new_paid,
        "message": "Payment recorded successfully",
    }


# ---------------------------------------------------------------------------
# Discount / Tax adjustments
# ---------------------------------------------------------------------------

class BillDiscountRequest(BaseModel):
    discount_amount: Optional[float] = Field(None, ge=0)
    discount_percentage: Optional[float] = Field(None, ge=0, le=100)
    reason: str = Field(..., min_length=2, max_length=500)


class BillTaxRequest(BaseModel):
    tax_percentage: float = Field(..., ge=0, le=100)
    reason: str = Field(..., min_length=2, max_length=500)


def _ensure_bill_editable(bill: Bill):
    if bill.status == "cancelled":
        raise HTTPException(status_code=400, detail="Cannot modify a cancelled bill")
    if bill.status == "paid":
        raise HTTPException(status_code=400, detail="Bill is already fully paid; reverse a payment first")
    paid = sum(float(p.amount_paid or 0) for p in (bill.payments or []))
    if paid > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot modify a bill with payments recorded (paid ₹{paid:.2f}). Reverse payments first.",
        )


@router.patch("/billing/bills/{bill_id}/discount")
async def apply_bill_discount(
    bill_id: int,
    req: BillDiscountRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Apply (or replace) a flat-amount or percentage discount on a bill.
    Either discount_amount or discount_percentage is required."""
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin']):
        raise HTTPException(status_code=403, detail="Only admins can apply discounts")
    if req.discount_amount is None and req.discount_percentage is None:
        raise HTTPException(status_code=400, detail="Provide discount_amount or discount_percentage")

    bill = db.query(Bill).filter(Bill.id == bill_id).first()
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    _ensure_bill_editable(bill)

    subtotal = float(bill.subtotal or 0)
    if req.discount_percentage is not None:
        discount = round(subtotal * (req.discount_percentage / 100), 2)
    else:
        discount = round(float(req.discount_amount or 0), 2)
    if discount > subtotal:
        raise HTTPException(status_code=400, detail=f"Discount ₹{discount:.2f} exceeds subtotal ₹{subtotal:.2f}")

    bill.discount_amount = discount
    bill.total_amount = round(subtotal + float(bill.tax_amount or 0) - discount, 2)
    note_line = f"[DISCOUNT by user {current_user.id} on {datetime.now().isoformat()}]: Rs. {discount:.2f} — {req.reason}"
    bill.notes = (bill.notes + "\n" if bill.notes else "") + note_line
    db.commit()

    from app.services.audit_service import log_action
    log_action(db, current_user, "apply_discount", "billing", "Bill", bill.id,
               f"Applied discount ₹{discount:.2f} to bill {bill.bill_number}: {req.reason}",
               details={"discount_amount": discount, "discount_percentage": req.discount_percentage, "reason": req.reason})

    return {
        "bill_id": bill.id,
        "subtotal": subtotal,
        "discount_amount": float(bill.discount_amount),
        "tax_amount": float(bill.tax_amount or 0),
        "total_amount": float(bill.total_amount),
        "message": "Discount applied",
    }


@router.patch("/billing/bills/{bill_id}/tax")
async def apply_bill_tax(
    bill_id: int,
    req: BillTaxRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Apply a percentage tax on (subtotal - discount). Replaces any prior tax_amount."""
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin']):
        raise HTTPException(status_code=403, detail="Only admins can apply tax")

    bill = db.query(Bill).filter(Bill.id == bill_id).first()
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    _ensure_bill_editable(bill)

    subtotal = float(bill.subtotal or 0)
    discount = float(bill.discount_amount or 0)
    tax = round((subtotal - discount) * (req.tax_percentage / 100), 2)
    bill.tax_amount = tax
    bill.total_amount = round(subtotal + tax - discount, 2)
    note_line = f"[TAX by user {current_user.id} on {datetime.now().isoformat()}]: {req.tax_percentage}% (Rs. {tax:.2f}) — {req.reason}"
    bill.notes = (bill.notes + "\n" if bill.notes else "") + note_line
    db.commit()

    from app.services.audit_service import log_action
    log_action(db, current_user, "apply_tax", "billing", "Bill", bill.id,
               f"Applied {req.tax_percentage}% tax (₹{tax:.2f}) to bill {bill.bill_number}: {req.reason}",
               details={"tax_percentage": req.tax_percentage, "tax_amount": tax, "reason": req.reason})

    return {
        "bill_id": bill.id,
        "subtotal": subtotal,
        "discount_amount": discount,
        "tax_amount": float(bill.tax_amount),
        "total_amount": float(bill.total_amount),
        "message": "Tax applied",
    }


# ---------------------------------------------------------------------------
# Payment refund / reversal
# ---------------------------------------------------------------------------

class PaymentRefundRequest(BaseModel):
    amount: Optional[float] = Field(None, gt=0, description="Partial refund amount; if omitted, full remaining amount is refunded")
    reason: str = Field(..., min_length=2, max_length=500)


@router.post("/billing/payments/{payment_id}/refund")
async def refund_payment(
    payment_id: int,
    req: PaymentRefundRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Refund (fully or partially) a recorded payment by creating a negative
    Payment row linked back via parent_payment_id. Recomputes bill status."""
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin']):
        raise HTTPException(status_code=403, detail="Only admins can issue refunds")

    original = db.query(Payment).filter(Payment.id == payment_id).first()
    if not original:
        raise HTTPException(status_code=404, detail="Payment not found")
    if (original.amount_paid or 0) <= 0:
        raise HTTPException(status_code=400, detail="Cannot refund a refund row")

    # Sum prior refunds against this original
    prior_refunded = db.query(Payment).filter(Payment.parent_payment_id == original.id).all()
    refunded_so_far = sum(-float(p.amount_paid or 0) for p in prior_refunded)  # refunds are negative
    remaining = float(original.amount_paid) - refunded_so_far
    if remaining <= 0.01:
        raise HTTPException(status_code=400, detail="Payment already fully refunded")

    refund_amount = float(req.amount) if req.amount is not None else remaining
    if refund_amount > remaining + 0.01:
        raise HTTPException(
            status_code=400,
            detail=f"Refund ₹{refund_amount:.2f} exceeds remaining refundable ₹{remaining:.2f}",
        )

    # Build refund payment_number
    today_str = datetime.now().strftime("%Y%m%d")
    rfd_prefix = f"RFD-{today_str}-"
    last = db.query(Payment).filter(Payment.payment_number.like(f"{rfd_prefix}%")).order_by(Payment.id.desc()).first()
    seq = (int(last.payment_number.split("-")[-1]) + 1) if last else 1
    refund_number = f"{rfd_prefix}{seq:04d}"

    refund = Payment(
        payment_number=refund_number,
        bill_id=original.bill_id,
        amount_paid=-round(refund_amount, 2),
        payment_method_name=original.payment_method_name,
        payment_date=datetime.now(),
        notes=f"Refund of payment {original.payment_number}: {req.reason}",
        received_by_id=current_user.id,
        parent_payment_id=original.id,
    )
    db.add(refund)

    # Mark original fully reversed if everything refunded now
    if refund_amount + refunded_so_far >= float(original.amount_paid) - 0.01:
        original.reversed_by_id = current_user.id
        original.reversed_at = datetime.now()
        original.reversal_reason = req.reason

    # Recompute bill status
    bill = db.query(Bill).filter(Bill.id == original.bill_id).first()
    db.flush()
    net_paid = sum(float(p.amount_paid or 0) for p in (bill.payments or []))
    total = float(bill.total_amount or 0)
    if bill.status != "cancelled":
        if net_paid <= 0.01:
            bill.status = "pending"
        elif net_paid >= total - 0.01:
            bill.status = "paid"
        else:
            bill.status = "partial"

    _sync_admission_item_payment_status(bill, db)

    db.commit()

    from app.services.audit_service import log_action
    log_action(db, current_user, "refund_payment", "billing", "Payment", refund.id,
               f"Refunded ₹{refund_amount:.2f} of payment {original.payment_number} (bill {bill.bill_number}): {req.reason}",
               details={
                   "original_payment_id": original.id,
                   "original_payment_number": original.payment_number,
                   "amount_refunded": round(refund_amount, 2),
                   "reason": req.reason,
                   "bill_id": bill.id,
                   "new_bill_status": bill.status,
               })

    return {
        "refund_id": refund.id,
        "refund_number": refund_number,
        "amount_refunded": round(refund_amount, 2),
        "original_payment_id": original.id,
        "bill_status": bill.status,
        "net_paid": net_paid,
        "balance_due": total - net_paid,
        "fully_reversed": original.reversed_at is not None,
        "message": "Refund recorded",
    }


@router.get("/billing/payments/{payment_id}/refund-receipt/pdf")
async def refund_receipt_pdf(
    payment_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Generate a refund receipt PDF for a refund Payment row (negative amount)."""
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin', 'receptionist']):
        raise HTTPException(status_code=403, detail="Not authorized")

    refund = db.query(Payment).filter(Payment.id == payment_id).first()
    if not refund or (refund.amount_paid or 0) >= 0:
        raise HTTPException(status_code=404, detail="Refund payment not found")

    bill = db.query(Bill).filter(Bill.id == refund.bill_id).first()
    patient = db.query(Patient).filter(Patient.id == bill.patient_id).first() if bill else None
    original = db.query(Payment).filter(Payment.id == refund.parent_payment_id).first()
    hospital = db.query(Hospital).filter(Hospital.id == bill.hospital_id).first() if bill else None

    from app.utils.pdf_service import pdf_service
    hospital_info = {
        "name": hospital.name if hospital else "HOSPITAL",
        "address": hospital.address if hospital else "",
        "phone": hospital.phone if hospital else "",
        "email": hospital.email if hospital else "",
    }
    refund_data = {
        "refund_number": refund.payment_number,
        "refund_date": refund.payment_date.strftime("%d/%m/%Y %H:%M") if refund.payment_date else "",
        "amount": abs(float(refund.amount_paid or 0)),
        "payment_method": refund.payment_method_name or "cash",
        "reason": (refund.notes or "").split(":", 1)[-1].strip() if refund.notes else "",
        "patient_name": f"{patient.first_name} {patient.last_name}" if patient else "Unknown",
        "patient_phone": patient.primary_phone if patient else "",
        "village": (patient.village or "") if patient else "",
        "district": (patient.district or "") if patient else "",
        "bill_number": bill.bill_number if bill else "",
        "original_payment_number": original.payment_number if original else "",
        "original_amount": float(original.amount_paid or 0) if original else 0,
    }
    pdf_buffer = pdf_service.generate_refund_receipt_pdf(refund_data, hospital_info, **pdf_gen_kwargs(db, current_user.hospital_id, 'refund_receipt'))
    from fastapi.responses import Response
    return Response(content=pdf_buffer.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="refund_{refund.payment_number}.pdf"'})


# ---------------------------------------------------------------------------
# Billing reports
# ---------------------------------------------------------------------------

def _report_auth(current_user: User):
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin']):
        raise HTTPException(status_code=403, detail="Only admins can view billing reports")


def _parse_date_range(date_from: Optional[str], date_to: Optional[str]):
    today = datetime.now().date()
    d_to = datetime.strptime(date_to, "%Y-%m-%d").date() if date_to else today
    d_from = datetime.strptime(date_from, "%Y-%m-%d").date() if date_from else (d_to - timedelta(days=30))
    return d_from, d_to


@router.get("/billing/reports/daily-collection")
async def report_daily_collection(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Per-day collection grouped by payment method. Refund rows
    (negative amount_paid) net against collections so the row reflects net cash in."""
    _report_auth(current_user)
    d_from, d_to = _parse_date_range(date_from, date_to)

    payments = db.query(Payment).join(Bill, Payment.bill_id == Bill.id).filter(
        Bill.hospital_id == current_user.hospital_id,
        sql_func.date(Payment.payment_date) >= d_from,
        sql_func.date(Payment.payment_date) <= d_to,
    ).all()

    by_day: dict = {}
    methods_seen = set()
    for p in payments:
        if not p.payment_date:
            continue
        day = p.payment_date.date().isoformat()
        method = (p.payment_method_name or "cash").lower()
        methods_seen.add(method)
        by_day.setdefault(day, {"date": day, "total": 0.0, "refunds": 0.0, "by_method": {}})
        amt = float(p.amount_paid or 0)
        by_day[day]["total"] += amt
        if amt < 0:
            by_day[day]["refunds"] += -amt
        by_day[day]["by_method"][method] = by_day[day]["by_method"].get(method, 0.0) + amt

    rows = sorted(by_day.values(), key=lambda r: r["date"])
    grand_total = round(sum(r["total"] for r in rows), 2)
    grand_refunds = round(sum(r["refunds"] for r in rows), 2)
    return {
        "date_from": d_from.isoformat(), "date_to": d_to.isoformat(),
        "methods": sorted(methods_seen),
        "rows": [{**r, "total": round(r["total"], 2), "refunds": round(r["refunds"], 2),
                  "by_method": {k: round(v, 2) for k, v in r["by_method"].items()}} for r in rows],
        "totals": {"net_collected": grand_total, "refunds": grand_refunds, "gross_collected": round(grand_total + grand_refunds, 2)},
    }


@router.get("/billing/reports/doctor-revenue")
async def report_doctor_revenue(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Revenue per doctor, summed across non-cancelled consultations and
    admissions. Lab orders are excluded (no single attributable doctor)."""
    _report_auth(current_user)
    d_from, d_to = _parse_date_range(date_from, date_to)

    by_doc: dict = {}

    # Consultations (Appointments)
    appts = db.query(Appointment).join(Patient).filter(
        Patient.hospital_id == current_user.hospital_id,
        Appointment.appointment_date >= d_from,
        Appointment.appointment_date <= datetime.combine(d_to, datetime.max.time()),
        Appointment.payment_status != "cancelled",
    ).all()
    for a in appts:
        doc_id = a.doctor_id or 0
        amt = float((a.consultation_fee or 0) + (a.registration_fee or 0))
        if amt <= 0:
            continue
        by_doc.setdefault(doc_id, {"doctor_id": doc_id, "doctor_name": "", "consultation_revenue": 0.0,
                                    "admission_revenue": 0.0, "consultation_count": 0, "admission_count": 0})
        by_doc[doc_id]["consultation_revenue"] += amt
        by_doc[doc_id]["consultation_count"] += 1

    # Admission bills
    adm_bills = db.query(Bill).join(Patient).filter(
        Patient.hospital_id == current_user.hospital_id,
        Bill.bill_type == "admission",
        sql_func.date(Bill.bill_date) >= d_from,
        sql_func.date(Bill.bill_date) <= d_to,
        Bill.status != "cancelled",
    ).all()
    for b in adm_bills:
        adm = db.query(Admission).filter(Admission.id == b.reference_id).first() if b.reference_id else None
        doc_id = adm.admitting_doctor_id if adm and adm.admitting_doctor_id else 0
        amt = float(b.total_amount or 0)
        by_doc.setdefault(doc_id, {"doctor_id": doc_id, "doctor_name": "", "consultation_revenue": 0.0,
                                    "admission_revenue": 0.0, "consultation_count": 0, "admission_count": 0})
        by_doc[doc_id]["admission_revenue"] += amt
        by_doc[doc_id]["admission_count"] += 1

    # Resolve doctor names
    for doc_id, row in by_doc.items():
        if doc_id:
            u = db.query(User).filter(User.id == doc_id).first()
            row["doctor_name"] = f"Dr. {u.first_name} {u.last_name}" if u else f"User #{doc_id}"
        else:
            row["doctor_name"] = "(Unassigned)"
        row["total_revenue"] = round(row["consultation_revenue"] + row["admission_revenue"], 2)
        row["consultation_revenue"] = round(row["consultation_revenue"], 2)
        row["admission_revenue"] = round(row["admission_revenue"], 2)

    rows = sorted(by_doc.values(), key=lambda r: r["total_revenue"], reverse=True)
    return {
        "date_from": d_from.isoformat(), "date_to": d_to.isoformat(),
        "rows": rows,
        "totals": {
            "grand_total": round(sum(r["total_revenue"] for r in rows), 2),
            "consultation_total": round(sum(r["consultation_revenue"] for r in rows), 2),
            "admission_total": round(sum(r["admission_revenue"] for r in rows), 2),
        },
    }


@router.get("/billing/reports/tax-summary")
async def report_tax_summary(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Per-day GST/tax register for non-cancelled bills."""
    _report_auth(current_user)
    d_from, d_to = _parse_date_range(date_from, date_to)

    bills = db.query(Bill).join(Patient).filter(
        Patient.hospital_id == current_user.hospital_id,
        sql_func.date(Bill.bill_date) >= d_from,
        sql_func.date(Bill.bill_date) <= d_to,
        Bill.status != "cancelled",
        Bill.bill_type != "credit_note",
    ).all()

    by_day: dict = {}
    for b in bills:
        day = b.bill_date.date().isoformat() if b.bill_date else "unknown"
        by_day.setdefault(day, {"date": day, "taxable_value": 0.0, "tax_amount": 0.0, "bill_count": 0})
        subtotal = float(b.subtotal or 0)
        discount = float(b.discount_amount or 0)
        tax = float(b.tax_amount or 0)
        by_day[day]["taxable_value"] += max(subtotal - discount, 0)
        by_day[day]["tax_amount"] += tax
        by_day[day]["bill_count"] += 1

    rows = sorted(by_day.values(), key=lambda r: r["date"])
    return {
        "date_from": d_from.isoformat(), "date_to": d_to.isoformat(),
        "rows": [{**r, "taxable_value": round(r["taxable_value"], 2), "tax_amount": round(r["tax_amount"], 2)} for r in rows],
        "totals": {
            "taxable_value": round(sum(r["taxable_value"] for r in rows), 2),
            "tax_amount": round(sum(r["tax_amount"] for r in rows), 2),
            "bill_count": sum(r["bill_count"] for r in rows),
        },
    }


# ---------------------------------------------------------------------------
# Consolidated billing — combine multiple OPD + lab charges into one Bill
# ---------------------------------------------------------------------------

@router.get("/billing/consolidate/preview")
async def consolidate_preview(
    patient_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """List a patient's unbilled / pending consultations and lab orders that
    can be folded into a single consolidated bill."""
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin', 'receptionist']):
        raise HTTPException(status_code=403, detail="Not authorized")

    patient = db.query(Patient).filter(Patient.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    # Consultations / appointments still owing payment
    appts = db.query(Appointment).filter(
        Appointment.patient_id == patient_id,
        Appointment.payment_status.in_(["pending", "partial"]),
    ).order_by(Appointment.appointment_date.desc()).all()
    consultations = [
        {
            "id": a.id,
            "appointment_number": a.appointment_number,
            "date": a.appointment_date.isoformat() if a.appointment_date else None,
            "consultation_fee": float(a.consultation_fee or 0),
            "registration_fee": float(a.registration_fee or 0),
            "total": float((a.consultation_fee or 0) + (a.registration_fee or 0)),
            "payment_status": a.payment_status,
        }
        for a in appts if (a.consultation_fee or 0) + (a.registration_fee or 0) > 0
    ]

    lab_orders = db.query(PatientLabOrder).filter(
        PatientLabOrder.patient_id == patient_id,
        PatientLabOrder.payment_status.in_(["pending", "partial"]),
    ).all()
    labs = []
    for o in lab_orders:
        test = db.query(LabTest).filter(LabTest.id == o.test_id).first()
        if not test or (test.cost or 0) <= 0:
            continue
        labs.append({
            "id": o.id,
            "order_number": o.order_number,
            "test_name": test.name,
            "test_code": test.test_code,
            "cost": float(test.cost or 0),
            "payment_status": o.payment_status,
        })

    total_consult = sum(c["total"] for c in consultations)
    total_lab = sum(l["cost"] for l in labs)
    return {
        "patient_id": patient.id,
        "patient_name": f"{patient.first_name} {patient.last_name}",
        "consultations": consultations,
        "lab_orders": labs,
        "totals": {
            "consultation": total_consult,
            "lab": total_lab,
            "grand": total_consult + total_lab,
        },
    }


class ConsolidateBillRequest(BaseModel):
    patient_id: int
    consultation_ids: List[int] = Field(default_factory=list)
    lab_order_ids: List[int] = Field(default_factory=list)
    notes: Optional[str] = None


@router.post("/billing/consolidate")
async def create_consolidated_bill(
    req: ConsolidateBillRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Create a single Bill row consolidating selected consultations + lab
    orders for a patient. Marks source rows as payment_status='consolidated'
    so they drop out of future consolidation previews."""
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin', 'receptionist']):
        raise HTTPException(status_code=403, detail="Not authorized")
    if not req.consultation_ids and not req.lab_order_ids:
        raise HTTPException(status_code=400, detail="Pick at least one consultation or lab order")

    patient = db.query(Patient).filter(Patient.id == req.patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    today_str = datetime.now().strftime("%Y%m%d")
    cb_prefix = f"CB-{today_str}-"
    last = db.query(Bill).filter(Bill.bill_number.like(f"{cb_prefix}%")).order_by(Bill.id.desc()).first()
    seq = (int(last.bill_number.split("-")[-1]) + 1) if last else 1
    bill_number = f"{cb_prefix}{seq:04d}"

    bill = Bill(
        bill_number=bill_number,
        patient_id=patient.id,
        bill_type="consolidated",
        bill_subtype="final",
        reference_id=0,
        subtotal=0,
        tax_amount=0,
        discount_amount=0,
        total_amount=0,
        status="pending",
        bill_date=datetime.now(),
        created_by_id=current_user.id,
        hospital_id=patient.hospital_id,
        notes=req.notes or None,
    )
    db.add(bill)
    db.flush()

    subtotal = 0.0
    appts_to_mark = []
    for aid in req.consultation_ids:
        a = db.query(Appointment).filter(Appointment.id == aid, Appointment.patient_id == patient.id).first()
        if not a:
            continue
        if a.payment_status not in ("pending", "partial"):
            continue
        amount = float((a.consultation_fee or 0) + (a.registration_fee or 0))
        if amount <= 0:
            continue
        db.add(BillItem(
            bill_id=bill.id,
            item_type="consultation",
            item_name=f"Consultation {a.appointment_number}",
            item_code=f"APT-{a.id}",
            quantity=1,
            unit_price=amount,
            total_price=amount,
        ))
        subtotal += amount
        appts_to_mark.append(a)

    labs_to_mark = []
    for lid in req.lab_order_ids:
        o = db.query(PatientLabOrder).filter(PatientLabOrder.id == lid, PatientLabOrder.patient_id == patient.id).first()
        if not o:
            continue
        if o.payment_status not in ("pending", "partial"):
            continue
        test = db.query(LabTest).filter(LabTest.id == o.test_id).first()
        if not test or (test.cost or 0) <= 0:
            continue
        amount = float(test.cost or 0)
        db.add(BillItem(
            bill_id=bill.id,
            item_type="lab_test",
            item_name=test.name,
            item_code=f"LAB-{o.id}",
            quantity=1,
            unit_price=amount,
            total_price=amount,
        ))
        subtotal += amount
        labs_to_mark.append(o)

    if subtotal <= 0:
        db.rollback()
        raise HTTPException(status_code=400, detail="Selected items resolved to zero billable amount")

    bill.subtotal = round(subtotal, 2)
    bill.total_amount = round(subtotal, 2)

    for a in appts_to_mark:
        a.payment_status = "consolidated"
    for o in labs_to_mark:
        o.payment_status = "consolidated"

    db.commit()
    db.refresh(bill)

    from app.services.audit_service import log_action
    log_action(db, current_user, "create_consolidated_bill", "billing", "Bill", bill.id,
               f"Created consolidated bill {bill_number} for {patient.first_name} {patient.last_name} "
               f"({len(appts_to_mark)} consultations + {len(labs_to_mark)} lab orders, ₹{subtotal:.2f})",
               details={
                   "patient_id": patient.id,
                   "consultation_ids": [a.id for a in appts_to_mark],
                   "lab_order_ids": [o.id for o in labs_to_mark],
                   "total": round(subtotal, 2),
               })

    return {
        "bill_id": bill.id,
        "bill_number": bill_number,
        "total_amount": bill.total_amount,
        "consultations_count": len(appts_to_mark),
        "lab_orders_count": len(labs_to_mark),
        "message": "Consolidated bill created",
    }


# ---------------------------------------------------------------------------
# Credit notes
# ---------------------------------------------------------------------------

class CreditNoteItemRequest(BaseModel):
    item_name: str = Field(..., min_length=1, max_length=200)
    item_code: Optional[str] = None
    quantity: int = Field(1, gt=0)
    unit_price: float = Field(..., gt=0)


class CreditNoteRequest(BaseModel):
    items: List[CreditNoteItemRequest] = Field(..., min_length=1)
    reason: str = Field(..., min_length=2, max_length=500)


@router.post("/billing/bills/{bill_id}/credit-note")
async def issue_credit_note(
    bill_id: int,
    req: CreditNoteRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Issue a credit note against a bill. Creates a new Bill row with
    bill_type='credit_note', negative total, parent_bill_id pointing to the
    original. Also records a 'credit_note' Payment on the original bill so
    the balance reduces and status recomputes."""
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin']):
        raise HTTPException(status_code=403, detail="Only admins can issue credit notes")

    parent = db.query(Bill).filter(Bill.id == bill_id).first()
    if not parent:
        raise HTTPException(status_code=404, detail="Bill not found")
    if parent.bill_type == "credit_note":
        raise HTTPException(status_code=400, detail="Cannot issue a credit note against another credit note")
    if parent.status == "cancelled":
        raise HTTPException(status_code=400, detail="Cannot issue a credit note for a cancelled bill")

    # Compute net amount available to credit: remaining balance OR remaining total minus prior CNs
    prior_cn_amount = sum(
        abs(float(cn.total_amount or 0))
        for cn in db.query(Bill).filter(Bill.parent_bill_id == parent.id, Bill.bill_type == "credit_note").all()
    )
    parent_total = float(parent.total_amount or 0)
    creditable = parent_total - prior_cn_amount
    if creditable <= 0.01:
        raise HTTPException(status_code=400, detail="Bill already fully credited")

    cn_subtotal = round(sum(it.quantity * it.unit_price for it in req.items), 2)
    if cn_subtotal > creditable + 0.01:
        raise HTTPException(
            status_code=400,
            detail=f"Credit note amount ₹{cn_subtotal:.2f} exceeds creditable balance ₹{creditable:.2f}",
        )

    # Build credit note bill_number
    today_str = datetime.now().strftime("%Y%m%d")
    cn_prefix = f"CN-{today_str}-"
    last = db.query(Bill).filter(Bill.bill_number.like(f"{cn_prefix}%")).order_by(Bill.id.desc()).first()
    seq = (int(last.bill_number.split("-")[-1]) + 1) if last else 1
    cn_number = f"{cn_prefix}{seq:04d}"

    cn = Bill(
        bill_number=cn_number,
        patient_id=parent.patient_id,
        bill_type="credit_note",
        bill_subtype="final",
        reference_id=parent.id,
        parent_bill_id=parent.id,
        subtotal=-cn_subtotal,
        tax_amount=0,
        discount_amount=0,
        total_amount=-cn_subtotal,
        status="paid",  # CN itself is "settled" immediately upon issue
        bill_date=datetime.now(),
        created_by_id=current_user.id,
        hospital_id=parent.hospital_id,
        notes=f"Credit note for bill {parent.bill_number}: {req.reason}",
    )
    db.add(cn)
    db.flush()

    for it in req.items:
        db.add(BillItem(
            bill_id=cn.id,
            item_type="credit",
            item_name=it.item_name,
            item_code=it.item_code,
            quantity=it.quantity,
            unit_price=it.unit_price,
            total_price=round(it.quantity * it.unit_price, 2),
        ))

    # Record a Payment-style entry on the parent so balance reduces and status recomputes.
    today_str2 = datetime.now().strftime("%Y%m%d")
    pay_prefix = f"PAY-{today_str2}-"
    last_pay = db.query(Payment).filter(Payment.payment_number.like(f"{pay_prefix}%")).order_by(Payment.id.desc()).first()
    pay_seq = (int(last_pay.payment_number.split("-")[-1]) + 1) if last_pay else 1
    db.add(Payment(
        payment_number=f"{pay_prefix}{pay_seq:04d}",
        bill_id=parent.id,
        amount_paid=cn_subtotal,
        payment_method_name="credit_note",
        payment_date=datetime.now(),
        notes=f"Credit note {cn_number}: {req.reason}",
        received_by_id=current_user.id,
    ))

    db.flush()
    # Recompute parent status
    parent_paid = sum(float(p.amount_paid or 0) for p in (parent.payments or []))
    if parent.status != "cancelled":
        if parent_paid <= 0.01:
            parent.status = "pending"
        elif parent_paid >= parent_total - 0.01:
            parent.status = "paid"
        else:
            parent.status = "partial"

    _sync_admission_item_payment_status(parent, db)

    db.commit()

    from app.services.audit_service import log_action
    log_action(db, current_user, "issue_credit_note", "billing", "Bill", cn.id,
               f"Issued credit note {cn_number} for ₹{cn_subtotal:.2f} against bill {parent.bill_number}: {req.reason}",
               details={"parent_bill_id": parent.id, "amount": cn_subtotal, "reason": req.reason})

    return {
        "credit_note_id": cn.id,
        "credit_note_number": cn_number,
        "amount": cn_subtotal,
        "parent_bill_id": parent.id,
        "parent_bill_status": parent.status,
        "parent_balance_due": parent_total - parent_paid,
        "message": "Credit note issued",
    }


@router.get("/billing/bills/{bill_id}/pdf")
async def get_bill_pdf(
    bill_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Generic PDF for any Bill row (procedure / consolidated / etc.).
    Uses the standard OPD bill template via `generate_bill_pdf`."""
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin', 'receptionist', 'doctor']):
        raise HTTPException(status_code=403, detail="Not authorized")

    bill = db.query(Bill).filter(Bill.id == bill_id).first()
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")

    patient = db.query(Patient).filter(Patient.id == bill.patient_id).first()
    hospital = db.query(Hospital).filter(Hospital.id == bill.hospital_id).first()
    items = db.query(BillItem).filter(BillItem.bill_id == bill.id).all()
    payments = db.query(Payment).filter(Payment.bill_id == bill.id).all()
    paid = sum(float(p.amount_paid or 0) for p in payments)
    created_by = db.query(User).filter(User.id == bill.created_by_id).first() if bill.created_by_id else None
    prepared_by = (f"{created_by.first_name} {created_by.last_name}".strip()
                   or created_by.username) if created_by else ""

    # Hide Paid/Balance for fresh unpaid bills (procedure / consolidated /
    # any bill that has no payments yet) — the printed bill should read as
    # an invoice, not a paid receipt.
    hide_payment_summary = (paid <= 0.01)
    bill_data = {
        "bill_number": bill.bill_number,
        "bill_date": bill.bill_date.isoformat() if bill.bill_date else "",
        "patient_name": f"{patient.first_name} {patient.last_name}" if patient else "Unknown",
        "patient_age": getattr(patient, 'age', None) if patient else None,
        "patient_gender": patient.gender if patient else "",
        "patient_phone": patient.primary_phone if patient else "",
        "mrn": (patient.mrn or "") if patient else "",
        "village": (patient.village or "") if patient else "",
        "district": (patient.district or "") if patient else "",
        "payment_method": (payments[0].payment_method_name if payments else "Cash"),
        "items": [
            {"item_name": it.item_name, "item_code": it.item_code or "",
             "quantity": it.quantity, "total_price": float(it.total_price or 0)}
            for it in items
        ],
        "subtotal": float(bill.subtotal or 0),
        "discount_amount": float(bill.discount_amount or 0),
        "tax_amount": float(bill.tax_amount or 0),
        "total_amount": float(bill.total_amount or 0),
        "amount_paid": paid,
        "balance_due": float(bill.total_amount or 0) - paid,
        "hide_payment_summary": hide_payment_summary,
        "prepared_by": prepared_by,
        "referred_by": bill.referred_by or "",
        "notes": bill.notes or "",
    }
    hospital_info = {
        "name": hospital.name if hospital else "HOSPITAL",
        "address": hospital.address if hospital else "",
        "phone": hospital.phone if hospital else "",
        "email": hospital.email if hospital else "",
        "logo_url": getattr(hospital, "logo_url", "") if hospital else "",
        "hospital_subname": getattr(hospital, "hospital_subname", "") if hospital else "",
    }
    from app.utils.pdf_service import pdf_service
    buf = pdf_service.generate_bill_pdf(bill_data, hospital_info, **bill_pdf_gen_kwargs(db, current_user.hospital_id, 'opd_bill'))
    from fastapi.responses import Response
    return Response(content=buf.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="{bill.bill_number}.pdf"'})


@router.get("/billing/bills/{bill_id}/credit-note/pdf")
async def credit_note_pdf(
    bill_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """PDF for a credit-note Bill row."""
    if not any(r in current_user.role_names for r in ['super_admin', 'hospital_admin', 'receptionist']):
        raise HTTPException(status_code=403, detail="Not authorized")

    cn = db.query(Bill).filter(Bill.id == bill_id, Bill.bill_type == "credit_note").first()
    if not cn:
        raise HTTPException(status_code=404, detail="Credit note not found")

    parent = db.query(Bill).filter(Bill.id == cn.parent_bill_id).first() if cn.parent_bill_id else None
    patient = db.query(Patient).filter(Patient.id == cn.patient_id).first()
    hospital = db.query(Hospital).filter(Hospital.id == cn.hospital_id).first()
    items = db.query(BillItem).filter(BillItem.bill_id == cn.id).all()

    from app.utils.pdf_service import pdf_service
    hospital_info = {
        "name": hospital.name if hospital else "HOSPITAL",
        "address": hospital.address if hospital else "",
        "phone": hospital.phone if hospital else "",
        "email": hospital.email if hospital else "",
    }
    cn_data = {
        "credit_note_number": cn.bill_number,
        "credit_note_date": cn.bill_date.strftime("%d/%m/%Y %H:%M") if cn.bill_date else "",
        "amount": abs(float(cn.total_amount or 0)),
        "reason": (cn.notes or "").split(":", 1)[-1].strip() if cn.notes else "",
        "patient_name": f"{patient.first_name} {patient.last_name}" if patient else "Unknown",
        "patient_phone": patient.primary_phone if patient else "",
        "village": (patient.village or "") if patient else "",
        "district": (patient.district or "") if patient else "",
        "parent_bill_number": parent.bill_number if parent else "",
        "parent_bill_total": float(parent.total_amount or 0) if parent else 0,
        "items": [
            {"name": it.item_name, "code": it.item_code or "", "qty": it.quantity,
             "unit_price": float(it.unit_price), "total": float(it.total_price)}
            for it in items
        ],
    }
    pdf_buffer = pdf_service.generate_credit_note_pdf(cn_data, hospital_info, **pdf_gen_kwargs(db, current_user.hospital_id, 'credit_note'))
    from fastapi.responses import Response
    return Response(content=pdf_buffer.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="{cn.bill_number}.pdf"'})