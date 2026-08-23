"""CA-ready GST audit workbook (GSTR-shaped Excel, not a GSTN upload)."""
from io import BytesIO
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.services import gst_report_service as reports
from app.services.gst_classification import MODULE_LABELS


HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONEY_FORMAT = '#,##0.00'
THIN = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _autosize(ws, min_width=10, max_width=28):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_width
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            width = max(width, min(len(val) + 2, max_width))
        ws.column_dimensions[letter].width = width


def _append_dict_sheet(ws, columns, rows, money_cols=None):
    """columns: list of (key, label)."""
    money_cols = set(money_cols or [])
    ws.append([label for _, label in columns])
    _style_header(ws)
    for row in rows:
        values = []
        for key, _label in columns:
            values.append(row.get(key, ""))
        ws.append(values)
        r_idx = ws.max_row
        for i, (key, _) in enumerate(columns, start=1):
            cell = ws.cell(r_idx, i)
            cell.border = THIN
            if key in money_cols:
                cell.number_format = MONEY_FORMAT
    _autosize(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_gst_audit_workbook(db, hospital, d_from: date, d_to: date, module=None) -> bytes:
    from app.services.gst_classification import (
        module_in_gst_scope, scope_includes_hospital_credit_notes,
        scope_includes_pharmacy_docs,
    )
    from app.services.gst_return_forms import has_inward_books, normalize_module

    hid = hospital.id
    selected = normalize_module(module)
    sales = reports.sales_summary(db, hid, d_from, d_to)
    if selected:
        by_mod = [
            row for row in (sales.get("by_module") or [])
            if module_in_gst_scope(selected, row.get("module"))
        ]
        sales = {**sales, "by_module": by_mod}
    if has_inward_books(selected):
        purchases = reports.purchase_summary(db, hid, d_from, d_to, group_by="supplier")
    else:
        purchases = {"invoices": [], "returns": [], "rows": []}
    pharm_outward = scope_includes_pharmacy_docs(selected)
    if pharm_outward:
        outward = reports.gst_outward_hsn(db, hid, d_from, d_to, module=selected)
        b2 = reports.gst_b2b_b2c(db, hid, d_from, d_to, module=selected)
    else:
        outward = {"rows": []}
        b2 = {"b2b": [], "b2c": []}
    exempt = reports.gst_exempt_register(db, hid, d_from, d_to, module=selected or "all")
    cdnr = reports.gst_cdnr(db, hid, d_from, d_to)
    legacy = reports.gst_legacy_tax(db, hid, d_from, d_to)
    gstr3b = reports.gstr3b_summary(db, hid, d_from, d_to, module=selected)
    invoices = reports.collect_sales_invoices(db, hid, d_from, d_to)
    if selected:
        invoices = [r for r in invoices if module_in_gst_scope(selected, r.get("module"))]
        cdnr_rows = cdnr.get("rows") or []
        if not scope_includes_pharmacy_docs(selected):
            cdnr_rows = [r for r in cdnr_rows if r.get("kind") != "sale_return"]
        if not scope_includes_hospital_credit_notes(selected):
            cdnr_rows = [r for r in cdnr_rows if r.get("kind") != "credit_note"]
        cdnr = {**cdnr, "rows": cdnr_rows}

    wb = Workbook()

    # Cover
    ws = wb.active
    ws.title = "Cover"
    ws["A1"] = "KT HEALTH ERP — GST audit pack"
    ws["A1"].font = Font(bold=True, size=16)
    meta = [
        ("Hospital", hospital.name),
        ("GSTIN", getattr(hospital, "gstin", None) or hospital.tax_id or ""),
        ("State code", getattr(hospital, "gst_state_code", None) or ""),
        ("Module", (gstr3b.get("hospital") or {}).get("module_label") or "All modules"),
        ("Period from", d_from.isoformat()),
        ("Period to", d_to.isoformat()),
        ("Generated for", "Chartered accountant / GSTR working papers"),
        ("Note", "This is not a GSTN JSON upload. Import figures into the GST portal or GSTR utility."),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, v)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 70

    # GSTR-1 B2B
    ws = wb.create_sheet("GSTR-1 B2B")
    _append_dict_sheet(ws, [
        ("date", "Invoice date"), ("number", "Invoice no."), ("party", "Customer"),
        ("gstin", "GSTIN"), ("hsn_code", "HSN"),
        ("taxable_value", "Taxable value"),
        ("sgst_pct", "SGST %"), ("cgst_pct", "CGST %"), ("igst_pct", "IGST %"),
        ("sgst_amount", "SGST"), ("cgst_amount", "CGST"), ("igst_amount", "IGST"),
        ("total_tax", "Total tax"),
    ], b2["b2b"], {"taxable_value", "sgst_amount", "cgst_amount", "igst_amount", "total_tax",
                   "sgst_pct", "cgst_pct", "igst_pct"})

    ws = wb.create_sheet("GSTR-1 B2CS")
    _append_dict_sheet(ws, [
        ("rate", "Rate %"), ("sgst_pct", "SGST %"), ("cgst_pct", "CGST %"), ("igst_pct", "IGST %"),
        ("invoice_count", "Invoices"), ("taxable_value", "Taxable value"),
        ("sgst_amount", "SGST"), ("cgst_amount", "CGST"), ("igst_amount", "IGST"),
        ("total_tax", "Total tax"),
    ], b2["b2c"], {"rate", "sgst_pct", "cgst_pct", "igst_pct", "taxable_value",
                   "sgst_amount", "cgst_amount", "igst_amount", "total_tax"})

    ws = wb.create_sheet("GSTR-1 HSN")
    _append_dict_sheet(ws, [
        ("hsn_code", "HSN"), ("qty", "Qty"),
        ("sgst_pct", "SGST %"), ("cgst_pct", "CGST %"), ("igst_pct", "IGST %"),
        ("taxable_value", "Taxable value"),
        ("sgst_amount", "SGST"), ("cgst_amount", "CGST"), ("igst_amount", "IGST"),
        ("total_tax", "Total tax"),
    ], outward["rows"], {"qty", "sgst_pct", "cgst_pct", "igst_pct", "taxable_value",
                         "sgst_amount", "cgst_amount", "igst_amount", "total_tax"})

    ws = wb.create_sheet("GSTR-1 CDNR")
    _append_dict_sheet(ws, [
        ("kind", "Type"), ("date", "Date"), ("number", "Number"),
        ("party", "Party"), ("gstin", "GSTIN"),
        ("taxable_value", "Taxable value"), ("tax", "Tax"), ("grand_total", "Grand total"),
    ], cdnr["rows"], {"taxable_value", "tax", "grand_total"})

    ws = wb.create_sheet("Exempt")
    _append_dict_sheet(ws, [
        ("date", "Date"), ("number", "Invoice / ref"), ("module_label", "Module"),
        ("party", "Patient"), ("gstin", "GSTIN"), ("hsn_sac", "SAC"),
        ("billed", "Exempt value"),
    ], exempt["rows"], {"billed"})

    ws = wb.create_sheet("Purchase register")
    _append_dict_sheet(ws, [
        ("date", "Date"), ("number", "GRN no."), ("invoice_number", "Supplier invoice"),
        ("supplier", "Supplier"), ("gstin", "GSTIN"),
        ("taxable", "Taxable"), ("sgst", "SGST"), ("cgst", "CGST"), ("igst", "IGST"),
        ("total_tax", "Total tax"), ("grand_total", "Grand total"),
    ], purchases["invoices"], {"taxable", "sgst", "cgst", "igst", "total_tax", "grand_total"})

    ws = wb.create_sheet("Purchase returns")
    _append_dict_sheet(ws, [
        ("date", "Date"), ("number", "Return no."), ("supplier", "Supplier"),
        ("gstin", "GSTIN"), ("taxable", "Taxable"), ("sgst", "SGST"),
        ("cgst", "CGST"), ("igst", "IGST"), ("total_tax", "Total tax"),
        ("grand_total", "Grand total"),
    ], purchases["returns"], {"taxable", "sgst", "cgst", "igst", "total_tax", "grand_total"})

    ws = wb.create_sheet("GSTR-3B summary")
    ws.append(["Particulars", "Amount"])
    _style_header(ws)
    for label, key in [
        ("Outward taxable value", "outward_taxable"),
        ("Outward SGST", "outward_sgst"),
        ("Outward CGST", "outward_cgst"),
        ("Outward IGST", "outward_igst"),
        ("Outward tax", "outward_tax"),
        ("Inward taxable (ITC base)", "inward_taxable"),
        ("ITC SGST", "itc_sgst"),
        ("ITC CGST", "itc_cgst"),
        ("ITC IGST", "itc_igst"),
        ("ITC total", "itc_total"),
        ("Exempt / nil (SAC 9993)", "exempt_value"),
        ("Net tax payable (outward − ITC)", "tax_payable"),
    ]:
        ws.append([label, gstr3b[key]])
        ws.cell(ws.max_row, 2).number_format = MONEY_FORMAT
    _autosize(ws)
    ws.column_dimensions["A"].width = 40

    ws = wb.create_sheet("Sales by module")
    _append_dict_sheet(ws, [
        ("module_label", "Module"), ("count", "Bills"),
        ("billed", "Billed"), ("discount", "Discount"), ("tax", "Tax"),
        ("collected", "Collected"), ("outstanding", "Outstanding"),
    ], sales["by_module"], {"billed", "discount", "tax", "collected", "outstanding"})

    money_inv = {"billed", "discount", "tax", "net", "collected", "outstanding"}
    inv_cols = [
        ("date", "Date"), ("number", "Number"), ("party", "Party"),
        ("gstin", "GSTIN"), ("billed", "Billed"), ("discount", "Discount"),
        ("tax", "Tax"), ("collected", "Collected"), ("outstanding", "Outstanding"),
        ("status", "Status"),
    ]
    used_names = {ws.title.lower() for ws in wb.worksheets}
    for module, label in MODULE_LABELS.items():
        mod_rows = [r for r in invoices if r["module"] == module]
        if not mod_rows:
            continue
        title = label[:31]
        # Excel sheet names must be unique and cannot contain : \ / ? * [ ]
        safe = "".join(ch if ch not in r':\/?*[]' else " " for ch in title).strip() or module
        base = safe
        n = 2
        while safe.lower() in used_names:
            safe = f"{base[:28]} {n}"
            n += 1
        used_names.add(safe.lower())
        ws = wb.create_sheet(safe)
        _append_dict_sheet(ws, inv_cols, mod_rows, money_inv)

    ws = wb.create_sheet("Purchase summary")
    _append_dict_sheet(ws, [
        ("bucket", "Supplier / day"), ("count", "Docs"),
        ("taxable", "Taxable"), ("sgst", "SGST"), ("cgst", "CGST"),
        ("igst", "IGST"), ("total_tax", "Total tax"), ("grand_total", "Grand total"),
    ], purchases["rows"], {"taxable", "sgst", "cgst", "igst", "total_tax", "grand_total"})

    ws = wb.create_sheet("Legacy tax")
    _append_dict_sheet(ws, [
        ("date", "Date"), ("number", "Bill no."), ("bill_type", "Type"),
        ("taxable_value", "Taxable value"), ("tax_amount", "Flat tax_amount"),
    ], legacy["rows"], {"taxable_value", "tax_amount"})

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


TITLE_FONT = Font(bold=True, size=12)
BOLD = Font(bold=True)
MONEY_COLS_COMMON = {
    "taxable_value", "taxable", "invoice_value", "note_value", "cess",
    "igst", "cgst", "sgst", "total_value", "qty", "rate", "total_tax",
    "nil", "exempt", "non_gst", "grand_total",
}


def _cover_sheet(wb, title: str, hospital_meta: dict, extra=None):
    ws = wb.active
    ws.title = "Cover"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16)
    rows = [
        ("Hospital", hospital_meta.get("name") or ""),
        ("GSTIN", hospital_meta.get("gstin") or ""),
        ("State code", hospital_meta.get("state_code") or ""),
        ("Place of supply", hospital_meta.get("place_of_supply") or ""),
        ("Period", hospital_meta.get("period_label") or ""),
        ("Module", hospital_meta.get("module_label") or "All modules"),
        ("From", hospital_meta.get("date_from") or ""),
        ("To", hospital_meta.get("date_to") or ""),
        ("Note", hospital_meta.get("disclaimer") or (
            "Working paper — not a GSTN JSON / offline-utility upload."
        )),
    ]
    for k, v in extra or []:
        rows.append((k, v))
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(i, 1, k).font = BOLD
        ws.cell(i, 2, v)
        ws.cell(i, 2).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80
    return ws


def _gstn_table_sheet(wb, name, title, columns, rows, summary_labels=None, summary_values=None,
                      money_cols=None):
    """GSTN-style sheet: title, optional summary row, then header + data."""
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    start = 4
    if summary_labels:
        for i, label in enumerate(summary_labels, start=1):
            ws.cell(2, i, label).font = BOLD
        if summary_values:
            for i, val in enumerate(summary_values, start=1):
                cell = ws.cell(3, i, val)
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    cell.number_format = MONEY_FORMAT
        start = 5
    money_cols = set(money_cols or MONEY_COLS_COMMON)
    for i, (_key, label) in enumerate(columns, start=1):
        cell = ws.cell(start, i, label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN
    for r_i, row in enumerate(rows or [], start=start + 1):
        for c_i, (key, _label) in enumerate(columns, start=1):
            val = row.get(key, "")
            cell = ws.cell(r_i, c_i, val)
            cell.border = THIN
            if key in money_cols and isinstance(val, (int, float)):
                cell.number_format = MONEY_FORMAT
    _autosize(ws, min_width=12, max_width=32)
    ws.freeze_panes = f"A{start + 1}"
    return ws


def _meta_with_disclaimer(data: dict) -> dict:
    h = dict(data.get("hospital") or {})
    h["disclaimer"] = data.get("disclaimer") or h.get("disclaimer")
    return h


def build_gstr1_workbook(data: dict) -> bytes:
    wb = Workbook()
    _cover_sheet(wb, "KT HEALTH ERP — GSTR-1 working paper", _meta_with_disclaimer(data))

    b2b = data.get("b2b") or {}
    s = b2b.get("summary") or {}
    _gstn_table_sheet(
        wb, "b2b,sez,de", "Summary For B2B(4)",
        [
            ("gstin", "GSTIN/UIN of Recipient"),
            ("receiver_name", "Receiver Name"),
            ("invoice_number", "Invoice Number"),
            ("invoice_date", "Invoice date"),
            ("invoice_value", "Invoice Value"),
            ("place_of_supply", "Place Of Supply"),
            ("reverse_charge", "Reverse Charge"),
            ("applicable_pct", "Applicable %Tax"),
            ("invoice_type", "Invoice Type"),
            ("ecommerce_gstin", "E-Commerce GSTIN"),
            ("rate", "Rate"),
            ("taxable_value", "Taxable Value"),
            ("cess", "Cess Amount"),
        ],
        b2b.get("rows") or [],
        summary_labels=["No. of Recipients", "", "No. of Invoices", "", "Total Invoice Value",
                        "", "", "", "", "", "", "Taxable Value", "Cess Amount"],
        summary_values=[s.get("recipients", 0), "", s.get("invoices", 0), "",
                        s.get("invoice_value", 0), "", "", "", "", "", "",
                        s.get("taxable_value", 0), s.get("cess", 0)],
    )

    b2cl = data.get("b2cl") or {}
    s = b2cl.get("summary") or {}
    _gstn_table_sheet(
        wb, "b2cl", "Summary For B2CL(5)",
        [
            ("invoice_number", "Invoice Number"),
            ("invoice_date", "Invoice date"),
            ("invoice_value", "Invoice Value"),
            ("place_of_supply", "Place Of Supply"),
            ("applicable_pct", "Applicable %Tax"),
            ("rate", "Rate"),
            ("taxable_value", "Taxable Value"),
            ("cess", "Cess Amount"),
            ("ecommerce_gstin", "E-Commerce GSTIN"),
        ],
        b2cl.get("rows") or [],
        summary_labels=["No. of Invoices", "", "Total Inv Value", "", "", "",
                        "Total Taxable Value", "Total Cess"],
        summary_values=[s.get("invoices", 0), "", s.get("invoice_value", 0), "", "", "",
                        s.get("taxable_value", 0), s.get("cess", 0)],
    )

    b2cs = data.get("b2cs") or {}
    s = b2cs.get("summary") or {}
    _gstn_table_sheet(
        wb, "b2cs", "Summary For B2CS(7)",
        [
            ("type", "Type"),
            ("place_of_supply", "Place of Supply"),
            ("applicable_pct", "Applicable %Tax"),
            ("rate", "Rate"),
            ("taxable_value", "Taxable Value"),
            ("cess", "Cess Amount"),
            ("ecommerce_gstin", "E-Commerce GSTIN"),
        ],
        b2cs.get("rows") or [],
        summary_labels=["", "", "", "", "Total Taxable  Value", "Total Cess"],
        summary_values=["", "", "", "", s.get("taxable_value", 0), s.get("cess", 0)],
    )

    cdnr = data.get("cdnr") or {}
    s = cdnr.get("summary") or {}
    _gstn_table_sheet(
        wb, "cdnr", "Summary For CDNR (9B)",
        [
            ("gstin", "GSTIN/UIN of Recipient"),
            ("receiver_name", "Receiver Name"),
            ("note_number", "Note Number"),
            ("note_date", "Note date"),
            ("note_type", "Note Type"),
            ("place_of_supply", "Place Of Supply"),
            ("reverse_charge", "Reverse Charge"),
            ("note_supply_type", "Note Supply Type"),
            ("note_value", "Note Value"),
            ("applicable_pct", "Applicable % of Tax Rate"),
            ("rate", "Rate"),
            ("taxable_value", "Taxable Value"),
            ("cess", "Cess Amount"),
        ],
        cdnr.get("rows") or [],
        summary_labels=["No. of Recipients", "", "No. of Notes", "", "", "", "", "",
                        "Total Note Value", "", "", "Total Taxable Value", "Total Cess"],
        summary_values=[s.get("recipients", 0), "", s.get("notes", 0), "", "", "", "", "",
                        s.get("note_value", 0), "", "", s.get("taxable_value", 0), s.get("cess", 0)],
    )

    cdnur = data.get("cdnur") or {}
    s = cdnur.get("summary") or {}
    _gstn_table_sheet(
        wb, "cdnur", "Summary For CDNUR (9B)",
        [
            ("ur_type", "UR Type"),
            ("note_number", "Note Number"),
            ("note_date", "Note date"),
            ("note_type", "Note Type"),
            ("place_of_supply", "Place Of Supply"),
            ("note_value", "Note Value"),
            ("applicable_pct", "Applicable % of Tax Rate"),
            ("rate", "Rate"),
            ("taxable_value", "Taxable Value"),
            ("cess", "Cess Amount"),
        ],
        cdnur.get("rows") or [],
        summary_labels=["", "No. of Notes/Vouchers", "", "", "", "Total Note Value", "", "",
                        "Total Taxable Value", "Total Cess"],
        summary_values=["", s.get("notes", 0), "", "", "", s.get("note_value", 0), "", "",
                        s.get("taxable_value", 0), s.get("cess", 0)],
    )

    _gstn_table_sheet(
        wb, "exp", "Summary For EXP(6)",
        [
            ("export_type", "Export Type"),
            ("invoice_number", "Invoice Number"),
            ("invoice_date", "Invoice date"),
            ("invoice_value", "Invoice Value"),
            ("port_code", "Port Code"),
            ("shipping_bill_number", "Shipping Bill Number"),
            ("shipping_bill_date", "Shipping Bill Date"),
            ("rate", "Rate"),
            ("taxable_value", "Taxable Value"),
            ("cess", "Cess Value"),
        ],
        [],
    )
    _gstn_table_sheet(
        wb, "at", "Summary For Advance Received(11B)",
        [
            ("place_of_supply", "Place of Supply"),
            ("applicable_pct", "Applicable %Tax"),
            ("rate", "Rate"),
            ("gross_advance", "Gross Advance Received"),
            ("cess", "Cess Amount"),
        ],
        [],
    )
    _gstn_table_sheet(
        wb, "atadj", "Summary For Advance Adjusted(11B)",
        [
            ("place_of_supply", "Place Of Supply"),
            ("applicable_pct", "Applicable %Tax"),
            ("rate", "Rate"),
            ("gross_advance", "Gross Advance Adjusted"),
            ("cess", "Cess Amount"),
        ],
        [],
    )

    exemp = data.get("exemp") or {}
    s = exemp.get("summary") or {}
    _gstn_table_sheet(
        wb, "exemp", "Summary For Nil rated, exempted and non GST outward supplies (8)",
        [
            ("description", "Description"),
            ("nil", "Nil Rated Supplies"),
            ("exempt", "Exempted (other than nil rated/non GST supply )"),
            ("non_gst", "Non-GST supplies"),
        ],
        exemp.get("rows") or [],
        summary_labels=["", "Total Nil Rated Supplies", "Total Exempted Supplies", "Total Non-GST Supplies"],
        summary_values=["", s.get("nil", 0), s.get("exempt", 0), s.get("non_gst", 0)],
        money_cols={"nil", "exempt", "non_gst"},
    )

    for key, sheet, title in (
        ("hsn_b2b", "hsn(b2b)", "Summary For HSN(12)"),
        ("hsn_b2c", "hsn(b2c)", "Summary For HSN(12)"),
    ):
        block = data.get(key) or {}
        s = block.get("summary") or {}
        _gstn_table_sheet(
            wb, sheet, title,
            [
                ("hsn", "HSN"),
                ("description", "Description"),
                ("uqc", "UQC"),
                ("qty", "Total Quantity"),
                ("total_value", "Total Value"),
                ("rate", "Rate"),
                ("taxable_value", "Taxable Value"),
                ("igst", "Integrated Tax Amount"),
                ("cgst", "Central Tax Amount"),
                ("sgst", "State/UT Tax Amount"),
                ("cess", "Cess Amount"),
            ],
            block.get("rows") or [],
            summary_labels=["No. of HSN", "", "", "", "Total Value", "", "Total Taxable Value",
                            "Total Integrated Tax", "Total Central Tax", "Total State/UT Tax", "Total Cess"],
            summary_values=[s.get("hsn_count", 0), "", "", "", s.get("total_value", 0), "",
                            s.get("taxable_value", 0), s.get("igst", 0), s.get("cgst", 0),
                            s.get("sgst", 0), s.get("cess", 0)],
        )

    docs = data.get("docs") or {}
    _gstn_table_sheet(
        wb, "docs", "Summary of documents issued during the tax period (13)",
        [
            ("nature", "Nature of Document"),
            ("from", "Sr.No. From"),
            ("to", "Sr.No. To"),
            ("total", "Total Number"),
            ("cancelled", "Cancelled"),
        ],
        docs.get("rows") or [],
        money_cols=set(),
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_gstr2_workbook(data: dict) -> bytes:
    wb = Workbook()
    _cover_sheet(wb, "KT HEALTH ERP — GSTR-2 / 2A-2B working paper", _meta_with_disclaimer(data), extra=[
        ("Match helper", "Use columns GSTIN + Invoice no. + Total tax to VLOOKUP a portal 2B export."),
    ])
    money = {"taxable", "igst", "cgst", "sgst", "cess", "invoice_value", "total_tax", "note_value"}
    _gstn_table_sheet(
        wb, "B2B inward", "Inward supplies from registered persons (books)",
        [
            ("gstin", "GSTIN of supplier"),
            ("supplier", "Supplier"),
            ("invoice_number", "Invoice number"),
            ("invoice_date", "Invoice date"),
            ("grn_number", "GRN no."),
            ("place_of_supply", "Place of supply"),
            ("taxable", "Taxable value"),
            ("igst", "IGST"),
            ("cgst", "CGST"),
            ("sgst", "SGST"),
            ("cess", "Cess"),
            ("invoice_value", "Invoice value"),
            ("total_tax", "Total tax"),
            ("match_key", "Match key (GSTIN|Invoice|Tax)"),
        ],
        (data.get("b2b") or {}).get("rows") or [],
        money_cols=money,
    )
    _gstn_table_sheet(
        wb, "Unregistered inward", "Inward supplies from unregistered persons",
        [
            ("supplier", "Supplier"),
            ("invoice_number", "Invoice number"),
            ("invoice_date", "Invoice date"),
            ("grn_number", "GRN no."),
            ("place_of_supply", "Place of supply"),
            ("taxable", "Taxable value"),
            ("igst", "IGST"),
            ("cgst", "CGST"),
            ("sgst", "SGST"),
            ("invoice_value", "Invoice value"),
            ("total_tax", "Total tax"),
        ],
        (data.get("unregistered") or {}).get("rows") or [],
        money_cols=money,
    )
    _gstn_table_sheet(
        wb, "CDN inward", "Credit / debit notes (inward)",
        [
            ("gstin", "GSTIN"),
            ("supplier", "Supplier"),
            ("note_number", "Note number"),
            ("note_date", "Note date"),
            ("note_type", "Note type"),
            ("place_of_supply", "Place of supply"),
            ("taxable", "Taxable value"),
            ("igst", "IGST"),
            ("cgst", "CGST"),
            ("sgst", "SGST"),
            ("note_value", "Note value"),
            ("total_tax", "Total tax"),
            ("match_key", "Match key"),
        ],
        (data.get("cdn") or {}).get("rows") or [],
        money_cols=money,
    )
    hsn = data.get("hsn") or {}
    s = hsn.get("summary") or {}
    _gstn_table_sheet(
        wb, "HSN inward", "HSN-wise inward supplies",
        [
            ("hsn", "HSN"),
            ("description", "Description"),
            ("uqc", "UQC"),
            ("qty", "Quantity"),
            ("rate", "Rate"),
            ("taxable_value", "Taxable value"),
            ("igst", "IGST"),
            ("cgst", "CGST"),
            ("sgst", "SGST"),
            ("cess", "Cess"),
            ("total_value", "Total value"),
        ],
        hsn.get("rows") or [],
        summary_labels=["No. of HSN", "", "", "", "", "Taxable", "IGST", "CGST", "SGST"],
        summary_values=[s.get("hsn_count", 0), "", "", "", "", s.get("taxable_value", 0),
                        s.get("igst", 0), s.get("cgst", 0), s.get("sgst", 0)],
    )
    itc = data.get("itc") or {}
    ws = wb.create_sheet("ITC totals")
    ws.append(["Particulars", "Amount"])
    _style_header(ws)
    for label, key in [
        ("Taxable (net of returns)", "taxable"),
        ("ITC IGST", "igst"),
        ("ITC CGST", "cgst"),
        ("ITC SGST", "sgst"),
        ("ITC total", "total"),
    ]:
        ws.append([label, itc.get(key, 0)])
        ws.cell(ws.max_row, 2).number_format = MONEY_FORMAT
    _autosize(ws)
    ws.column_dimensions["A"].width = 36

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_3b_money_table(ws, title, headers, rows):
    ws.append([title])
    ws.cell(ws.max_row, 1).font = TITLE_FONT
    ws.append(headers)
    _style_header(ws, ws.max_row)
    for row in rows:
        ws.append(row)
        for i, val in enumerate(row, start=1):
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                ws.cell(ws.max_row, i).number_format = MONEY_FORMAT
            ws.cell(ws.max_row, i).border = THIN
    ws.append([])


def build_gstr3b_workbook(data: dict) -> bytes:
    wb = Workbook()
    _cover_sheet(wb, "KT HEALTH ERP — Form GSTR-3B working paper", _meta_with_disclaimer(data), extra=[
        ("ITC note", (data.get("table_4") or {}).get("footnote") or ""),
    ])
    t31 = data.get("table_3_1") or {}
    ws = wb.create_sheet("3.1 Outward supplies")
    _write_3b_money_table(ws, "3.1 Details of Outward Supplies and inward supplies liable to reverse charge", [
        "Nature of Supplies", "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess",
    ], [
        ["(a) Outward taxable supplies (other than zero rated, Nil rated and exempted)",
         t31.get("a", {}).get("taxable", 0), t31.get("a", {}).get("igst", 0),
         t31.get("a", {}).get("cgst", 0), t31.get("a", {}).get("sgst", 0), t31.get("a", {}).get("cess", 0)],
        ["(b) Outward taxable supplies (zero rated)",
         t31.get("b", {}).get("taxable", 0), t31.get("b", {}).get("igst", 0),
         t31.get("b", {}).get("cgst", 0), t31.get("b", {}).get("sgst", 0), t31.get("b", {}).get("cess", 0)],
        ["(c) Other Outward supplies (Nil rated, exempted)",
         t31.get("c", {}).get("taxable", 0), t31.get("c", {}).get("igst", 0),
         t31.get("c", {}).get("cgst", 0), t31.get("c", {}).get("sgst", 0), t31.get("c", {}).get("cess", 0)],
        ["(d) Inward supplies (liable to reverse charge)",
         t31.get("d", {}).get("taxable", 0), t31.get("d", {}).get("igst", 0),
         t31.get("d", {}).get("cgst", 0), t31.get("d", {}).get("sgst", 0), t31.get("d", {}).get("cess", 0)],
        ["(e) Non-GST outward supplies",
         t31.get("e", {}).get("taxable", 0), t31.get("e", {}).get("igst", 0),
         t31.get("e", {}).get("cgst", 0), t31.get("e", {}).get("sgst", 0), t31.get("e", {}).get("cess", 0)],
    ])
    ws.column_dimensions["A"].width = 70
    _autosize(ws, min_width=14, max_width=70)

    ws = wb.create_sheet("3.2 Inter-state B2C")
    rows = [
        [r.get("place_of_supply"), r.get("taxable", 0), r.get("igst", 0)]
        for r in (data.get("table_3_2") or [])
    ]
    _write_3b_money_table(ws, "3.2 Inter-state supplies to unregistered persons, composition taxable persons and UIN holders", [
        "Place of supply (State/UT)", "Total Taxable Value", "Amount of Integrated tax",
    ], rows)
    _autosize(ws, min_width=18, max_width=40)

    t4 = data.get("table_4") or {}
    ws = wb.create_sheet("4 Eligible ITC")
    def _t4(label, key):
        r = t4.get(key) or {}
        return [label, r.get("igst", 0), r.get("cgst", 0), r.get("sgst", 0), r.get("cess", 0)]
    _write_3b_money_table(ws, "4. Eligible ITC", [
        "Details", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess",
    ], [
        _t4("(A)(1) Import of Goods", "a1_import_goods"),
        _t4("(A)(2) Import of Services", "a2_import_services"),
        _t4("(A)(3) Inward supplies liable to reverse charge", "a3_rcm"),
        _t4("(A)(4) Inward supplies from ISD", "a4_isd"),
        _t4("(A)(5) All other ITC", "a5_all_other"),
        _t4("(B)(1) As per rules 38, 42 & 43 of CGST rules and section 17(5)", "b1_rules"),
        _t4("(B)(2) Others", "b2_others"),
        _t4("(C) Net ITC available (A)-(B)", "c_net"),
        _t4("(D)(1) ITC reclaimed which was reversed under Table 4(B)(2) in earlier tax period", "d1_reclaimed"),
        _t4("(D)(2) Ineligible ITC under section 16(4) & ITC restricted due to PoS rules", "d2_ineligible"),
    ])
    ws.append([t4.get("footnote") or ""])
    ws.column_dimensions["A"].width = 78
    _autosize(ws, min_width=14, max_width=78)

    t5 = data.get("table_5") or {}
    ws = wb.create_sheet("5 Exempt inward")
    comp = t5.get("composition_nil_exempt") or {}
    non = t5.get("non_gst") or {}
    _write_3b_money_table(ws, "5. Values of Exempt, Nil-rated and Non-GST inward supplies", [
        "Nature of Supplies", "Inter-State Supplies", "Intra-State Supplies",
    ], [
        ["From a supplier under composition scheme, Exempt and Nil rated supply",
         comp.get("inter", 0), comp.get("intra", 0)],
        ["Non GST supply", non.get("inter", 0), non.get("intra", 0)],
    ])
    _autosize(ws, min_width=18, max_width=70)

    p = data.get("table_6_1") or {}
    ws = wb.create_sheet("6.1 Payment of tax")
    def _p(label, key):
        r = p.get(key) or {}
        return [label, r.get("payable", 0), r.get("itc", 0), r.get("cash", 0)]
    _write_3b_money_table(ws, "6.1 Payment of Tax", [
        "Description", "Tax payable", "Paid through ITC", "Paid in cash",
    ], [
        _p("Integrated Tax", "igst"),
        _p("Central Tax", "cgst"),
        _p("State/UT Tax", "sgst"),
        _p("Cess", "cess"),
    ])
    _autosize(ws, min_width=16, max_width=28)

    ws = wb.create_sheet("6.2 TDS TCS")
    _write_3b_money_table(ws, "6.2 TDS/TCS Credit", [
        "Details", "Integrated Tax", "Central Tax", "State/UT Tax",
    ], [
        ["TDS", 0, 0, 0],
        ["TCS", 0, 0, 0],
    ])

    chk = data.get("check") or {}
    ws = wb.create_sheet("GSTR-1 vs 3B")
    ws.append(["Check", "OK"])
    _style_header(ws)
    ws.append(["3.1(a) taxable equals GSTR-1 outward taxable (net of CDNs)",
               "Yes" if chk.get("gstr1_3b_taxable_match") else "Review"])
    ws.append(["3.1(a) tax equals GSTR-1 outward tax",
               "Yes" if chk.get("gstr1_3b_tax_match") else "Review"])
    ws.append(["4(C) ITC equals GSTR-2 books",
               "Yes" if chk.get("itc_matches_gstr2") else "Review"])
    _autosize(ws)
    ws.column_dimensions["A"].width = 70

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_gstr9_workbook(data: dict) -> bytes:
    wb = Workbook()
    _cover_sheet(wb, "KT HEALTH ERP — GSTR-9 annual working paper", _meta_with_disclaimer(data))
    t4 = data.get("table_4") or {}
    ws = wb.create_sheet("Table 4 Outward")
    def _r(label, block):
        b = block or {}
        return [label, b.get("taxable", 0), b.get("igst", 0), b.get("cgst", 0), b.get("sgst", 0)]
    _write_3b_money_table(ws, "Table 4 — Details of advances, inward and outward supplies on which tax is payable", [
        "Nature", "Taxable value", "IGST", "CGST", "SGST",
    ], [
        _r("B2B (supplies to registered persons)", t4.get("b2b")),
        _r("B2C (supplies to unregistered persons)", t4.get("b2c")),
        _r("Nil rated / exempt", t4.get("nil_exempt")),
        _r("Net outward (as in GSTR-3B 3.1(a))", t4.get("net_outward")),
    ])
    cn = t4.get("credit_notes") or {}
    ws.append(["Credit notes (taxable / tax)", cn.get("taxable", 0), cn.get("tax", 0)])
    ws.column_dimensions["A"].width = 55
    _autosize(ws, min_width=14, max_width=55)

    t6 = data.get("table_6") or {}
    ws = wb.create_sheet("Table 6 ITC")
    avail = t6.get("itc_available") or {}
    rev = t6.get("itc_reversed") or {}
    net = t6.get("net_itc") or {}
    _write_3b_money_table(ws, "Table 6 — Details of ITC availed during the financial year", [
        "Particulars", "IGST", "CGST", "SGST", "Cess",
    ], [
        ["ITC available (all other ITC from books)",
         avail.get("igst", 0), avail.get("cgst", 0), avail.get("sgst", 0), avail.get("cess", 0)],
        ["ITC reversed",
         rev.get("igst", 0), rev.get("cgst", 0), rev.get("sgst", 0), rev.get("cess", 0)],
        ["Net ITC",
         net.get("igst", 0), net.get("cgst", 0), net.get("sgst", 0), net.get("cess", 0)],
    ])
    ws.column_dimensions["A"].width = 50
    _autosize(ws, min_width=14, max_width=50)

    t9 = data.get("table_9") or {}
    ws = wb.create_sheet("Table 9 Tax paid")
    def _pay(label, key):
        r = t9.get(key) or {}
        return [label, r.get("payable", 0), r.get("itc", 0), r.get("cash", 0)]
    _write_3b_money_table(ws, "Table 9 — Details of tax paid as declared in returns filed during the financial year", [
        "Description", "Tax payable", "Paid through ITC", "Paid in cash",
    ], [
        _pay("Integrated Tax", "igst"),
        _pay("Central Tax", "cgst"),
        _pay("State/UT Tax", "sgst"),
        _pay("Cess", "cess"),
    ])
    _autosize(ws)

    hsn_cols = [
        ("hsn", "HSN"),
        ("description", "Description"),
        ("uqc", "UQC"),
        ("qty", "Quantity"),
        ("rate", "Rate"),
        ("taxable_value", "Taxable value"),
        ("igst", "IGST"),
        ("cgst", "CGST"),
        ("sgst", "SGST"),
        ("cess", "Cess"),
    ]
    t17 = data.get("table_17_hsn_outward") or {}
    _gstn_table_sheet(wb, "Table 17 HSN outward", "Table 17 — HSN wise summary of outward supplies",
                      hsn_cols, t17.get("rows") or [])
    t18 = data.get("table_18_hsn_inward") or {}
    _gstn_table_sheet(wb, "Table 18 HSN inward", "Table 18 — HSN wise summary of inward supplies",
                      hsn_cols, t18.get("rows") or [])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

