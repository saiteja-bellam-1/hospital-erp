"""Bulk Excel/CSV importers used by the guided hospital onboarding wizard."""
from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass
from typing import Any, Optional

from sqlalchemy.orm import Session

from app.models.hospital import Hospital
from app.models.inpatient import (
    AncillaryServiceCatalog,
    Bed,
    DoctorRoomTypeRate,
    RoomManagement,
    RoomTypeRateConfig,
)
from app.models.outpatient import OutpatientProcedure
from app.models.user import User, UserRole


ROOM_TYPES = {
    "general", "semi_private", "private", "suite", "icu", "hdu", "nicu", "picu",
    "isolation", "labour", "recovery", "daycare", "emergency", "operation",
}
AMENITY_OPTIONS = {
    "ac", "tv", "wifi", "attached_bath", "refrigerator", "locker",
    "oxygen_point", "suction_point", "call_bell", "visitor_chair",
    "cardiac_monitor", "pulse_oximeter", "ventilator_support",
    "infusion_pump", "dialysis_point",
}
ANCILLARY_CATEGORIES = {
    "imaging", "physiotherapy", "dialysis", "oxygen", "equipment",
    "consumable", "procedure", "other",
}
ANCILLARY_UNITS = {"per_session", "per_hour", "per_day", "per_unit"}


@dataclass
class RowError:
    sheet: str
    row: int
    message: str

    def as_dict(self) -> dict:
        return {"sheet": self.sheet, "row": self.row, "message": self.message}


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def _cell_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text != "" else None


def _cell_float(value: Any, *, field: str) -> Optional[float]:
    text = _cell_str(value)
    if text is None:
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        raise ValueError(f"{field} must be a number")


def _cell_int(value: Any, *, field: str, default: int | None = None) -> Optional[int]:
    text = _cell_str(value)
    if text is None:
        return default
    try:
        return int(float(text))
    except (TypeError, ValueError):
        raise ValueError(f"{field} must be a whole number")


def _cell_bool(value: Any, default: bool = True) -> bool:
    text = _cell_str(value)
    if text is None:
        return default
    return text.lower() in {"1", "true", "yes", "y", "on"}


def _row_empty(values: dict) -> bool:
    return not any(_cell_str(v) for k, v in values.items() if k != "_row")


def read_workbook_sheets(content: bytes, sheet_names: list[str]) -> dict[str, list[dict]]:
    """Return {sheet_name: [rowdict...]} for preferred sheet names."""
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    lower_map = {name.lower(): name for name in workbook.sheetnames}
    out: dict[str, list[dict]] = {}
    for preferred in sheet_names:
        actual = lower_map.get(preferred.lower())
        if not actual:
            out[preferred] = []
            continue
        rows = list(workbook[actual].iter_rows(values_only=True))
        header_idx = next(
            (i for i, row in enumerate(rows) if any(c is not None and str(c).strip() != "" for c in row)),
            None,
        )
        if header_idx is None:
            out[preferred] = []
            continue
        headers = [_norm_header(c) for c in rows[header_idx]]
        parsed = []
        for index in range(header_idx + 1, len(rows)):
            raw = rows[index]
            if not any(c is not None and str(c).strip() != "" for c in raw):
                continue
            rowdict = {"_row": index + 1}
            for col, header in enumerate(headers):
                if header:
                    rowdict[header] = raw[col] if col < len(raw) else None
            parsed.append(rowdict)
        out[preferred] = parsed
    return out


def read_csv_or_xlsx(content: bytes, filename: str, sheet_names: list[str]) -> dict[str, list[dict]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return read_workbook_sheets(content, sheet_names)
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for i, raw in enumerate(reader):
        rowdict = {_norm_header(k): v for k, v in raw.items() if k is not None}
        rowdict["_row"] = i + 2
        rows.append(rowdict)
    # CSV maps onto the first sheet name.
    result = {sheet: [] for sheet in sheet_names}
    if sheet_names:
        result[sheet_names[0]] = rows
    return result


def _hospital(db: Session) -> Hospital:
    hospital = db.query(Hospital).filter(Hospital.is_active.is_(True)).first()
    if not hospital:
        raise ValueError("No active hospital found")
    return hospital


def import_rooms(db: Session, content: bytes, filename: str) -> dict:
    sheets = read_csv_or_xlsx(content, filename, ["Rooms", "Beds"])
    room_rows = sheets.get("Rooms") or []
    bed_rows = sheets.get("Beds") or []
    errors: list[RowError] = []
    hospital = _hospital(db)

    created_rooms = 0
    created_beds = 0
    skipped = 0
    room_by_number: dict[str, RoomManagement] = {
        r.room_number.lower(): r
        for r in db.query(RoomManagement).filter(
            RoomManagement.hospital_id == hospital.id,
            RoomManagement.is_active.is_(True),
        ).all()
    }

    for row in room_rows:
        if _row_empty(row):
            continue
        line = row.get("_row", 0)
        try:
            room_number = _cell_str(row.get("room_number"))
            room_type = (_cell_str(row.get("room_type")) or "").lower()
            charge = _cell_float(row.get("room_charge_per_day"), field="room_charge_per_day")
            bed_count = _cell_int(row.get("bed_count"), field="bed_count", default=1) or 1
            nursing = _cell_float(row.get("nursing_charge_per_visit"), field="nursing_charge_per_visit") or 0.0
            gender = (_cell_str(row.get("gender_policy")) or "mixed").lower()
            amenities_raw = _cell_str(row.get("amenities")) or ""
            amenities = [a.strip() for a in amenities_raw.replace(";", ",").split(",") if a.strip()]
        except ValueError as exc:
            errors.append(RowError("Rooms", line, str(exc)))
            continue

        if not room_number:
            errors.append(RowError("Rooms", line, "room_number is required"))
            continue
        if room_type not in ROOM_TYPES:
            errors.append(RowError("Rooms", line, f"Invalid room_type '{room_type}'"))
            continue
        if charge is None:
            errors.append(RowError("Rooms", line, "room_charge_per_day is required"))
            continue
        if gender not in {"mixed", "male", "female"}:
            errors.append(RowError("Rooms", line, "gender_policy must be mixed, male or female"))
            continue
        bad_amenities = [a for a in amenities if a not in AMENITY_OPTIONS]
        if bad_amenities:
            errors.append(RowError("Rooms", line, f"Unknown amenities: {', '.join(bad_amenities)}"))
            continue
        if room_number.lower() in room_by_number:
            skipped += 1
            continue

        room = RoomManagement(
            room_number=room_number,
            room_type=room_type,
            floor=_cell_str(row.get("floor")),
            department=_cell_str(row.get("department")),
            ward=_cell_str(row.get("ward")),
            bed_count=bed_count,
            available_beds=bed_count,
            room_charge_per_day=charge,
            nursing_charge_per_visit=nursing,
            amenities=json.dumps(amenities) if amenities else None,
            is_isolation=_cell_bool(row.get("is_isolation"), default=False),
            gender_policy=gender,
            hospital_id=hospital.id,
            is_active=True,
        )
        db.add(room)
        db.flush()
        room_by_number[room_number.lower()] = room
        created_rooms += 1

        # Auto-create beds when no Beds sheet row targets this room.
        has_explicit_beds = any(
            (_cell_str(b.get("room_number")) or "").lower() == room_number.lower()
            for b in bed_rows
        )
        if not has_explicit_beds:
            for index in range(1, bed_count + 1):
                label = f"Bed-{index}" if bed_count > 1 else "Bed-1"
                db.add(Bed(room_id=room.id, bed_label=label, status="available"))
                created_beds += 1

    for row in bed_rows:
        if _row_empty(row):
            continue
        line = row.get("_row", 0)
        room_number = _cell_str(row.get("room_number"))
        bed_label = _cell_str(row.get("bed_label"))
        if not room_number or not bed_label:
            errors.append(RowError("Beds", line, "room_number and bed_label are required"))
            continue
        room = room_by_number.get(room_number.lower())
        if not room:
            errors.append(RowError("Beds", line, f"Room '{room_number}' not found"))
            continue
        exists = (
            db.query(Bed)
            .filter(Bed.room_id == room.id, Bed.bed_label == bed_label)
            .first()
        )
        if exists:
            skipped += 1
            continue
        db.add(Bed(room_id=room.id, bed_label=bed_label, status="available"))
        created_beds += 1

    # Sync bed counts for rooms we touched.
    for room in room_by_number.values():
        count = db.query(Bed).filter(Bed.room_id == room.id).count()
        available = db.query(Bed).filter(Bed.room_id == room.id, Bed.status == "available").count()
        if count:
            room.bed_count = count
            room.available_beds = available

    if errors:
        db.rollback()
        return {
            "ok": False,
            "created_rooms": 0,
            "created_beds": 0,
            "skipped": 0,
            "errors": [e.as_dict() for e in errors],
        }

    db.commit()
    return {
        "ok": True,
        "created_rooms": created_rooms,
        "created_beds": created_beds,
        "skipped": skipped,
        "errors": [],
    }


def import_nursing_rates(db: Session, content: bytes, filename: str) -> dict:
    sheets = read_csv_or_xlsx(content, filename, ["Data"])
    rows = sheets.get("Data") or []
    hospital = _hospital(db)
    errors: list[RowError] = []
    updated = 0

    for row in rows:
        if _row_empty(row):
            continue
        line = row.get("_row", 0)
        room_type = (_cell_str(row.get("room_type")) or "").lower()
        try:
            rate = _cell_float(row.get("nursing_charge_per_visit"), field="nursing_charge_per_visit")
        except ValueError as exc:
            errors.append(RowError("Data", line, str(exc)))
            continue
        if room_type not in ROOM_TYPES:
            errors.append(RowError("Data", line, f"Invalid room_type '{room_type}'"))
            continue
        if rate is None:
            errors.append(RowError("Data", line, "nursing_charge_per_visit is required"))
            continue
        existing = (
            db.query(RoomTypeRateConfig)
            .filter(
                RoomTypeRateConfig.hospital_id == hospital.id,
                RoomTypeRateConfig.room_type == room_type,
            )
            .first()
        )
        if existing:
            existing.nursing_charge_per_visit = rate
        else:
            db.add(
                RoomTypeRateConfig(
                    hospital_id=hospital.id,
                    room_type=room_type,
                    nursing_charge_per_visit=rate,
                )
            )
        updated += 1

    if errors:
        db.rollback()
        return {"ok": False, "updated": 0, "errors": [e.as_dict() for e in errors]}
    db.commit()
    return {"ok": True, "updated": updated, "errors": []}


def import_ancillary(db: Session, content: bytes, filename: str) -> dict:
    sheets = read_csv_or_xlsx(content, filename, ["Data"])
    rows = sheets.get("Data") or []
    hospital = _hospital(db)
    errors: list[RowError] = []
    created = 0
    skipped = 0

    existing_names = {
        (s.service_name or "").lower()
        for s in db.query(AncillaryServiceCatalog).filter(
            AncillaryServiceCatalog.hospital_id == hospital.id,
            AncillaryServiceCatalog.is_active.is_(True),
        ).all()
    }

    for row in rows:
        if _row_empty(row):
            continue
        line = row.get("_row", 0)
        name = _cell_str(row.get("service_name"))
        category = (_cell_str(row.get("category")) or "").lower()
        unit = (_cell_str(row.get("charge_unit")) or "per_session").lower()
        try:
            charge = _cell_float(row.get("default_charge"), field="default_charge")
        except ValueError as exc:
            errors.append(RowError("Data", line, str(exc)))
            continue
        if not name:
            errors.append(RowError("Data", line, "service_name is required"))
            continue
        if category not in ANCILLARY_CATEGORIES:
            errors.append(RowError("Data", line, f"Invalid category '{category}'"))
            continue
        if unit not in ANCILLARY_UNITS:
            errors.append(RowError("Data", line, f"Invalid charge_unit '{unit}'"))
            continue
        if charge is None:
            errors.append(RowError("Data", line, "default_charge is required"))
            continue
        if name.lower() in existing_names:
            skipped += 1
            continue
        db.add(
            AncillaryServiceCatalog(
                hospital_id=hospital.id,
                service_name=name,
                service_code=_cell_str(row.get("service_code")),
                category=category,
                default_charge=charge,
                charge_unit=unit,
                description=_cell_str(row.get("description")),
                is_active=True,
            )
        )
        existing_names.add(name.lower())
        created += 1

    if errors:
        db.rollback()
        return {"ok": False, "created": 0, "skipped": 0, "errors": [e.as_dict() for e in errors]}
    db.commit()
    return {"ok": True, "created": created, "skipped": skipped, "errors": []}


def import_doctor_room_rates(db: Session, content: bytes, filename: str) -> dict:
    sheets = read_csv_or_xlsx(content, filename, ["Data"])
    rows = sheets.get("Data") or []
    hospital = _hospital(db)
    errors: list[RowError] = []
    upserted = 0

    users_by_username = {
        (u.username or "").lower(): u
        for u in db.query(User).filter(User.hospital_id == hospital.id, User.is_active.is_(True)).all()
    }
    doctor_role = db.query(UserRole).filter(UserRole.name == "doctor").first()

    for row in rows:
        if _row_empty(row):
            continue
        line = row.get("_row", 0)
        username = (_cell_str(row.get("doctor_username")) or "").lower()
        room_type = (_cell_str(row.get("room_type")) or "").lower()
        try:
            visit_rate = _cell_float(row.get("visit_rate"), field="visit_rate")
        except ValueError as exc:
            errors.append(RowError("Data", line, str(exc)))
            continue
        doctor = users_by_username.get(username)
        if not doctor:
            errors.append(RowError("Data", line, f"Unknown doctor_username '{username}'"))
            continue
        if doctor_role and doctor.role_id != doctor_role.id and "doctor" not in doctor.role_names:
            errors.append(RowError("Data", line, f"User '{username}' is not a doctor"))
            continue
        if room_type not in ROOM_TYPES:
            errors.append(RowError("Data", line, f"Invalid room_type '{room_type}'"))
            continue
        if visit_rate is None:
            errors.append(RowError("Data", line, "visit_rate is required"))
            continue
        existing = (
            db.query(DoctorRoomTypeRate)
            .filter(
                DoctorRoomTypeRate.hospital_id == hospital.id,
                DoctorRoomTypeRate.doctor_id == doctor.id,
                DoctorRoomTypeRate.room_type == room_type,
            )
            .first()
        )
        if existing:
            existing.visit_rate = visit_rate
        else:
            db.add(
                DoctorRoomTypeRate(
                    hospital_id=hospital.id,
                    doctor_id=doctor.id,
                    room_type=room_type,
                    visit_rate=visit_rate,
                )
            )
        upserted += 1

    if errors:
        db.rollback()
        return {"ok": False, "upserted": 0, "errors": [e.as_dict() for e in errors]}
    db.commit()
    return {"ok": True, "upserted": upserted, "errors": []}


def import_opd_procedures(db: Session, content: bytes, filename: str, *, user_id: int | None = None) -> dict:
    sheets = read_csv_or_xlsx(content, filename, ["Data"])
    rows = sheets.get("Data") or []
    hospital = _hospital(db)
    errors: list[RowError] = []
    created = 0
    skipped = 0

    existing = {
        (p.name or "").lower()
        for p in db.query(OutpatientProcedure).filter(
            OutpatientProcedure.hospital_id == hospital.id,
            OutpatientProcedure.is_active.is_(True),
        ).all()
    }

    for row in rows:
        if _row_empty(row):
            continue
        line = row.get("_row", 0)
        name = _cell_str(row.get("name"))
        try:
            price = _cell_float(row.get("default_price"), field="default_price")
        except ValueError as exc:
            errors.append(RowError("Data", line, str(exc)))
            continue
        if not name:
            errors.append(RowError("Data", line, "name is required"))
            continue
        if price is None:
            errors.append(RowError("Data", line, "default_price is required"))
            continue
        if name.lower() in existing:
            skipped += 1
            continue
        db.add(
            OutpatientProcedure(
                name=name,
                code=_cell_str(row.get("code")),
                category=_cell_str(row.get("category")),
                default_price=price,
                description=_cell_str(row.get("description")),
                is_active=_cell_bool(row.get("is_active"), default=True),
                hospital_id=hospital.id,
                created_by_id=user_id,
            )
        )
        existing.add(name.lower())
        created += 1

    if errors:
        db.rollback()
        return {"ok": False, "created": 0, "skipped": 0, "errors": [e.as_dict() for e in errors]}
    db.commit()
    return {"ok": True, "created": created, "skipped": skipped, "errors": []}
