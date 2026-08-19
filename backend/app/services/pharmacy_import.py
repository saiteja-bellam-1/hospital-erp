"""Pharmacy bulk import/export — medicines, suppliers, masters, opening stock, purchases."""
from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime, timedelta
from typing import Callable, Dict, List, Optional, Tuple

import openpyxl
from fastapi import HTTPException
from sqlalchemy import func as sa_func
from sqlalchemy.orm import Session

from app.models.pharmacy import (
    Medicine,
    MedicineCategory,
    PharmacyCompany,
    PharmacyHSN,
    PharmacyInventory,
    PharmacyMedicineImportAlias,
    PharmacyPurchase,
    PharmacyPurchaseImportMapping,
    PharmacyPurchaseItem,
    PharmacyRack,
    PharmacySalt,
    PharmacyStockAdjustment,
    PharmacyStockLedger,
    PharmacyStore,
    PharmacySupplier,
    PharmacyUoM,
)
from app.models.user import User
from app.services.pharmacy_store_service import get_master_store_id
from app.utils.pharmacy_pricing import (
    apply_cost_pcs_from_mrp,
    apply_medicine_price_rounding,
    compute_line_tax,
    round_money,
)

# ---------------------------------------------------------------------------
# Column headers (import / export shape)
# ---------------------------------------------------------------------------

MEDICINE_HEADERS = [
    "medicine_code", "name", "category", "generic_name", "dosage_form", "strength",
    "mrp", "purchase_rate", "rate_a", "rate_b", "unit_price", "default_discount_pct",
    "item_discount_pct", "barcode", "packaging", "strip_conversion_factor", "rate_unit",
    "decimal_supported", "requires_prescription", "is_narcotic", "is_high_alert",
    "is_schedule_h", "is_schedule_h1", "is_tramadol", "is_controlled", "is_active", "is_hidden",
    "min_qty", "max_qty", "reorder_qty", "description", "side_effects", "contraindications",
    "storage_conditions", "manufacturer", "company", "rack_code", "salt", "uom", "hsn_code",
    "sgst_pct", "cgst_pct",
]

SUPPLIER_HEADERS = [
    "name", "mobile", "phone_office", "phone", "email", "address", "pin_code", "state",
    "state_code", "country", "gstin_no", "gstin", "gst_heading", "dl_number", "dl_expiry",
    "pan_number", "contact_person", "designation", "ledger_type", "ledger_category",
    "account_group", "opening_balance", "opening_balance_dr_cr", "is_active", "is_hidden",
    "website", "vat_number", "food_license_no", "narco_sch_h_billing", "bill_import",
    "color_tag", "station", "balancing_method",
]

MASTER_SHEETS: Dict[str, List[str]] = {
    "Categories": ["name", "description", "is_active"],
    "Companies": ["name", "contact", "is_active"],
    "Salts": ["name", "description", "is_active"],
    "Racks": ["code", "location", "description", "is_active"],
    "UoMs": ["name", "abbreviation", "decimal_supported", "is_active"],
    "HSN": ["code", "description", "sgst_pct", "cgst_pct", "igst_pct", "is_active"],
    "Stores": [
        "code", "name", "store_type", "parent_code", "location", "description",
        "can_receive_supplier_purchase", "is_active", "is_default",
    ],
}

OPENING_STOCK_HEADERS = [
    "medicine_code", "batch_number", "expiry_date", "quantity", "store_code",
    "mrp", "purchase_rate", "rate_a", "rate_b", "cost_price", "selling_price",
    "supplier_name", "hsn_code", "strip_conversion_factor", "free_quantity",
    "discount_pct", "notes",
]

# Flat named-header template (one row per line item; header fields repeat).
# Vendor distributor CSVs (H/T/F + CL1..CL31) are also accepted — see
# `_parse_vendor_purchase_blocks`.
PURCHASE_HEADERS = [
    "supplier_name", "invoice_number", "entry_date", "bill_date",
    "payment_type", "purchase_type", "tax_mode", "store_code", "notes",
    "medicine_code", "medicine_name", "batch_number", "expiry_date",
    "quantity", "free_quantity", "mrp", "purchase_rate", "rate_a", "rate_b",
    "strip_conversion_factor", "discount_pct", "pack_size", "hsn_code",
]

# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def _norm_header(h) -> str:
    if h is None:
        return ""
    return str(h).strip().lower().replace(" ", "_")


def _cell_str(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def _cell_float(v) -> Optional[float]:
    s = _cell_str(v)
    if s is None:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        raise ValueError(f"'{s}' is not a valid number")


def _cell_int(v) -> Optional[int]:
    f = _cell_float(v)
    if f is None:
        return None
    if f != int(f):
        raise ValueError(f"'{v}' is not a valid integer")
    return int(f)


def _cell_bool(v) -> bool:
    s = (_cell_str(v) or "").lower()
    return s in {"1", "true", "yes", "y"}


def _cell_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _cell_str(v)
    if s is None:
        return None
    try:
        return date.fromisoformat(s[:10])
    except ValueError as exc:
        raise ValueError(f"'{s}' is not a valid date (use YYYY-MM-DD)") from exc


def _row_is_empty(r) -> bool:
    return not any(c is not None and str(c).strip() != "" for c in r)


def _key_skipped(key: Optional[str]) -> bool:
    return not key or key.startswith("#")


def _read_xlsx_sheet(
    wb, preferred_names: List[str], *,
    fallback_first: bool = False,
    header_row: Optional[int] = None,
    row_end: Optional[int] = None,
) -> List[dict]:
    lower_map = {name.lower(): name for name in wb.sheetnames}
    ws = None
    for pn in preferred_names:
        if pn.lower() in lower_map:
            ws = wb[lower_map[pn.lower()]]
            break
    if ws is None:
        if fallback_first and wb.sheetnames:
            ws = wb[wb.sheetnames[0]]
        else:
            return []
    rows = list(ws.iter_rows(values_only=True))
    if header_row is not None:
        header_idx = int(header_row) - 1
        if header_idx < 0 or header_idx >= len(rows) or _row_is_empty(rows[header_idx]):
            raise HTTPException(
                status_code=400,
                detail=f"Start row {header_row} is not a valid header line in this file",
            )
    else:
        header_idx = next((i for i, r in enumerate(rows) if not _row_is_empty(r)), None)
    if header_idx is None:
        return []
    headers = [_norm_header(c) for c in rows[header_idx]]
    out: List[dict] = []
    end_idx = len(rows) - 1
    if row_end is not None:
        end_idx = min(end_idx, int(row_end) - 1)
    for j in range(header_idx + 1, end_idx + 1):
        r = rows[j]
        if _row_is_empty(r):
            continue
        rowdict: dict = {}
        for k, h in enumerate(headers):
            if not h:
                continue
            rowdict[h] = r[k] if k < len(r) else None
        rowdict["_row"] = j + 1
        out.append(rowdict)
    return out


def _parse_xlsx_sheet(
    content: bytes, sheet_names: List[str], *,
    fallback_first: bool = False,
    header_row: Optional[int] = None,
    row_end: Optional[int] = None,
) -> List[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    return _read_xlsx_sheet(
        wb, sheet_names,
        fallback_first=fallback_first,
        header_row=header_row,
        row_end=row_end,
    )


def _parse_xlsx_multi(content: bytes, sheet_name_list: List[str]) -> Dict[str, List[dict]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    return {name: _read_xlsx_sheet(wb, [name]) for name in sheet_name_list}


def _parse_csv_rows(
    content: bytes, *,
    header_row: Optional[int] = None,
    row_end: Optional[int] = None,
) -> List[dict]:
    """Parse CSV. `header_row` is 1-based file line of column names (default: 1).
    `row_end` is optional 1-based last data line to include (inclusive).
    Each row dict includes `_row` = 1-based file line number.
    """
    text = content.decode("utf-8-sig", errors="replace")
    raw_rows = list(csv.reader(io.StringIO(text)))
    if not raw_rows:
        return []
    hi = (int(header_row) - 1) if header_row is not None else 0
    if hi < 0 or hi >= len(raw_rows):
        raise HTTPException(
            status_code=400,
            detail=f"Start row {header_row} is out of range for this file ({len(raw_rows)} lines)",
        )
    header_cells = raw_rows[hi]
    if not any(str(c or "").strip() for c in header_cells):
        raise HTTPException(
            status_code=400,
            detail=f"Start row {hi + 1} is empty — pick the line that contains column names",
        )
    # Preserve display order; normalize keys; make duplicates unique
    headers: List[str] = []
    seen: Dict[str, int] = {}
    for c in header_cells:
        h = _norm_header(c)
        if not h:
            headers.append("")
            continue
        if h in seen:
            seen[h] += 1
            h = f"{h}_{seen[h]}"
        else:
            seen[h] = 1
        headers.append(h)
    out: List[dict] = []
    last = len(raw_rows) - 1
    if row_end is not None:
        last = min(last, int(row_end) - 1)
    for i in range(hi + 1, last + 1):
        cells = raw_rows[i]
        if not any(str(c or "").strip() for c in cells):
            continue
        rowdict: dict = {}
        for k, h in enumerate(headers):
            if not h:
                continue
            rowdict[h] = cells[k] if k < len(cells) else None
        rowdict["_row"] = i + 1
        out.append(rowdict)
    return out


def _parse_upload(
    content: bytes, filename: str, sheet_names: List[str], *,
    multi: bool = False,
    header_row: Optional[int] = None,
    row_end: Optional[int] = None,
):
    fn = (filename or "").lower()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    if fn.endswith(".csv"):
        if multi:
            raise HTTPException(status_code=400, detail="CSV supports a single data sheet only. Use .xlsx for multi-sheet imports.")
        return _parse_csv_rows(content, header_row=header_row, row_end=row_end)
    if fn.endswith(".xlsx"):
        if multi:
            return _parse_xlsx_multi(content, sheet_name_list=sheet_names)
        return _parse_xlsx_sheet(
            content, sheet_names,
            fallback_first=True,
            header_row=header_row,
            row_end=row_end,
        )
    raise HTTPException(status_code=400, detail="Unsupported file type. Upload a .xlsx or .csv file.")


def _excel_col_letter(idx: int) -> str:
    """0 → A, 25 → Z, 26 → AA."""
    n = int(idx) + 1
    if n < 1:
        return "A"
    letters: List[str] = []
    while n:
        n, rem = divmod(n - 1, 26)
        letters.append(chr(65 + rem))
    return "".join(reversed(letters))


def _parse_excel_col_letter(raw) -> Optional[str]:
    """Return a normalized Excel column letter (A, B, … AA) or None."""
    if raw is None:
        return None
    s = str(raw).strip().upper()
    if not s or not re.fullmatch(r"[A-Z]{1,3}", s):
        return None
    return s


def _read_grid_matrix(content: bytes, filename: str) -> List[list]:
    """Raw cell rows for purchase import (no header interpretation)."""
    fn = (filename or "").lower()
    if fn.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        return list(csv.reader(io.StringIO(text)))
    if fn.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        if not wb.sheetnames:
            return []
        ws = wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    raise HTTPException(status_code=400, detail="Unsupported file type. Upload a .xlsx or .csv file.")


def _parse_purchase_grid(
    content: bytes, filename: str, *,
    row_start: Optional[int] = None,
    row_end: Optional[int] = None,
) -> List[dict]:
    """Parse file as positional columns A, B, C, … Start row is the first *data* line (inclusive)."""
    matrix = _read_grid_matrix(content, filename)
    if not matrix:
        return []
    start_i = (int(row_start) - 1) if row_start is not None else 0
    if start_i < 0:
        start_i = 0
    if start_i >= len(matrix):
        raise HTTPException(
            status_code=400,
            detail=f"Start row {row_start} is out of range for this file ({len(matrix)} lines)",
        )
    last = len(matrix) - 1
    if row_end is not None:
        last = min(last, int(row_end) - 1)
    out: List[dict] = []
    for i in range(start_i, last + 1):
        cells = matrix[i] or []
        if _row_is_empty(cells):
            continue
        rowdict: dict = {"_row": i + 1}
        for k, val in enumerate(cells):
            rowdict[_excel_col_letter(k)] = val
        out.append(rowdict)
    return out


def _normalize_hsn_tax(sgst_pct: float, cgst_pct: float) -> Tuple[float, float, float]:
    sgst = float(sgst_pct or 0)
    cgst = float(cgst_pct or 0)
    return sgst, cgst, round(sgst + cgst, 4)


def _workbook_bytes(build_fn: Callable[[openpyxl.Workbook], None]) -> bytes:
    wb = openpyxl.Workbook()
    build_fn(wb)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _append_instructions(wb: openpyxl.Workbook, lines: List[str]) -> None:
    if "Instructions" in wb.sheetnames:
        ws = wb["Instructions"]
    else:
        ws = wb.create_sheet("Instructions")
    for line in lines:
        ws.append([line])


def _export_bool(v: bool) -> str:
    return "true" if v else "false"


def _export_date(v: Optional[date]) -> str:
    if v is None:
        return ""
    return v.isoformat()


def _empty_summary(*, dry_run: bool) -> dict:
    return {
        "dry_run": dry_run,
        "total_rows": 0,
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "error_count": 0,
        "masters_created": [],
        "errors": [],
        "preview": [],
        "form": None,
        "unmatched_medicines": [],
    }


# ---------------------------------------------------------------------------
# Master resolver (auto-create FK targets during medicine import)
# ---------------------------------------------------------------------------

class _MasterResolver:
    def __init__(self, db: Session, hospital_id: int):
        self.db = db
        self.hospital_id = hospital_id
        self.masters_created: List[str] = []
        self._categories = {
            c.name.strip().lower(): c
            for c in db.query(MedicineCategory).filter(MedicineCategory.hospital_id == hospital_id).all()
        }
        self._companies = {
            c.name.strip().lower(): c
            for c in db.query(PharmacyCompany).filter(PharmacyCompany.hospital_id == hospital_id).all()
        }
        self._salts = {
            s.name.strip().lower(): s
            for s in db.query(PharmacySalt).filter(PharmacySalt.hospital_id == hospital_id).all()
        }
        self._racks = {
            r.code.strip().lower(): r
            for r in db.query(PharmacyRack).filter(PharmacyRack.hospital_id == hospital_id).all()
        }
        self._uoms = {
            u.name.strip().lower(): u
            for u in db.query(PharmacyUoM).filter(PharmacyUoM.hospital_id == hospital_id).all()
        }
        self._hsn: Dict[Tuple[str, float, float], PharmacyHSN] = {}
        for h in db.query(PharmacyHSN).filter(PharmacyHSN.hospital_id == hospital_id).all():
            self._hsn[(h.code.strip().lower(), h.sgst_pct or 0, h.cgst_pct or 0)] = h
        self._hsn_by_code: Dict[str, List[PharmacyHSN]] = {}
        for h in db.query(PharmacyHSN).filter(
            PharmacyHSN.hospital_id == hospital_id,
            PharmacyHSN.is_active == True,  # noqa: E712
        ).all():
            self._hsn_by_code.setdefault(h.code.strip().lower(), []).append(h)

    def _track(self, kind: str, key: str) -> None:
        label = f"{kind}:{key}"
        if label not in self.masters_created:
            self.masters_created.append(label)

    def category(self, name: str) -> MedicineCategory:
        key = name.strip().lower()
        row = self._categories.get(key)
        if row:
            return row
        row = MedicineCategory(name=name.strip(), hospital_id=self.hospital_id, is_active=True)
        self.db.add(row)
        self.db.flush()
        self._categories[key] = row
        self._track("category", name.strip())
        return row

    def company(self, name: str) -> PharmacyCompany:
        key = name.strip().lower()
        row = self._companies.get(key)
        if row:
            return row
        row = PharmacyCompany(name=name.strip(), hospital_id=self.hospital_id, is_active=True)
        self.db.add(row)
        self.db.flush()
        self._companies[key] = row
        self._track("company", name.strip())
        return row

    def salt(self, name: str) -> PharmacySalt:
        key = name.strip().lower()
        row = self._salts.get(key)
        if row:
            return row
        row = PharmacySalt(name=name.strip(), hospital_id=self.hospital_id, is_active=True)
        self.db.add(row)
        self.db.flush()
        self._salts[key] = row
        self._track("salt", name.strip())
        return row

    def rack(self, code: str) -> PharmacyRack:
        key = code.strip().lower()
        row = self._racks.get(key)
        if row:
            return row
        row = PharmacyRack(code=code.strip(), hospital_id=self.hospital_id, is_active=True)
        self.db.add(row)
        self.db.flush()
        self._racks[key] = row
        self._track("rack", code.strip())
        return row

    def uom(self, name: str) -> PharmacyUoM:
        key = name.strip().lower()
        row = self._uoms.get(key)
        if row:
            return row
        row = PharmacyUoM(name=name.strip(), hospital_id=self.hospital_id, is_active=True)
        self.db.add(row)
        self.db.flush()
        self._uoms[key] = row
        self._track("uom", name.strip())
        return row

    def hsn(self, code: str, sgst: Optional[float], cgst: Optional[float]) -> PharmacyHSN:
        code_key = code.strip().lower()
        if sgst is not None or cgst is not None:
            s, c, igst = _normalize_hsn_tax(sgst or 0, cgst or 0)
        else:
            matches = self._hsn_by_code.get(code_key) or []
            if matches:
                return matches[0]
            s, c, igst = 0.0, 0.0, 0.0
        cache_key = (code_key, s, c)
        row = self._hsn.get(cache_key)
        if row:
            return row
        row = PharmacyHSN(
            code=code.strip(), sgst_pct=s, cgst_pct=c, igst_pct=igst,
            hospital_id=self.hospital_id, is_active=True,
        )
        self.db.add(row)
        self.db.flush()
        self._hsn[cache_key] = row
        self._hsn_by_code.setdefault(code_key, []).append(row)
        self._track("hsn", f"{code.strip()}({s}+{c})")
        return row


# ---------------------------------------------------------------------------
# Medicine import / export
# ---------------------------------------------------------------------------

def _find_medicine(db: Session, hospital_id: int, code: str) -> Tuple[Optional[Medicine], Optional[Medicine]]:
    active = db.query(Medicine).filter(
        Medicine.medicine_code == code,
        Medicine.hospital_id == hospital_id,
        Medicine.is_active == True,  # noqa: E712
    ).first()
    inactive = db.query(Medicine).filter(
        Medicine.medicine_code == code,
        Medicine.hospital_id == hospital_id,
        Medicine.is_active == False,  # noqa: E712
    ).first()
    return active, inactive


def _apply_medicine_row(med: Medicine, row: dict, resolver: _MasterResolver) -> None:
    med.name = _cell_str(row.get("name")) or med.name
    cat_name = _cell_str(row.get("category"))
    if cat_name:
        med.category_id = resolver.category(cat_name).id

    for field in (
        "generic_name", "dosage_form", "strength", "description", "side_effects",
        "contraindications", "storage_conditions", "manufacturer", "barcode", "packaging",
    ):
        val = _cell_str(row.get(field))
        if val is not None:
            setattr(med, field, val)

    for field in ("mrp", "purchase_rate", "rate_a", "rate_b", "unit_price",
                  "default_discount_pct", "item_discount_pct"):
        val = _cell_float(row.get(field))
        if val is not None:
            setattr(med, field, val)

    rate_a = med.rate_a or 0
    unit_price = med.unit_price or 0
    if rate_a and not unit_price:
        med.unit_price = rate_a
    elif unit_price and not rate_a:
        med.rate_a = unit_price

    scf = _cell_int(row.get("strip_conversion_factor"))
    if scf is not None:
        med.strip_conversion_factor = max(1, scf)

    rate_unit = (_cell_str(row.get("rate_unit")) or med.rate_unit or "tablet").lower()
    if rate_unit in ("tablet", "strip"):
        med.rate_unit = rate_unit

    for field in (
        "decimal_supported", "requires_prescription", "is_narcotic", "is_high_alert",
        "is_schedule_h", "is_schedule_h1", "is_tramadol", "is_controlled", "is_active", "is_hidden",
    ):
        if row.get(field) is not None and _cell_str(row.get(field)) is not None:
            setattr(med, field, _cell_bool(row.get(field)))

    for field in ("min_qty", "max_qty", "reorder_qty"):
        val = _cell_int(row.get(field))
        if val is not None:
            setattr(med, field, val)

    company_name = _cell_str(row.get("company"))
    if company_name:
        med.company_id = resolver.company(company_name).id

    rack_code = _cell_str(row.get("rack_code"))
    if rack_code:
        med.rack_id = resolver.rack(rack_code).id

    salt_name = _cell_str(row.get("salt"))
    if salt_name:
        med.salt_id = resolver.salt(salt_name).id

    uom_name = _cell_str(row.get("uom"))
    if uom_name:
        med.uom_id = resolver.uom(uom_name).id

    hsn_code = _cell_str(row.get("hsn_code"))
    if hsn_code:
        sgst = _cell_float(row.get("sgst_pct"))
        cgst = _cell_float(row.get("cgst_pct"))
        med.hsn_id = resolver.hsn(hsn_code, sgst, cgst).id

    apply_medicine_price_rounding(med)
    apply_cost_pcs_from_mrp(med)


def import_medicines(
    db: Session, user: User, content: bytes, filename: str,
    *, dry_run: bool, on_duplicate: str,
    column_mapping: Optional[dict] = None,
    row_start: Optional[int] = None,
    row_end: Optional[int] = None,
) -> dict:
    summary = _empty_summary(dry_run=dry_run)
    try:
        rows = resolve_mapped_rows(
            content, filename,
            column_mapping=column_mapping,
            valid_targets=_VALID_MEDICINE_TARGETS,
            required_fields=REQUIRED_MEDICINE_LETTER_FIELDS,
            named_sheet_names=["Medicines"],
            row_start=row_start,
            row_end=row_end,
        )
    except HTTPException as exc:
        summary["errors"].append({
            "sheet": "Medicines", "row": 0, "message": str(exc.detail),
        })
        summary["error_count"] = 1
        return summary
    hospital_id = user.hospital_id
    resolver = _MasterResolver(db, hospital_id)

    for row in rows:
        rownum = row.get("_row", 0)
        code = _cell_str(row.get("medicine_code"))
        if _key_skipped(code):
            continue
        summary["total_rows"] += 1
        name = _cell_str(row.get("name"))
        category = _cell_str(row.get("category"))
        row_errs: List[str] = []
        if not name:
            row_errs.append("Missing name")
        if not category:
            row_errs.append("Missing category")
        try:
            for f in ("mrp", "purchase_rate", "rate_a", "rate_b", "unit_price",
                      "default_discount_pct", "item_discount_pct", "sgst_pct", "cgst_pct"):
                if row.get(f) is not None and _cell_str(row.get(f)) is not None:
                    _cell_float(row.get(f))
            for f in ("min_qty", "max_qty", "reorder_qty", "strip_conversion_factor"):
                if row.get(f) is not None and _cell_str(row.get(f)) is not None:
                    _cell_int(row.get(f))
        except ValueError as exc:
            row_errs.append(str(exc))

        if row_errs:
            msg = "; ".join(row_errs)
            summary["errors"].append({"sheet": "Medicines", "row": rownum, "message": msg})
            summary["preview"].append({
                "row": rownum, "key": code or "", "name": name or "", "status": "error",
                "message": msg, "sheet": "Medicines",
            })
            continue

        active, inactive = _find_medicine(db, hospital_id, code)
        target = active or inactive

        if active and on_duplicate == "skip":
            summary["skipped"] += 1
            summary["preview"].append({
                "row": rownum, "key": code, "name": name, "status": "skip",
                "message": "Medicine code already exists (active)", "sheet": "Medicines",
            })
            continue

        if target and on_duplicate == "update":
            if not active and inactive:
                inactive.is_active = True
            _apply_medicine_row(target, row, resolver)
            summary["updated"] += 1
            summary["preview"].append({
                "row": rownum, "key": code, "name": name, "status": "update", "sheet": "Medicines",
            })
        elif not target:
            cat = resolver.category(category)
            med = Medicine(
                medicine_code=code, name=name, category_id=cat.id,
                hospital_id=hospital_id, unit_price=0.0,
            )
            _apply_medicine_row(med, row, resolver)
            db.add(med)
            db.flush()
            summary["created"] += 1
            summary["preview"].append({
                "row": rownum, "key": code, "name": name, "status": "new", "sheet": "Medicines",
            })
        else:
            summary["skipped"] += 1
            summary["preview"].append({
                "row": rownum, "key": code, "name": name, "status": "skip",
                "message": "Inactive medicine exists; use on_duplicate=update to reactivate",
                "sheet": "Medicines",
            })

    summary["masters_created"] = resolver.masters_created
    summary["error_count"] = len(summary["errors"])
    return summary


def export_medicines_xlsx(db: Session, hospital_id: int) -> bytes:
    meds = (
        db.query(Medicine)
        .filter(Medicine.hospital_id == hospital_id, Medicine.is_active == True)  # noqa: E712
        .order_by(Medicine.name)
        .all()
    )

    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        ws.title = "Medicines"
        ws.append(MEDICINE_HEADERS)
        for med in meds:
            cat = db.query(MedicineCategory).filter(MedicineCategory.id == med.category_id).first()
            company = db.query(PharmacyCompany).filter(PharmacyCompany.id == med.company_id).first() if med.company_id else None
            rack = db.query(PharmacyRack).filter(PharmacyRack.id == med.rack_id).first() if med.rack_id else None
            salt = db.query(PharmacySalt).filter(PharmacySalt.id == med.salt_id).first() if med.salt_id else None
            uom = db.query(PharmacyUoM).filter(PharmacyUoM.id == med.uom_id).first() if med.uom_id else None
            hsn = db.query(PharmacyHSN).filter(PharmacyHSN.id == med.hsn_id).first() if med.hsn_id else None
            ws.append([
                med.medicine_code, med.name, cat.name if cat else "",
                med.generic_name or "", med.dosage_form or "", med.strength or "",
                med.mrp or 0, med.purchase_rate or 0, med.rate_a or 0, med.rate_b or 0,
                med.unit_price or 0, med.default_discount_pct or 0, med.item_discount_pct or 0,
                med.barcode or "", med.packaging or "", med.strip_conversion_factor or 1,
                med.rate_unit or "tablet",
                _export_bool(bool(med.decimal_supported)), _export_bool(bool(med.requires_prescription)),
                _export_bool(bool(med.is_narcotic)), _export_bool(bool(med.is_high_alert)),
                _export_bool(bool(med.is_schedule_h)), _export_bool(bool(med.is_schedule_h1)),
                _export_bool(bool(med.is_tramadol)), _export_bool(bool(med.is_controlled)),
                _export_bool(bool(med.is_active)), _export_bool(bool(med.is_hidden)),
                med.min_qty or 0, med.max_qty or 0, med.reorder_qty or 0,
                med.description or "", med.side_effects or "", med.contraindications or "",
                med.storage_conditions or "", med.manufacturer or "",
                company.name if company else "", rack.code if rack else "",
                salt.name if salt else "", uom.name if uom else "",
                hsn.code if hsn else "", hsn.sgst_pct if hsn else "", hsn.cgst_pct if hsn else "",
            ])

    return _workbook_bytes(build)


def build_medicines_template() -> bytes:
    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        ws.title = "Medicines"
        ws.append(MEDICINE_HEADERS)
        ws.append([
            "PCM500", "Paracetamol 500mg", "General", "Paracetamol", "tablet", "500mg",
            25, 18, 20, 0, 20, 0, 0, "8901234567890", "10x10", 10, "tablet",
            "false", "false", "false", "false", "false", "false", "false", "false",
            "true", "false", 10, 500, 50, "", "", "", "Store below 25C",
            "", "Cipla Ltd", "R-A1", "Paracetamol", "TAB", "30049099", 6, 6,
        ])
        ws.append([
            "AMOX250", "Amoxicillin 250mg", "Antibiotics", "Amoxicillin", "capsule", "250mg",
            80, 55, 65, 0, 65, 5, 0, "", "10 caps", 10, "strip",
            "false", "true", "false", "false", "false", "false", "false", "false",
            "true", "false", 5, 200, 20, "", "", "", "",
            "", "Sun Pharma", "R-B2", "Amoxicillin", "CAP", "30041010", 6, 6,
        ])
        _append_instructions(wb, [
            "KT HEALTH ERP — Medicine Import Template",
            "",
            "Required: medicine_code, name, category",
            "Upsert key: medicine_code (per hospital)",
            "on_duplicate=skip (default) skips active duplicates; on_duplicate=update updates or reactivates",
            "Categories, companies, salts, racks, UoMs, HSN auto-created when missing",
            "Bool columns: true/false, 1/0, yes/no",
            "rate_a and unit_price are kept in sync when only one is provided",
        ])

    return _workbook_bytes(build)


# ---------------------------------------------------------------------------
# Supplier import / export
# ---------------------------------------------------------------------------

def _apply_supplier_row(sup: PharmacySupplier, row: dict) -> None:
    for field in (
        "mobile", "phone_office", "phone", "email", "address", "pin_code", "state",
        "state_code", "country", "gstin_no", "gstin", "gst_heading", "dl_number",
        "pan_number", "contact_person", "designation", "ledger_type", "ledger_category",
        "account_group", "website", "vat_number", "food_license_no",
        "narco_sch_h_billing", "bill_import", "color_tag", "station", "balancing_method",
    ):
        val = _cell_str(row.get(field))
        if val is not None:
            setattr(sup, field, val)

    if row.get("opening_balance") is not None and _cell_str(row.get("opening_balance")) is not None:
        sup.opening_balance = _cell_float(row.get("opening_balance")) or 0

    dr_cr = _cell_str(row.get("opening_balance_dr_cr"))
    if dr_cr in ("Dr", "Cr"):
        sup.opening_balance_dr_cr = dr_cr

    if row.get("dl_expiry") is not None and _cell_str(row.get("dl_expiry")) is not None:
        sup.dl_expiry = _cell_date(row.get("dl_expiry"))

    for field in ("is_active", "is_hidden"):
        if row.get(field) is not None and _cell_str(row.get(field)) is not None:
            setattr(sup, field, _cell_bool(row.get(field)))


def import_suppliers(
    db: Session, user: User, content: bytes, filename: str,
    *, dry_run: bool, on_duplicate: str,
) -> dict:
    summary = _empty_summary(dry_run=dry_run)
    rows = _parse_upload(content, filename, ["Suppliers"])
    hospital_id = user.hospital_id

    for row in rows:
        rownum = row.get("_row", 0)
        name = _cell_str(row.get("name"))
        if _key_skipped(name):
            continue
        summary["total_rows"] += 1
        if not name:
            msg = "Missing name"
            summary["errors"].append({"sheet": "Suppliers", "row": rownum, "message": msg})
            summary["preview"].append({
                "row": rownum, "key": "", "name": "", "status": "error", "message": msg, "sheet": "Suppliers",
            })
            continue

        existing = db.query(PharmacySupplier).filter(
            PharmacySupplier.hospital_id == hospital_id,
            PharmacySupplier.is_active == True,  # noqa: E712
            sa_func.lower(PharmacySupplier.name) == name.lower(),
        ).first()

        if existing and on_duplicate == "skip":
            summary["skipped"] += 1
            summary["preview"].append({
                "row": rownum, "key": name, "name": name, "status": "skip",
                "message": "Supplier already exists", "sheet": "Suppliers",
            })
            continue

        try:
            if existing and on_duplicate == "update":
                _apply_supplier_row(existing, row)
                summary["updated"] += 1
                summary["preview"].append({
                    "row": rownum, "key": name, "name": name, "status": "update", "sheet": "Suppliers",
                })
            elif not existing:
                sup = PharmacySupplier(name=name, hospital_id=hospital_id)
                _apply_supplier_row(sup, row)
                db.add(sup)
                db.flush()
                summary["created"] += 1
                summary["preview"].append({
                    "row": rownum, "key": name, "name": name, "status": "new", "sheet": "Suppliers",
                })
        except ValueError as exc:
            msg = str(exc)
            summary["errors"].append({"sheet": "Suppliers", "row": rownum, "message": msg})
            summary["preview"].append({
                "row": rownum, "key": name, "name": name, "status": "error", "message": msg, "sheet": "Suppliers",
            })

    summary["error_count"] = len(summary["errors"])
    return summary


def export_suppliers_xlsx(db: Session, hospital_id: int) -> bytes:
    rows = (
        db.query(PharmacySupplier)
        .filter(PharmacySupplier.hospital_id == hospital_id, PharmacySupplier.is_active == True)  # noqa: E712
        .order_by(PharmacySupplier.name)
        .all()
    )

    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        ws.title = "Suppliers"
        ws.append(SUPPLIER_HEADERS)
        for sup in rows:
            ws.append([
                sup.name, sup.mobile or "", sup.phone_office or "", sup.phone or "",
                sup.email or "", sup.address or "", sup.pin_code or "", sup.state or "",
                sup.state_code or "", sup.country or "India", sup.gstin_no or sup.gstin or "",
                sup.gstin or "", sup.gst_heading or "local", sup.dl_number or "",
                _export_date(sup.dl_expiry), sup.pan_number or "", sup.contact_person or "",
                sup.designation or "", sup.ledger_type or "unregistered",
                sup.ledger_category or "OTHERS", sup.account_group or "Sundry Creditors",
                sup.opening_balance or 0, sup.opening_balance_dr_cr or "Dr",
                _export_bool(bool(sup.is_active)), _export_bool(bool(sup.is_hidden)),
                sup.website or "", sup.vat_number or "", sup.food_license_no or "",
                sup.narco_sch_h_billing or "allow_all", sup.bill_import or "mobile",
                sup.color_tag or "normal", sup.station or "", sup.balancing_method or "bill_by_bill",
            ])

    return _workbook_bytes(build)


def build_suppliers_template() -> bytes:
    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        ws.title = "Suppliers"
        ws.append(SUPPLIER_HEADERS)
        ws.append([
            "MedSupply Co", "9876543210", "040-1234567", "", "sales@medsupply.example",
            "12 Industrial Area", "500001", "Telangana", "36", "India",
            "36AABCM1234A1Z5", "", "local", "DL-TS-12345", "2027-12-31",
            "AABCM1234A", "Ravi Kumar", "Manager", "registered", "OTHERS",
            "Sundry Creditors", 0, "Dr", "true", "false",
            "", "", "", "allow_all", "mobile", "normal", "Hyderabad", "bill_by_bill",
        ])
        _append_instructions(wb, [
            "KT HEALTH ERP — Supplier Import Template",
            "",
            "Required: name",
            "Upsert key: name (case-insensitive, active suppliers only)",
            "on_duplicate=skip | update",
        ])

    return _workbook_bytes(build)


# ---------------------------------------------------------------------------
# Masters import / export
# ---------------------------------------------------------------------------

def _upsert_master_row(
    db: Session, hospital_id: int, sheet: str, row: dict,
    caches: dict, summary: dict, on_duplicate: str,
) -> None:
    rownum = row.get("_row", 0)

    if sheet == "Categories":
        name = _cell_str(row.get("name"))
        if _key_skipped(name):
            return
        summary["total_rows"] += 1
        if not name:
            _add_error(summary, sheet, rownum, "Missing name", key="")
            return
        key = name.lower()
        existing = caches["categories"].get(key)
        if existing and on_duplicate == "skip":
            summary["skipped"] += 1
            summary["preview"].append({"row": rownum, "key": name, "name": name, "status": "skip", "sheet": sheet})
            return
        if existing:
            if _cell_str(row.get("description")) is not None:
                existing.description = _cell_str(row.get("description"))
            if row.get("is_active") is not None and _cell_str(row.get("is_active")) is not None:
                existing.is_active = _cell_bool(row.get("is_active"))
            summary["updated"] += 1
            summary["preview"].append({"row": rownum, "key": name, "name": name, "status": "update", "sheet": sheet})
        else:
            obj = MedicineCategory(
                name=name, hospital_id=hospital_id,
                description=_cell_str(row.get("description")),
                is_active=_cell_bool(row.get("is_active")) if _cell_str(row.get("is_active")) else True,
            )
            db.add(obj)
            db.flush()
            caches["categories"][key] = obj
            summary["created"] += 1
            summary["preview"].append({"row": rownum, "key": name, "name": name, "status": "new", "sheet": sheet})

    elif sheet == "Companies":
        name = _cell_str(row.get("name"))
        if _key_skipped(name):
            return
        summary["total_rows"] += 1
        if not name:
            _add_error(summary, sheet, rownum, "Missing name", key="")
            return
        key = name.lower()
        existing = caches["companies"].get(key)
        if existing and on_duplicate == "skip":
            summary["skipped"] += 1
            summary["preview"].append({"row": rownum, "key": name, "name": name, "status": "skip", "sheet": sheet})
            return
        if existing:
            if _cell_str(row.get("contact")) is not None:
                existing.contact = _cell_str(row.get("contact"))
            if row.get("is_active") is not None and _cell_str(row.get("is_active")) is not None:
                existing.is_active = _cell_bool(row.get("is_active"))
            summary["updated"] += 1
            summary["preview"].append({"row": rownum, "key": name, "name": name, "status": "update", "sheet": sheet})
        else:
            obj = PharmacyCompany(
                name=name, hospital_id=hospital_id,
                contact=_cell_str(row.get("contact")),
                is_active=_cell_bool(row.get("is_active")) if _cell_str(row.get("is_active")) else True,
            )
            db.add(obj)
            db.flush()
            caches["companies"][key] = obj
            summary["created"] += 1
            summary["preview"].append({"row": rownum, "key": name, "name": name, "status": "new", "sheet": sheet})

    elif sheet == "Salts":
        name = _cell_str(row.get("name"))
        if _key_skipped(name):
            return
        summary["total_rows"] += 1
        if not name:
            _add_error(summary, sheet, rownum, "Missing name", key="")
            return
        key = name.lower()
        existing = caches["salts"].get(key)
        if existing and on_duplicate == "skip":
            summary["skipped"] += 1
            summary["preview"].append({"row": rownum, "key": name, "name": name, "status": "skip", "sheet": sheet})
            return
        if existing:
            if _cell_str(row.get("description")) is not None:
                existing.description = _cell_str(row.get("description"))
            if row.get("is_active") is not None and _cell_str(row.get("is_active")) is not None:
                existing.is_active = _cell_bool(row.get("is_active"))
            summary["updated"] += 1
            summary["preview"].append({"row": rownum, "key": name, "name": name, "status": "update", "sheet": sheet})
        else:
            obj = PharmacySalt(
                name=name, hospital_id=hospital_id,
                description=_cell_str(row.get("description")),
                is_active=_cell_bool(row.get("is_active")) if _cell_str(row.get("is_active")) else True,
            )
            db.add(obj)
            db.flush()
            caches["salts"][key] = obj
            summary["created"] += 1
            summary["preview"].append({"row": rownum, "key": name, "name": name, "status": "new", "sheet": sheet})

    elif sheet == "Racks":
        code = _cell_str(row.get("code"))
        if _key_skipped(code):
            return
        summary["total_rows"] += 1
        if not code:
            _add_error(summary, sheet, rownum, "Missing code", key="")
            return
        key = code.lower()
        existing = caches["racks"].get(key)
        if existing and on_duplicate == "skip":
            summary["skipped"] += 1
            summary["preview"].append({"row": rownum, "key": code, "name": code, "status": "skip", "sheet": sheet})
            return
        if existing:
            for f in ("location", "description"):
                if _cell_str(row.get(f)) is not None:
                    setattr(existing, f, _cell_str(row.get(f)))
            if row.get("is_active") is not None and _cell_str(row.get("is_active")) is not None:
                existing.is_active = _cell_bool(row.get("is_active"))
            summary["updated"] += 1
            summary["preview"].append({"row": rownum, "key": code, "name": code, "status": "update", "sheet": sheet})
        else:
            obj = PharmacyRack(
                code=code, hospital_id=hospital_id,
                location=_cell_str(row.get("location")),
                description=_cell_str(row.get("description")),
                is_active=_cell_bool(row.get("is_active")) if _cell_str(row.get("is_active")) else True,
            )
            db.add(obj)
            db.flush()
            caches["racks"][key] = obj
            summary["created"] += 1
            summary["preview"].append({"row": rownum, "key": code, "name": code, "status": "new", "sheet": sheet})

    elif sheet == "UoMs":
        name = _cell_str(row.get("name"))
        if _key_skipped(name):
            return
        summary["total_rows"] += 1
        if not name:
            _add_error(summary, sheet, rownum, "Missing name", key="")
            return
        key = name.lower()
        existing = caches["uoms"].get(key)
        if existing and on_duplicate == "skip":
            summary["skipped"] += 1
            summary["preview"].append({"row": rownum, "key": name, "name": name, "status": "skip", "sheet": sheet})
            return
        if existing:
            if _cell_str(row.get("abbreviation")) is not None:
                existing.abbreviation = _cell_str(row.get("abbreviation"))
            if row.get("decimal_supported") is not None and _cell_str(row.get("decimal_supported")) is not None:
                existing.decimal_supported = _cell_bool(row.get("decimal_supported"))
            if row.get("is_active") is not None and _cell_str(row.get("is_active")) is not None:
                existing.is_active = _cell_bool(row.get("is_active"))
            summary["updated"] += 1
            summary["preview"].append({"row": rownum, "key": name, "name": name, "status": "update", "sheet": sheet})
        else:
            obj = PharmacyUoM(
                name=name, hospital_id=hospital_id,
                abbreviation=_cell_str(row.get("abbreviation")),
                decimal_supported=_cell_bool(row.get("decimal_supported")) if _cell_str(row.get("decimal_supported")) else False,
                is_active=_cell_bool(row.get("is_active")) if _cell_str(row.get("is_active")) else True,
            )
            db.add(obj)
            db.flush()
            caches["uoms"][key] = obj
            summary["created"] += 1
            summary["preview"].append({"row": rownum, "key": name, "name": name, "status": "new", "sheet": sheet})

    elif sheet == "HSN":
        code = _cell_str(row.get("code"))
        if _key_skipped(code):
            return
        summary["total_rows"] += 1
        if not code:
            _add_error(summary, sheet, rownum, "Missing code", key="")
            return
        try:
            sgst, cgst, igst = _normalize_hsn_tax(
                _cell_float(row.get("sgst_pct")) or 0,
                _cell_float(row.get("cgst_pct")) or 0,
            )
            if row.get("igst_pct") is not None and _cell_str(row.get("igst_pct")) is not None:
                igst = _cell_float(row.get("igst_pct")) or igst
        except ValueError as exc:
            _add_error(summary, sheet, rownum, str(exc), key=code)
            return
        cache_key = (code.lower(), sgst, cgst)
        existing = caches["hsn"].get(cache_key)
        if existing and on_duplicate == "skip":
            summary["skipped"] += 1
            summary["preview"].append({"row": rownum, "key": code, "name": code, "status": "skip", "sheet": sheet})
            return
        if existing:
            if _cell_str(row.get("description")) is not None:
                existing.description = _cell_str(row.get("description"))
            existing.igst_pct = igst
            if row.get("is_active") is not None and _cell_str(row.get("is_active")) is not None:
                existing.is_active = _cell_bool(row.get("is_active"))
            summary["updated"] += 1
            summary["preview"].append({"row": rownum, "key": code, "name": code, "status": "update", "sheet": sheet})
        else:
            obj = PharmacyHSN(
                code=code, hospital_id=hospital_id, sgst_pct=sgst, cgst_pct=cgst, igst_pct=igst,
                description=_cell_str(row.get("description")),
                is_active=_cell_bool(row.get("is_active")) if _cell_str(row.get("is_active")) else True,
            )
            db.add(obj)
            db.flush()
            caches["hsn"][cache_key] = obj
            summary["created"] += 1
            summary["preview"].append({"row": rownum, "key": code, "name": code, "status": "new", "sheet": sheet})

    elif sheet == "Stores":
        code = _cell_str(row.get("code"))
        if _key_skipped(code):
            return
        summary["total_rows"] += 1
        if not code:
            _add_error(summary, sheet, rownum, "Missing code", key="")
            return
        name = _cell_str(row.get("name")) or code
        key = code.lower()
        existing = caches["stores"].get(key)
        if existing and on_duplicate == "skip":
            summary["skipped"] += 1
            summary["preview"].append({"row": rownum, "key": code, "name": name, "status": "skip", "sheet": sheet})
            return

        store_type = (_cell_str(row.get("store_type")) or "master").lower()
        if store_type not in ("master", "satellite"):
            _add_error(summary, sheet, rownum, f"Invalid store_type '{store_type}'", key=code)
            return

        parent_code = _cell_str(row.get("parent_code"))
        parent_id = None
        if parent_code:
            parent = caches["stores"].get(parent_code.lower())
            if not parent:
                _add_error(summary, sheet, rownum, f"Unknown parent_code '{parent_code}'", key=code)
                return
            parent_id = parent.id

        is_default = _cell_bool(row.get("is_default")) if _cell_str(row.get("is_default")) else False
        if is_default:
            for s in caches["stores"].values():
                s.is_default = False

        if existing:
            existing.name = name
            existing.store_type = store_type
            existing.parent_store_id = parent_id
            if _cell_str(row.get("location")) is not None:
                existing.location = _cell_str(row.get("location"))
            if _cell_str(row.get("description")) is not None:
                existing.description = _cell_str(row.get("description"))
            if row.get("can_receive_supplier_purchase") is not None and _cell_str(row.get("can_receive_supplier_purchase")) is not None:
                existing.can_receive_supplier_purchase = _cell_bool(row.get("can_receive_supplier_purchase"))
            if row.get("is_active") is not None and _cell_str(row.get("is_active")) is not None:
                existing.is_active = _cell_bool(row.get("is_active"))
            existing.is_default = is_default
            summary["updated"] += 1
            summary["preview"].append({"row": rownum, "key": code, "name": name, "status": "update", "sheet": sheet})
        else:
            obj = PharmacyStore(
                code=code, name=name, store_type=store_type, parent_store_id=parent_id,
                location=_cell_str(row.get("location")),
                description=_cell_str(row.get("description")),
                can_receive_supplier_purchase=(
                    _cell_bool(row.get("can_receive_supplier_purchase"))
                    if _cell_str(row.get("can_receive_supplier_purchase")) else False
                ),
                is_active=_cell_bool(row.get("is_active")) if _cell_str(row.get("is_active")) else True,
                is_default=is_default,
                hospital_id=hospital_id,
            )
            db.add(obj)
            db.flush()
            caches["stores"][key] = obj
            summary["created"] += 1
            summary["preview"].append({"row": rownum, "key": code, "name": name, "status": "new", "sheet": sheet})


def _add_error(summary: dict, sheet: str, rownum: int, message: str, *, key: str) -> None:
    summary["errors"].append({"sheet": sheet, "row": rownum, "message": message})
    summary["preview"].append({
        "row": rownum, "key": key, "name": key, "status": "error", "message": message, "sheet": sheet,
    })


def _load_master_caches(db: Session, hospital_id: int) -> dict:
    return {
        "categories": {c.name.strip().lower(): c for c in db.query(MedicineCategory).filter(MedicineCategory.hospital_id == hospital_id).all()},
        "companies": {c.name.strip().lower(): c for c in db.query(PharmacyCompany).filter(PharmacyCompany.hospital_id == hospital_id).all()},
        "salts": {s.name.strip().lower(): s for s in db.query(PharmacySalt).filter(PharmacySalt.hospital_id == hospital_id).all()},
        "racks": {r.code.strip().lower(): r for r in db.query(PharmacyRack).filter(PharmacyRack.hospital_id == hospital_id).all()},
        "uoms": {u.name.strip().lower(): u for u in db.query(PharmacyUoM).filter(PharmacyUoM.hospital_id == hospital_id).all()},
        "hsn": {(h.code.strip().lower(), h.sgst_pct or 0, h.cgst_pct or 0): h for h in db.query(PharmacyHSN).filter(PharmacyHSN.hospital_id == hospital_id).all()},
        "stores": {s.code.strip().lower(): s for s in db.query(PharmacyStore).filter(PharmacyStore.hospital_id == hospital_id).all()},
    }


def import_masters(
    db: Session, user: User, content: bytes, filename: str,
    *, dry_run: bool, on_duplicate: str,
) -> dict:
    summary = _empty_summary(dry_run=dry_run)
    hospital_id = user.hospital_id
    caches = _load_master_caches(db, hospital_id)

    parsed = _parse_upload(content, filename, list(MASTER_SHEETS.keys()), multi=True)
    if isinstance(parsed, list):
        parsed = {"Categories": parsed}

    for sheet in MASTER_SHEETS:
        for row in parsed.get(sheet, []):
            try:
                _upsert_master_row(db, hospital_id, sheet, row, caches, summary, on_duplicate)
            except ValueError as exc:
                _add_error(summary, sheet, row.get("_row", 0), str(exc), key=_cell_str(row.get("name")) or _cell_str(row.get("code")) or "")

    summary["error_count"] = len(summary["errors"])
    return summary


def export_masters_xlsx(db: Session, hospital_id: int) -> bytes:
    def build(wb: openpyxl.Workbook) -> None:
        wb.remove(wb.active)
        for sheet_name, headers in MASTER_SHEETS.items():
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
            if sheet_name == "Categories":
                for r in db.query(MedicineCategory).filter(MedicineCategory.hospital_id == hospital_id, MedicineCategory.is_active == True).order_by(MedicineCategory.name).all():  # noqa: E712
                    ws.append([r.name, r.description or "", _export_bool(bool(r.is_active))])
            elif sheet_name == "Companies":
                for r in db.query(PharmacyCompany).filter(PharmacyCompany.hospital_id == hospital_id, PharmacyCompany.is_active == True).order_by(PharmacyCompany.name).all():  # noqa: E712
                    ws.append([r.name, r.contact or "", _export_bool(bool(r.is_active))])
            elif sheet_name == "Salts":
                for r in db.query(PharmacySalt).filter(PharmacySalt.hospital_id == hospital_id, PharmacySalt.is_active == True).order_by(PharmacySalt.name).all():  # noqa: E712
                    ws.append([r.name, r.description or "", _export_bool(bool(r.is_active))])
            elif sheet_name == "Racks":
                for r in db.query(PharmacyRack).filter(PharmacyRack.hospital_id == hospital_id, PharmacyRack.is_active == True).order_by(PharmacyRack.code).all():  # noqa: E712
                    ws.append([r.code, r.location or "", r.description or "", _export_bool(bool(r.is_active))])
            elif sheet_name == "UoMs":
                for r in db.query(PharmacyUoM).filter(PharmacyUoM.hospital_id == hospital_id, PharmacyUoM.is_active == True).order_by(PharmacyUoM.name).all():  # noqa: E712
                    ws.append([r.name, r.abbreviation or "", _export_bool(bool(r.decimal_supported)), _export_bool(bool(r.is_active))])
            elif sheet_name == "HSN":
                for r in db.query(PharmacyHSN).filter(PharmacyHSN.hospital_id == hospital_id, PharmacyHSN.is_active == True).order_by(PharmacyHSN.code).all():  # noqa: E712
                    ws.append([r.code, r.description or "", r.sgst_pct or 0, r.cgst_pct or 0, r.igst_pct or 0, _export_bool(bool(r.is_active))])
            elif sheet_name == "Stores":
                for r in db.query(PharmacyStore).filter(PharmacyStore.hospital_id == hospital_id, PharmacyStore.is_active == True).order_by(PharmacyStore.code).all():  # noqa: E712
                    parent_code = ""
                    if r.parent_store_id:
                        p = db.query(PharmacyStore).filter(PharmacyStore.id == r.parent_store_id).first()
                        parent_code = p.code if p else ""
                    ws.append([
                        r.code, r.name, r.store_type, parent_code, r.location or "", r.description or "",
                        _export_bool(bool(r.can_receive_supplier_purchase)), _export_bool(bool(r.is_active)),
                        _export_bool(bool(r.is_default)),
                    ])

    return _workbook_bytes(build)


def build_masters_template() -> bytes:
    def build(wb: openpyxl.Workbook) -> None:
        wb.remove(wb.active)
        samples = {
            "Categories": [["General", "General medicines", "true"], ["Antibiotics", "", "true"]],
            "Companies": [["Cipla Ltd", "Mumbai", "true"]],
            "Salts": [["Paracetamol", "", "true"]],
            "Racks": [["R-A1", "Aisle A", "", "true"]],
            "UoMs": [["Tablet", "TAB", "false", "true"]],
            "HSN": [["30049099", "Medicaments", 6, 6, 12, "true"]],
            "Stores": [["MAIN", "Main Pharmacy", "master", "", "Ground floor", "", "true", "true", "true"]],
        }
        for sheet_name, headers in MASTER_SHEETS.items():
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
            for sample in samples.get(sheet_name, []):
                ws.append(sample)
        _append_instructions(wb, [
            "KT HEALTH ERP — Pharmacy Masters Import",
            "",
            "Sheets: Categories, Companies, Salts, Racks, UoMs, HSN, Stores",
            "Each sheet upserts independently; total_rows counts all sheets",
            "HSN upsert key: code + sgst_pct + cgst_pct (igst defaults to sgst+cgst if blank)",
            "Stores: only one is_default=true per hospital — importing default clears others",
        ])

    return _workbook_bytes(build)


# ---------------------------------------------------------------------------
# Opening stock import / export
# ---------------------------------------------------------------------------

def _resolve_store_id(db: Session, hospital_id: int, store_code: Optional[str], store_cache: dict) -> Optional[int]:
    if store_code:
        key = store_code.strip().lower()
        row = store_cache.get(key)
        if not row:
            row = db.query(PharmacyStore).filter(
                PharmacyStore.hospital_id == hospital_id,
                sa_func.lower(PharmacyStore.code) == key,
            ).first()
            if row:
                store_cache[key] = row
        return row.id if row else None
    return get_master_store_id(db, hospital_id)


def import_opening_stock(
    db: Session, user: User, content: bytes, filename: str,
    *, dry_run: bool, on_duplicate: str,
) -> dict:
    summary = _empty_summary(dry_run=dry_run)
    rows = _parse_upload(content, filename, ["OpeningStock"])
    hospital_id = user.hospital_id
    store_cache: Dict[str, PharmacyStore] = {
        s.code.strip().lower(): s
        for s in db.query(PharmacyStore).filter(PharmacyStore.hospital_id == hospital_id).all()
    }
    medicine_cache: Dict[str, Medicine] = {}

    for row in rows:
        rownum = row.get("_row", 0)
        med_code = _cell_str(row.get("medicine_code"))
        if _key_skipped(med_code):
            continue
        summary["total_rows"] += 1
        batch_number = _cell_str(row.get("batch_number"))
        row_errs: List[str] = []
        if not batch_number:
            row_errs.append("Missing batch_number")

        expiry = None
        qty = None
        try:
            expiry = _cell_date(row.get("expiry_date"))
            if expiry is None:
                row_errs.append("Missing or invalid expiry_date")
            qty_val = _cell_float(row.get("quantity"))
            if qty_val is None:
                row_errs.append("Missing quantity")
            elif qty_val <= 0:
                row_errs.append("quantity must be > 0")
            else:
                qty = int(qty_val) if qty_val == int(qty_val) else qty_val
        except ValueError as exc:
            row_errs.append(str(exc))

        if row_errs:
            msg = "; ".join(row_errs)
            summary["errors"].append({"sheet": "OpeningStock", "row": rownum, "message": msg})
            summary["preview"].append({
                "row": rownum, "key": med_code or "", "name": batch_number or "",
                "status": "error", "message": msg, "sheet": "OpeningStock",
            })
            continue

        med_key = med_code.lower()
        med = medicine_cache.get(med_key)
        if not med:
            med = db.query(Medicine).filter(
                Medicine.hospital_id == hospital_id,
                Medicine.medicine_code == med_code,
                Medicine.is_active == True,  # noqa: E712
            ).first()
            if med:
                medicine_cache[med_key] = med
        if not med:
            msg = f"Medicine '{med_code}' not found — import medicines first"
            summary["errors"].append({"sheet": "OpeningStock", "row": rownum, "message": msg})
            summary["preview"].append({
                "row": rownum, "key": med_code, "name": batch_number, "status": "error",
                "message": msg, "sheet": "OpeningStock",
            })
            continue

        store_code = _cell_str(row.get("store_code"))
        store_id = _resolve_store_id(db, hospital_id, store_code, store_cache)
        if not store_id:
            msg = f"Store not found: '{store_code or '(default master)'}'"
            summary["errors"].append({"sheet": "OpeningStock", "row": rownum, "message": msg})
            summary["preview"].append({
                "row": rownum, "key": med_code, "name": batch_number, "status": "error",
                "message": msg, "sheet": "OpeningStock",
            })
            continue

        existing = db.query(PharmacyInventory).filter(
            PharmacyInventory.medicine_id == med.id,
            PharmacyInventory.batch_number == batch_number,
            PharmacyInventory.store_id == store_id,
            PharmacyInventory.hospital_id == hospital_id,
            PharmacyInventory.is_active == True,  # noqa: E712
        ).first()

        if existing and on_duplicate == "skip":
            summary["skipped"] += 1
            summary["preview"].append({
                "row": rownum, "key": med_code, "name": batch_number, "status": "skip",
                "message": "Batch already exists", "sheet": "OpeningStock",
            })
            continue

        mrp = _cell_float(row.get("mrp"))
        purchase_rate = _cell_float(row.get("purchase_rate"))
        rate_a = _cell_float(row.get("rate_a"))
        rate_b = _cell_float(row.get("rate_b"))
        cost_price = _cell_float(row.get("cost_price"))
        selling_price = _cell_float(row.get("selling_price"))
        scf = _cell_int(row.get("strip_conversion_factor"))
        free_qty = _cell_int(row.get("free_quantity")) or 0
        discount_pct = _cell_float(row.get("discount_pct")) or 0.0
        notes = _cell_str(row.get("notes")) or "Opening stock import"

        if rate_a is None:
            rate_a = med.rate_a or 0
        if mrp is None:
            mrp = med.mrp or 0
        if purchase_rate is None:
            purchase_rate = med.purchase_rate or 0
        if selling_price is None:
            selling_price = rate_a or mrp or med.rate_a or 0
        if cost_price is None:
            cost_price = purchase_rate or med.purchase_rate or 0
        if scf is None:
            scf = max(1, int(med.strip_conversion_factor or 1))

        supplier_id = None
        supplier_name = _cell_str(row.get("supplier_name"))
        if supplier_name:
            sup = db.query(PharmacySupplier).filter(
                PharmacySupplier.hospital_id == hospital_id,
                sa_func.lower(PharmacySupplier.name) == supplier_name.lower(),
                PharmacySupplier.is_active == True,  # noqa: E712
            ).first()
            if sup:
                supplier_id = sup.id

        hsn_id = med.hsn_id
        hsn_code = _cell_str(row.get("hsn_code"))
        if hsn_code:
            resolver = _MasterResolver(db, hospital_id)
            hsn_id = resolver.hsn(hsn_code, _cell_float(row.get("sgst_pct")), _cell_float(row.get("cgst_pct"))).id
            summary["masters_created"].extend(resolver.masters_created)

        if existing and on_duplicate == "update":
            old_qty = existing.quantity_in_stock or 0
            new_qty = int(qty)
            if new_qty < 0:
                msg = "Resulting quantity cannot be negative"
                summary["errors"].append({"sheet": "OpeningStock", "row": rownum, "message": msg})
                summary["preview"].append({
                    "row": rownum, "key": med_code, "name": batch_number, "status": "error",
                    "message": msg, "sheet": "OpeningStock",
                })
                continue
            delta = new_qty - old_qty
            existing.quantity_in_stock = new_qty
            existing.expiry_date = expiry
            existing.mrp = mrp or existing.mrp
            existing.purchase_rate = purchase_rate or existing.purchase_rate
            existing.rate_a = rate_a or existing.rate_a
            existing.rate_b = rate_b if rate_b is not None else existing.rate_b
            existing.cost_price = cost_price
            existing.selling_price = selling_price
            existing.strip_conversion_factor = scf
            existing.free_quantity = free_qty
            existing.discount_pct = discount_pct
            existing.supplier_id = supplier_id or existing.supplier_id
            existing.hsn_id = hsn_id or existing.hsn_id
            reason = f"Opening stock import (set qty to {new_qty}, was {old_qty})"
            adj = PharmacyStockAdjustment(
                medicine_id=med.id, batch_id=existing.id, qty_change=delta,
                reason=reason, performed_by=user.id, store_id=store_id, hospital_id=hospital_id,
            )
            db.add(adj)
            db.flush()
            db.add(PharmacyStockLedger(
                medicine_id=med.id, batch_id=existing.id, txn_type="opening_stock",
                qty_delta=delta, reference_type="opening_stock_import", reference_id=adj.id,
                performed_by=user.id, store_id=store_id, hospital_id=hospital_id, notes=notes,
            ))
            summary["updated"] += 1
            summary["preview"].append({
                "row": rownum, "key": med_code, "name": batch_number, "status": "update", "sheet": "OpeningStock",
            })
        else:
            inv = PharmacyInventory(
                medicine_id=med.id, batch_number=batch_number, expiry_date=expiry,
                quantity_in_stock=int(qty), cost_price=cost_price, selling_price=selling_price,
                mrp=mrp or 0, purchase_rate=purchase_rate or 0, rate_a=rate_a or 0,
                rate_b=rate_b or 0, strip_conversion_factor=scf, free_quantity=free_qty,
                discount_pct=discount_pct, hsn_id=hsn_id, supplier_id=supplier_id,
                store_id=store_id, is_active=True, hospital_id=hospital_id,
            )
            db.add(inv)
            db.flush()
            reason = "Opening stock import"
            adj = PharmacyStockAdjustment(
                medicine_id=med.id, batch_id=inv.id, qty_change=int(qty),
                reason=reason, performed_by=user.id, store_id=store_id, hospital_id=hospital_id,
            )
            db.add(adj)
            db.flush()
            db.add(PharmacyStockLedger(
                medicine_id=med.id, batch_id=inv.id, txn_type="opening_stock",
                qty_delta=int(qty), reference_type="opening_stock_import", reference_id=adj.id,
                performed_by=user.id, store_id=store_id, hospital_id=hospital_id, notes=notes,
            ))
            summary["created"] += 1
            summary["preview"].append({
                "row": rownum, "key": med_code, "name": batch_number, "status": "new", "sheet": "OpeningStock",
            })

    summary["error_count"] = len(summary["errors"])
    return summary


def export_opening_stock_xlsx(db: Session, hospital_id: int) -> bytes:
    batches = (
        db.query(PharmacyInventory, Medicine, PharmacyStore, PharmacySupplier)
        .join(Medicine, Medicine.id == PharmacyInventory.medicine_id)
        .outerjoin(PharmacyStore, PharmacyStore.id == PharmacyInventory.store_id)
        .outerjoin(PharmacySupplier, PharmacySupplier.id == PharmacyInventory.supplier_id)
        .filter(
            PharmacyInventory.hospital_id == hospital_id,
            PharmacyInventory.is_active == True,  # noqa: E712
            PharmacyInventory.quantity_in_stock > 0,
        )
        .order_by(Medicine.medicine_code, PharmacyInventory.batch_number)
        .all()
    )

    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        ws.title = "OpeningStock"
        ws.append(OPENING_STOCK_HEADERS)
        for inv, med, store, sup in batches:
            hsn_code = ""
            if inv.hsn_id:
                hsn = db.query(PharmacyHSN).filter(PharmacyHSN.id == inv.hsn_id).first()
                hsn_code = hsn.code if hsn else ""
            ws.append([
                med.medicine_code, inv.batch_number, _export_date(inv.expiry_date),
                inv.quantity_in_stock, store.code if store else "",
                inv.mrp or 0, inv.purchase_rate or 0, inv.rate_a or 0, inv.rate_b or 0,
                inv.cost_price or 0, inv.selling_price or 0,
                sup.name if sup else "", hsn_code, inv.strip_conversion_factor or 1,
                inv.free_quantity or 0, inv.discount_pct or 0, "",
            ])

    return _workbook_bytes(build)


def build_opening_stock_template() -> bytes:
    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        ws.title = "OpeningStock"
        ws.append(OPENING_STOCK_HEADERS)
        ws.append([
            "PCM500", "BATCH001", "2027-06-30", 100, "MAIN",
            25, 18, 20, 0, 18, 20, "MedSupply Co", "30049099", 10, 0, 0,
            "Initial stock",
        ])
        _append_instructions(wb, [
            "KT HEALTH ERP — Opening Stock Import",
            "",
            "Required: medicine_code, batch_number, expiry_date (YYYY-MM-DD), quantity (>0)",
            "Medicine must already exist — this import does NOT create medicines",
            "store_code optional — defaults to master store",
            "on_duplicate=update sets quantity_in_stock to the NEW absolute value (not a delta)",
            "Ledger records the difference between old and new quantity",
        ])

    return _workbook_bytes(build)


# ---------------------------------------------------------------------------
# Purchases (vendor H/T/F CSV or named-header template)
# ---------------------------------------------------------------------------

# Vendor tax-invoice CSV column map (CL1..CL31), aligned to PDF line headers:
# Item Name, Pack Size, Batch No., Expiry, QTY, Package, MRP, PTR, Rate,
# Taxable, CGST%/Amt, IGST%/Amt, SGST%/Amt, HSN Code.
#
# Pricing policy for vendor imports:
#   PTR (CL11) → purchase_rate
#   MRP (CL13) → mrp, rate_a, and rate_b
#   Rate (CL12) is ignored (distributor net rate; not used for our purchase entry)
_VENDOR_DEFAULT_COLS = {
    "record_type": "cl1",
    # CL3 is invoice # on H rows and supplier name on T rows
    "supplier_or_invoice": "cl3",
    "bill_date": "cl4",
    "purchase_type": "cl7",
    "medicine_code": "cl5",
    "medicine_name": "cl6",
    "pack_size": "cl7",  # T rows; on H CL7 is purchase_type (resolved by record type)
    "manufacturer": "cl8",
    "batch_number": "cl9",
    "expiry_date": "cl10",
    "purchase_rate": "cl11",  # PTR on T rows
    "mrp": "cl13",
    "quantity": "cl16",
    "free_quantity": "cl17",
    "discount_pct": "cl19",
    "hsn_code": "cl31",
    "cgst_pct": "cl23",
    "sgst_pct": "cl27",
    # Footer (F row) — positional defaults
    "footer_taxable": "cl2",
    "footer_cgst": "cl5",
    "footer_sgst": "cl7",
    "footer_round_off": "cl21",
    "footer_invoice_value": "cl22",
    "transporter": "cl11",  # H rows only
    "place_of_supply": "cl14",
}

# Target fields shown in the import column-mapper UI.
PURCHASE_IMPORT_TARGETS = [
    {"key": "ignore", "label": "— Ignore —", "group": ""},
    {"key": "record_type", "label": "Record type (H / T / F)", "group": "Vendor format"},
    {"key": "supplier_or_invoice", "label": "Supplier (T) / Invoice # (H)", "group": "Vendor format"},
    {"key": "supplier_name", "label": "Supplier name", "group": "Purchase header"},
    {"key": "invoice_number", "label": "Invoice number", "group": "Purchase header"},
    {"key": "entry_date", "label": "Entry date", "group": "Purchase header"},
    {"key": "bill_date", "label": "Bill / invoice date", "group": "Purchase header"},
    {"key": "payment_type", "label": "Payment type (cash/credit)", "group": "Purchase header"},
    {"key": "purchase_type", "label": "Purchase type", "group": "Purchase header"},
    {"key": "tax_mode", "label": "Tax mode", "group": "Purchase header"},
    {"key": "store_code", "label": "Store code", "group": "Purchase header"},
    {"key": "notes", "label": "Notes", "group": "Purchase header"},
    {"key": "medicine_code", "label": "Medicine code", "group": "Line item"},
    {"key": "medicine_name", "label": "Medicine / item name", "group": "Line item"},
    {"key": "batch_number", "label": "Batch number", "group": "Line item"},
    {"key": "expiry_date", "label": "Expiry date", "group": "Line item"},
    {"key": "quantity", "label": "Quantity", "group": "Line item"},
    {"key": "free_quantity", "label": "Free quantity", "group": "Line item"},
    {"key": "mrp", "label": "MRP", "group": "Line item"},
    {"key": "purchase_rate", "label": "Purchase rate (PTR)", "group": "Line item"},
    {"key": "rate_a", "label": "Rate A (sale)", "group": "Line item"},
    {"key": "rate_b", "label": "Rate B (sale)", "group": "Line item"},
    {"key": "strip_conversion_factor", "label": "Strip conversion factor", "group": "Line item"},
    {"key": "discount_pct", "label": "Discount % / amount", "group": "Line item"},
    {"key": "pack_size", "label": "Pack size", "group": "Line item"},
    {"key": "hsn_code", "label": "HSN code", "group": "Line item"},
    {"key": "cgst_pct", "label": "CGST %", "group": "Line item"},
    {"key": "sgst_pct", "label": "SGST %", "group": "Line item"},
    {"key": "manufacturer", "label": "Manufacturer", "group": "Line item"},
    {"key": "transporter", "label": "Transporter", "group": "Optional"},
    {"key": "place_of_supply", "label": "Place of supply", "group": "Optional"},
]

_VALID_PURCHASE_TARGETS = {t["key"] for t in PURCHASE_IMPORT_TARGETS}

REQUIRED_PURCHASE_LETTER_FIELDS = [
    "medicine_name",
    "batch_number",
    "quantity",
    "purchase_rate",
    "mrp",
    "discount_pct",
    "expiry_date",
]


def _strip_excel_formula(v) -> Optional[str]:
    """Normalize Excel-exported cells like `=\"000008\"` or `=\"3338.06\"`."""
    s = _cell_str(v)
    if s is None:
        return None
    if s.startswith("="):
        s = s[1:].strip()
        if len(s) >= 2 and s[0] == s[-1] and s[0] in ("\"", "'"):
            s = s[1:-1]
    return s.strip() if s.strip() else None


def _safe_date(year, month, day) -> Optional[date]:
    try:
        return date(int(year), int(month), int(day))
    except (TypeError, ValueError):
        return None


def _parse_ddmmyyyy(v) -> Optional[date]:
    """Parse vendor expiry / bill dates.

    Accepts datetime/date, Excel serials, packed Marg/Vasu integers
    (DDMMYYYY, DMMYYYY, YYYYMMDD), ISO, and MM/YYYY or MM/YY.
    """
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, bool):
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    if isinstance(v, int):
        # Typical Excel serial range for 1954–2119
        if 20000 <= v <= 80000:
            try:
                return date(1899, 12, 30) + timedelta(days=v)
            except OverflowError:
                return None
        s = str(v)
    else:
        s = _strip_excel_formula(v)
    if not s:
        return None

    digits = re.sub(r"\D", "", s)
    if len(digits) == 8:
        # YYYYMMDD (ISO stripped) vs DDMMYYYY (Vasu / Marg day>=10)
        if 1990 <= int(digits[0:4]) <= 2100:
            parsed = _safe_date(digits[0:4], digits[4:6], digits[6:8])
            if parsed:
                return parsed
        parsed = _safe_date(digits[4:8], digits[2:4], digits[0:2])
        if parsed:
            return parsed
    if len(digits) == 7:
        # DMMYYYY — Marg ERP omits the leading zero on day 1–9 (1032028 → 1 Mar 2028)
        parsed = _safe_date(digits[3:7], digits[1:3], digits[0])
        if parsed:
            return parsed
    if len(digits) == 6:
        parsed = _safe_date(digits[2:6], digits[0:2], 1)  # MMYYYY
        if parsed:
            return parsed

    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        pass
    m = re.match(r"^(\d{1,2})[/\-.](\d{4})$", s)
    if m:
        parsed = _safe_date(m.group(2), m.group(1), 1)
        if parsed:
            return parsed
    m = re.match(r"^(\d{1,2})[/\-.](\d{2})$", s)
    if m:
        yy = int(m.group(2))
        year = 2000 + yy if yy < 100 else yy
        parsed = _safe_date(year, m.group(1), 1)
        if parsed:
            return parsed
    raise ValueError(f"'{s}' is not a valid date")


def _parse_pack_scf(pack: Optional[str]) -> Optional[int]:
    """`1*10` → 10, `5*1` → 1. Non-numeric pack units (e.g. 30ML) → None."""
    if not pack:
        return None
    parts = re.split(r"[*xX×]", pack.strip())
    if len(parts) < 2:
        return None
    m = re.match(r"^(\d+)", parts[-1].strip())
    if not m:
        return None
    n = int(m.group(1))
    return n if n >= 1 else None


def _normalize_column_mapping(raw: Optional[dict], valid_targets: Optional[set] = None) -> Dict[str, str]:
    """Convert {source_header: target_field} → {target_field: normalized_source}."""
    targets = valid_targets if valid_targets is not None else _VALID_PURCHASE_TARGETS
    if not raw:
        return {}
    out: Dict[str, str] = {}
    for src, tgt in raw.items():
        if src is None or tgt is None:
            continue
        target = str(tgt).strip()
        if not target or target == "ignore" or target not in targets:
            continue
        out[target] = _norm_header(src)
    return out


def _normalize_letter_mapping(raw: Optional[dict], valid_targets: Optional[set] = None) -> Dict[str, str]:
    """Return {erp_field: Excel column letter}.

    Accepts the new UI shape `{medicine_name: "F"}` and the older inverted
    `{CL6: "medicine_name"}` / `{F: "medicine_name"}` presets.
    """
    targets = valid_targets if valid_targets is not None else _VALID_PURCHASE_TARGETS
    if not raw or not isinstance(raw, dict):
        return {}
    keys = [str(k) for k in raw.keys()]
    looks_new = any(k in targets for k in keys)
    out: Dict[str, str] = {}
    if looks_new:
        for k, v in raw.items():
            field = str(k).strip()
            if field not in targets or field == "ignore":
                continue
            letter = _parse_excel_col_letter(v)
            if letter:
                out[field] = letter
        return out
    inverted = _normalize_column_mapping(raw, targets)
    for field, src in inverted.items():
        letter = _parse_excel_col_letter(src)
        if letter:
            out[field] = letter
            continue
        m = re.fullmatch(r"cl(\d+)", src or "")
        if m:
            out[field] = _excel_col_letter(int(m.group(1)) - 1)
    return out


def _line_items_from_letter_grid(grid: List[dict], letter_map: Dict[str, str]) -> List[dict]:
    """Turn positional A/B/C rows into purchase line dicts using field → letter map."""

    def _opt_float(v):
        try:
            return _cell_float(_strip_excel_formula(v) or v)
        except (ValueError, TypeError):
            return None

    def _opt_int(v):
        try:
            return _cell_int(_strip_excel_formula(v) or v)
        except (ValueError, TypeError):
            return None

    items: List[dict] = []
    for row in grid:
        def cell(field: str):
            letter = letter_map.get(field)
            if not letter:
                return None
            return row.get(letter)

        typ = (_strip_excel_formula(cell("record_type")) or "").upper()
        if typ in ("H", "F"):
            continue
        name = _strip_excel_formula(cell("medicine_name"))
        code = _strip_excel_formula(cell("medicine_code"))
        if not name and not code:
            continue
        expiry = None
        try:
            expiry = _parse_ddmmyyyy(cell("expiry_date"))
        except ValueError:
            expiry = None
        qty = _opt_float(cell("quantity"))
        free = _opt_float(cell("free_quantity"))
        if free is None:
            free = 0.0
        rate = _opt_float(cell("purchase_rate"))
        mrp = _opt_float(cell("mrp"))
        rate_a_raw = cell("rate_a")
        rate_b_raw = cell("rate_b")
        rate_a = _opt_float(rate_a_raw) if rate_a_raw is not None else None
        rate_b = _opt_float(rate_b_raw) if rate_b_raw is not None else None
        if rate_a is None:
            rate_a = mrp
        if rate_b is None:
            rate_b = mrp
        scf_raw = cell("strip_conversion_factor")
        scf = _opt_int(scf_raw) if scf_raw is not None else None
        pack = _strip_excel_formula(cell("pack_size"))
        if scf is None:
            scf = _parse_pack_scf(pack)
        items.append({
            "_row": row.get("_row"),
            "medicine_name": name,
            "medicine_code": code,
            "batch_number": _strip_excel_formula(cell("batch_number")) or "",
            "expiry_date": expiry,
            "quantity": qty,
            "free_quantity": free,
            "purchase_rate": rate,
            "mrp": mrp,
            "rate_a": rate_a,
            "rate_b": rate_b,
            "pack_size": pack,
            "manufacturer": _strip_excel_formula(cell("manufacturer")),
            "hsn_code": _strip_excel_formula(cell("hsn_code")),
            "discount_pct": _opt_float(cell("discount_pct")) or 0.0,
            "strip_conversion_factor": scf,
            "cgst_pct": _opt_float(cell("cgst_pct")),
            "sgst_pct": _opt_float(cell("sgst_pct")),
        })
    return items


def _suggest_purchase_mapping(headers: List[str]) -> Dict[str, str]:
    """Return {original_header: target_field} suggestions for the mapper UI."""
    norms = {_norm_header(h): h for h in headers}
    # Alias lists: first match wins per target
    aliases: List[Tuple[str, List[str]]] = [
        ("record_type", ["record_type", "type", "row_type", "cl1"]),
        ("supplier_or_invoice", ["supplier_or_invoice", "cl3"]),
        ("supplier_name", ["supplier_name", "supplier", "party", "party_name"]),
        ("invoice_number", ["invoice_number", "invoice_no", "invoice", "bill_no", "bill_number"]),
        ("entry_date", ["entry_date"]),
        ("bill_date", ["bill_date", "invoice_date", "date", "cl4"]),
        ("payment_type", ["payment_type", "payment"]),
        ("purchase_type", ["purchase_type"]),
        ("tax_mode", ["tax_mode"]),
        ("store_code", ["store_code", "store"]),
        ("notes", ["notes", "remark", "remarks"]),
        ("medicine_code", ["medicine_code", "item_code", "product_code", "sku", "cl5"]),
        ("medicine_name", ["medicine_name", "item_name", "item", "product", "product_name", "cl6"]),
        ("pack_size", ["pack_size", "pack", "packing", "cl7"]),
        ("manufacturer", ["manufacturer", "mfg", "company", "cl8"]),
        ("batch_number", ["batch_number", "batch", "batch_no", "lot", "cl9"]),
        ("expiry_date", ["expiry_date", "expiry", "exp", "exp_date", "cl10"]),
        ("purchase_rate", ["purchase_rate", "ptr", "p_rate", "prate", "cl11"]),
        ("mrp", ["mrp", "cl13"]),
        ("rate_a", ["rate_a", "sale_rate", "selling_rate"]),
        ("rate_b", ["rate_b"]),
        ("quantity", ["quantity", "qty", "cl16"]),
        ("free_quantity", ["free_quantity", "free_qty", "free", "package", "cl17"]),
        ("discount_pct", ["discount_pct", "discount", "disc", "cl19"]),
        ("strip_conversion_factor", ["strip_conversion_factor", "scf", "conversion"]),
        ("hsn_code", ["hsn_code", "hsn", "cl31"]),
        ("cgst_pct", ["cgst_pct", "cgst", "cl23"]),
        ("sgst_pct", ["sgst_pct", "sgst", "cl27"]),
        ("transporter", ["transporter", "transport"]),
        ("place_of_supply", ["place_of_supply", "place"]),
    ]
    used_sources: set = set()
    mapping: Dict[str, str] = {h: "ignore" for h in headers}

    # Strong vendor CL* preset when headers look like that format
    is_vendor = all(re.fullmatch(r"cl\d+", _norm_header(h) or "") for h in headers if h) and len(headers) >= 10
    if is_vendor or ("cl1" in norms and "cl6" in norms and "cl11" in norms):
        for field, src in _VENDOR_DEFAULT_COLS.items():
            if field.startswith("footer_"):
                continue
            orig = norms.get(src)
            if orig and orig not in used_sources:
                # Prefer purchase_rate over transporter for CL11 in suggestions
                if field == "transporter" and mapping.get(orig) == "purchase_rate":
                    continue
                if field == "pack_size" and mapping.get(orig) == "purchase_type":
                    # CL7: show as pack_size (dominant on T rows); purchase_type still defaulted in parser
                    pass
                mapping[orig] = field
                used_sources.add(orig)
        # Ensure CL7 marked pack_size (T) — purchase_type resolved from H via same col in parser default
        if "cl7" in norms:
            mapping[norms["cl7"]] = "pack_size"
        if "cl11" in norms:
            mapping[norms["cl11"]] = "purchase_rate"
        if "cl3" in norms:
            mapping[norms["cl3"]] = "supplier_or_invoice"
        return mapping

    for field, names in aliases:
        for name in names:
            if name in norms and norms[name] not in used_sources:
                mapping[norms[name]] = field
                used_sources.add(norms[name])
                break
    return mapping


def inspect_purchase_import(
    content: bytes, filename: str, *,
    row_start: Optional[int] = None,
    row_end: Optional[int] = None,
) -> dict:
    """Return file line count and mapper field catalog. Does not read cell values for mapping."""
    start = int(row_start) if row_start is not None else 1
    if start < 1:
        raise HTTPException(status_code=400, detail="row_start must be >= 1")
    if row_end is not None and int(row_end) < start:
        raise HTTPException(status_code=400, detail="row_end cannot be before row_start")

    file_line_count = _count_file_lines(content, filename)
    try:
        grid = _parse_purchase_grid(
            content, filename, row_start=start, row_end=row_end,
        )
    except HTTPException:
        grid = []
    row_nums = [int(r.get("_row") or 0) for r in grid]
    return {
        "headers": [],
        "suggested_mapping": {},
        "format_hint": "letter_columns",
        "targets": PURCHASE_IMPORT_TARGETS,
        "row_count": len(grid),
        "min_row": min(row_nums) if row_nums else start,
        "max_row": max(row_nums) if row_nums else file_line_count,
        "header_row": start,
        "file_line_count": file_line_count,
        "required_fields": list(REQUIRED_PURCHASE_LETTER_FIELDS),
    }


def _alias_norm_set(aliases: List[Tuple[str, List[str]]]) -> set:
    out: set = set()
    for _, names in aliases:
        out.update(names)
    return out


def _row_looks_like_headers(cells, aliases: List[Tuple[str, List[str]]]) -> bool:
    names = _alias_norm_set(aliases)
    hits = 0
    for c in cells or []:
        if _norm_header(c) in names:
            hits += 1
    return hits >= 2


def _suggest_letter_mapping_from_cells(
    cells, aliases: List[Tuple[str, List[str]]],
) -> Dict[str, str]:
    """Return {erp_field: Excel letter} from a header row."""
    by_norm: Dict[str, str] = {}
    for i, cell in enumerate(cells or []):
        n = _norm_header(cell)
        if n and n not in by_norm:
            by_norm[n] = _excel_col_letter(i)
    suggested: Dict[str, str] = {}
    used: set = set()
    for field, names in aliases:
        for name in names:
            letter = by_norm.get(name)
            if letter and letter not in used:
                suggested[field] = letter
                used.add(letter)
                break
    return suggested


def _header_preview(cells, limit: int = 20) -> List[dict]:
    out: List[dict] = []
    for i, cell in enumerate(cells or []):
        if i >= limit:
            break
        text = _strip_excel_formula(cell) or _cell_str(cell) or ""
        out.append({"letter": _excel_col_letter(i), "value": text})
    return out


def inspect_letter_import(
    content: bytes, filename: str, *,
    row_start: Optional[int] = None,
    row_end: Optional[int] = None,
    required_fields: Optional[List[str]] = None,
    aliases: Optional[List[Tuple[str, List[str]]]] = None,
) -> dict:
    """Inspect a spreadsheet for letter-column mapping (medicines, sales, …)."""
    start = int(row_start) if row_start is not None else 1
    if start < 1:
        raise HTTPException(status_code=400, detail="row_start must be >= 1")
    if row_end is not None and int(row_end) < start:
        raise HTTPException(status_code=400, detail="row_end cannot be before row_start")

    file_line_count = _count_file_lines(content, filename)
    matrix = _read_grid_matrix(content, filename)
    alias_list = aliases or []

    def cells_at(zero_idx: int) -> list:
        if 0 <= zero_idx < len(matrix):
            return matrix[zero_idx] or []
        return []

    suggested: Dict[str, str] = {}
    header_detected = False
    suggested_row_start = start
    preview_cells: list = []

    header_idx = None
    if start >= 2:
        header_idx = start - 2
    elif _row_looks_like_headers(cells_at(0), alias_list):
        header_idx = 0
        header_detected = True
        suggested_row_start = 2

    if header_idx is not None:
        preview_cells = cells_at(header_idx)
        suggested = _suggest_letter_mapping_from_cells(preview_cells, alias_list)
        if suggested:
            header_detected = True
            if start == 1:
                suggested_row_start = 2
    elif matrix:
        preview_cells = cells_at(max(0, start - 1))

    try:
        grid = _parse_purchase_grid(
            content, filename, row_start=start, row_end=row_end,
        )
    except HTTPException:
        grid = []
    row_nums = [int(r.get("_row") or 0) for r in grid]
    return {
        "headers": [],
        "suggested_mapping": {},
        "suggested_letter_mapping": suggested,
        "format_hint": "letter_columns",
        "row_count": len(grid),
        "min_row": min(row_nums) if row_nums else start,
        "max_row": max(row_nums) if row_nums else file_line_count,
        "header_row": start,
        "file_line_count": file_line_count,
        "required_fields": list(required_fields or []),
        "header_detected": header_detected,
        "suggested_row_start": suggested_row_start,
        "header_preview": _header_preview(preview_cells),
    }


def _dict_rows_from_letter_grid(grid: List[dict], letter_map: Dict[str, str]) -> List[dict]:
    """Copy letter-keyed grid rows into {erp_field: value, _row: n} dicts."""
    rows: List[dict] = []
    for row in grid:
        out: dict = {"_row": row.get("_row")}
        any_val = False
        for field, letter in letter_map.items():
            raw = row.get(letter)
            stripped = _strip_excel_formula(raw)
            out[field] = stripped if stripped is not None else raw
            if stripped is not None or (raw is not None and str(raw).strip() != ""):
                any_val = True
        if any_val:
            rows.append(out)
    return rows


def _is_headerish_mapped_row(row: dict, fields: List[str]) -> bool:
    hits = 0
    for field in fields:
        val = _norm_header(row.get(field))
        if val and (val == field or val.replace("_", "") == field.replace("_", "")):
            hits += 1
    return hits >= 2


def resolve_mapped_rows(
    content: bytes, filename: str, *,
    column_mapping: Optional[dict],
    valid_targets: set,
    required_fields: List[str],
    named_sheet_names: List[str],
    row_start: Optional[int] = None,
    row_end: Optional[int] = None,
    require_any: Optional[List[List[str]]] = None,
) -> List[dict]:
    """Letter-map rows when ``column_mapping`` is set; otherwise named-header parse."""
    mapping = column_mapping if isinstance(column_mapping, dict) and column_mapping else None
    if not mapping:
        return _parse_upload(content, filename, named_sheet_names)

    letter_map = _normalize_letter_mapping(mapping, valid_targets)
    missing = [f for f in required_fields if f not in letter_map]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Map required columns: {', '.join(missing)}",
        )
    for group in require_any or []:
        if not any(f in letter_map for f in group):
            raise HTTPException(
                status_code=400,
                detail=f"Map at least one of: {', '.join(group)}",
            )
    start_row = int(row_start) if row_start is not None else 1
    if start_row < 1:
        raise HTTPException(status_code=400, detail="row_start must be >= 1")
    if row_end is not None and int(row_end) < start_row:
        raise HTTPException(status_code=400, detail="row_end cannot be before row_start")
    grid = _parse_purchase_grid(
        content, filename, row_start=start_row, row_end=row_end,
    )
    rows = _dict_rows_from_letter_grid(grid, letter_map)
    keep_fields = list(letter_map.keys())
    return [r for r in rows if not _is_headerish_mapped_row(r, keep_fields)]


REQUIRED_MEDICINE_LETTER_FIELDS = ["medicine_code", "name", "category"]
_VALID_MEDICINE_TARGETS = set(MEDICINE_HEADERS)

MEDICINE_IMPORT_ALIASES: List[Tuple[str, List[str]]] = [
    ("medicine_code", ["medicine_code", "item_code", "product_code", "sku", "code", "itemcode"]),
    ("name", ["name", "medicine_name", "item_name", "product_name", "item", "product"]),
    ("category", ["category", "category_name", "cat"]),
    ("generic_name", ["generic_name", "generic"]),
    ("dosage_form", ["dosage_form", "form"]),
    ("strength", ["strength"]),
    ("mrp", ["mrp"]),
    ("purchase_rate", ["purchase_rate", "ptr", "p_rate", "prate"]),
    ("rate_a", ["rate_a", "sale_rate", "selling_rate", "rate"]),
    ("rate_b", ["rate_b"]),
    ("unit_price", ["unit_price"]),
    ("default_discount_pct", ["default_discount_pct"]),
    ("item_discount_pct", ["item_discount_pct"]),
    ("barcode", ["barcode"]),
    ("packaging", ["packaging", "pack_size", "pack", "packing"]),
    ("strip_conversion_factor", ["strip_conversion_factor", "scf", "conversion", "tabs_per_strip"]),
    ("rate_unit", ["rate_unit"]),
    ("decimal_supported", ["decimal_supported"]),
    ("requires_prescription", ["requires_prescription"]),
    ("is_narcotic", ["is_narcotic"]),
    ("is_high_alert", ["is_high_alert"]),
    ("is_schedule_h", ["is_schedule_h"]),
    ("is_schedule_h1", ["is_schedule_h1"]),
    ("is_tramadol", ["is_tramadol"]),
    ("is_controlled", ["is_controlled"]),
    ("is_active", ["is_active"]),
    ("is_hidden", ["is_hidden"]),
    ("min_qty", ["min_qty"]),
    ("max_qty", ["max_qty"]),
    ("reorder_qty", ["reorder_qty"]),
    ("description", ["description"]),
    ("side_effects", ["side_effects"]),
    ("contraindications", ["contraindications"]),
    ("storage_conditions", ["storage_conditions"]),
    ("manufacturer", ["manufacturer", "mfg"]),
    ("company", ["company", "company_name"]),
    ("rack_code", ["rack_code", "rack"]),
    ("salt", ["salt", "salt_name"]),
    ("uom", ["uom", "unit"]),
    ("hsn_code", ["hsn_code", "hsn"]),
    ("sgst_pct", ["sgst_pct", "sgst"]),
    ("cgst_pct", ["cgst_pct", "cgst"]),
]


def inspect_medicines_import(
    content: bytes, filename: str, *,
    row_start: Optional[int] = None,
    row_end: Optional[int] = None,
) -> dict:
    return inspect_letter_import(
        content, filename,
        row_start=row_start, row_end=row_end,
        required_fields=REQUIRED_MEDICINE_LETTER_FIELDS,
        aliases=MEDICINE_IMPORT_ALIASES,
    )


def _count_file_lines(content: bytes, filename: str) -> int:
    fn = (filename or "").lower()
    if fn.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        return sum(1 for _ in io.StringIO(text))
    if fn.endswith(".xlsx"):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            if not wb.sheetnames:
                return 0
            ws = wb[wb.sheetnames[0]]
            return sum(1 for _ in ws.iter_rows())
        except Exception:
            return 0
    return 0


def _looks_like_vendor_purchase(rows: List[dict]) -> bool:
    if not rows:
        return False
    keys = {k for k in rows[0].keys() if k != "_row"}
    if "cl1" in keys and "cl6" in keys:
        return True
    for row in rows[:5]:
        typ = (_strip_excel_formula(row.get("cl1")) or "").upper()
        if typ in ("H", "T", "F"):
            return True
    return False


def _colmap_get(row: dict, colmap: Dict[str, str], *fields: str, default_key: Optional[str] = None):
    """Read a cell via user/default column map. `fields` tried in order."""
    for f in fields:
        src = colmap.get(f)
        if src and src in row:
            return row.get(src)
    if default_key:
        src = _VENDOR_DEFAULT_COLS.get(default_key) or default_key
        return row.get(src)
    # Fall back to first field's default
    for f in fields:
        src = _VENDOR_DEFAULT_COLS.get(f)
        if src:
            return row.get(src)
    return None


def _parse_vendor_purchase_blocks(
    rows: List[dict], column_mapping: Optional[dict] = None,
) -> List[dict]:
    """Group H / T / F rows into one block per invoice.

    `column_mapping` is {source_header: target_field} from the UI.
    """
    field_cols = dict(_VENDOR_DEFAULT_COLS)
    # User overrides (target → source)
    user_cols = _normalize_column_mapping(column_mapping)
    field_cols.update(user_cols)
    # If user mapped purchase_rate onto CL11, don't also treat it as transporter default
    # unless they explicitly mapped transporter elsewhere.
    if "purchase_rate" in user_cols and user_cols.get("transporter") == user_cols.get("purchase_rate"):
        pass

    blocks: List[dict] = []
    current: Optional[dict] = None

    def _finish():
        nonlocal current
        if current is not None:
            blocks.append(current)
            current = None

    for row in rows:
        typ = (_strip_excel_formula(
            _colmap_get(row, field_cols, "record_type", default_key="record_type")
        ) or "").upper()
        rownum = row.get("_row", 0)
        if typ == "H":
            _finish()
            inv = _strip_excel_formula(
                _colmap_get(row, field_cols, "invoice_number", "supplier_or_invoice", default_key="supplier_or_invoice")
            )
            bill_date = None
            try:
                bill_date = _parse_ddmmyyyy(
                    _colmap_get(row, field_cols, "bill_date", "entry_date", default_key="bill_date")
                )
            except ValueError:
                bill_date = None
            # purchase_type: prefer dedicated mapping; else default CL7 on H
            ptype = _strip_excel_formula(
                _colmap_get(row, field_cols, "purchase_type", default_key="purchase_type")
            )
            if not ptype and not user_cols.get("purchase_type"):
                ptype = _strip_excel_formula(row.get("cl7"))
            transporter = _strip_excel_formula(
                _colmap_get(row, field_cols, "transporter")
            )
            if transporter is None and "transporter" not in user_cols and "purchase_rate" not in user_cols:
                transporter = _strip_excel_formula(row.get("cl11"))
            current = {
                "invoice_number": inv,
                "bill_date": bill_date,
                "entry_date": bill_date or date.today(),
                "purchase_type": ptype,
                "transporter": transporter,
                "place_of_supply": _strip_excel_formula(
                    _colmap_get(row, field_cols, "place_of_supply", default_key="place_of_supply")
                ),
                "supplier_name": _strip_excel_formula(
                    _colmap_get(row, field_cols, "supplier_name")
                ),
                "payment_type": (
                    _strip_excel_formula(_colmap_get(row, field_cols, "payment_type")) or "credit"
                ),
                "tax_mode": (
                    _strip_excel_formula(_colmap_get(row, field_cols, "tax_mode")) or "exclusive"
                ),
                "store_code": _strip_excel_formula(_colmap_get(row, field_cols, "store_code")),
                "notes": _strip_excel_formula(_colmap_get(row, field_cols, "notes")),
                "footer": None,
                "items": [],
                "_header_row": rownum,
            }
        elif typ == "T":
            if current is None:
                current = {
                    "invoice_number": _strip_excel_formula(
                        _colmap_get(row, field_cols, "invoice_number")
                    ),
                    "bill_date": None,
                    "entry_date": date.today(),
                    "purchase_type": None,
                    "transporter": None,
                    "place_of_supply": None,
                    "supplier_name": None,
                    "payment_type": "credit",
                    "tax_mode": "exclusive",
                    "store_code": None,
                    "notes": None,
                    "footer": None,
                    "items": [],
                    "_header_row": rownum,
                }
            supplier = _strip_excel_formula(
                _colmap_get(row, field_cols, "supplier_name", "supplier_or_invoice", default_key="supplier_or_invoice")
            )
            if supplier and not current.get("supplier_name"):
                current["supplier_name"] = supplier
            item_name = _strip_excel_formula(
                _colmap_get(row, field_cols, "medicine_name", default_key="medicine_name")
            )
            batch = _strip_excel_formula(
                _colmap_get(row, field_cols, "batch_number", default_key="batch_number")
            )
            if not item_name and not batch:
                continue
            expiry = None
            try:
                expiry = _parse_ddmmyyyy(
                    _colmap_get(row, field_cols, "expiry_date", default_key="expiry_date")
                )
            except ValueError:
                expiry = None
            qty_raw = _colmap_get(row, field_cols, "quantity", default_key="quantity")
            qty = _cell_float(_strip_excel_formula(qty_raw) or qty_raw)
            free_raw = _colmap_get(row, field_cols, "free_quantity", default_key="free_quantity")
            free = _cell_float(_strip_excel_formula(free_raw) or free_raw) or 0.0
            rate_raw = _colmap_get(row, field_cols, "purchase_rate", default_key="purchase_rate")
            purchase_rate = _cell_float(_strip_excel_formula(rate_raw) or rate_raw)
            mrp_raw = _colmap_get(row, field_cols, "mrp", default_key="mrp")
            mrp = _cell_float(_strip_excel_formula(mrp_raw) or mrp_raw)
            rate_a_raw = _colmap_get(row, field_cols, "rate_a")
            rate_b_raw = _colmap_get(row, field_cols, "rate_b")
            rate_a = _cell_float(_strip_excel_formula(rate_a_raw) or rate_a_raw) if rate_a_raw is not None else None
            rate_b = _cell_float(_strip_excel_formula(rate_b_raw) or rate_b_raw) if rate_b_raw is not None else None
            if rate_a is None:
                rate_a = mrp
            if rate_b is None:
                rate_b = mrp
            disc_raw = _colmap_get(row, field_cols, "discount_pct", default_key="discount_pct")
            disc_val = _cell_float(_strip_excel_formula(disc_raw) or disc_raw) or 0.0
            discount_pct = 0.0
            if disc_val and qty and purchase_rate:
                base = qty * purchase_rate
                src = field_cols.get("discount_pct") or ""
                # Vendor CL19 stores amount; named-template discount_pct is already a percent
                if src == "cl19" or disc_val > 100:
                    discount_pct = round((disc_val / base) * 100.0, 4) if base else 0.0
                else:
                    discount_pct = disc_val
            pack = _strip_excel_formula(
                _colmap_get(row, field_cols, "pack_size", default_key="pack_size")
            )
            scf_raw = _colmap_get(row, field_cols, "strip_conversion_factor")
            scf = _cell_int(_strip_excel_formula(scf_raw) or scf_raw) if scf_raw is not None else None
            if scf is None:
                scf = _parse_pack_scf(pack)
            cgst_raw = _colmap_get(row, field_cols, "cgst_pct", default_key="cgst_pct")
            sgst_raw = _colmap_get(row, field_cols, "sgst_pct", default_key="sgst_pct")
            current["items"].append({
                "_row": rownum,
                "medicine_code": _strip_excel_formula(
                    _colmap_get(row, field_cols, "medicine_code", default_key="medicine_code")
                ),
                "medicine_name": item_name,
                "pack_size": pack,
                "manufacturer": _strip_excel_formula(
                    _colmap_get(row, field_cols, "manufacturer", default_key="manufacturer")
                ),
                "batch_number": batch,
                "expiry_date": expiry,
                "quantity": qty,
                "free_quantity": free,
                "mrp": mrp,
                "purchase_rate": purchase_rate,
                "rate_a": rate_a,
                "rate_b": rate_b,
                "strip_conversion_factor": scf,
                "discount_pct": discount_pct,
                "hsn_code": _strip_excel_formula(
                    _colmap_get(row, field_cols, "hsn_code", default_key="hsn_code")
                ),
                "cgst_pct": _cell_float(_strip_excel_formula(cgst_raw) or cgst_raw),
                "sgst_pct": _cell_float(_strip_excel_formula(sgst_raw) or sgst_raw),
            })
        elif typ == "F":
            if current is None:
                continue
            taxable = _cell_float(_strip_excel_formula(
                _colmap_get(row, field_cols, "footer_taxable", default_key="footer_taxable")
            ) or _colmap_get(row, field_cols, "footer_taxable", default_key="footer_taxable"))
            cgst = _cell_float(_strip_excel_formula(
                _colmap_get(row, field_cols, "footer_cgst", default_key="footer_cgst")
            ) or _colmap_get(row, field_cols, "footer_cgst", default_key="footer_cgst"))
            sgst = _cell_float(_strip_excel_formula(
                _colmap_get(row, field_cols, "footer_sgst", default_key="footer_sgst")
            ) or _colmap_get(row, field_cols, "footer_sgst", default_key="footer_sgst"))
            round_off = _cell_float(_strip_excel_formula(
                _colmap_get(row, field_cols, "footer_round_off", default_key="footer_round_off")
            ) or _colmap_get(row, field_cols, "footer_round_off", default_key="footer_round_off"))
            invoice_value = _cell_float(_strip_excel_formula(
                _colmap_get(row, field_cols, "footer_invoice_value", default_key="footer_invoice_value")
            ) or _colmap_get(row, field_cols, "footer_invoice_value", default_key="footer_invoice_value"))
            current["footer"] = {
                "taxable": taxable,
                "cgst": cgst,
                "sgst": sgst,
                "round_off": round_off,
                "invoice_value": invoice_value,
            }
            note_bits = []
            if invoice_value is not None:
                note_bits.append(f"Invoice value ₹{invoice_value:.2f}")
            if taxable is not None:
                note_bits.append(f"taxable ₹{taxable:.2f}")
            if cgst is not None or sgst is not None:
                note_bits.append(f"CGST ₹{(cgst or 0):.2f} / SGST ₹{(sgst or 0):.2f}")
            if round_off:
                note_bits.append(f"round-off {round_off}")
            if current.get("transporter"):
                note_bits.append(f"transporter {current['transporter']}")
            if current.get("place_of_supply"):
                note_bits.append(f"place of supply {current['place_of_supply']}")
            if note_bits and not current.get("notes"):
                current["notes"] = "Imported from vendor CSV — " + "; ".join(note_bits)
            _finish()
        else:
            continue

    _finish()
    return blocks


def _group_named_purchase_rows(rows: List[dict]) -> List[dict]:
    """Group flat named-header rows by (supplier_name, invoice_number)."""
    groups: Dict[Tuple[str, str], dict] = {}
    order: List[Tuple[str, str]] = []
    for row in rows:
        supplier = _cell_str(row.get("supplier_name"))
        invoice = _cell_str(row.get("invoice_number"))
        med_code = _cell_str(row.get("medicine_code"))
        med_name = _cell_str(row.get("medicine_name") or row.get("item_name"))
        if _key_skipped(med_code) and _key_skipped(med_name):
            # Allow skip only when both empty; treat #-prefixed as skip
            if (med_code and med_code.startswith("#")) or (med_name and med_name.startswith("#")):
                continue
            if not med_code and not med_name:
                continue
        key = ((supplier or "").lower(), (invoice or "").lower())
        if key not in groups:
            bill_date = None
            entry_date = None
            try:
                bill_date = _cell_date(row.get("bill_date"))
            except ValueError:
                bill_date = None
            try:
                entry_date = _cell_date(row.get("entry_date"))
            except ValueError:
                entry_date = None
            groups[key] = {
                "invoice_number": invoice,
                "bill_date": bill_date,
                "entry_date": entry_date or bill_date or date.today(),
                "purchase_type": _cell_str(row.get("purchase_type")),
                "supplier_name": supplier,
                "payment_type": (_cell_str(row.get("payment_type")) or "credit").lower(),
                "tax_mode": (_cell_str(row.get("tax_mode")) or "exclusive").lower(),
                "store_code": _cell_str(row.get("store_code")),
                "notes": _cell_str(row.get("notes")),
                "footer": None,
                "items": [],
                "_header_row": row.get("_row", 0),
            }
            order.append(key)
        try:
            expiry = _cell_date(row.get("expiry_date"))
        except ValueError:
            try:
                expiry = _parse_ddmmyyyy(row.get("expiry_date"))
            except ValueError:
                expiry = None
        pack = _cell_str(row.get("pack_size"))
        scf = _cell_int(row.get("strip_conversion_factor"))
        if scf is None:
            scf = _parse_pack_scf(pack)
        groups[key]["items"].append({
            "_row": row.get("_row", 0),
            "medicine_code": med_code,
            "medicine_name": med_name,
            "pack_size": pack,
            "manufacturer": _cell_str(row.get("manufacturer")),
            "batch_number": _cell_str(row.get("batch_number")),
            "expiry_date": expiry,
            "quantity": _cell_float(row.get("quantity")),
            "free_quantity": _cell_float(row.get("free_quantity")) or 0.0,
            "mrp": _cell_float(row.get("mrp")),
            "purchase_rate": _cell_float(row.get("purchase_rate")),
            "rate_a": _cell_float(row.get("rate_a")),
            "rate_b": _cell_float(row.get("rate_b")),
            "strip_conversion_factor": scf,
            "discount_pct": _cell_float(row.get("discount_pct")) or 0.0,
            "hsn_code": _cell_str(row.get("hsn_code")),
            "cgst_pct": _cell_float(row.get("cgst_pct")),
            "sgst_pct": _cell_float(row.get("sgst_pct")),
        })
    return [groups[k] for k in order]


def _remap_flat_purchase_rows(rows: List[dict], column_mapping: Optional[dict]) -> List[dict]:
    """Rewrite row keys from source headers → ERP field names using the UI mapping."""
    field_cols = _normalize_column_mapping(column_mapping)
    if not field_cols:
        return rows
    remapped: List[dict] = []
    for row in rows:
        out: dict = {"_row": row.get("_row")}
        for field, src in field_cols.items():
            if field in ("record_type", "supplier_or_invoice"):
                continue
            out[field] = row.get(src)
        # supplier_or_invoice with no separate supplier → treat as supplier on flat rows
        if "supplier_name" not in out and "supplier_or_invoice" in field_cols:
            out["supplier_name"] = row.get(field_cols["supplier_or_invoice"])
        if "invoice_number" not in out and "supplier_or_invoice" in field_cols:
            # Flat files shouldn't use supplier_or_invoice for both — leave invoice empty
            pass
        mrp = out.get("mrp")
        if out.get("rate_a") in (None, "") and mrp not in (None, ""):
            out["rate_a"] = mrp
        if out.get("rate_b") in (None, "") and mrp not in (None, ""):
            out["rate_b"] = mrp
        remapped.append(out)
    return remapped


def _resolve_purchase_store_id(
    db: Session, hospital_id: int, store_code: Optional[str], store_cache: dict,
) -> Optional[int]:
    if store_code:
        return _resolve_store_id(db, hospital_id, store_code, store_cache)
    # Prefer a purchase-capable store, then master.
    purchase_store = db.query(PharmacyStore).filter(
        PharmacyStore.hospital_id == hospital_id,
        PharmacyStore.is_active == True,  # noqa: E712
        PharmacyStore.can_receive_supplier_purchase == True,  # noqa: E712
    ).order_by(PharmacyStore.is_default.desc(), PharmacyStore.id.asc()).first()
    if purchase_store:
        return purchase_store.id
    return get_master_store_id(db, hospital_id)


def _find_supplier_by_name(db: Session, hospital_id: int, name: str) -> Optional[PharmacySupplier]:
    return db.query(PharmacySupplier).filter(
        PharmacySupplier.hospital_id == hospital_id,
        sa_func.lower(PharmacySupplier.name) == name.strip().lower(),
        PharmacySupplier.is_active == True,  # noqa: E712
    ).first()


def upsert_medicine_import_alias(
    db: Session, hospital_id: int, medicine_id: int, alias: str,
) -> PharmacyMedicineImportAlias:
    name = (alias or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="alias is required")
    norm = name.lower()
    row = db.query(PharmacyMedicineImportAlias).filter(
        PharmacyMedicineImportAlias.hospital_id == hospital_id,
        PharmacyMedicineImportAlias.alias_norm == norm,
    ).first()
    if row:
        row.alias_name = name
        row.medicine_id = medicine_id
    else:
        row = PharmacyMedicineImportAlias(
            alias_name=name,
            alias_norm=norm,
            medicine_id=medicine_id,
            hospital_id=hospital_id,
        )
        db.add(row)
    db.flush()
    return row


def _apply_name_aliases_to_cache(
    db: Session, hospital_id: int, cache: dict, extra_aliases: Optional[dict],
) -> None:
    """Seed the medicine cache so vendor spellings resolve to catalog items."""
    pairs: List[Tuple[str, int]] = []
    if isinstance(extra_aliases, dict):
        for raw, mid in extra_aliases.items():
            name = str(raw or "").strip()
            try:
                mid_int = int(mid)
            except (TypeError, ValueError):
                continue
            if name:
                pairs.append((name, mid_int))
    for row in db.query(PharmacyMedicineImportAlias).filter(
        PharmacyMedicineImportAlias.hospital_id == hospital_id,
    ).all():
        if row.alias_name:
            pairs.append((row.alias_name, row.medicine_id))
    med_ids = {mid for _, mid in pairs}
    if not med_ids:
        return
    meds = {
        m.id: m
        for m in db.query(Medicine).filter(
            Medicine.hospital_id == hospital_id,
            Medicine.id.in_(med_ids),
            Medicine.is_active == True,  # noqa: E712
        ).all()
    }
    for name, mid in pairs:
        med = meds.get(mid)
        if not med:
            continue
        cache[f"name:{name.strip().lower()}"] = med


def _normalize_name_aliases(raw: Optional[dict]) -> Dict[str, int]:
    if not isinstance(raw, dict):
        return {}
    out: Dict[str, int] = {}
    for k, v in raw.items():
        name = str(k or "").strip()
        try:
            mid = int(v)
        except (TypeError, ValueError):
            continue
        if name and mid > 0:
            out[name] = mid
    return out


def _find_medicine_for_purchase(
    db: Session, hospital_id: int, *,
    medicine_code: Optional[str], medicine_name: Optional[str], pack_size: Optional[str],
    cache: dict,
) -> Tuple[Optional[Medicine], Optional[str]]:
    """Return (medicine, error_message). Match code first, then name (+ pack)."""
    if medicine_code:
        key = f"code:{medicine_code.lower()}"
        if key in cache:
            return cache[key], None
        med = db.query(Medicine).filter(
            Medicine.hospital_id == hospital_id,
            Medicine.medicine_code == medicine_code,
            Medicine.is_active == True,  # noqa: E712
        ).first()
        if med:
            cache[key] = med
            return med, None
        # Vendor item codes rarely match our medicine_code — fall through to name.

    if not medicine_name:
        return None, "Missing medicine_code / medicine_name"

    key = f"name:{medicine_name.lower()}"
    if key in cache and cache[key] is not None and not isinstance(cache[key], list):
        return cache[key], None

    matches = db.query(Medicine).filter(
        Medicine.hospital_id == hospital_id,
        sa_func.lower(Medicine.name) == medicine_name.strip().lower(),
        Medicine.is_active == True,  # noqa: E712
    ).all()
    if not matches:
        alias = db.query(PharmacyMedicineImportAlias).filter(
            PharmacyMedicineImportAlias.hospital_id == hospital_id,
            PharmacyMedicineImportAlias.alias_norm == medicine_name.strip().lower(),
        ).first()
        if alias:
            med = db.query(Medicine).filter(
                Medicine.id == alias.medicine_id,
                Medicine.hospital_id == hospital_id,
                Medicine.is_active == True,  # noqa: E712
            ).first()
            if med:
                cache[key] = med
                return med, None
        return None, f"Medicine '{medicine_name}' not found"
    if len(matches) == 1:
        cache[key] = matches[0]
        return matches[0], None

    if pack_size:
        pack_norm = pack_size.replace(" ", "").lower()
        narrowed = [
            m for m in matches
            if (m.packaging or "").replace(" ", "").lower() == pack_norm
            or pack_norm in (m.packaging or "").replace(" ", "").lower()
            or pack_norm in (m.strength or "").replace(" ", "").lower()
        ]
        if len(narrowed) == 1:
            cache[key] = narrowed[0]
            return narrowed[0], None

    codes = ", ".join(m.medicine_code for m in matches[:5])
    return None, (
        f"Ambiguous medicine name '{medicine_name}' "
        f"({len(matches)} matches: {codes}). Set medicine_code or packaging."
    )


def _allocate_purchase_medicine_code(
    db: Session, hospital_id: int, preferred: Optional[str], name: str,
) -> str:
    """Pick an unused medicine_code (≤20 chars). Prefer vendor code when free."""
    preferred = (preferred or "").strip()
    if preferred and len(preferred) <= 20:
        clash = db.query(Medicine.id).filter(
            Medicine.hospital_id == hospital_id,
            Medicine.medicine_code == preferred,
        ).first()
        if not clash:
            return preferred

    base = re.sub(r"[^A-Za-z0-9]", "", (name or "MED").upper())[:12] or "MED"
    candidate = base[:20]
    if not db.query(Medicine.id).filter(
        Medicine.hospital_id == hospital_id, Medicine.medicine_code == candidate,
    ).first():
        return candidate
    for n in range(2, 1000):
        suffix = str(n)
        cand = (base[: max(1, 20 - len(suffix))] + suffix)[:20]
        if not db.query(Medicine.id).filter(
            Medicine.hospital_id == hospital_id, Medicine.medicine_code == cand,
        ).first():
            return cand
    return f"M{int(datetime.now().timestamp()) % 10_000_000_000}"[:20]


def _get_or_create_medicine_for_purchase(
    db: Session, hospital_id: int, item: dict, *,
    resolver: _MasterResolver, cache: dict,
) -> Tuple[Optional[Medicine], bool, Optional[str]]:
    """Find medicine or auto-create from purchase line details.

    Returns (medicine, created, error_message).
    Auto-creates company / HSN / category "General" as needed.
    """
    med, err = _find_medicine_for_purchase(
        db, hospital_id,
        medicine_code=item.get("medicine_code"),
        medicine_name=item.get("medicine_name"),
        pack_size=item.get("pack_size"),
        cache=cache,
    )
    if not med and item.get("medicine_code") and item.get("medicine_name"):
        med, err = _find_medicine_for_purchase(
            db, hospital_id,
            medicine_code=None,
            medicine_name=item.get("medicine_name"),
            pack_size=item.get("pack_size"),
            cache=cache,
        )

    hsn_code = _cell_str(item.get("hsn_code"))
    cgst = item.get("cgst_pct")
    sgst = item.get("sgst_pct")
    if cgst is not None and not isinstance(cgst, (int, float)):
        cgst = _cell_float(cgst)
    if sgst is not None and not isinstance(sgst, (int, float)):
        sgst = _cell_float(sgst)

    if med:
        if hsn_code and not med.hsn_id:
            med.hsn_id = resolver.hsn(hsn_code, sgst, cgst).id
            db.flush()
        return med, False, None

    if err and "Ambiguous" in err:
        return None, False, err

    name = _cell_str(item.get("medicine_name"))
    if not name:
        return None, False, err or "Missing medicine_name — cannot auto-create"

    mrp = item.get("mrp") or 0.0
    purchase_rate = item.get("purchase_rate") or 0.0
    rate_a = item.get("rate_a") if item.get("rate_a") is not None else mrp
    rate_b = item.get("rate_b") if item.get("rate_b") is not None else mrp
    pack = _cell_str(item.get("pack_size"))
    scf = item.get("strip_conversion_factor") or _parse_pack_scf(pack) or 1
    scf = max(1, int(scf))
    manufacturer = _cell_str(item.get("manufacturer"))

    cat = resolver.category("General")
    code = _allocate_purchase_medicine_code(
        db, hospital_id, item.get("medicine_code"), name,
    )
    unit_price = float(rate_a or mrp or purchase_rate or 0)

    med = Medicine(
        medicine_code=code,
        name=name,
        category_id=cat.id,
        hospital_id=hospital_id,
        unit_price=unit_price,
        mrp=float(mrp or 0),
        purchase_rate=float(purchase_rate or 0),
        rate_a=float(rate_a or 0),
        rate_b=float(rate_b or 0),
        packaging=pack,
        strip_conversion_factor=scf,
        manufacturer=manufacturer,
        is_active=True,
        requires_prescription=True,
    )
    if manufacturer:
        med.company_id = resolver.company(manufacturer).id
    if hsn_code:
        med.hsn_id = resolver.hsn(hsn_code, sgst, cgst).id

    apply_medicine_price_rounding(med)
    apply_cost_pcs_from_mrp(med)
    db.add(med)
    db.flush()
    resolver._track("medicine", f"{code} ({name})")

    cache[f"code:{code.lower()}"] = med
    cache[f"name:{name.lower()}"] = med
    if item.get("medicine_code"):
        vendor_code = str(item["medicine_code"]).strip().lower()
        if vendor_code == code.lower():
            cache[f"code:{vendor_code}"] = med

    return med, True, None


def _next_purchase_number_import(db: Session, hospital_id: int) -> str:
    today = date.today()
    prefix = f"PURCH-{today.strftime('%y%m%d')}-"
    last = db.query(PharmacyPurchase).filter(
        PharmacyPurchase.purchase_number.like(prefix + "%"),
        PharmacyPurchase.hospital_id == hospital_id,
    ).order_by(PharmacyPurchase.purchase_number.desc()).first()
    seq = 1
    if last:
        try:
            seq = int(last.purchase_number.rsplit("-", 1)[-1]) + 1
        except Exception:
            seq = 1
    return f"{prefix}{seq:04d}"


def _recompute_purchase_totals_import(purchase: PharmacyPurchase, db: Session) -> None:
    subtotal = 0.0
    disc = 0.0
    tax = 0.0
    grand = 0.0
    tax_mode = getattr(purchase, "tax_mode", None) or "exclusive"
    for it in purchase.items:
        hsn_row = None
        if it.hsn_id:
            hsn_row = db.query(PharmacyHSN).filter(PharmacyHSN.id == it.hsn_id).first()
        elif it.medicine_id:
            med = db.query(Medicine).filter(Medicine.id == it.medicine_id).first()
            if med and med.hsn_id:
                it.hsn_id = med.hsn_id
                hsn_row = db.query(PharmacyHSN).filter(PharmacyHSN.id == med.hsn_id).first()
        qty = float(it.quantity or 0)
        rate = float(it.purchase_rate or 0)
        disc_pct = float(it.discount_pct or 0)
        base = qty * rate
        base_after = base * (1 - disc_pct / 100.0)
        tax_pct = 0.0
        if hsn_row:
            tax_pct = float(hsn_row.sgst_pct or 0) + float(hsn_row.cgst_pct or 0)
            if not tax_pct:
                tax_pct = float(hsn_row.igst_pct or 0)
        _taxable, tax_amt, line_total = compute_line_tax(base_after, tax_pct, tax_mode=tax_mode)
        it.tax_amount = tax_amt
        it.line_total = line_total
        it.sgst_pct = (hsn_row.sgst_pct or 0) if hsn_row else 0.0
        it.cgst_pct = (hsn_row.cgst_pct or 0) if hsn_row else 0.0
        it.igst_pct = (hsn_row.igst_pct if hsn_row else 0.0) or (it.sgst_pct + it.cgst_pct)
        subtotal += base
        disc += round(base - base_after, 2)
        tax += tax_amt
        grand += line_total
    pct = float(getattr(purchase, "bill_discount_pct", None) or 0)
    requested = float(getattr(purchase, "bill_discount_amount", None) or 0)
    if pct > 0:
        bill_disc = round(min(grand * pct / 100.0, grand), 2)
    else:
        bill_disc = round(min(max(requested, 0.0), grand), 2)
    purchase.bill_discount_amount = bill_disc
    purchase.subtotal = round(subtotal, 2)
    purchase.total_discount = round(disc + bill_disc, 2)
    purchase.total_tax = round(tax, 2)
    purchase.grand_total = round(grand - bill_disc, 2)


def _find_existing_purchase(
    db: Session, hospital_id: int, supplier_id: int, invoice_number: Optional[str],
) -> Optional[PharmacyPurchase]:
    inv = (invoice_number or "").strip()
    if not inv:
        return None
    return db.query(PharmacyPurchase).filter(
        PharmacyPurchase.hospital_id == hospital_id,
        PharmacyPurchase.supplier_id == supplier_id,
        PharmacyPurchase.invoice_number == inv,
    ).first()


def _filter_rows_by_line(
    rows: List[dict],
    row_start: Optional[int],
    row_end: Optional[int],
) -> List[dict]:
    """Keep rows whose `_row` (1-based file line) is within [row_start, row_end]."""
    if row_start is None and row_end is None:
        return rows
    out: List[dict] = []
    for r in rows:
        n = int(r.get("_row") or 0)
        if row_start is not None and n < int(row_start):
            continue
        if row_end is not None and n > int(row_end):
            continue
        out.append(r)
    return out


def _purchase_form_item(resolved: dict) -> dict:
    """Serialize a resolved import line into PurchaseEntry form shape."""
    med = resolved["medicine"]
    exp = resolved.get("expiry_date")
    if hasattr(exp, "isoformat"):
        exp = exp.isoformat()
    elif exp is not None:
        exp = str(exp)[:10]
    return {
        "medicine_id": med.id,
        "medicine_name": med.name or "",
        "medicine_code": med.medicine_code or "",
        "batch_number": resolved.get("batch_number") or "",
        "expiry_date": exp,
        "mrp": resolved.get("mrp") or 0,
        "quantity": resolved.get("quantity") or 0,
        "free_quantity": resolved.get("free_quantity") or 0,
        "purchase_rate": resolved.get("purchase_rate") or 0,
        "rate_a": resolved.get("rate_a") or 0,
        "rate_b": resolved.get("rate_b") or 0,
        "strip_conversion_factor": resolved.get("strip_conversion_factor") or 1,
        "discount_pct": resolved.get("discount_pct") or 0,
        "hsn_id": med.hsn_id,
    }


def _unmatched_catalog_entry(item: dict, name: str, rownum: int) -> dict:
    scf = item.get("strip_conversion_factor")
    try:
        scf_int = int(scf) if scf is not None else None
    except (TypeError, ValueError):
        scf_int = None
    return {
        "name": name,
        "medicine_code": _cell_str(item.get("medicine_code")) or None,
        "pack_size": _cell_str(item.get("pack_size")) or None,
        "manufacturer": _cell_str(item.get("manufacturer")) or None,
        "mrp": item.get("mrp"),
        "purchase_rate": item.get("purchase_rate"),
        "hsn_code": _cell_str(item.get("hsn_code")) or None,
        "strip_conversion_factor": scf_int,
        "row": rownum,
    }


def _merge_unmatched_medicines(existing: List, new_entries: List[dict]) -> List[dict]:
    out: List[dict] = []
    seen: set = set()
    for entry in list(existing or []) + list(new_entries or []):
        if isinstance(entry, str):
            entry = {"name": entry}
        name = (entry.get("name") or "").strip()
        key = name.lower()
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(entry)
    return out


def _iso_date(v) -> Optional[str]:
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()
    s = str(v).strip()
    return s[:10] if s else None


def import_purchases(
    db: Session, user: User, content: bytes, filename: str,
    *, dry_run: bool, on_duplicate: str,
    column_mapping: Optional[dict] = None,
    supplier_id: Optional[int] = None,
    invoice_number: Optional[str] = None,
    entry_date: Optional[date] = None,
    bill_date: Optional[date] = None,
    payment_type: Optional[str] = None,
    row_start: Optional[int] = None,
    row_end: Optional[int] = None,
    name_aliases: Optional[dict] = None,
) -> dict:
    """Parse a spreadsheet by Excel column letters and return a purchase form payload.

    Does **not** create a PharmacyPurchase. Medicines must already exist in the
    catalog — unmatched names are returned in `unmatched_medicines`.
    """
    summary = _empty_summary(dry_run=dry_run)
    start_row = int(row_start) if row_start is not None else 1
    if start_row < 1:
        summary["errors"].append({
            "sheet": "Purchases", "row": 0,
            "message": "row_start must be >= 1",
        })
        summary["error_count"] = 1
        return summary
    if row_end is not None and int(row_end) < start_row:
        summary["errors"].append({
            "sheet": "Purchases", "row": 0,
            "message": "row_end cannot be before row_start",
        })
        summary["error_count"] = 1
        return summary

    letter_map = _normalize_letter_mapping(column_mapping)
    missing_req = [f for f in REQUIRED_PURCHASE_LETTER_FIELDS if f not in letter_map]
    if missing_req:
        labels = ", ".join(missing_req)
        summary["errors"].append({
            "sheet": "Purchases", "row": 0,
            "message": f"Map required columns: {labels}",
        })
        summary["error_count"] = 1
        return summary

    try:
        grid = _parse_purchase_grid(
            content, filename, row_start=start_row, row_end=row_end,
        )
    except HTTPException as exc:
        summary["errors"].append({
            "sheet": "Purchases", "row": 0,
            "message": str(exc.detail),
        })
        summary["error_count"] = 1
        return summary

    hospital_id = user.hospital_id
    items = _line_items_from_letter_grid(grid, letter_map)
    if not items:
        summary["errors"].append({
            "sheet": "Purchases", "row": start_row,
            "message": "No purchase rows found in the selected range. Adjust start/end rows or column mapping.",
        })
        summary["error_count"] = 1
        return summary

    override_supplier = None
    if supplier_id is not None:
        override_supplier = db.query(PharmacySupplier).filter(
            PharmacySupplier.id == supplier_id,
            PharmacySupplier.hospital_id == hospital_id,
            PharmacySupplier.is_active == True,  # noqa: E712
        ).first()
        if not override_supplier:
            summary["errors"].append({
                "sheet": "Purchases", "row": 0,
                "message": f"Invalid supplier_id {supplier_id}",
            })
            summary["error_count"] = 1
            return summary

    blocks = [{
        "invoice_number": (invoice_number or "").strip() or None,
        "entry_date": entry_date,
        "bill_date": bill_date or entry_date,
        "payment_type": payment_type or "credit",
        "purchase_type": "local",
        "tax_mode": "exclusive",
        "notes": None,
        "supplier_name": override_supplier.name if override_supplier else None,
        "_supplier_id": override_supplier.id if override_supplier else None,
        "items": items,
        "_header_row": items[0].get("_row") if items else start_row,
    }]

    # Apply header overrides from the wizard (take precedence over file values)
    inv_override = (invoice_number or "").strip() or None
    for block in blocks:
        if override_supplier is not None:
            block["supplier_name"] = override_supplier.name
            block["_supplier_id"] = override_supplier.id
        if inv_override is not None:
            block["invoice_number"] = inv_override
        if entry_date is not None:
            block["entry_date"] = entry_date
        if bill_date is not None:
            block["bill_date"] = bill_date
        elif entry_date is not None and not block.get("bill_date"):
            block["bill_date"] = entry_date
        if payment_type in ("cash", "credit"):
            block["payment_type"] = payment_type

    if not blocks:
        summary["errors"].append({
            "sheet": "Purchases", "row": 0,
            "message": "No purchase rows found in the selected range. Adjust start/end rows or column mapping.",
        })
        summary["error_count"] = 1
        return summary

    medicine_cache: dict = {}
    _apply_name_aliases_to_cache(
        db, hospital_id, medicine_cache, _normalize_name_aliases(name_aliases),
    )
    supplier_cache: Dict[str, PharmacySupplier] = {}
    form_header: Optional[dict] = None
    form_items: List[dict] = []
    form_warnings: List[str] = []

    for block in blocks:
        items = block.get("items") or []
        if not items:
            continue

        # Count each line toward total_rows for preview parity with other imports
        for _ in items:
            summary["total_rows"] += 1

        supplier_name = block.get("supplier_name")
        invoice_number_b = block.get("invoice_number")
        header_row = block.get("_header_row") or (items[0].get("_row") if items else 0)
        preview_key = invoice_number_b or f"row-{header_row}"

        if block.get("_supplier_id"):
            supplier = override_supplier
        else:
            if not supplier_name:
                msg = "Missing supplier — select a supplier in the import wizard"
                summary["errors"].append({"sheet": "Purchases", "row": header_row, "message": msg})
                summary["preview"].append({
                    "row": header_row, "key": preview_key, "name": "",
                    "status": "error", "message": msg, "sheet": "Purchases",
                })
                continue

            sup_key = supplier_name.lower()
            supplier = supplier_cache.get(sup_key)
            if not supplier:
                supplier = _find_supplier_by_name(db, hospital_id, supplier_name)
                if supplier:
                    supplier_cache[sup_key] = supplier
            if not supplier:
                msg = f"Supplier '{supplier_name}' not found — import suppliers first"
                summary["errors"].append({"sheet": "Purchases", "row": header_row, "message": msg})
                summary["preview"].append({
                    "row": header_row, "key": preview_key, "name": supplier_name,
                    "status": "error", "message": msg, "sheet": "Purchases",
                })
                continue

        payment_type = (block.get("payment_type") or "credit").lower()
        if payment_type not in ("cash", "credit"):
            payment_type = "credit"
        tax_mode = (block.get("tax_mode") or "exclusive").lower()
        if tax_mode not in ("exclusive", "inclusive"):
            tax_mode = "exclusive"

        existing = _find_existing_purchase(db, hospital_id, supplier.id, invoice_number_b)
        if existing:
            form_warnings.append(
                f"Invoice #{invoice_number_b} already exists as {existing.purchase_number} "
                f"({existing.status}). Change the invoice number before saving, or edit that purchase instead."
            )

        # Resolve all line items first; fail the whole invoice if any line errors
        resolved: List[dict] = []
        line_errors: List[str] = []
        unmatched_entries: List[dict] = []
        seen_unmatched: set = set()

        # Pass 1 — structural validation (no catalog writes)
        for item in items:
            rownum = item.get("_row") or header_row
            batch = (item.get("batch_number") or "").strip()
            qty = item.get("quantity")
            expiry = item.get("expiry_date")
            rate = item.get("purchase_rate")
            errs: List[str] = []
            if not batch:
                errs.append("Missing batch_number")
            if not expiry:
                errs.append("Missing or invalid expiry_date")
            if qty is None or qty <= 0:
                errs.append("quantity must be > 0")
            if rate is None or rate < 0:
                errs.append("Missing purchase_rate (PTR)")
            if item.get("mrp") is None or item.get("mrp") < 0:
                errs.append("Missing MRP")
            if not _cell_str(item.get("medicine_name")) and not _cell_str(item.get("medicine_code")):
                errs.append("Missing medicine_name / medicine_code")
            if errs:
                msg = f"Line {rownum}: " + "; ".join(errs)
                line_errors.append(msg)
                summary["errors"].append({"sheet": "Purchases", "row": rownum, "message": msg})
                summary["preview"].append({
                    "row": rownum,
                    "key": item.get("medicine_code") or item.get("medicine_name") or "",
                    "name": batch or "",
                    "status": "error", "message": msg, "sheet": "Purchases",
                })

        # Pass 2 — match existing catalog medicines only (never auto-create)
        for item in items:
            rownum = item.get("_row") or header_row
            batch = (item.get("batch_number") or "").strip()
            qty = item.get("quantity")
            expiry = item.get("expiry_date")
            rate = item.get("purchase_rate")
            name = _cell_str(item.get("medicine_name")) or _cell_str(item.get("medicine_code")) or ""

            med, med_err = _find_medicine_for_purchase(
                db, hospital_id,
                medicine_code=item.get("medicine_code"),
                medicine_name=item.get("medicine_name"),
                pack_size=item.get("pack_size"),
                cache=medicine_cache,
            )
            if not med and item.get("medicine_code") and item.get("medicine_name"):
                med, med_err = _find_medicine_for_purchase(
                    db, hospital_id,
                    medicine_code=None,
                    medicine_name=item.get("medicine_name"),
                    pack_size=item.get("pack_size"),
                    cache=medicine_cache,
                )
            if not med:
                msg = f"Line {rownum}: {med_err or 'Medicine not found in catalog'}"
                line_errors.append(msg)
                err_l = (med_err or "").lower()
                is_missing = "not found" in err_l or "missing medicine" in err_l or not med_err
                is_ambiguous = "ambiguous" in err_l
                if is_missing and not is_ambiguous and name:
                    key = name.strip().lower()
                    if key not in seen_unmatched:
                        seen_unmatched.add(key)
                        unmatched_entries.append(_unmatched_catalog_entry(item, name, rownum))
                summary["errors"].append({"sheet": "Purchases", "row": rownum, "message": msg})
                summary["preview"].append({
                    "row": rownum,
                    "key": item.get("medicine_code") or name,
                    "name": name,
                    "status": "error",
                    "message": med_err or "Not in catalog",
                    "sheet": "Purchases",
                })
                continue

            scf = item.get("strip_conversion_factor") or med.strip_conversion_factor or 1
            scf = max(1, int(scf))
            resolved.append({
                "row": rownum,
                "medicine": med,
                "batch_number": batch,
                "expiry_date": expiry,
                "quantity": float(qty or 0),
                "free_quantity": float(item.get("free_quantity") or 0),
                "mrp": round_money(item.get("mrp") or med.mrp or 0),
                "purchase_rate": round_money(rate or 0),
                "rate_a": round_money(
                    item.get("rate_a") if item.get("rate_a") is not None else (med.rate_a or 0)
                ),
                "rate_b": round_money(
                    item.get("rate_b") if item.get("rate_b") is not None else (med.rate_b or 0)
                ),
                "strip_conversion_factor": scf,
                "discount_pct": float(item.get("discount_pct") or 0),
            })

        if unmatched_entries:
            summary["unmatched_medicines"] = _merge_unmatched_medicines(
                summary.get("unmatched_medicines") or [], unmatched_entries,
            )

        if line_errors or unmatched_entries:
            continue
        if not resolved:
            continue

        entry = block.get("entry_date") or date.today()
        bill = block.get("bill_date") or entry
        if form_header is None:
            form_header = {
                "supplier_id": supplier.id,
                "invoice_number": invoice_number_b or "",
                "entry_date": _iso_date(entry),
                "bill_date": _iso_date(bill),
                "payment_type": payment_type,
                "purchase_type": block.get("purchase_type") or "local",
                "tax_mode": tax_mode,
                "notes": block.get("notes") or None,
            }
        elif block.get("notes") and not form_header.get("notes"):
            form_header["notes"] = block.get("notes")

        for r in resolved:
            form_items.append(_purchase_form_item(r))
            summary["preview"].append({
                "row": r["row"],
                "key": r["medicine"].medicine_code or "",
                "name": r["medicine"].name,
                "status": "new",
                "message": r["batch_number"],
                "sheet": "Purchases",
            })
        summary["created"] += len(resolved)
        summary["preview"].append({
            "row": header_row,
            "key": preview_key,
            "name": f"{supplier.name} — {len(resolved)} items",
            "status": "new",
            "message": f"{len(resolved)} line(s) matched catalog medicines",
            "sheet": "Purchases",
        })

    if form_header and form_items and not summary.get("unmatched_medicines"):
        summary["form"] = {
            "header": form_header,
            "items": form_items,
            "warnings": form_warnings,
        }

    summary["masters_created"] = []
    summary["error_count"] = len(summary["errors"])
    return summary


def build_purchases_template() -> bytes:
    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        ws.title = "Purchases"
        ws.append(PURCHASE_HEADERS)
        ws.append([
            "VASU PHARMA", "2026-27/TAX/1808", "2026-08-01", "2026-08-01",
            "credit", "Direct", "exclusive", "", "Sample import row",
            "AF200", "AF-200", "HAH022603", "2028-01-31",
            20, 0, 77.20, 58.82, 77.20, 77.20, 1, 0, "5*1", "30049029",
        ])
        _append_instructions(wb, [
            "KT HEALTH ERP — Purchase Import",
            "",
            "Loads lines into the New Purchase form (Save / Submit there).",
            "Supplier must already exist (chosen in the import wizard).",
            "Map Excel columns by letter (A, B, C…). Start row is the first data row.",
            "Required columns: medicine name, batch, quantity, PTR, MRP, discount, expiry.",
            "Every medicine name must already exist in the catalog — unmatched items block import.",
            "",
            "Option A — Named template (this sheet):",
            "  Required: supplier_name, medicine_code OR medicine_name, batch_number,",
            "  expiry_date (YYYY-MM-DD), quantity, purchase_rate",
            "  Rows with the same supplier_name + invoice_number become one purchase.",
            "",
            "Option B — Vendor distributor CSV (e.g. Vasu Pharma tax-invoice export):",
            "  Upload the H/T/F CSV as-is (CL1..CL31). Header (H) has invoice # + date;",
            "  detail (T) rows map: Item Name→medicine, Pack Size, Batch, Expiry,",
            "  QTY, Package(free), MRP→mrp/rate_a/rate_b, PTR→purchase_rate, HSN.",
            "  Footer (F) invoice value is stored in purchase notes for reconciliation.",
            "",
            "Tax % is taken from each medicine's HSN master (not from the CSV GST columns).",
            "on_duplicate=skip leaves existing invoices alone; update replaces draft lines only.",
        ])

    return _workbook_bytes(build)
